"""Assignments router — competency-based work assignment API (GAP-W09)."""

from datetime import date

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session

from api.database.connection import get_db
from api.dependencies.auth import get_current_user, require_role
from api.services import assignment_service

router = APIRouter(prefix="/assignments", tags=["assignments"], dependencies=[Depends(get_current_user)])


# D1 Tanda D (David 2026-04-28, Magda transcript): carga masiva de Team via Excel.
# Pedido por Jorge para arranque Goldfields formal — no manual user-by-user.
@router.post("/import-team-excel")
async def import_team_excel(
    file: UploadFile = File(...),
    plant_id: str = "OCP-JFC1",
    user=Depends(require_role("admin", "manager")),
    db: Session = Depends(get_db),
):
    """Import workforce from Excel.
    Columns expected: name, specialty, shift, shift_pattern, shift_cycle_start, skills, certifications.
    Returns: { created, skipped, errors[], total }.
    """
    import io as _io
    import openpyxl
    from api.database.models import WorkforceModel

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Archivo debe ser .xlsx")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")

    # Header row
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip().lower())
        else:
            headers.append("")
    required = {"name", "specialty"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(status_code=400, detail=f"Faltan columnas obligatorias: {', '.join(missing)}")

    def col(row, name):
        try:
            idx = headers.index(name)
            v = row[idx].value
            return str(v).strip() if v is not None else ""
        except ValueError:
            return ""

    created = 0
    skipped = 0
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if not any(cell.value for cell in row):
            continue
        name = col(row, "name")
        if not name:
            errors.append(f"fila {row_idx}: sin nombre")
            continue
        # Skip if exists by name + plant
        existing = db.query(WorkforceModel).filter(
            WorkforceModel.plant_id == plant_id,
            WorkforceModel.name == name,
        ).first()
        if existing:
            skipped += 1
            continue
        skills_str = col(row, "skills")
        certs_str = col(row, "certifications")
        worker = WorkforceModel(
            plant_id=plant_id,
            name=name,
            specialty=col(row, "specialty") or "OTRO",
            shift=col(row, "shift") or "day",
            shift_pattern=col(row, "shift_pattern") or "5x2",
            shift_cycle_start=col(row, "shift_cycle_start") or None,
            skills=[s.strip() for s in skills_str.split(",") if s.strip()] if skills_str else [],
            certifications=[c.strip() for c in certs_str.split(",") if c.strip()] if certs_str else [],
            available=True,
        )
        db.add(worker)
        created += 1
    db.commit()
    return {
        "ok": True,
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "total": created + skipped + len(errors),
    }


@router.get("/technicians")
def list_technicians(
    plant_id: str | None = None,
    shift: str | None = None,
    specialty: str | None = None,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List technician profiles with optional filters.
    C7 Tanda C: si el usuario tiene scoped_specialty (supervisor mec/elec/inst),
    se fuerza el filtro a esa especialidad — supervisores no ven equipos de
    otras disciplinas. Admin/manager/planner pasan sin restricción.
    """
    scoped = getattr(user, "scoped_specialty", None)
    if scoped:
        # Si el llamador pidió otra specialty, la sobrescribimos al scope.
        specialty = scoped
    return assignment_service.get_technician_profiles(
        db, plant_id=plant_id, shift=shift, specialty=specialty,
    )


@router.post("/rank-for-operation")
def rank_for_operation(
    data: dict,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """SF-568 — Smart Assignment IA: ranking de técnicos para una operación.

    Body: {
      plant_id, specialty (op specialty), shift?, planned_hours,
      target_date (ISO), exclude_worker_ids? (list)
    }

    Score (0-100):
      +40 si specialty match exacto en su perfil
      +30 si specialty match en sus skills (hidráulica, mecánica, neumática, eléctrica)
      +20 si shift coincide con el turno objetivo
      +10 * (HH disponibles / 8) hasta 10 pts
      -10 si está en absence_reason

    Devuelve top 10 técnicos ordenados con score y breakdown.
    """
    plant_id = data.get("plant_id")
    op_spec = (data.get("specialty") or "").strip()
    shift = (data.get("shift") or "").strip().lower()
    planned_hours = float(data.get("planned_hours", 0) or 0)
    exclude_ids = set(data.get("exclude_worker_ids") or [])

    techs = assignment_service.get_technician_profiles(db, plant_id=plant_id)
    op_norm = op_spec.lower()
    ranked = []
    for tn in techs:
        wid = tn.get("id") or tn.get("worker_id")
        if not wid or wid in exclude_ids: continue
        if not tn.get("available", True):
            continue
        score = 0.0
        breakdown = {}
        tn_spec = (tn.get("specialty") or "").lower()
        tn_skills = [str(s).lower() for s in (tn.get("skills") or [])]
        # Spec match (substring tolerante: MEC ~ Mecánico ~ PMEC)
        if op_norm and (op_norm in tn_spec or tn_spec[:3] == op_norm[:3]):
            score += 40; breakdown["specialty_match"] = 40
        # Skills match
        if op_norm and any(op_norm in sk or sk in op_norm for sk in tn_skills):
            score += 30; breakdown["skill_match"] = 30
        # Shift match
        tn_shift = (tn.get("shift") or "").lower()
        if shift and tn_shift and (shift == tn_shift or shift in tn_shift):
            score += 20; breakdown["shift_match"] = 20
        # HH disponibles (cap a 8h/turno)
        hh_avail = float(tn.get("hh_available", tn.get("hours_available", 8)) or 8)
        if planned_hours <= hh_avail:
            hh_pts = min(10.0, hh_avail)
            score += hh_pts; breakdown["hh_available_pts"] = hh_pts
        else:
            score -= 5; breakdown["hh_overload_penalty"] = -5
        if tn.get("absence_reason"):
            score -= 10; breakdown["absence_penalty"] = -10
        ranked.append({
            "worker_id": wid,
            "name": tn.get("name"),
            "specialty": tn.get("specialty"),
            "shift": tn.get("shift"),
            "skills": tn.get("skills") or [],
            "hh_available": hh_avail,
            "score": round(score, 1),
            "breakdown": breakdown,
        })
    ranked.sort(key=lambda x: x["score"], reverse=True)
    return {"plant_id": plant_id, "specialty": op_spec, "shift": shift, "candidates": ranked[:10]}


@router.post("/optimize")
def optimize_assignments(data: dict, db: Session = Depends(get_db)):
    """Generate optimized technician-to-task assignments.

    Body: { tasks, plant_id, date (ISO), shift, shift_hours? }
    """
    try:
        return assignment_service.optimize_assignments(
            db,
            tasks=data.get("tasks", []),
            plant_id=data["plant_id"],
            target_date=date.fromisoformat(data["date"]),
            target_shift=data["shift"],
            shift_hours=data.get("shift_hours", 8.0),
        )
    except KeyError as e:
        raise HTTPException(status_code=422, detail=f"Missing required field: {e}")


@router.post("/reoptimize")
def reoptimize_assignments(data: dict, db: Session = Depends(get_db)):
    """Re-optimize assignments when workers are absent.

    Body: { existing_assignments, absent_worker_ids, tasks, plant_id, date, shift }
    """
    try:
        return assignment_service.reoptimize_assignments(
            db,
            existing_assignments=data.get("existing_assignments", []),
            absent_worker_ids=data.get("absent_worker_ids", []),
            tasks=data.get("tasks", []),
            plant_id=data["plant_id"],
            target_date=date.fromisoformat(data["date"]),
            target_shift=data["shift"],
            shift_hours=data.get("shift_hours", 8.0),
        )
    except KeyError as e:
        raise HTTPException(status_code=422, detail=f"Missing required field: {e}")


@router.post("/summary")
def get_summary(data: dict, db: Session = Depends(get_db)):
    """Generate supervisor-friendly summary from AssignmentSummary data."""
    return assignment_service.get_assignment_summary(db, data)


# ── Fase 3 Jorge 2026-04-21 — edit técnico (especialidad, turno, pattern, skills)
@router.patch("/technicians/{worker_id}")
def update_technician(
    worker_id: str,
    data: dict,
    db: Session = Depends(get_db),
):
    """Update a technician's profile fields. Safe partial update — only
    fields present in the body are touched."""
    from api.database.models import WorkforceModel
    w = db.query(WorkforceModel).filter(WorkforceModel.worker_id == worker_id).first()
    if not w:
        raise HTTPException(status_code=404, detail="Technician not found")
    ALLOWED = {
        "name", "specialty", "shift", "shift_pattern", "shift_cycle_start",
        "skills", "certifications", "available", "absence_reason",
        "absence_until", "competency_level", "years_experience",
        "equipment_expertise", "safety_training_current", "competencies",
    }
    for k, v in data.items():
        if k in ALLOWED:
            setattr(w, k, v)
    db.commit()
    db.refresh(w)
    return {
        "worker_id": w.worker_id,
        "name": w.name,
        "specialty": w.specialty,
        "shift": w.shift,
        "shift_pattern": getattr(w, "shift_pattern", None),
        "shift_cycle_start": getattr(w, "shift_cycle_start", None),
        "skills": getattr(w, "skills", None) or [],
        "certifications": w.certifications or [],
        "available": w.available,
        "years_experience": w.years_experience,
    }
