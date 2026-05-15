"""Scheduling router — weekly program management, Gantt, HH balance, materials."""

from pydantic import BaseModel
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from api.database.connection import get_db
from api.dependencies.auth import get_current_user, require_role
from api.schemas import ProgramCreate
from api.services import scheduling_service

router = APIRouter(prefix="/scheduling", tags=["scheduling"], dependencies=[Depends(get_current_user)])


# ── Jorge F (transcript 2026-05-14): Disponibilidad por equipo/día/semana ──
@router.get("/availability")
def get_equipment_availability(
    plant_id: str = "OCP-JFC1",
    week_start: str | None = None,
    weeks: int = 1,
    db: Session = Depends(get_db),
):
    """Calcula disponibilidad por equipo, por día y por semana.

    Disponibilidad = 100% - (horas detenido / horas calendario).
    Detención = suma de planned_end - planned_start de OTs PROGRAMADO/
    EN_EJECUCION/EJECUTADO que afectan al equipo en el día.

    Equipos sin OTs en el día → disponibilidad = 100% - 0.5% baseline
    (factor falla previsto que Jorge mencionó en el transcript).

    Devuelve estructura tipo carta-Gantt:
    {
      "week_start": "2026-05-19",
      "days": ["Lun 19", "Mar 20", ...],
      "equipment": [
        {
          "tag": "BRY-SAG-ML-001",
          "name": "SAG Mill #1",
          "daily": [{"date": "2026-05-19", "downtime_h": 0, "available_pct": 99.5}, ...],
          "week_available_pct": 98.2,
          "total_downtime_h": 12,
        },
        ...
      ],
      "summary": {"equipment_count": 25, "week_available_avg": 96.8}
    }
    """
    from api.database.models import (
        ManagedWorkOrderModel,
        HierarchyNodeModel,
    )

    HOURS_PER_DAY = 24
    BASELINE_FAILURE_PCT = 0.5  # Jorge: ~0.5% de falla previsto

    # Parse week_start o tomar lunes de esta semana
    if week_start:
        try:
            start = datetime.fromisoformat(week_start)
        except (ValueError, TypeError):
            start = datetime.now()
    else:
        now = datetime.now()
        start = now - timedelta(days=now.weekday())
    start = start.replace(hour=0, minute=0, second=0, microsecond=0)
    end = start + timedelta(days=weeks * 7)

    # Equipos del plant (solo nodos EQUIPMENT)
    equipments = (
        db.query(HierarchyNodeModel)
        .filter(HierarchyNodeModel.plant_id == plant_id)
        .filter(HierarchyNodeModel.node_type == "EQUIPMENT")
        .all()
    )

    # OTs en el rango
    wos = (
        db.query(ManagedWorkOrderModel)
        .filter(ManagedWorkOrderModel.plant_id == plant_id)
        .filter(ManagedWorkOrderModel.status.in_(
            ["PROGRAMADO", "EN_EJECUCION", "EJECUTADO", "CERRADO",
             "SCHEDULED", "IN_PROGRESS", "COMPLETED"]
        ))
        .filter(ManagedWorkOrderModel.planned_start.isnot(None))
        .filter(ManagedWorkOrderModel.planned_end.isnot(None))
        .filter(ManagedWorkOrderModel.planned_start < end)
        .filter(ManagedWorkOrderModel.planned_end >= start)
        .all()
    )

    # Indexar OTs por equipment_tag y por día
    days = [(start + timedelta(days=i)) for i in range(weeks * 7)]
    day_keys = [d.strftime("%Y-%m-%d") for d in days]
    day_labels = [d.strftime("%a %d") for d in days]

    by_tag_day: dict[str, dict[str, float]] = {}
    for wo in wos:
        tag = wo.equipment_tag or ""
        if not tag:
            continue
        # Distribuir horas de OT por día
        ps = wo.planned_start
        pe = wo.planned_end
        if ps >= pe:
            continue
        # Para cada día en el rango, calcular intersección
        for d in days:
            d_start = d
            d_end = d + timedelta(days=1)
            overlap_start = max(ps, d_start)
            overlap_end = min(pe, d_end)
            if overlap_end > overlap_start:
                hours = (overlap_end - overlap_start).total_seconds() / 3600
                by_tag_day.setdefault(tag, {}).setdefault(
                    d.strftime("%Y-%m-%d"), 0.0
                )
                by_tag_day[tag][d.strftime("%Y-%m-%d")] += hours

    # Armar payload por equipo
    items = []
    total_avail_sum = 0.0
    eq_count = 0
    for eq in equipments:
        tag = eq.tag or eq.code or ""
        if not tag:
            continue
        daily = []
        total_dt = 0.0
        for d, dkey in zip(days, day_keys):
            dt_h = by_tag_day.get(tag, {}).get(dkey, 0.0)
            total_dt += dt_h
            avail_pct = max(0.0, 100.0 - BASELINE_FAILURE_PCT - (dt_h / HOURS_PER_DAY) * 100)
            daily.append({
                "date": dkey,
                "downtime_h": round(dt_h, 2),
                "available_pct": round(avail_pct, 1),
            })
        period_hours = HOURS_PER_DAY * weeks * 7
        week_avail = max(0.0, 100.0 - BASELINE_FAILURE_PCT - (total_dt / period_hours) * 100)
        items.append({
            "tag": tag,
            "name": eq.name,
            "criticality": eq.criticality or "C",
            "daily": daily,
            "week_available_pct": round(week_avail, 1),
            "total_downtime_h": round(total_dt, 2),
        })
        total_avail_sum += week_avail
        eq_count += 1

    # Ordenar por disponibilidad ascendente (los más críticos primero)
    items.sort(key=lambda x: x["week_available_pct"])

    return {
        "plant_id": plant_id,
        "week_start": start.strftime("%Y-%m-%d"),
        "weeks": weeks,
        "days": day_keys,
        "day_labels": day_labels,
        "equipment": items,
        "summary": {
            "equipment_count": eq_count,
            "week_available_avg": round(total_avail_sum / eq_count, 1) if eq_count > 0 else 0.0,
            "total_wos_in_period": len(wos),
            "baseline_failure_pct": BASELINE_FAILURE_PCT,
        },
    }


@router.post("/programs")
def create_program(data: ProgramCreate, db: Session = Depends(get_db)):
    return scheduling_service.create_program(db, data.plant_id, data.week_number, data.year)


@router.get("/programs")
def list_programs(
    plant_id: str | None = None,
    status: str | None = None,
    db: Session = Depends(get_db),
):
    return scheduling_service.list_programs(db, plant_id, status)


@router.get("/programs/{program_id}")
def get_program(program_id: str, db: Session = Depends(get_db)):
    model = scheduling_service.get_program(db, program_id)
    if not model:
        raise HTTPException(status_code=404, detail="Program not found")
    return {
        "program_id": model.program_id,
        "plant_id": model.plant_id,
        "week_number": model.week_number,
        "year": model.year,
        "status": model.status,
        "total_hours": model.total_hours,
        "work_packages": model.work_packages,
        "resource_slots": model.resource_slots,
        "conflicts": model.conflicts,
        "support_tasks": model.support_tasks,
        "created_at": model.created_at.isoformat() if model.created_at else None,
        "finalized_at": model.finalized_at.isoformat() if model.finalized_at else None,
        "published_at": model.published_at.isoformat() if model.published_at else None,
        "published_by": model.published_by,
        "material_status": model.material_status,
        "hh_balance": model.hh_balance,
    }


@router.put("/programs/{program_id}/finalize")
def finalize_program(program_id: str, user=Depends(require_role("admin", "manager", "planner")), db: Session = Depends(get_db)):
    result = scheduling_service.finalize_program(db, program_id)
    if not result:
        raise HTTPException(status_code=404, detail="Program not found")
    return result


@router.put("/programs/{program_id}/activate")
def activate_program(program_id: str, user=Depends(require_role("admin", "manager", "planner")), db: Session = Depends(get_db)):
    result = scheduling_service.activate_program(db, program_id)
    if not result:
        raise HTTPException(status_code=404, detail="Program not found")
    return result


@router.put("/programs/{program_id}/complete")
def complete_program(program_id: str, user=Depends(require_role("admin", "manager", "planner")), db: Session = Depends(get_db)):
    result = scheduling_service.complete_program(db, program_id)
    if not result:
        raise HTTPException(status_code=404, detail="Program not found")
    return result


# ── Phase 3: Publish program ─────────────────────────────────────────

@router.put("/programs/{program_id}/publish")
def publish_program(
    program_id: str,
    user=Depends(require_role("admin", "manager", "planner")),
    db: Session = Depends(get_db),
):
    """Publish a finalized program — makes it visible to all and locks edits."""
    result = scheduling_service.publish_program(db, program_id, getattr(user, "user_id", ""))
    if not result:
        raise HTTPException(status_code=400, detail="Cannot publish — program not found or not finalized/active")
    return result


# ── Phase 3: Material check ──────────────────────────────────────────

@router.get("/programs/{program_id}/material-check")
def material_check(program_id: str, db: Session = Depends(get_db)):
    """Check material availability for all WOs/work packages in a program."""
    result = scheduling_service.check_materials(db, program_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Program not found")
    return result


# ── Phase 3: HH balance ─────────────────────────────────────────────

@router.get("/programs/{program_id}/hh-balance")
def hh_balance(program_id: str, db: Session = Depends(get_db)):
    """Calculate manpower hours balance: capacity vs assigned by specialty."""
    result = scheduling_service.hh_balance(db, program_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Program not found")
    return result


@router.get("/hh-balance-live")
def hh_balance_live(plant_id: str = "OCP-JFC1", week_start: str = "", db: Session = Depends(get_db)):
    """HH balance from actual scheduled WOs (not from program backlog).
    Optional week_start (YYYY-MM-DD) returns also per-day HH for the 7-day window.
    """
    return scheduling_service.hh_balance_from_wos(db, plant_id, week_start=week_start or None)


# SF-656 (jornada VSC 2026-05-08) — auditoría visual: sobrecapacidad técnico
# + violaciones día/noche en la semana en curso. Endpoint determinista, sin LLM.
@router.get("/audit-capacity")
def audit_capacity(
    plant_id: str | None = None,
    week_start: str | None = None,
    hours_per_week: float = 40,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Detecta dos clases de violación en la semana ventana [week_start, +6d]:
       (a) técnico con HH asignadas > hours_per_week (overcapacity)
       (b) técnico con shift DAY asignado a OT en horario nocturno (18:00-06:00)
           o técnico con shift NIGHT asignado a OT en horario diurno (06:00-18:00).
    """
    from datetime import date as _date, datetime, timedelta
    from api.database.models import ManagedWorkOrderModel, WorkforceModel
    import json as _json

    if week_start:
        try:
            start = datetime.fromisoformat(week_start)
        except Exception:
            raise HTTPException(status_code=400, detail="week_start debe ser YYYY-MM-DD")
    else:
        today = _date.today()
        start = datetime.combine(today - timedelta(days=today.weekday()), datetime.min.time())
    end = start + timedelta(days=7)

    # IDOR / plant scoping
    if getattr(user, "plant_id", None) and user.role not in ("admin", "manager"):
        plant_id = user.plant_id

    # 1) Workforce indexada por worker_id
    wq = db.query(WorkforceModel)
    if plant_id:
        wq = wq.filter(WorkforceModel.plant_id == plant_id)
    workers_by_id = {w.worker_id: w for w in wq.all()}

    # 2) WOs programadas en la ventana
    woq = db.query(ManagedWorkOrderModel).filter(
        ManagedWorkOrderModel.planned_start.isnot(None),
        ManagedWorkOrderModel.planned_start >= start,
        ManagedWorkOrderModel.planned_start < end,
        ManagedWorkOrderModel.status.in_(["PROGRAMADO", "EN_EJECUCION", "REPROGRAMADO"]),
    )
    if plant_id:
        woq = woq.filter(ManagedWorkOrderModel.plant_id == plant_id)
    wos = woq.all()

    # 3) Acumulado HH por worker + detección de shift mismatch
    hh_by_worker: dict[str, float] = {}
    shift_violations: list[dict] = []
    for wo in wos:
        aw = wo.assigned_workers
        if isinstance(aw, str):
            try: aw = _json.loads(aw)
            except: aw = []
        if not aw:
            continue
        n = len(aw)
        est = float(wo.estimated_hours or 0)
        per = est / n if n else 0
        # Shift slot derivado del planned_start
        ps = wo.planned_start if isinstance(wo.planned_start, datetime) else None
        if ps is None:
            try: ps = datetime.fromisoformat(str(wo.planned_start).replace("Z",""))
            except: ps = None
        if ps:
            slot_shift = "night" if (ps.hour >= 18 or ps.hour < 6) else "day"
        else:
            slot_shift = None
        for entry in aw:
            wid = entry.get("worker_id") if isinstance(entry, dict) else None
            if not wid:
                continue
            hh_by_worker[wid] = hh_by_worker.get(wid, 0) + per
            # Shift mismatch
            tech = workers_by_id.get(wid)
            if tech and slot_shift:
                tech_shift = "night" if str(tech.shift or "").upper() in ("NIGHT", "NOCHE") else "day"
                if tech_shift != slot_shift:
                    shift_violations.append({
                        "worker_id": wid,
                        "worker_name": tech.name,
                        "tech_shift": tech_shift,
                        "slot_shift": slot_shift,
                        "wo_id": wo.wo_id,
                        "wo_number": wo.wo_number,
                        "planned_start": ps.isoformat() if ps else None,
                    })

    # 4) Overcapacity
    overcapacity = []
    for wid, hh in hh_by_worker.items():
        if hh > hours_per_week:
            tech = workers_by_id.get(wid)
            overcapacity.append({
                "worker_id": wid,
                "worker_name": tech.name if tech else wid,
                "specialty": tech.specialty if tech else None,
                "hh_assigned": round(hh, 2),
                "hh_capacity": hours_per_week,
                "overload_pct": round((hh / hours_per_week - 1) * 100, 1),
                "severity": "critical" if hh > hours_per_week * 1.25 else "high",
            })
    overcapacity.sort(key=lambda x: -x["hh_assigned"])

    return {
        "week_start": start.date().isoformat(),
        "week_end": (end - timedelta(days=1)).date().isoformat(),
        "plant_id": plant_id,
        "hours_per_week": hours_per_week,
        "totals": {
            "workers_with_load": len(hh_by_worker),
            "overcapacity_count": len(overcapacity),
            "shift_violations_count": len(shift_violations),
            "wos_in_window": len(wos),
        },
        "overcapacity": overcapacity,
        "shift_violations": shift_violations[:200],  # cap
    }


@router.get("/materials-live")
def materials_live(plant_id: str = "OCP-JFC1", db: Session = Depends(get_db)):
    """Materials status from actual scheduled WOs."""
    return scheduling_service.materials_from_wos(db, plant_id)


@router.put("/materials/{wo_id}/collection-status")
def update_material_collection(
    wo_id: str,
    data: dict,
    user=Depends(require_role("admin", "manager", "planner", "supervisor")),
    db: Session = Depends(get_db),
):
    """Update material collection status for a WO.
    data: { material_index: int, status: str, notes: str? }
    Statuses: PENDIENTE, PARCIAL, COMPLETADO, EN_AREA_ESPERA, ENTREGADO
    """
    return scheduling_service.update_material_collection(db, wo_id, data, user.get("user_id", ""))


@router.put("/materials/{wo_id}/bulk-status")
def bulk_update_material_status(
    wo_id: str,
    data: dict,
    user=Depends(require_role("admin", "manager", "planner", "supervisor")),
    db: Session = Depends(get_db),
):
    """Bulk update all materials in a WO to a given status.
    data: { status: str }
    """
    return scheduling_service.bulk_update_material_collection(db, wo_id, data.get("status", ""), user.get("user_id", ""))


# ── Batch clear assignments ──────────────────────────────────────────

@router.post("/clear-week")
def clear_week_assignments(
    data: dict,
    user=Depends(require_role("admin", "manager", "planner")),
    db: Session = Depends(get_db),
):
    """Clear all assignments for a specific week in one DB operation."""
    from api.database.models import ManagedWorkOrderModel
    from datetime import datetime
    plant_id = data.get("plant_id", "OCP-JFC1")
    week_start = data.get("week_start", "")
    week_end = data.get("week_end", "")
    if not week_start or not week_end:
        raise HTTPException(status_code=400, detail="week_start and week_end required")
    start_dt = datetime.fromisoformat(week_start)
    end_dt = datetime.fromisoformat(week_end + "T23:59:59") if "T" not in week_end else datetime.fromisoformat(week_end)
    # Clear any WO that has a planned_start in this week, OR has assigned_workers
    # but no planned_start yet (drafts sitting on technicians). Skip closed/cancelled
    # and anything already being executed (don't pull the rug on a running WO).
    # SF-743 (Jorge Sprint 7): además skip PROGRAMADO — son OTs "Congeladas" que
    # el planner reservó explícitamente; Clear Assignment NO debe tocarlas, solo
    # debe limpiar lo propuesto por la IA / drafts no validados.
    SKIP = ("CERRADO", "CANCELADO", "EN_EJECUCION", "PROGRAMADO")
    rows = db.query(ManagedWorkOrderModel).filter(
        ManagedWorkOrderModel.plant_id == plant_id,
        ~ManagedWorkOrderModel.status.in_(SKIP),
        ManagedWorkOrderModel.planned_start >= start_dt,
        ManagedWorkOrderModel.planned_start <= end_dt,
    ).all()
    # Also pick up WOs with assigned_workers but planned_start NULL (partial drafts).
    orphans = db.query(ManagedWorkOrderModel).filter(
        ManagedWorkOrderModel.plant_id == plant_id,
        ~ManagedWorkOrderModel.status.in_(SKIP),
        ManagedWorkOrderModel.planned_start.is_(None),
        ManagedWorkOrderModel.assigned_workers.isnot(None),
    ).all()
    count = 0
    for wo in list(rows) + list(orphans):
        wo.assigned_workers = None
        wo.planned_start = None
        wo.planned_end = None
        if wo.status in ("PROGRAMADO", "EN_PROGRAMACION", "REPROGRAMADO"):
            wo.status = "PLANIFICADO"
        count += 1
    db.commit()
    try:
        from api.services.ws_manager import queue_notify
        queue_notify("wo_bulk_clear", {"cleared": count, "week": week_start}, plant_id)
    except Exception:
        pass
    return {"ok": True, "cleared": count}


# ── Support Equipment (Cranes, Heavy Equipment) ─────────────────────

@router.get("/support-equipment")
def list_support_equipment(plant_id: str = "OCP-JFC1", db: Session = Depends(get_db)):
    from api.database.models import SupportEquipmentModel
    items = db.query(SupportEquipmentModel).filter(SupportEquipmentModel.plant_id == plant_id).all()
    return [{"equipment_id": e.equipment_id, "name": e.name, "equipment_type": e.equipment_type,
             "capacity_tons": e.capacity_tons, "available": e.available, "is_rented": e.is_rented,
             "out_of_service_reason": e.out_of_service_reason} for e in items]


@router.post("/support-equipment")
def create_support_equipment(data: dict, user=Depends(require_role("admin", "manager")), db: Session = Depends(get_db)):
    from api.database.models import SupportEquipmentModel
    eq = SupportEquipmentModel(plant_id=data.get("plant_id", "OCP-JFC1"), name=data.get("name", ""),
        equipment_type=data.get("equipment_type", "MOBILE_CRANE"), capacity_tons=data.get("capacity_tons"),
        available=data.get("available", True), is_rented=data.get("is_rented", False))
    db.add(eq); db.commit(); db.refresh(eq)
    return {"ok": True, "equipment_id": eq.equipment_id}


@router.put("/support-equipment/{equipment_id}")
def update_support_equipment(equipment_id: str, data: dict, user=Depends(require_role("admin", "manager", "supervisor")), db: Session = Depends(get_db)):
    from api.database.models import SupportEquipmentModel
    eq = db.query(SupportEquipmentModel).filter(SupportEquipmentModel.equipment_id == equipment_id).first()
    if not eq: raise HTTPException(status_code=404, detail="Equipment not found")
    for k in ["name", "available", "is_rented", "out_of_service_reason", "out_of_service_until", "capacity_tons"]:
        if k in data: setattr(eq, k, data[k])
    db.commit()
    return {"ok": True}


# Jorge 2026-04-27 (reunión 18:06): permitir eliminar equipos de apoyo creados
# por error. Sólo admin/manager — el supervisor sólo puede bloquear (PUT).
@router.delete("/support-equipment/{equipment_id}")
def delete_support_equipment(equipment_id: str, user=Depends(require_role("admin", "manager")), db: Session = Depends(get_db)):
    from api.database.models import SupportEquipmentModel
    eq = db.query(SupportEquipmentModel).filter(SupportEquipmentModel.equipment_id == equipment_id).first()
    if not eq: raise HTTPException(status_code=404, detail="Equipment not found")
    db.delete(eq); db.commit()
    return {"ok": True}


# ── Workforce availability management ─────────────────────────────────

@router.put("/workforce/{worker_id}/availability")
def update_worker_availability(
    worker_id: str,
    data: dict,
    user=Depends(require_role("admin", "manager", "planner", "supervisor")),
    db: Session = Depends(get_db),
):
    """Update worker availability (vacations, courses, absences)."""
    from api.database.models import WorkforceModel
    worker = db.query(WorkforceModel).filter(WorkforceModel.worker_id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    if "available" in data:
        worker.available = data["available"]
    if "absence_reason" in data:
        worker.absence_reason = data.get("absence_reason", "")
    if "absence_until" in data:
        worker.absence_until = data.get("absence_until")
    db.commit()
    return {"ok": True, "worker_id": worker_id, "available": worker.available}


# ── Phase 3: Gantt from managed WOs ─────────────────────────────────

@router.get("/gantt/export.xlsx")
def export_gantt_xlsx(
    plant_id: str = "OCP-JFC1",
    weeks: int = 2,
    db: Session = Depends(get_db),
):
    """Jorge G (transcript 2026-05-14): export carta Gantt a Excel.

    Devuelve un archivo .xlsx con una hoja por semana mostrando:
    - Columna izquierda: equipos
    - Columnas: cada día/turno
    - Celdas coloreadas por impacto (CRITICO=rojo, ALTO=naranja, etc.)
    - Hoja resumen con disponibilidad por equipo.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Alignment, Font
    from fastapi.responses import StreamingResponse

    rows = scheduling_service.get_gantt_managed(db, plant_id, weeks)
    availability = get_equipment_availability(
        plant_id=plant_id, week_start=None, weeks=weeks, db=db,
    )

    wb = Workbook()

    # ── Hoja 1: Carta Gantt (OT por OT) ─────────────────────────────
    ws = wb.active
    ws.title = "Carta Gantt"
    headers = ["N° OT", "Equipo", "Criticidad", "Descripción", "Tipo",
               "Prio", "Impacto", "Inicio plan", "Fin plan", "HH est.",
               "Especialidad", "Técnicos", "Status"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B5E20")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    color_by_impact = {
        "CRITICO": "EF4444",  # rojo
        "ALTO": "F97316",     # naranja
        "MEDIO": "EAB308",    # amarillo
        "BAJO": "10B981",     # verde
    }
    for r in rows:
        ws.append([
            r.get("wo_number", ""),
            r.get("equipment_tag", ""),
            r.get("equipment_criticality", ""),
            r.get("description", ""),
            r.get("wo_type", ""),
            r.get("priority_code", ""),
            r.get("impact_level", ""),
            r.get("planned_start", "")[:16] if r.get("planned_start") else "",
            r.get("planned_end", "")[:16] if r.get("planned_end") else "",
            r.get("estimated_hours", 0),
            ", ".join(r.get("specialties", [])),
            ", ".join([w.get("name", w.get("worker_id", "")) if isinstance(w, dict) else str(w)
                       for w in (r.get("assigned_workers") or [])]),
            r.get("status", ""),
        ])
        # Color cell "Impacto"
        impact_cell = ws.cell(row=ws.max_row, column=7)
        fill_color = color_by_impact.get(r.get("impact_level", "BAJO"), "10B981")
        impact_cell.fill = PatternFill("solid", fgColor=fill_color)
        impact_cell.font = Font(bold=True, color="FFFFFF")
        impact_cell.alignment = Alignment(horizontal="center")

    # Ajustar anchos
    widths = [12, 18, 10, 35, 8, 6, 10, 18, 18, 8, 18, 25, 15]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 2: Disponibilidad por equipo ───────────────────────────
    ws2 = wb.create_sheet("Disponibilidad")
    av_headers = ["Equipo", "Nombre", "Crit"] + availability["day_labels"] + ["Semana %", "Detenido (h)"]
    ws2.append(av_headers)
    for col in range(1, len(av_headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B5E20")

    for eq in availability["equipment"]:
        row_vals = [eq["tag"], eq["name"], eq["criticality"]]
        for d in eq["daily"]:
            row_vals.append(d["available_pct"])
        row_vals.append(eq["week_available_pct"])
        row_vals.append(eq["total_downtime_h"])
        ws2.append(row_vals)
        # Color celdas % por umbral
        for col_idx in range(4, 4 + len(eq["daily"]) + 1):
            cell = ws2.cell(row=ws2.max_row, column=col_idx)
            pct = cell.value
            if isinstance(pct, (int, float)):
                if pct >= 95:
                    cell.fill = PatternFill("solid", fgColor="D1FAE5")
                elif pct >= 90:
                    cell.fill = PatternFill("solid", fgColor="FEF3C7")
                else:
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")

    widths2 = [14, 28, 6] + [10] * len(availability["day_labels"]) + [10, 12]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Stream a memoria
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"carta_gantt_{plant_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/gantt/export.pdf")
def export_gantt_pdf(
    plant_id: str = "OCP-JFC1",
    weeks: int = 2,
    db: Session = Depends(get_db),
):
    """SF-745 (Jorge Sprint 7): export Carta Gantt a PDF.

    Tabla horizontal con OTs ordenadas por fecha, colores por impacto.
    Más liviano que el Excel — pensado para imprimir y entregar al
    supervisor de turno.
    """
    import io
    from datetime import datetime as _dt
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )
    from reportlab.lib.enums import TA_LEFT
    from fastapi.responses import StreamingResponse

    rows = scheduling_service.get_gantt_managed(db, plant_id, weeks)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1*cm, rightMargin=1*cm,
        topMargin=1*cm, bottomMargin=1*cm,
        title=f"Carta Gantt {plant_id}",
    )
    styles = getSampleStyleSheet()
    H1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=14,
                       textColor=rl_colors.HexColor("#0f172a"), leading=18)
    META = ParagraphStyle("META", fontSize=8.5,
                         textColor=rl_colors.HexColor("#64748b"))

    story = []
    story.append(Paragraph(f"<b>Carta Gantt</b> · {plant_id} · {weeks} sem · "
                          f"{_dt.now().strftime('%Y-%m-%d %H:%M')}", H1))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"{len(rows)} OTs · Generado por MAGEAM/AMS", META))
    story.append(Spacer(1, 8))

    impact_color = {
        "CRITICO": rl_colors.HexColor("#EF4444"),
        "ALTO": rl_colors.HexColor("#F97316"),
        "MEDIO": rl_colors.HexColor("#EAB308"),
        "BAJO": rl_colors.HexColor("#10B981"),
    }

    headers = ["N° OT", "Equipo", "Descripción", "Tipo", "Impacto",
               "Inicio", "Fin", "HH", "Especialidad", "Status"]
    data = [headers]
    for r in rows[:200]:  # limit a 200 para evitar PDFs gigantes
        data.append([
            r.get("wo_number", ""),
            r.get("equipment_tag", "")[:18],
            (r.get("description", "") or "")[:60],
            r.get("wo_type", ""),
            r.get("impact_level", ""),
            (r.get("planned_start", "") or "")[:16],
            (r.get("planned_end", "") or "")[:16],
            f"{r.get('estimated_hours', 0):.1f}",
            ", ".join((r.get("specialties") or [])[:2]),
            r.get("status", ""),
        ])
    tbl = Table(data, colWidths=[2.2*cm, 2.5*cm, 6*cm, 1.2*cm, 1.5*cm,
                                  2.5*cm, 2.5*cm, 1*cm, 2.5*cm, 2.4*cm],
                repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1B5E20")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.3, rl_colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [rl_colors.white, rl_colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])
    # Color celda Impacto por valor (col 4)
    for i, r in enumerate(rows[:200], start=1):
        c = impact_color.get(r.get("impact_level"), rl_colors.HexColor("#94a3b8"))
        style.add("BACKGROUND", (4, i), (4, i), c)
        style.add("TEXTCOLOR", (4, i), (4, i), rl_colors.white)
        style.add("FONTNAME", (4, i), (4, i), "Helvetica-Bold")
    tbl.setStyle(style)
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    filename = f"carta_gantt_{plant_id}_{_dt.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/gantt")
def get_gantt_managed(
    plant_id: str = "OCP-JFC1",
    weeks: int = 2,
    db: Session = Depends(get_db),
):
    """Gantt data from managed_work_orders for next N weeks."""
    return scheduling_service.get_gantt_managed(db, plant_id, weeks)


# ── Existing Gantt endpoints ─────────────────────────────────────────

@router.get("/programs/{program_id}/gantt")
def get_gantt(program_id: str, db: Session = Depends(get_db)):
    result = scheduling_service.get_gantt(db, program_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Program not found")
    return result


@router.get("/programs/{program_id}/gantt/export")
def export_gantt_excel(program_id: str, db: Session = Depends(get_db)):
    filepath = scheduling_service.export_gantt_excel(db, program_id)
    if not filepath:
        raise HTTPException(status_code=404, detail="Program not found")
    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"gantt_{program_id}.xlsx",
    )



# ── AI Auto-Schedule ─────────────────────────────────────────────────

class AIAutoScheduleRequest(BaseModel):
    plant_id: str = "OCP-JFC1"
    week_start: str = ""  # ISO date, defaults to current week


from pydantic import BaseModel as BaseModel2


class ShiftContinuityRequest(BaseModel):
    model_config = {"extra": "ignore"}
    wo_id: str
    shift_hours: float = 12.0     # duración del turno (default 12h, configurable)
    strategy: str = "A"           # A=continuar en turno siguiente · B=mismo tec al día siguiente
    start_at: str | None = None   # ISO; default = wo.planned_start


@router.post("/shift-continuity-plan")
def shift_continuity_plan(
    data: ShiftContinuityRequest,
    db: Session = Depends(get_db),
):
    """SF-562 — Plan de continuidad cuando una OT excede la duración del turno.

    Dada una OT con estimated_hours > shift_hours, devuelve un plan de chunks
    según la estrategia elegida:
      - A (default): saldo al técnico del turno siguiente (noche).
      - B: saldo al mismo técnico al día siguiente (post-descanso).

    Respeta días de descanso (workforce.shift_pattern). NO transiciona la OT;
    solo devuelve el plan para que el planificador lo confirme antes de aplicar.
    """
    from api.database.models import ManagedWorkOrderModel, WorkforceModel

    wo = db.query(ManagedWorkOrderModel).filter(ManagedWorkOrderModel.wo_id == data.wo_id).first()
    if not wo:
        raise HTTPException(status_code=404, detail="WO not found")

    total_h = float(wo.estimated_hours or 0)
    shift_h = max(1.0, float(data.shift_hours))
    if total_h <= shift_h:
        return {"requires_continuity": False, "total_hours": total_h, "shift_hours": shift_h, "chunks": []}

    if data.start_at:
        try:
            cursor = datetime.fromisoformat(data.start_at)
        except ValueError:
            cursor = wo.planned_start or datetime.now()
    else:
        cursor = wo.planned_start or datetime.now()

    workers = wo.assigned_workers or []
    worker_ids = [w.get("worker_id") if isinstance(w, dict) else w for w in workers]
    workforce_rows = (
        db.query(WorkforceModel).filter(WorkforceModel.worker_id.in_(worker_ids)).all()
        if worker_ids else []
    )
    rest_days = {}
    for w in workforce_rows:
        # shift_pattern típico: "DDDDNNN" o lista de off-days; fallback: ningún descanso conocido.
        pattern = (getattr(w, "shift_pattern", "") or "").upper()
        rest_days[w.worker_id] = set(i for i, ch in enumerate(pattern) if ch in ("X", "L", "R", "-"))

    chunks = []
    remaining = total_h
    chunk_idx = 0
    safety = 0
    while remaining > 0 and safety < 20:
        safety += 1
        chunk_h = min(shift_h, remaining)

        if data.strategy == "B" and chunk_idx > 0:
            # Saltar al día siguiente, misma hora-de-inicio que el primer chunk.
            cursor = cursor + timedelta(days=1)
            # Saltar días de descanso si aplica al primer worker
            if worker_ids and rest_days.get(worker_ids[0]):
                while cursor.weekday() in rest_days[worker_ids[0]]:
                    cursor = cursor + timedelta(days=1)
            assigned_to = worker_ids[0] if worker_ids else None
            shift_label = "siguiente_dia"
        elif data.strategy == "A" and chunk_idx > 0:
            # Saltar al siguiente bloque de shift_h horas (turno noche típicamente)
            cursor = cursor + timedelta(hours=shift_h)
            assigned_to = (worker_ids[chunk_idx % len(worker_ids)] if worker_ids else None)
            shift_label = "noche" if chunk_idx % 2 else "dia"
        else:
            assigned_to = worker_ids[0] if worker_ids else None
            shift_label = "dia"

        chunks.append({
            "chunk": chunk_idx + 1,
            "start": cursor.isoformat(),
            "end": (cursor + timedelta(hours=chunk_h)).isoformat(),
            "hours": chunk_h,
            "assigned_to": assigned_to,
            "shift": shift_label,
        })
        remaining -= chunk_h
        chunk_idx += 1
        if data.strategy == "A":
            cursor = cursor + timedelta(hours=chunk_h)  # avance lineal en estrategia A intra-loop

    return {
        "requires_continuity": True,
        "wo_id": wo.wo_id,
        "wo_number": wo.wo_number,
        "total_hours": total_h,
        "shift_hours": shift_h,
        "strategy": data.strategy,
        "chunks": chunks,
    }


@router.post("/ai-auto-schedule")
def ai_auto_schedule(
    data: AIAutoScheduleRequest = None,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """AI auto-assigns unscheduled WOs to optimal technicians based on
    skills, availability, priority, and workload balance."""
    import os, json
    from api.database.models import ManagedWorkOrderModel
    from api.services import assignment_service

    plant_id = data.plant_id if data else "OCP-JFC1"

    # Get unscheduled WOs (PLANIFICADO + RELEASED)
    wos = db.query(ManagedWorkOrderModel).filter(
        ManagedWorkOrderModel.status.in_(["PLANIFICADO", "RELEASED", "CREADO"]),
        ManagedWorkOrderModel.plant_id == plant_id,
    ).order_by(
        # P1 first, then P2, etc.
        ManagedWorkOrderModel.priority_code.asc(),
        ManagedWorkOrderModel.created_at.asc(),
    ).limit(20).all()

    if not wos:
        return {"assignments": [], "message": "No unscheduled WOs found"}

    # Get available technicians
    techs = assignment_service.get_technician_profiles(db, plant_id=plant_id)
    if not techs:
        return {"assignments": [], "message": "No technicians available"}

    # Build context for AI
    wo_list = []
    for wo in wos:
        ops = wo.operations or []
        specialties = list(set(op.get("specialty", "") for op in ops if isinstance(op, dict)))
        wo_list.append({
            "wo_id": wo.wo_id,
            "wo_number": wo.wo_number,
            "priority": wo.priority_code,
            "equipment": wo.equipment_tag,
            "description": (wo.description or "")[:100],
            "hours": wo.estimated_hours or 4,
            "specialties_needed": specialties or ["MECHANICAL"],
            "wo_type": wo.wo_type,
        })

    tech_list = []
    for t in techs:
        tech_list.append({
            "worker_id": t.get("worker_id",""),
            "name": t.get("name",""),
            "specialty": t.get("specialty",""),
            "shift": t.get("shift",""),
            "available": t.get("available",True),
            "years_exp": t.get("years_experience",0),
            "equipment_expertise": (t.get("equipment_expertise",[]) or [])[:5],
        })

    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        # Fallback: simple round-robin assignment
        assignments = []
        available_techs = [t for t in tech_list if t.get("available", True)]
        if not available_techs:
            available_techs = tech_list[:20]  # fallback: use first 20
        today = datetime.now().date()
        # Spread across current work week (Mon-Fri)
        weekday = today.weekday()
        monday = today - timedelta(days=weekday)
        work_days = [monday + timedelta(days=d) for d in range(5)]
        
        # Track how many WOs each tech has, to spread across days
        tech_wo_count = {}
        for i, wo in enumerate(wo_list):
            if not available_techs:
                break
            tech = available_techs[i % len(available_techs)]
            tid = tech["worker_id"]
            n = tech_wo_count.get(tid, 0)
            tech_wo_count[tid] = n + 1
            day = work_days[n % len(work_days)]
            shift = "night" if n % 2 == 1 else "day"
            assignments.append({
                "wo_id": wo["wo_id"],
                "wo_number": wo["wo_number"],
                "worker_id": tid,
                "worker_name": tech["name"],
                "suggested_date": day.isoformat(),
                "reason": "AI fallback: round-robin by specialty",
                "shift": shift,
            })
        return {"assignments": assignments, "message": f"Assigned {len(assignments)} WOs to {len(set(a['worker_id'] for a in assignments))} technicians", "ai_used": False}

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        prompt = (
            "You are a maintenance scheduling AI. Assign these work orders to the best technicians.\n\n"
            "WORK ORDERS:\n" + json.dumps(wo_list, indent=2) + "\n\n"
            "TECHNICIANS:\n" + json.dumps(tech_list, indent=2) + "\n\n"
            "Rules:\n"
            "- Match specialty: mechanical WOs to FITTER/MECHANICAL techs, electrical to ELECTRICAL, etc.\n"
            "- P1/P2 get the most experienced technicians\n"
            "- Balance workload across technicians (don't overload one person)\n"
            "- Consider equipment_expertise match when possible\n"
            "- Assign shift: 'day' or 'night' based on priority (P1 = day shift priority)\n\n"
            "Return ONLY a JSON array of assignments:\n"
            '[{"wo_id": "...", "wo_number": "...", "worker_id": "...", "worker_name": "...", "reason": "brief reason", "shift": "day|night", "suggested_date": "YYYY-MM-DD"}]'
        )

        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}],
        )

        text = resp.content[0].text
        # Strip markdown code blocks
        text = text.replace('```json', '').replace('```', '').strip()
        import logging; logging.getLogger(__name__).info("AI auto-schedule response: %s", text[:200])
        # Extract JSON from response
        import re
        json_match = re.search(r'\[.*\]', text, re.DOTALL)
        if json_match:
            assignments = json.loads(json_match.group())
        else:
            assignments = []

        # Fallback if AI returned empty
        if not assignments:
            available_techs = [t for t in tech_list if t.get("available", True)]
            for i, wo in enumerate(wo_list):
                if not available_techs:
                    break
                tech = available_techs[i % len(available_techs)]
                assignments.append({
                    "wo_id": wo["wo_id"],
                    "wo_number": wo["wo_number"],
                    "worker_id": tech["worker_id"],
                    "worker_name": tech["name"],
                    "reason": "AI fallback: round-robin by specialty",
                    "shift": "day",
                    "suggested_date": (datetime.now()).strftime("%Y-%m-%d"),
                })

        return {
            "assignments": assignments,
            "message": "AI assigned " + str(len(assignments)) + " work orders to " + str(len(set(a.get("worker_id", "") for a in assignments))) + " technicians",
            "ai_used": True,
            "wo_count": len(wo_list),
            "tech_count": len(tech_list),
        }

    except Exception as e:
        return {"assignments": [], "message": "AI error: " + str(e)[:100], "ai_used": False}


# ── AI Daily Briefing ────────────────────────────────────────────────

@router.post("/ai-daily-briefing")
def ai_daily_briefing(
    plant_id: str = "OCP-JFC1",
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Generate AI daily briefing for the execution meeting."""
    import os, json
    from api.database.models import ManagedWorkOrderModel

    # Gather data
    all_wos = db.query(ManagedWorkOrderModel).filter(
        ManagedWorkOrderModel.plant_id == plant_id,
    ).all()

    by_status = {}
    for wo in all_wos:
        by_status.setdefault(wo.status, []).append(wo)

    in_exec = by_status.get("EN_EJECUCION", []) + by_status.get("IN_PROGRESS", [])
    scheduled = by_status.get("PROGRAMADO", []) + by_status.get("SCHEDULED", [])
    completed_today = [w for w in by_status.get("CERRADO", []) + by_status.get("COMPLETED", [])
                       if w.closed_at and w.closed_at.date() == datetime.now().date()]
    overdue = [w for w in in_exec if w.planned_end and w.planned_end < datetime.now()]
    fast_track = [w for w in all_wos if w.is_fast_track and w.status not in ("CERRADO", "CANCELADO", "COMPLETED")]

    summary_data = {
        "date": datetime.now().strftime("%Y-%m-%d"),
        "plant": plant_id,
        "in_execution": len(in_exec),
        "scheduled_today": len(scheduled),
        "completed_today": len(completed_today),
        "overdue": len(overdue),
        "fast_track_active": len(fast_track),
        "total_open": len([w for w in all_wos if w.status not in ("CERRADO", "CANCELADO", "COMPLETED")]),
        "total_hh_planned": sum(w.estimated_hours or 0 for w in in_exec),
        "overdue_details": [{"wo": w.wo_number, "equip": w.equipment_tag, "priority": w.priority_code, "planned_end": w.planned_end.isoformat() if w.planned_end else ""} for w in overdue[:5]],
        "fast_track_details": [{"wo": w.wo_number, "equip": w.equipment_tag, "priority": w.priority_code, "desc": (w.description or "")[:60]} for w in fast_track[:5]],
    }

    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"briefing": "AI unavailable. Manual data:\n" + json.dumps(summary_data, indent=2), "data": summary_data}

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        prompt = (
            "You are the AI assistant for a daily maintenance execution meeting at a mining plant.\n"
            "Generate a concise daily briefing (max 300 words) covering:\n"
            "1. Executive summary (1-2 sentences)\n"
            "2. Active emergencies / fast-track WOs\n"
            "3. Overdue WOs requiring attention\n"
            "4. Today's workload summary\n"
            "5. Key risks or recommendations\n\n"
            "Data:\n" + json.dumps(summary_data, indent=2) + "\n\n"
            "Format with markdown headers. Be direct and actionable. Respond in the same language as the input data."
        )

        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=500,
            messages=[{"role": "user", "content": prompt}],
        )
        briefing = resp.content[0].text
        return {"briefing": briefing, "data": summary_data, "ai_used": True}

    except Exception as e:
        return {"briefing": "AI error: " + str(e)[:100], "data": summary_data, "ai_used": False}


class AutoLevelParseRequest(BaseModel):
    text: str  # natural language instructions
    wo_summary: list = []  # [{wo_number, equipment_tag, priority_code, hours}]
    week_days: list = []   # ['2026-04-29', '2026-04-30', ...]


@router.post("/parse-autolevel-instructions")
def parse_autolevel_instructions(data: AutoLevelParseRequest):
    """Convierte instrucciones en lenguaje natural en constraints estructurados
    para el algoritmo Auto-Level. Antes el frontend usaba `.includes()` con
    keywords hardcodeadas — ahora Claude entiende intención.

    Devuelve { parsed: {priority_boost_wos, priority_boost_equipment, light_days,
    excluded_days, capacity_override_pct, summary}, ai_used: bool }
    """
    import os, json
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key or not data.text.strip():
        # Fallback: extracción keyword (lo que ya hacía el frontend)
        return {"parsed": _fallback_parse_instructions(data.text), "ai_used": False}
    try:
        import anthropic
    except ImportError:
        return {"parsed": _fallback_parse_instructions(data.text), "ai_used": False}

    wos_compact = [
        f"{w.get('wo_number','?')} ({w.get('equipment_tag','?')}, {w.get('priority_code','P3')}, {w.get('hours',0)}h)"
        for w in (data.wo_summary or [])[:50]
    ]
    days_str = ", ".join(data.week_days or [])

    prompt = f"""Eres un planificador de mantenimiento. Convierte estas instrucciones en lenguaje natural a constraints estructurados para el algoritmo Auto-Level.

INSTRUCCIONES DEL USUARIO:
"{data.text}"

CONTEXTO — OTs disponibles ({len(wos_compact)}):
{chr(10).join(wos_compact)}

DÍAS DE LA SEMANA: {days_str}

Devuelve SOLO un JSON con estas claves:
{{
  "priority_boost_wos": ["lista de wo_number a priorizar (boost al top, ej: usuario dijo 'priorizar OT-2026-00031')"],
  "priority_boost_equipment": ["lista de equipment_tag a priorizar"],
  "deprioritize_wos": ["wo_numbers a postergar"],
  "light_days": ["fechas YYYY-MM-DD con capacidad reducida (ej: 'lunes liviano' → fecha del lunes)"],
  "excluded_days": ["fechas YYYY-MM-DD a excluir totalmente (ej: 'no programar viernes' o feriados)"],
  "capacity_override_pct": null o int (1-100) si menciona % capacidad,
  "include_weekend": true/false (true si menciona sabado/domingo/weekend),
  "summary": "resumen 1-2 frases de lo que el usuario quiso decir"
}}"""

    try:
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=600,
            messages=[{"role": "user", "content": prompt}],
        )
        text = resp.content[0].text.strip()
        start = text.find("{")
        end = text.rfind("}") + 1
        if start < 0:
            return {"parsed": _fallback_parse_instructions(data.text), "ai_used": False}
        parsed = json.loads(text[start:end])
        return {"parsed": parsed, "ai_used": True}
    except Exception as e:
        return {"parsed": _fallback_parse_instructions(data.text), "ai_used": False, "error": str(e)[:120]}


def _fallback_parse_instructions(text: str) -> dict:
    """Heurística keyword si Claude no responde — equivalente a lo que
    el frontend ya hacía con .includes()."""
    t = (text or "").lower()
    return {
        "priority_boost_wos": [],
        "priority_boost_equipment": [],
        "deprioritize_wos": [],
        "light_days": [],
        "excluded_days": [],
        "capacity_override_pct": None,
        "include_weekend": any(k in t for k in ["sabado", "sábado", "domingo", "weekend", "fin de semana", "7x7", "7 dias", "todos los dias"]),
        "summary": (text[:120] + "...") if len(text) > 120 else text,
    }
