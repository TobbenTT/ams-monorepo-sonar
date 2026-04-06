"""Migrate work_orders → managed_work_orders for GOLDFIELDS-SN."""
import sys, uuid, json
from datetime import datetime
sys.path.insert(0, '/app')
from api.database.connection import SessionLocal
from sqlalchemy import text

db = SessionLocal()
PLANT = 'GOLDFIELDS-SN'
now = datetime.utcnow().isoformat()

# ═══ 1. work_orders → managed_work_orders ═══
print("=== work_orders → managed_work_orders ===")
wos = db.execute(text("""
    SELECT work_order_id, order_type, equipment_id, equipment_tag, priority, status,
           created_date, actual_duration_hours, description
    FROM work_orders WHERE plant_id = :p
    LIMIT 300
"""), {'p': PLANT}).fetchall()
print(f"  Found {len(wos)} work_orders for {PLANT}")

mwo_count = 0
for wo in wos:
    wo_id = str(uuid.uuid4())
    wo_num = wo[0] or f"OT-GF-{str(uuid.uuid4())[:6]}"
    raw_status = (wo[5] or '').upper()
    if any(x in raw_status for x in ['CECO', 'CLSD', 'TECO', 'CLOSED']):
        status = 'CERRADO'
    elif any(x in raw_status for x in ['REL', 'RELEASED']):
        status = 'PROGRAMADO'
    elif any(x in raw_status for x in ['CRTD', 'CREATED']):
        status = 'CREADO'
    else:
        status = 'CERRADO'
    prio = wo[4] or 'P3'
    if prio not in ('P1','P2','P3','P4'):
        prio = 'P3'
    try:
        db.execute(text("""
            INSERT INTO managed_work_orders (wo_id, wo_number, wo_type, description, equipment_tag,
                equipment_id, priority_code, status, plant_id, created_at, actual_hours, planned_by)
            VALUES (:id, :num, :type, :desc, :tag, :eid, :prio, :status, :plant, :at, :hours, :by)
        """), {
            'id': wo_id, 'num': wo_num, 'type': wo[1] or 'PM01',
            'desc': wo[8] or '', 'tag': wo[3] or '', 'eid': wo[2] or '',
            'prio': prio, 'status': status, 'plant': PLANT,
            'at': wo[6] or now, 'hours': wo[7] or 0, 'by': 'import',
        })
        mwo_count += 1
    except Exception as e:
        db.rollback()
        if mwo_count == 0:
            print(f"  First error: {str(e)[:150]}")
        continue

db.commit()
print(f"  Migrated {mwo_count} → managed_work_orders")

# ═══ 2. Create work_requests from Excel notifications data ═══
# The notifications were imported into the wrong table.
# Read from the Excel directly to create proper work_requests
print("\n=== Creating work_requests from 24_notifications.xlsx ===")
try:
    from openpyxl import load_workbook
    wb = load_workbook('/app/seed_data/24_notifications.xlsx', read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h else f'col_{i}' for i, h in enumerate(rows[0])]
    data = rows[1:201]  # First 200
    wb.close()
    print(f"  Excel has {len(rows)-1} rows, importing first 200")

    wr_count = 0
    for row in data:
        d = dict(zip(headers, row))
        wr_id = f"WR-GF-{str(uuid.uuid4())[:8]}"
        desc = d.get('qmtxt') or d.get('description') or d.get('short_text') or 'Imported notification'
        tag = d.get('sap_func_loc_short') or d.get('equipment_tag') or d.get('tplnr') or ''
        prio_raw = d.get('priokx') or d.get('priority') or '3'
        prio_map = {'1': 'P1', '2': 'P2', '3': 'P3', '4': 'P4', 'Very High': 'P1', 'High': 'P2', 'Medium': 'P3', 'Low': 'P4'}
        prio = prio_map.get(str(prio_raw), 'P3')
        notif_type = d.get('qmart') or d.get('notification_type') or 'M1'
        created = d.get('erdat') or d.get('created_date') or now
        if hasattr(created, 'isoformat'):
            created = created.isoformat()

        ai_class = json.dumps({"plant_id": PLANT, "source": "excel_import"})

        try:
            db.execute(text("""
                INSERT INTO work_requests (request_id, status, equipment_tag, problem_description,
                    priority_code, created_by, created_at, notification_type, ai_classification, aviso_number)
                VALUES (:id, :status, :tag, :desc, :prio, :by, :at, :ntype, :ai, :anum)
            """), {
                'id': wr_id, 'status': 'APROBADO', 'tag': str(tag),
                'desc': str(desc), 'prio': prio,
                'by': 'import', 'at': str(created), 'ntype': str(notif_type),
                'ai': ai_class, 'anum': d.get('qmnum') or '',
            })
            wr_count += 1
        except Exception as e:
            db.rollback()
            if wr_count == 0:
                print(f"  First error: {str(e)[:150]}")
            continue

    db.commit()
    print(f"  Created {wr_count} work_requests")
except Exception as e:
    print(f"  Error: {e}")

# ═══ Stats ═══
print("\n=== Final stats for GOLDFIELDS-SN ===")
for t, col in [('managed_work_orders', 'plant_id'), ('work_requests', 'ai_classification')]:
    if col == 'plant_id':
        c = db.execute(text(f"SELECT count(*) FROM {t} WHERE plant_id = :p"), {'p': PLANT}).scalar()
    else:
        c = db.execute(text(f"SELECT count(*) FROM {t} WHERE ai_classification LIKE :p"), {'p': f'%{PLANT}%'}).scalar()
    print(f"  {t}: {c}")

print("\n  managed_work_orders by status:")
for r in db.execute(text("SELECT status, count(*) FROM managed_work_orders WHERE plant_id = :p GROUP BY status"), {'p': PLANT}).fetchall():
    print(f"    {r[0]}: {r[1]}")

db.close()
print("\nDone!")
