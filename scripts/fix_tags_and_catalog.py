"""
1. Fix equipment_tag in 03_failure_modes.xlsx (use sap_func_loc_short instead of equnr)
2. Add equipment_name column to 15_catalog_profiles.xlsx Catálogos sheet
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl import Workbook
from collections import defaultdict

# ============================================================
# LOAD HIERARCHY DATA
# ============================================================
print("Loading hierarchy...", flush=True)
wb_hier = openpyxl.load_workbook('seed_data/01_equipment_hierarchy.xlsx', data_only=True)
ws_hier = wb_hier['Equipment Hierarchy']
h_headers = [ws_hier.cell(1, c).value for c in range(1, ws_hier.max_column + 1)]
h_idx = {h: i + 1 for i, h in enumerate(h_headers)}

# Build equnr → tag/name and eqart → profile mapping
equnr_to_tag = {}
equnr_to_pltxt = {}
eqart_to_profile_name = {}

EQART_TO_PROFILE = {
    'BLAX': 'VENT_AX', 'BLSC': 'COMP_TOR', 'BULL': 'BULL_BUL',
    'CCTV': 'CÁM_CCTV', 'CELO': 'CELOSÌA', 'ELHE': 'CAL_ELÉC',
    'EMBT': 'MOT_ELÉC', 'HBFF': 'SOP_CENT', 'HECQ': 'INT_ENFR',
    'HLCN': 'PUEN_GRÚ', 'LDLU': 'SIST_LUB', 'MSST': 'AGIT_MEC',
    'MXAG': 'AGIT_MEC', 'PCDT': 'DET_META', 'POOL': 'PIS_SUM',
    'PUCE': 'BOMB_CEN', 'PURE': 'BOMB_REC', 'PURO': 'BOMB_ROT',
    'PUTU': 'BOMB_TUR', 'PUVA': 'BOMB_VAC', 'SERV': 'SERVIDOR',
    'SFPR': 'FIL_PREN', 'SHEW': 'ESPESADO', 'SRCV': 'CORR_TRA',
    'SRFE': 'ALIM_BAN', 'VEBN': 'TOLVA', 'VEBX': 'CAJÒN',
    'VECH': 'CHUTE', 'VEDR': 'SECA_AIR', 'VELN': 'LIN_PROC',
    'VETA': 'EST_AGIT', 'VIBR': 'VIBRADOR',
    'ACOU': 'UNID_AC', 'AEID': 'INCH_DRI', 'AESS': 'SIST_LAN',
    'AETS': 'SOPORTES', 'AEWM': 'MON_LAVA', 'BCOM': 'TAB_COMU',
    'BLEH': 'CAMP_EXT', 'BLFC': 'VENT_CEN', 'BSCI': 'DET_FH',
    'BUVA': 'LIN_PROC', 'CEDE': 'BOMB_CEN', 'CFOF': 'DIAL_ACE',
    'CMCS': 'CORE_SWI', 'COHI': 'COMP_HID', 'CPLC': 'GAB_PLC',
    'CRIO': 'GAB_RIO', 'CSSP': 'PANE_SCA', 'CTDP': 'TAB_CEL',
    'DECH': 'GAB_MANG', 'DECO': 'CONS_ESC', 'DEFI': 'EST_MANG',
    'DEHY': 'HIDRANTE', 'EABA': 'BANC_BAT', 'EABC': 'CARG_BAT',
    'EAUP': 'UPS', 'ELCB': 'CEL_IPOT', 'ELDI': 'CEL_IPOT',
    'ELIN': 'CEL_PMOT', 'ELPS': 'CEL_PMOT', 'EMRE': 'CARG_FRO',
    'EYWA': 'DUCH_EME', 'FIPS': 'TAB_SCI', 'FIWA': 'FIREWALL',
    'GAVA': 'LIN_PROC', 'HBEH': 'HORN_ELÉ', 'HEPL': 'INT_PLAC',
    'HIWA': 'HIDR_AGU', 'HMIB': 'GAB_PLC', 'HTEL': 'TECL_ELE',
    'HTHA': 'TECL_ELE', 'HYPU': 'UNID_HID', 'INDP': 'TAB_DIST',
    'INPR': 'IND_PROC', 'KNVA': 'VALVULAS', 'KVMS': 'KVMSWITC',
    'LEBT': 'CAM_PLUM', 'LEFL': 'MONT_CAR', 'LIDP': 'TAB_DIST',
    'LIIN': 'TECL_ELE', 'LOFR': 'CARG_FRO', 'LUME': 'PAQ_QUIM',
    'MBAY': 'MUES_TRA', 'MBSC': 'BASCULA', 'MCSC': 'HARN_EST',
    'MFSU': 'SUPR_MF', 'MIRI': 'IND_PROC', 'MOPL': 'PLAT_MÓV',
    'MTTR': 'CARR_MNT', 'OCRD': 'PUER_MOT', 'PCCC': 'CICL_POT',
    'PCCO': 'COL_POLV', 'PCPT': 'TRAN_POT', 'PREU': 'UNI_PRES',
    'PRLI': 'DET_FH', 'PTWW': 'PTA_AR', 'PUSC': 'BOMB_HAR',
    'PWDP': 'TAB_DIST', 'ROVA': 'VALVULAS', 'RPHA': 'REM_PERN',
    'SAPR': 'MUES_ROT', 'SASE': 'SEN_PROC', 'SAWS': 'MUES_ROT',
    'SBDP': 'TAB_DIST', 'SBLR': 'ARRA_LÍQ', 'SECT': 'CÁM_CCTV',
    'SEFL': 'SEN_PROC', 'SFCF': 'FIL_COAL', 'SFCL': 'FIL_HOJA',
    'SMST': 'CHIMENEA', 'SRRT': 'MESA_ROD', 'SSEP': 'TAB_DIST',
    'STIT': 'CAM_RIEG', 'SUPP': 'SOPORTES', 'TAPR': 'EST_ALMA',
    'TOLH': 'MANI_COR', 'TRDE': 'LIN_PROC', 'TRFL': 'LIN_PROC',
    'TRLE': 'TOLVA', 'TRPR': 'TRAN_SEÑ', 'TRTE': 'TRAN_SEÑ',
    'UPDP': 'TAB_DIST', 'USAP': 'TAB_CEL', 'USCT': 'TAB_CEL',
    'USPS': 'TAB_CEL', 'VABU': 'VALVULAS', 'VACO': 'VALVULAS',
    'VAKN': 'VALVULAS', 'VAPI': 'VALVULAS', 'VEAL': 'EST_ALMA',
    'VEAR': 'ACUM_AIR', 'VEMO': 'CASC_MOL', 'VEPC': 'COL_PROC',
    'VETR': 'RACK_BAN', 'VEWR': 'PLAN_POT', 'VFBT': 'CELD_VDF',
    'VVCP': 'SUPR_POL', 'DUCT': 'DUCTOS', 'HETU': 'BOMB_TUR',
    'ROBL': 'MESA_ROD', 'SRPL': 'PLAT_MÓV', 'MCTR': 'TRN_PROC',
    'LUBR': 'SIST_LUB',
}

# Reverse mapping: profile → list of equipment names
profile_to_equipment = defaultdict(list)

for row in range(2, ws_hier.max_row + 1):
    if ws_hier.cell(row, h_idx['level']).value != 4:
        continue
    equnr = ws_hier.cell(row, h_idx['equnr']).value
    fl_short = ws_hier.cell(row, h_idx['sap_func_loc_short']).value
    pltxt = ws_hier.cell(row, h_idx['pltxt']).value
    eqart = ws_hier.cell(row, h_idx['eqart']).value
    eqart_desc = ws_hier.cell(row, h_idx['eqart_desc']).value

    equnr_to_tag[equnr] = fl_short
    equnr_to_pltxt[equnr] = pltxt

    profile = EQART_TO_PROFILE.get(eqart)
    if profile:
        name = eqart_desc or pltxt or eqart
        if name not in profile_to_equipment.get(profile, []):
            profile_to_equipment[profile].append(name)

print(f"  Equipment loaded: {len(equnr_to_tag)}", flush=True)
print(f"  Profile→equipment mappings: {len(profile_to_equipment)}", flush=True)

# ============================================================
# FIX 1: 03_failure_modes.xlsx - Fix equipment_tag
# ============================================================
print("\n" + "=" * 60, flush=True)
print("FIX 1: Correcting equipment_tag in 03_failure_modes.xlsx", flush=True)
print("=" * 60, flush=True)

# Read all data from existing file
wb_fm = openpyxl.load_workbook('seed_data/03_failure_modes.xlsx', read_only=True)
ws_fm = wb_fm['failure_modes']
fm_headers = [cell.value for cell in next(ws_fm.iter_rows(min_row=1, max_row=1))]

# Write to new file with write_only mode
wb_out = Workbook(write_only=True)
ws_out = wb_out.create_sheet('failure_modes')
ws_out.append(fm_headers)

row_count = 0
fixed_count = 0
for row in ws_fm.iter_rows(min_row=2, values_only=True):
    row_list = list(row)
    equnr = row_list[1]  # equnr is col 2

    # Fix equipment_tag (col 1) - use sap_func_loc_short
    tag = equnr_to_tag.get(equnr)
    if tag:
        row_list[0] = tag
        fixed_count += 1

    ws_out.append(row_list)
    row_count += 1

    if row_count % 50000 == 0:
        print(f"  Processed {row_count} rows...", flush=True)

wb_fm.close()
wb_out.save('seed_data/03_failure_modes.xlsx')
print(f"  Total rows: {row_count}", flush=True)
print(f"  Tags fixed: {fixed_count}", flush=True)
print("  SAVED!", flush=True)

# ============================================================
# FIX 2: 15_catalog_profiles.xlsx - Add equipment_name column
# ============================================================
print("\n" + "=" * 60, flush=True)
print("FIX 2: Adding equipment_name to 15_catalog_profiles.xlsx", flush=True)
print("=" * 60, flush=True)

wb_cat = openpyxl.load_workbook('seed_data/15_catalog_profiles.xlsx')

# --- Fix Perfil_Catálogo sheet: add equipment_name column ---
ws_perf = wb_cat['Perfil_Catálogo']
perf_headers = [ws_perf.cell(1, c).value for c in range(1, ws_perf.max_column + 1)]
print(f"  Perfil_Catálogo current headers: {perf_headers}", flush=True)

# Add equipment_name as column 5
ws_perf.insert_cols(5)
ws_perf.cell(1, 5, 'equipment_name')

for r in range(2, ws_perf.max_row + 1):
    pc = ws_perf.cell(r, 1).value
    if pc and pc in profile_to_equipment:
        names = profile_to_equipment[pc]
        ws_perf.cell(r, 5, names[0])  # First equipment name for this profile

print(f"  Added equipment_name to Perfil_Catálogo", flush=True)

# --- Fix Catálogos sheet: add equipment_name column ---
ws_cat = wb_cat['Catálogos']
cat_headers = [ws_cat.cell(1, c).value for c in range(1, ws_cat.max_column + 1)]
print(f"  Catálogos current headers: {cat_headers}", flush=True)

# Add equipment_name as column 6
ws_cat.insert_cols(6)
ws_cat.cell(1, 6, 'equipment_name')

for r in range(2, ws_cat.max_row + 1):
    pc = ws_cat.cell(r, 1).value
    if pc and pc in profile_to_equipment:
        names = profile_to_equipment[pc]
        ws_cat.cell(r, 6, names[0])

print(f"  Added equipment_name to Catálogos", flush=True)

wb_cat.save('seed_data/15_catalog_profiles.xlsx')
print("  SAVED!", flush=True)

# ============================================================
# VERIFICATION
# ============================================================
print("\n" + "=" * 60, flush=True)
print("VERIFICATION", flush=True)
print("=" * 60, flush=True)

# Verify 03
wb_v = openpyxl.load_workbook('seed_data/03_failure_modes.xlsx', read_only=True)
ws_v = wb_v['failure_modes']
print("\n03_failure_modes.xlsx sample (first 5):", flush=True)
for i, row in enumerate(ws_v.iter_rows(min_row=1, max_row=6, values_only=True)):
    if i == 0:
        print(f"  Headers: {list(row[:5])}", flush=True)
    else:
        print(f"  Row {i+1}: tag={row[0]}, equnr={row[1]}, fl={row[2]}", flush=True)
wb_v.close()

# Verify 15
wb_v2 = openpyxl.load_workbook('seed_data/15_catalog_profiles.xlsx', data_only=True)
ws_v2 = wb_v2['Catálogos']
v2_headers = [ws_v2.cell(1, c).value for c in range(1, ws_v2.max_column + 1)]
print(f"\n15_catalog_profiles Catálogos headers: {v2_headers}", flush=True)
for r in range(2, min(7, ws_v2.max_row + 1)):
    row = [ws_v2.cell(r, c).value for c in range(1, ws_v2.max_column + 1)]
    print(f"  Row {r}: {row}", flush=True)

ws_v3 = wb_v2['Perfil_Catálogo']
v3_headers = [ws_v3.cell(1, c).value for c in range(1, ws_v3.max_column + 1)]
print(f"\n15_catalog_profiles Perfil_Catálogo headers: {v3_headers}", flush=True)
for r in range(2, min(7, ws_v3.max_row + 1)):
    row = [ws_v3.cell(r, c).value for c in range(1, ws_v3.max_column + 1)]
    print(f"  Row {r}: {row}", flush=True)

print("\nDONE!", flush=True)
