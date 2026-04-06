"""Migrate work_orders → managed_work_orders and create work_requests for GOLDFIELDS-SN."""
import sys, uuid, json
from datetime import datetime
sys.path.insert(0, '/app')
from api.database.connection import SessionLocal
from sqlalchemy import text

db = SessionLocal()
PLANT = 'GOLDFIELDS-SN'
now = datetime.utcnow().isoformat()

# ═══ 1. work_orders → managed_work_orders (first 300) ═══
print("=== work_orders → managed_work_orders ===")
wos = db.execute(text(
    "SELECT work_order_id, order_type, equipment_id, equipment_tag, priority, status, created_date, actual_duration_hours, description FROM work_orders LIMIT 300"
)).fetchall()
print(f"  Found {len(wos)} work_orders, migrating...")

mwo_count = 0
for wo in wos:
    wo_id = str(uuid.uuid4())
    wo_num = wo[0] or f"OT-GF-{mwo_count}"
    raw_status = (wo[5] or '').upper()
    if any(x in raw_status for x in ['CECO', 'CLSD', 'TECO', 'CLOSED', 'COMPLETADO']):
        status = 'CERRADO'
    elif any(x in raw_status for x in ['REL', 'RELEASED', 'LIBERADO']):
        status = 'PROGRAMADO'
    elif any(x in raw_status for x in ['CRTD', 'CREATED', 'CREADO']):
        status = 'CREADO'
    elif any(x in raw_status for x in ['EJEC', 'PROGRESS']):
        status = 'EN_EJECUCION'
    elif any(x in raw_status for x in ['PLAN']):
        status = 'PLANIFICADO'
    else:
        status = 'CERRADO'
    prio = wo[4] or 'P3'
    if prio not in ('P1','P2','P3','P4'):
        prio = 'P3'
    created = wo[6] or now
    if hasattr(created, 'isoformat'):
        created = created.isoformat()
    try:
        db.execute(text("""
            INSERT INTO managed_work_orders (wo_id, wo_number, wo_type, description, equipment_tag,
                equipment_id, priority_code, status, plant_id, created_at, actual_hours, planned_by,
                work_class, estimated_hours, completion_pct, budget_approved)
            VALUES (:id, :num, :type, :desc, :tag, :eid, :prio, :status, :plant, :at, :hours, :by,
                :wc, :est, :pct, :ba)
        """), {
            'id': wo_id, 'num': wo_num, 'type': wo[1] or 'PM01',
            'desc': wo[8] or '', 'tag': wo[3] or '', 'eid': wo[2] or '',
            'prio': prio, 'status': status, 'plant': PLANT,
            'at': str(created), 'hours': wo[7] or 0, 'by': 'import', 'wc': 'NO_PROGRAMADO',
            'est': wo[7] or 4.0, 'pct': 100.0 if status == 'CERRADO' else 0.0, 'ba': False,
        })
        mwo_count += 1
    except Exception as e:
        db.rollback()
        if mwo_count == 0:
            print(f"  Error: {str(e)[:150]}")
        continue

db.commit()
print(f"  Migrated {mwo_count}")

# ═══ 2. Create work_requests from Excel 24_notifications ═══
print("\n=== 24_notifications.xlsx → work_requests ===")
try:
    from openpyxl import load_workbook
    wb = load_workbook('/app/seed_data/24_notifications.xlsx', read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h else f'col_{i}' for i, h in enumerate(rows[0])]
    data = rows[1:201]
    wb.close()
    print(f"  Excel: {len(rows)-1} rows, importing 200")
    print(f"  Headers: {headers[:10]}")

    wr_count = 0
    for row in data:
        d = dict(zip(headers, row))
        wr_id = f"WR-GF-{str(uuid.uuid4())[:8]}"
        desc = str(d.get('qmtxt', '') or d.get('short_text', '') or d.get('description', '') or 'Notification')
        tag = str(d.get('sap_func_loc_short', '') or d.get('equipment_tag', '') or d.get('tplnr', '') or '')
        prio_raw = str(d.get('priokx', '') or d.get('priority', '') or '3')
        prio_map = {'1': 'P1', '2': 'P2', '3': 'P3', '4': 'P4', 'Very High': 'P1', 'High': 'P2', 'Medium': 'P3', 'Low': 'P4'}
        prio = prio_map.get(prio_raw, 'P3')
        created = d.get('erdat', '') or d.get('created_date', '') or now
        if hasattr(created, 'isoformat'):
            created = created.isoformat()
        ai_class = json.dumps({"plant_id": PLANT, "source": "excel_import"})
        aviso = str(d.get('qmnum', '') or '')

        try:
            db.execute(text("""
                INSERT INTO work_requests (request_id, status, equipment_tag, equipment_id, equipment_confidence,
                    resolution_method, problem_description,
                    priority_code, created_by, created_at, notification_type, ai_classification, aviso_number)
                VALUES (:id, 'PENDIENTE', :tag, :eid, 0.7, 'AI', :desc, :prio, 'import', :at, 'M1', :ai, :anum)
            """), {
                'id': wr_id, 'tag': tag, 'eid': str(d.get('equnr', '') or tag or 'unknown'), 'desc': desc, 'prio': prio,
                'at': str(created), 'ai': ai_class, 'anum': aviso,
            })
            wr_count += 1
        except Exception as e:
            db.rollback()
            if wr_count == 0:
                print(f"  Error: {str(e)[:150]}")
            continue

    db.commit()
    print(f"  Created {wr_count} work_requests")
except Exception as e:
    print(f"  Error: {e}")

# ═══ Stats ═══
print("\n=== Stats ===")
c1 = db.execute(text("SELECT count(*) FROM managed_work_orders WHERE plant_id = :p"), {'p': PLANT}).scalar()
c2 = db.execute(text("SELECT count(*) FROM work_requests WHERE ai_classification LIKE :p"), {'p': f'%{PLANT}%'}).scalar()
print(f"  managed_work_orders (GF): {c1}")
print(f"  work_requests (GF): {c2}")
print("\n  MWO by status:")
for r in db.execute(text("SELECT status, count(*) FROM managed_work_orders WHERE plant_id = :p GROUP BY status"), {'p': PLANT}).fetchall():
    print(f"    {r[0]}: {r[1]}")
db.close()
print("Done!")
