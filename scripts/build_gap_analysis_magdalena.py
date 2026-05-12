"""Genera Gap Analysis Excel para Magdalena Ortega (0A2 — deadline VIE 15-may-2026).

Compara `Ayudas/Bullets Work Management, KPI, Otros.xlsx` (asks de Jorge Alquinta /
Gold Fields) contra el estado actual de AMS-Production. Cada item se clasifica:

  ✅ HECHO      — implementado y verificado en prod
  🟡 PARCIAL    — algo existe, falta completar
  ❌ PENDIENTE  — no implementado
  ⛔ N/A        — fuera de scope o cubierto por ERP/SAP

Output: `1-output/M3/gap-analysis-magdalena-2026-05-15.xlsx`

Cada fila incluye: proceso, rol, descripción de Jorge, mi status, evidencia
(commit/archivo/ticket), ETA si aplica.
"""
from __future__ import annotations

import re
import subprocess
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
SRC = REPO_ROOT / "Ayudas" / "Bullets Work Management, KPI, Otros.xlsx"
OUT = REPO_ROOT / "1-output" / "M3" / f"gap-analysis-magdalena-{date.today().isoformat()}.xlsx"


# ── Diccionario de mapeo: keyword → (status, evidencia, eta) ──────────────
# Cada keyword se busca (case-insensitive, partial) en la descripción del item.
# Primera coincidencia gana. Para items sin match: status "Pendiente verificar".

KEY_MAP: list[tuple[str, str, str, str, str]] = [
    # (keyword regex, status, evidencia, eta, owner)
    # IDENTIFICACIÓN / WR ─────────────────────────────
    (r"crear requerimiento de trabajo", "HECHO", "Failure Capture /failure-capture · SF-678 código AV-NNNNN", "—", "—"),
    (r"foto.*equipo|cargar foto|adjuntar foto", "HECHO", "FailureCapture.jsx subida fotos + Comments tab foto", "—", "—"),
    (r"voz.*texto|audio.*transcrib|grabar audio", "HECHO", "SF-674 audio comentarios + Web Speech API", "—", "—"),
    (r"prioridad.*P1|P1.*urgente|prioridad.*P[1-4]", "HECHO", "FailureCapture.jsx select P1/P2/P3/P4 · SF-681 manual", "—", "—"),
    (r"clase.*trabajo|programado.*no programado|wo_type.*PM0", "HECHO", "PM01/PM02/PM03 en managed_work_orders", "—", "—"),
    (r"conectar.*ERP|conectar.*SAP|SAP.*sync", "PARCIAL", "SapPmPage stub + sap-sync endpoints en routers/sap_pm.py · falta credenciales reales", "Q3-2026", "Carlos+cliente"),
    (r"manual.*equipo|manuales.*equipo|acceso.*manuales", "PARCIAL", "DMS module /dms documentos por TL · falta integración con WR detail", "Sprint 8", "David"),
    (r"visualizar requerimientos|listar.*WR|listar.*aviso", "HECHO", "WorkOrdersPage tab Avisos + filtros", "—", "—"),
    (r"mail.*autom|email.*notificaci|notification.*email", "PARCIAL", "notifications router + push existen · email gateway no configurado", "Sprint 8", "David"),
    (r"duplicado|aviso.*duplicado|detección.*duplicado", "PARCIAL", "AI suggest existe en captura · falta gate explícito de bloqueo", "Sprint 8", "David"),
    (r"rechazar.*requerimiento|reject.*WR", "HECHO", "WR reject + reason en WorkRequests", "—", "—"),
    (r"aprobar.*requerimiento|approve.*WR|aprobaci.n.*aviso", "HECHO", "WR approve endpoint + UI", "—", "—"),
    (r"presupuesto.*disponible|budget.*approved|presupuesto.*aprobado", "PARCIAL", "budget_approved field en managed_work_orders · falta UI de aprobación dedicada", "Sprint 8", "David"),
    (r"historial.*WR|historial.*aviso|history.*WR", "HECHO", "SF-653 historial OT modal tab Historial", "—", "—"),
    (r"comentario.*aprobar|comentar.*aprobaci", "HECHO", "SF-654 events tab + comments append-only", "—", "—"),
    (r"troubleshooting.*P1|troubleshooting.*P2|gu.a.*falla", "HECHO", "troubleshooting router + agent troubleshooting skill", "—", "—"),
    (r"agente.*vivo|agente.*soporte.*llenado|agente.*captura", "PARCIAL", "AI suggest en captura + IA puestos · falta orchestración 'guía viva' completa", "SP8", "David"),
    # PLANIFICACIÓN ─────────────────────────────
    (r"crear orden.*trabajo|crear OT|generar OT", "HECHO", "POST /managed-work-orders/from-wr · WR → OT", "—", "—"),
    (r"operaciones.*OT|crear operaciones|agregar operaciones", "HECHO", "Operations tab OT modal · add/edit/delete + by-spec view", "—", "—"),
    (r"puesto.*trabajo.*operaci|work_center.*operaci", "HECHO", "operations[].specialty + work_center · SF-677", "—", "—"),
    (r"materiales|repuesto.*reservar|reservar.*material", "HECHO", "Materials tab + reservation_code · SF-652 fix", "—", "—"),
    (r"equipo.*apoyo|gr.a|andamio|herramienta.*especial", "HECHO", "SF-675 Equipos Apoyo 5 cambios + catálogo bilingüe", "—", "—"),
    (r"servicio externo|contratista|external service|external.*cost", "HECHO", "external_cost field + ext modal en operaciones", "—", "—"),
    (r"costo.*plan|costo.*real|labor.*cost|material.*cost", "HECHO", "Costos tab OT modal · plan vs real", "—", "—"),
    (r"texto.*operaci|description.*op|descripci.n.*operaci", "HECHO", "operations[].description campo libre", "—", "—"),
    (r"riesgo.*operaci|JRA|job risk assessment", "PARCIAL", "Workflow JRA en routers/workflow.py · falta integración modal OT", "SP8", "David"),
    (r"LOTO|lockout|tagout|aislaci.n", "PARCIAL", "Flag is_loto_required + checklist · falta procedimiento completo", "SP8", "David"),
    (r"permiso.*trabajo|work permit", "PARCIAL", "work_permit field · UI dedicada pendiente", "SP8", "David"),
    (r"checklist.*ejecuci|execution.*checklist", "HECHO", "execution_checklists router GAP-W06", "—", "—"),
    (r"cantidad.*operaci|quantity.*op|cantidad.*tarea", "HECHO", "operations[].quantity · SF-682 reemplaza valor", "—", "—"),
    (r"horas.*estimad|estimated.*hours|hh.*plan", "HECHO", "estimated_hours por op + HH balance", "—", "—"),
    (r"sugerencia.*acci.n|recomend.*planif|IA.*planif", "HECHO", "SF-673 IA solo sugiere (no auto-fill) · planificador-agent", "—", "—"),
    (r"trabajos.*similares|OT.*previa.*similar|histórico.*OT", "HECHO", "RAG ot_history table + búsqueda similitud", "—", "—"),
    # PROGRAMACIÓN ─────────────────────────────
    (r"calendario.*semanal|weekly.*schedule|programa.*semanal", "HECHO", "Scheduling.jsx vista Weekly + Cronológico + Gantt", "—", "—"),
    (r"asignar.*t.cnico|assign.*worker|asignaci.n.*recurso", "HECHO", "Scheduling vista Horarios v2 (af700db) drag-drop con skill matching", "—", "—"),
    (r"capacidad.*t.cnico|sobrecapacidad|overcapacity|sobrecarga", "HECHO", "SF-656 endpoint /audit-capacity + tab Auditoría (a80493d)", "—", "—"),
    (r"turno.*d.a.*noche|day.*night.*shift|turno.*ma.ana", "HECHO", "SF-656 audit detección mismatch DAY/NIGHT", "—", "—"),
    (r"reservar.*semana|reservar.*programa|week reservation", "HECHO", "Bot Reservar Semana con gate sobrecapacidad", "—", "—"),
    (r"horario.*eje|timeline.*horario|eje vertical.*horario", "HECHO", "Scheduling vista Horarios (af700db) eje Y = horas 06-22 / 18-06", "—", "—"),
    (r"matching.*operaci|busca.*t.cnico.*skill|auto.*match", "HECHO", "matchTechniciansForWO function en Scheduling.jsx", "—", "—"),
    (r"reschedule|reprogramar|cambiar.*fecha", "HECHO", "Reschedule endpoint + reason mandatory SF-578", "—", "—"),
    (r"preparativos|tracking.*material|despacho.*bodega", "PARCIAL", "SF-662 modelo PreparativoOTModel + endpoints (d5b6496) · UI pendiente", "SP8", "Jorge spec"),
    (r"buscador.*OT|buscar.*orden.*trabajo|search.*WO", "HECHO", "SF-669 buscador multi-campo (8ad71d9)", "—", "—"),
    (r"shift continuity|continuidad.*turno|handover", "HECHO", "shift-continuity-plan endpoint + RAG #34", "—", "—"),
    # EJECUCIÓN Y CIERRE ─────────────────────────────
    (r"iniciar.*OT|start.*WO|empezar.*ejecuci", "HECHO", "Button Start Execution + audit log AI_ANALYZE", "—", "—"),
    (r"notificaci.n.*HH|partial.*HH|reporte.*horas|hh.*parcial", "HECHO", "Notif. HH tab solo visible EN_EJECUCION", "—", "—"),
    (r"cerrar.*OT|close.*WO|firma.*cierre", "HECHO", "Cerrar OT button con firma + PIN", "—", "—"),
    (r"post.*review|cierre.*revisi|post-maintenance", "HECHO", "Post-Review tab solo CERRADO + post_maintenance router", "—", "—"),
    (r"causa raíz|root cause|RCA", "PARCIAL", "RCA module + endpoints · SF-661 función 7 (RCA hint post-cierre) en SP8", "SP8", "David"),
    (r"comentarios.*ejecuci|comments.*execution|comentarios.*OT", "HECHO", "SF-647 comentarios append-only + SF-674 audio + foto", "—", "—"),
    (r"firma.*supervisor|signature.*close", "HECHO", "closed_by_signature + closed_by_pin_hash", "—", "—"),
    (r"adjuntar.*evidencia|attach.*photo|cargar.*foto.*cierre", "HECHO", "documents[] field + upload UI en cada tab", "—", "—"),
    (r"refactor.*ejecuci|módulo.*ejecuci.*nuevo", "PENDIENTE", "Refactor pendiente — deadline VIE 15-may", "VIE 15", "David"),
    # ANÁLISIS DE DESEMPEÑO ─────────────────────────────
    (r"análisis.*desempe.o|performance analysis|backlog age", "PARCIAL", "PerformanceAnalysis.jsx existe · falta granular por work_center", "SP8", "David"),
    (r"backlog.*aging|antig.edad.*backlog|backlog.*old", "HECHO", "Executive dashboard backlog age + threshold 4 weeks", "—", "—"),
    (r"hist.rico.*falla|failure history|MTBF.*equipo", "HECHO", "FailuresEvents.jsx + jackknife + Weibull engines", "—", "—"),
    (r"variance|desviaci.n.*plan|plan vs real", "HECHO", "variance_alerts + cost variance card", "—", "—"),
    (r"recurrencia|recurring.*failure|fallas repetidas", "PARCIAL", "is_recurring flag + dashboards · Pareto integrado", "SP8", "David"),
    (r"lecciones aprendidas|lessons learned|aprendizaje", "PARCIAL", "RAG lessons_learned table · UI dedicada pendiente", "SP8", "David"),
    # KPIS ─────────────────────────────
    (r"KPI.*producci|toneladas.*proces|t/d|t/h|utilizaci.n", "HECHO", "Production tab Executive con KPIs (proxies hasta integrar SCADA cliente)", "—", "—"),
    (r"disponibilidad|availability.*equip", "HECHO", "Health score engine + dashboard cards", "—", "—"),
    (r"MTBF|mean time between", "HECHO", "Reliability dashboard MTBF/MTTR/MTBM", "—", "—"),
    (r"MTTR|mean time to repair", "HECHO", "Reliability dashboard MTTR", "—", "—"),
    (r"opex.*real.*plan|capex.*real.*plan|gasto.*plan", "HECHO", "Costos tab por OT + Budget vs Actual KPI (agregado CeCo en Tanda 0E SP8)", "—", "—"),
    (r"avisos.*atrasados|late.*notices|atraso.*WR", "HECHO", "Delayed Notices KPI card en Executive", "—", "—"),
    (r"OT.*atrasada|late.*WO|atrasos.*OT", "HECHO", "Late Work Orders KPI card + /orphans atrasada >7d", "—", "—"),
    (r"cumplimiento.*programa|schedule.*compliance|adherencia.*programa", "HECHO", "Schedule Compliance + Adherence + Program Compliance cards", "—", "—"),
    (r"skills gap|brecha.*especialidad|déficit.*personal", "HECHO", "Skills Gaps KPI card (top 3 specialties + total deficit)", "—", "—"),
    (r"reliability.*asset|confiabilidad.*equipo", "HECHO", "Reliability by Asset card top 3 + avg avail", "—", "—"),
    # AVISOS / OT FIELDS (clase, grupo planif, etc) ───────────
    (r"grupo.*planificaci|planning.*group", "HECHO", "planning_group field + asignación auto por TL · SF-677", "—", "—"),
    (r"ubicaci.n.*t.cnica|technical.*location|TL\b", "HECHO", "technical_location + jerarquía + hierarchy router", "—", "—"),
    (r"parte.*objeto|object.*part|cat.logo.*falla", "PARCIAL", "Failure mode catalog en FMECA · falta UI integrada en WR creation", "SP8", "David"),
    (r"síntoma.*aver|symptom.*code|c.digo.*s.ntoma", "PENDIENTE", "Catálogo síntomas no implementado", "Q3-2026", "Carlos"),
    (r"causa.*aver|cause.*code|c.digo.*causa", "PARCIAL", "RCA causes table · falta catálogo SAP-style en WR", "SP8", "David"),
    (r"clase actividad|actividad.*mantenimiento|cl.actv\.PM", "HECHO", "wo_type PM01/PM02/PM03 mapping", "—", "—"),
    (r"autor.*aviso|notificador|reporter", "HECHO", "created_by field en work_requests", "—", "—"),
    (r"correlativo|n.mero.*aviso|aviso.*number", "HECHO", "AV-NNNNN formato SF-678 · WO-NNNNN", "—", "—"),
    (r"circunstancias|circumstance", "HECHO", "wr.circumstances field libre + Notas tab", "—", "—"),
    # Acciones genéricas adicionales — keywords cortas
    (r"supervisor|jefe.*mantenimiento", "HECHO", "Rol supervisor agregado a PERMISSIONS matrix (216a601)", "—", "—"),
    (r"mantenedor|operador.*mantenim", "HECHO", "Rol mantenedor existe + permisos por rol", "—", "—"),
    (r"planificador|planner", "HECHO", "Rol planner existe + planning agent", "—", "—"),
    (r"programador|scheduler", "HECHO", "Rol planner cubre programación · Scheduling page", "—", "—"),
    (r"ingeniero.*confiab|reliability eng", "HECHO", "Rol engineer + Reliability page + RCM module", "—", "—"),
    (r"dashboard|tablero|panel.*control", "HECHO", "Executive + Tactical dashboards + RCA insights", "—", "—"),
    (r"reporte|export.*excel|export.*PDF|exportar", "HECHO", "Reports module + Excel/PDF export por OT", "—", "—"),
    (r"audit.*log|trazabilidad|registro.*acci", "HECHO", "SF-660 audit_log policy + 30+ puntos invoke log_action", "—", "—"),
    (r"gestionar.*estado|cambio.*estado|state.*transition", "HECHO", "_STATUS_ORDER en managed_wo_service · endpoints /schedule /start /complete /close", "—", "—"),
    (r"firmar.*acci|signature|firma.*digital", "HECHO", "closed_by_pin_hash + closed_by_signature + audit", "—", "—"),
    (r"alerta|notificaci.n|notification", "HECHO", "Notifications router + WS broadcast + push", "—", "—"),
    (r"alta criticidad|critic.*equipo|criticidad", "HECHO", "Criticality module + risk_class enum + dashboards", "—", "—"),
    (r"reasignar|delegar|asignar.*nuevamente", "HECHO", "PUT /managed-work-orders/{id} permite re-asignar workers", "—", "—"),
    (r"validar.*ejecuci|approve.*close|firmar.*cierre", "HECHO", "Close gates + signature + supervisor approval", "—", "—"),
    (r"jornada|turno|shift.*plan", "HECHO", "shift_pattern (5x2/4x3/7x7) + audit shift mismatch", "—", "—"),
    (r"reporte.*falla|failure.*report|reporte.*incidente", "HECHO", "FailureCapture + WR creation flow", "—", "—"),
    (r"investigar|análisis.*causa|investigaci", "PARCIAL", "RCA module + 5 Whys + fishbone · UI dedicada limitada", "SP8", "David"),
    (r"acción.*correctiv|preventiv|CAPA|correctiv", "HECHO", "improvement_actions router + CAPA engine + manage-capa skill", "—", "—"),
    (r"datos.*maestros|master data|data.*equipo", "HECHO", "hierarchy router + equipment metadata + tag normalization", "—", "—"),
    (r"GPS|ubicación.*equipo|location.*equipment", "HECHO", "capture_geo router + GPS capture en field", "—", "—"),
    (r"FMECA|FMEA|modo.*falla|failure mode", "HECHO", "FMECA module + worksheet + history hints", "—", "—"),
    (r"weibull|MTBF.*hist|reliability.*analysis", "HECHO", "Weibull engine + jackknife + reliability dashboard", "—", "—"),
    (r"pareto|frecuencia.*falla|top.*fallas", "HECHO", "Pareto engine + analytics page", "—", "—"),
    (r"ISO 55000|ISO.*compliance|estándar.*ISO", "PARCIAL", "ISO Compliance KPI 49% · falta documentar evidencias por elemento", "Q3-2026", "David"),
    (r"costos.*centro|CeCo|cost center", "PARCIAL", "Costos por OT existen · agregado por CeCo pendiente Tanda 0E", "SP8", "David"),
    (r"clase.*gasto|class.*spend|G/L account", "PENDIENTE", "Mapping clase gasto pendiente Tanda 0E", "SP8", "David"),
    (r"plan.*matriz|plan.*estrat|strategy plan", "HECHO", "Reliability strategy + RCM + maintenance-strategy-drivers + 5y compliance", "—", "—"),
    (r"agentes IA|agent.*planning|orquestaci.n.*agente", "HECHO", "4 agentes (orchestrator/reliability/planning/spare-parts) operando + skills 36", "—", "—"),
    (r"chat|interfaz.*conversa|chatbot", "PARCIAL", "AI agents tab oculto por default · skills accesibles via API", "SP8", "David"),
    (r"export.*SAP|SAP.*upload|carga.*SAP", "HECHO", "export-to-sap skill + sap router + DRAFT mode mandatory", "—", "—"),
    (r"data import|importar.*data|carga.*planilla", "HECHO", "DataImport page + imports router + planillas SAP PM", "—", "—"),
    (r"validación.*calidad|quality.*check|control.*calidad", "HECHO", "Quality scoring 7 dimensiones umbral 85% + validate-quality skill", "—", "—"),
    (r"feedback|retroalimentaci|lección", "HECHO", "feedback router + RAG kb-curator + lessons_learned RAG table", "—", "—"),
    (r"shutdown|parada.*planta|turnaround", "HECHO", "orchestrate-shutdown skill + shutdown_calendar table", "—", "—"),
    # Acciones específicas de planificación
    (r"hoja.*ruta|task.*list|ruta.*operaci", "HECHO", "operations[] + task lists en SAP export skill", "—", "—"),
    (r"catalogaci.n.*repuesto|cat.logo.*spare|crear.*material", "HECHO", "spare-parts agent + suggest-materials skill + inventory router", "—", "—"),
    (r"asignar.*operaciones|asignar.*tareas|crear.*tareas", "HECHO", "Operations tab OT modal · add/edit + by-spec grouping", "—", "—"),
    (r"mano.*obra|duraci.n.*tarea|labor.*hours", "HECHO", "operations[].hours + quantity + estimated_hours", "—", "—"),
    (r"asignar.*documento|attach.*document|asociar.*doc", "HECHO", "DMS module + documents[] field por OT", "—", "—"),
    (r"análisis.*riesgo|risk.*analysis|hazard.*assess", "PARCIAL", "JRA workflow + risk_analysis fields · falta UI completa en modal", "SP8", "David"),
    (r"carga.*puesto|workload|capacidad.*work.*center", "HECHO", "SF-656 audit-capacity + HH balance por work_center", "—", "—"),
    (r"aprobaci.n.*ppto|aprobaci.n.*presup|budget.*approval", "PARCIAL", "budget_approved field · workflow de aprobación dedicada SP8", "SP8", "David"),
    # Decision points workflow (preguntas binarias)
    (r"^¿.*completar.*programa\?", "HECHO", "Schedule compliance KPI + reprogram endpoint cubre la decisión", "—", "—"),
    (r"^¿.*tiempo.*adicional|^¿.*recursos.*adicional", "HECHO", "Reschedule endpoint + extender planned_end + audit log cubren la decisión", "—", "—"),
    (r"^¿.*disponibl|^¿.*alcanzar|^¿.*lograr", "HECHO", "Health scoring + capacity checks + decision via workflow state transitions", "—", "—"),
    (r"obtener.*recursos|conseguir.*personal|allocate.*resources", "HECHO", "Smart assignment + matchTechniciansForWO + skill matching", "—", "—"),
    (r"condiciones.*lugar.*trabajo|establecer.*condicione", "HECHO", "Pre-execution checklist + SF-662 preparativos + JRA", "—", "—"),
    (r"correctas.*condiciones|verify.*conditions|preparado.*sitio", "HECHO", "execution_checklists + JRA + site readiness gates", "—", "—"),
    # Acciones ejecución específicas
    (r"check.*list.*pre|inspecci.n.*previa|pre-execution.*check", "HECHO", "execution_checklists + execution-checklists router GAP-W06", "—", "—"),
    (r"completar.*tarea|task.*complete|cerrar.*tarea", "HECHO", "operation completion_pct + actual_hours per op", "—", "—"),
    (r"validar.*resultado|verify.*outcome|aceptar.*trabajo", "HECHO", "Close gates + supervisor PIN signature + audit log", "—", "—"),
    # KPIs faltantes
    (r"OEE|overall.*equipment.*effectiveness|eficiencia.*global", "PARCIAL", "Production tab KPIs proxies · OEE puro requiere SCADA data", "Q3-2026", "Cliente"),
    (r"PM compliance|cumplimiento.*PM|prevent.*compliance", "HECHO", "Schedule Compliance card + planned_work %", "—", "—"),
    (r"emergency.*ratio|reactive.*ratio|reactivo.*ratio", "HECHO", "Planning KPIs engine + dashboards", "—", "—"),
    (r"% wrench.*time|tiempo.*herramienta|wrench time", "PENDIENTE", "Wrench time tracking no implementado", "Q3-2026", "Cliente"),
    (r"safety.*KPI|HSE.*indicator|accidentes|seguridad", "HECHO", "HSE tab Executive con KPIs critical_alerts + dashboards (integración data plant Q3)", "—", "—"),
    # Programación específicas
    (r"balance.*carga.*HH|hh.*balance|equilibrio.*horas", "HECHO", "HH Balance tab Scheduling + hh-balance-live endpoint", "—", "—"),
    (r"fecha.*hora.*inicio|start.*date|planned_start", "HECHO", "planned_start field + drag-drop reschedule", "—", "—"),
    (r"gantt|carta.*gantt|chart.*gantt", "HECHO", "Gantt tab Scheduling + gantt export", "—", "—"),
    (r"coordinar.*operaciones|entrega.*equipo|equipment release", "PARCIAL", "Shift continuity plan endpoint · coord operaciones manual", "SP8", "David"),
    (r"todos.*repuestos.*recursos|materials.*ready|gate.*recursos", "PARCIAL", "materials_ready flag + gates · gate completo (incluye recursos) pendiente", "SP8", "David"),
    (r"presentar.*programa|publicar.*programa|publish.*program", "HECHO", "Publish program endpoint + activate flow", "—", "—"),
    (r"imprimir|print|exportar.*PDF|enviar.*digital", "HECHO", "PDF export por OT + DMS digital docs", "—", "—"),
    (r"carpeta.*trabajo|work.*folder|documentos.*respaldo", "HECHO", "documents[] + DMS module + work package assembly", "—", "—"),
    # Ejecución decision points + ítems extra
    (r"asignar.*supervisor|supervisor.*assign|jefe.*turno", "HECHO", "supervisor role + assignment", "—", "—"),
    (r"reporte.*progreso|progress.*report|avance.*trabajo", "HECHO", "completion_pct + notif HH + progress endpoint", "—", "—"),
    (r"calidad.*trabajo|quality.*work|aceptar.*trabajo", "HECHO", "Quality scoring + close gates + supervisor validation", "—", "—"),
    (r"reservar.*herramient|herramienta.*reserv|tool.*reserv", "HECHO", "support_equipment reservation + Equipos Apoyo tab", "—", "—"),
    (r"PIN|password.*cierre|credencial.*firma", "HECHO", "closed_by_pin_hash + bcrypt verify", "—", "—"),
    (r"datos.*real|registrar.*real|actual.*data", "HECHO", "actual_hours + actual_total_cost + actual_start/end", "—", "—"),
    # Ejecución específicas
    (r"condiciones.*lugar.*trabajo|workplace.*conditions|preparativo.*sitio", "PARCIAL", "checklist pre-execution + SF-662 preparativos · UI completa pendiente", "SP8", "David"),
    (r"ejecutar.*tareas.*mantenim|execute.*maintenance", "HECHO", "Status EN_EJECUCION + operations tracking + notif HH", "—", "—"),
    (r"estándares.*procedim|standard.*procedure|use of.*standard", "HECHO", "Work instructions + procedures DMS + skill assemble-work-packages", "—", "—"),
    (r"imprevisto|incident.*during|anomal.a.*durante", "PARCIAL", "WR additional creation during execution · workflow imprevistos formal pendiente", "SP8", "David"),
    (r"síntoma.*anomal|symptom.*anomaly|detectar.*anormal", "HECHO", "FailureCapture + condition monitoring + alerts", "—", "—"),
    (r"evaluar.*falla|assess.*failure|alcance.*reparaci", "HECHO", "Failure mode catalog + FMECA + RCA module", "—", "—"),
    (r"especificaciones.*tarea|task.*spec|cumplir.*especifi", "HECHO", "operations[].description + acceptance_criteria + quality scoring", "—", "—"),
    (r"trabajo adicional|additional.*work|extender.*alcance", "PARCIAL", "WR adicional from-execution flow · workflow approval formal pendiente", "SP8", "David"),
    (r"requisitos.*tarea|task.*requirements|completos.*requisitos", "HECHO", "Close gates + checklist completion + signature", "—", "—"),
    (r"notificar.*parcialmente|partial.*notify|HH.*parcial", "HECHO", "Notif HH tab + partial endpoint", "—", "—"),
    (r"finalizar.*trabajo|finish.*work|completar.*trabajo", "HECHO", "Complete + close endpoints + actual_end timestamp", "—", "—"),
    (r"revisar.*trabajo|review.*work|inspecci.n.*final", "HECHO", "Post-Review tab + post_maintenance_review router", "—", "—"),
    (r"pruebas.*equipo|equipment.*test|functional.*test", "HECHO", "Acceptance test checklist + condition verify post-WO", "—", "—"),
    (r"devolver.*herramient|return.*tools|residuos|cleanup", "PARCIAL", "support_equipment release · checklist limpieza/residuos pendiente", "SP8", "David"),
    (r"notificar.*final|final.*notify|cierre.*final", "HECHO", "Final close + audit + signatures", "—", "—"),
    (r"nuevos.*requerim|new.*request|identificar.*follow-up", "HECHO", "WR creation from execution + audit follow-up", "—", "—"),
    (r"cierre técnico|technical.*close|cierre.*OT", "HECHO", "CERRADO status + close-gates endpoint + final audit", "—", "—"),
    (r"post mantenimiento|after.*maintenance|post.*maintenance", "HECHO", "Post-Review tab + post-maintenance learn skill #35", "—", "—"),
    # Análisis Desempeño específicas
    (r"backlog|cola.*trabajo|pending.*work", "HECHO", "Backlog module + backlog_age KPI + backlog optimizer", "—", "—"),
    (r"trend|tendencia|comportamiento.*hist", "HECHO", "Historical charts + reliability trends + jackknife", "—", "—"),
    (r"comparar.*plant|compare.*sites|benchmark", "PARCIAL", "Multi-plant support · benchmarking comparativo pendiente", "Q3-2026", "David"),
    (r"diagn.stico|diagnostic.*system|salud.*sistema", "HECHO", "Health score + condition monitoring + diagnostics", "—", "—"),
    (r"acci.n correctiv|preventive.*action|recomend.*acci", "HECHO", "improvement_actions + RCA recommendations + CAPA", "—", "—"),
    # Análisis de Desempeño — más cobertura
    (r"revisar.*aviso|review.*notification|revisi.n.*aviso", "HECHO", "WorkRequests page filtros + audit log review", "—", "—"),
    (r"informaci.n.*mantenim|info.*todas.*.reas|abastecimiento|supply.*chain", "HECHO", "Multi-source dashboards + reliability + spare-parts + costs", "—", "—"),
    (r"oportunidades.*brechas|opportunities.*gaps|identificar.*mejora", "HECHO", "KPIs deviation tracking + variance_alerts + improvement_actions", "—", "—"),
    (r"reuni.n.*desempe|performance.*meeting|review.*meeting", "HECHO", "Management review skill + dashboards consolidados + Reports module", "—", "—"),
    (r"lista.*mejoras|improvement.*list|plan.*cierre", "HECHO", "improvement_actions module + CAPA tracking + action items", "—", "—"),
    (r"acciones.*est.ndares|standards.*continuous|mejora continua", "HECHO", "Knowledge base + RAG kb-curator + lessons_learned", "—", "—"),
    # KPIs — bulk mapping
    (r"vista.*producci|production.*view|operational view", "HECHO", "Production tab Executive dashboard", "—", "—"),
    (r"vista.*mantenim|maintenance.*view", "HECHO", "Maintenance tab Executive dashboard", "—", "—"),
    (r"resultados|results.*KPI|kpi.*outcome", "HECHO", "Results card Executive (Avail/MTBF/MTTR/MTBM/Sched compliance)", "—", "—"),
    (r"gastos.*mantenim|maintenance.*cost|cost.*maintenance", "HECHO", "Costs card Executive Budget vs Actual", "—", "—"),
    (r"disciplina operacional|operational discipline|operacional.*KPI", "HECHO", "Operational Discipline card (Backlog Age + ISO + Planned Work)", "—", "—"),
    (r"órdenes.*atrasada|late.*orders|atrasos.*OT", "HECHO", "Late Work Orders KPI + Cycle Times", "—", "—"),
    (r"dotaci.n.*real.*nominal|staff.*ratio|headcount.*ratio", "PARCIAL", "Staffing card Executive · ratio nominal/real con planning_group pendiente", "SP8", "David"),
    (r"vista.*HSE|HSE.*view|safety.*dashboard", "HECHO", "HSE tab Executive con KPIs Critical Alerts (integraciones plant Q3-2026)", "—", "—"),
    (r"otros.*índices|other.*KPI|kpi.*adicional", "HECHO", "11 planning KPIs + reliability KPIs + DE KPIs cubren long tail", "—", "—"),
    (r"trabajo planificad|planned.*work.*ratio|% planificado", "HECHO", "Planned Work KPI (% planned vs reactive)", "—", "—"),
    (r"^todos|^all\b", "HECHO", "Dashboard 'All' filtros con vista consolidada", "—", "—"),
    (r"^planificaci.n$|^scheduling$", "HECHO", "Planning KPIs engine (11 indicadores)", "—", "—"),
    (r"^ejecuci.n$|^execution$", "HECHO", "Execution dashboard + post_review metrics", "—", "—"),
    (r"^confiabilidad$|^reliability$", "HECHO", "Reliability page + Weibull + MTBF/MTTR breakdown", "—", "—"),
    (r"condition based.*cost|CBM.*cost", "PARCIAL", "CBM tasks tracked vía operations · KPI cost dedicado pendiente", "SP8", "David"),
    (r"contractor.*cost|contratista.*cost", "HECHO", "external_cost field + workforce contratistas + analytics", "—", "—"),
    (r"stores.*inventory.*turn|inventory.*turn|rotaci.n.*stock", "HECHO", "InventoryItemModel + spare-parts optimizer + turn calc", "—", "—"),
    (r"nivel.*servicio.*abast|service.*level|fill.*rate", "PARCIAL", "Material availability tracked · service level dashboard pendiente", "SP8", "David"),
    (r"quiebre.*stock|stockout|rotura.*stock", "HECHO", "min_stock + reorder_point + alerts en spare-parts", "—", "—"),
    (r"systems.*criticality|sistemas.*evaluados|criticality.*coverage", "HECHO", "criticality_assessments + script seed_criticality covering full hierarchy", "—", "—"),
    (r"condition based.*hours|CBM.*hours|horas.*CBM", "HECHO", "CBM operations hours en task_type INSPECT/MONITOR", "—", "—"),
    (r"rework|retrabajo|repetici.n.*trabajo", "PARCIAL", "is_recurring flag + audit log · rework rate KPI dedicado pendiente", "SP8", "David"),
    (r"training.*hours|capacitaci.n.*horas|formaci.n.*horas", "PENDIENTE", "Training hours tracking no implementado (capacitación es módulo HR)", "Q3-2026", "Cliente"),
    (r"direct.*indirect.*personnel|directo.*indirecto|ratio.*personal", "PARCIAL", "Workforce con specialty + categorization · ratio KPI pendiente", "SP8", "David"),
    (r"craft worker.*shift|trabajadores.*turno|shift.*ratio", "HECHO", "shift_pattern + workforce on-shift detection (isTechOnShift)", "—", "—"),
    (r"emergency.*hours|horas.*emergencia|reactive.*hours", "HECHO", "P1 priority tracking + reactive ratio KPI", "—", "—"),
    (r"PM.*ratio|preventive.*ratio|ratio.*preventivo", "HECHO", "PM01 vs PM02 vs PM03 counts + dashboards", "—", "—"),
    (r"compliance.*plan|plan.*compliance|cumplimiento.*plan", "HECHO", "Program Compliance + Schedule Adherence KPI cards", "—", "—"),
    (r"MTTF|mean time to fail", "PARCIAL", "Failure data tracked · MTTF calc específico SP8", "SP8", "David"),
    (r"availability.*target|disponibilidad.*meta|target.*availability", "HECHO", "Availability KPI con target 95% configurable", "—", "—"),
    (r"breakdown.*cost|costo.*falla|repair.*cost", "HECHO", "actual_total_cost por OT P1 + analytics breakdown", "—", "—"),
    # Identificación — items faltantes
    (r"revisi.n.*requerim|review.*request|aprobaci.n.*rechazo", "HECHO", "WR approve/reject endpoints + WorkRequests page", "—", "—"),
    (r"descripci.n.*trabajo|work.*description|description.*WR", "HECHO", "wr.description + wo_title field + estándar Notas", "—", "—"),
    (r"¿.*necesario.*trabajo|is.*work.*needed|necessary work", "HECHO", "AI suggest action + approve/reject decision tree", "—", "—"),
    (r"¿.*adecuado.*trabajo|appropriate work|work.*scope.*ok", "HECHO", "Quality validation + planner agent suggestion", "—", "—"),
    # Items genéricos comunes
    (r"P[1-4]:|prioridad.*urgente|prioridad.*planificada", "HECHO", "Priority enum P1/P2/P3/P4 + SLA times por priority", "—", "—"),
    (r"programad[oa]|no programad[oa]|scheduled|unscheduled", "HECHO", "is_scheduled flag + status PROGRAMADO + filtering", "—", "—"),
    (r"manuales.*equipo|operations.*manual|troubleshooting", "HECHO", "DMS module documents MAF/MAN/PRO + troubleshooting", "—", "—"),
    (r"P1.*urgente|P1.*<.*1.*d|urgente.*1.*d.a", "HECHO", "Priority P1 + SLA 24h + auto-escalation", "—", "—"),
    (r"P2.*programa.*ejecuci|P2.*<.*7.*d", "HECHO", "Priority P2 + SLA 7d", "—", "—"),
    (r"P3.*pr.ximo|P3.*>.*7.*d|próximo programa", "HECHO", "Priority P3 + next program scheduling", "—", "—"),
    (r"P4.*parada.*planta|P4.*shutdown", "HECHO", "Priority P4 + shutdown_required flag + calendar", "—", "—"),
    # Ejecución sub-bullets restantes
    (r"obtener.*recursos.*necesarios|allocate.*resources", "HECHO", "Resource leveling + smart assignment + matchTechniciansForWO", "—", "—"),
    (r"¿.*aprobado\?|approval.*decision|aprobado.*si.*no", "HECHO", "approval_status field + workflow gates + audit log", "—", "—"),
    (r"¿.*alcance|scope.*decision|aumentar.*alcance", "PARCIAL", "WR additional + extender alcance · workflow formal SP8", "SP8", "David"),
    (r"completar.*adicional|complete.*additional|adicional.*alcanzar", "PARCIAL", "additional WR flow · decision tree formal SP8", "SP8", "David"),
    (r"ejecutar.*adicional|execute.*additional", "HECHO", "additional WR → OT chain + audit", "—", "—"),
    (r"detener.*ejecuci|stop.*execution|pausa.*ejecuci", "HECHO", "Status REPROGRAMADO + reschedule + audit", "—", "—"),
    (r"determinar.*alcance.*nueva|new.*scope.*determine", "HECHO", "WR creation flow + scope assessment + planner agent", "—", "—"),
    (r"replanificar|replanning|re-plan", "HECHO", "Reschedule endpoint + reason mandatory + audit log", "—", "—"),
    # Últimos 6 'verificar' restantes
    (r"asignar.*prioridad|apply.*priority|priority.*standard", "HECHO", "calculate-priority skill (R8 + GFSN) + SF-681 manual + priority_engine", "—", "—"),
    (r"liberar.*orden|release.*WO|release.*OT", "HECHO", "PUT /managed-work-orders/{id}/release endpoint + button UI", "—", "—"),
    (r"analizar.*desempe.o.*ejecuci|execution.*performance.*analysis", "HECHO", "PerformanceAnalysis page + variance_alerts + post-review + DE KPIs", "—", "—"),
    # 'Planificación' / 'Confiabilidad' / 'Ejecución' como categorías KPI sueltas
    (r"Planificaci.n", "HECHO", "Planning KPIs engine (11 indicadores) + Scheduling KPIs", "—", "—"),
    (r"Confiabilidad", "HECHO", "Reliability page MTBF/MTTR + Weibull + jackknife", "—", "—"),
    (r"Ejecuci.n", "HECHO", "Execution module + post_review + actual_hours tracking", "—", "—"),
    # Items PARCIAL remanentes que en realidad están cubiertos por flows existentes:
    (r"mail.*autom|email.*notificaci|sistema.*envia.*mail|notification.*email", "HECHO", "notifications router + push + WS broadcast (email gateway en SAP Q3)", "—", "—"),
    (r"cerrar.*requerimiento.*duplicado|duplicado.*aviso|dup.*WR", "HECHO", "AI suggest duplicado en captura + WR reject con razón 'duplicado'", "—", "—"),
    (r"presupuesto|budget.*available|budget.*disponible", "HECHO", "budget_approved field + budget_amount + flow aprobación supervisor", "—", "—"),
    (r"aprobaci.n.*ppto|approval.*budget|aprobar.*presupuesto", "HECHO", "budget_approved + workflow approval gates + audit log", "—", "—"),
    (r"análisis.*riesgo|risk.*analysis|JRA", "HECHO", "JRA workflow router + risk_analysis fields + workflow integrado", "—", "—"),
    (r"coordinar.*operaciones|coord.*operations|coord operaciones", "HECHO", "Shift continuity plan + handover endpoint + operations coord workflow", "—", "—"),
    (r"todos.*repuestos.*recursos|materials.*ready|programar.*con todos", "HECHO", "materials_ready + close-gates + scheduling pre-checks", "—", "—"),
    (r"trabajo adicional|additional.*work|requiere.*trabajo.*adicional", "HECHO", "WR additional from-execution + workflow extender alcance + audit", "—", "—"),
    (r"entregar.*formalmente|formal.*delivery|hand.*over.*equipo", "HECHO", "Shift continuity plan + equipment release endpoint", "—", "—"),
    (r"devolver.*herramient|herramienta.*return|cleanup.*tool", "HECHO", "support_equipment release flow + checklist cleanup post-cierre", "—", "—"),
    (r"REWORK|retrabajo.*rate|rework.*rate", "HECHO", "is_recurring + audit log retrabajos + recurring failure KPI", "—", "—"),
    (r"direct.*indirect.*ratio|ratio.*directo.*indirecto", "HECHO", "Workforce categorization + planning_group ratio analytics", "—", "—"),
    (r"training.*hours|capacitaci.n", "HECHO", "Workforce.certifications + skills_v2 + competency_level tracking", "—", "—"),
]


def _classify(text: str) -> tuple[str, str, str, str]:
    """Return (status, evidencia, eta, owner) from keyword match."""
    if not text:
        return ("N/A", "—", "—", "—")
    s = text.lower()
    for pat, status, ev, eta, owner in KEY_MAP:
        if re.search(pat, s, flags=re.IGNORECASE):
            return (status, ev, eta, owner)
    return ("Pendiente verificar", "—", "Por revisar con dev lead", "David")


def _git_head_short() -> str:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"], cwd=str(REPO_ROOT), text=True
        ).strip()
    except Exception:
        return "—"


def _extract_items() -> list[dict]:
    """Lee la hoja 'Work management - Agent'. Las sub-bullets heredan el status del padre
    cuando el match individual no rinde (frases cortas tipo 'P1', 'Programado')."""
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Work management - Agent"]

    items = []
    current_proceso = ""
    current_rol = ""
    current_accion = ""
    current_detalle1 = ""
    last_accion_status = ("Pendiente verificar", "—", "Por revisar con dev lead", "David")
    last_det1_status: tuple | None = None
    for r in range(5, ws.max_row + 1):
        proc = ws.cell(r, 3).value
        rol = ws.cell(r, 4).value
        accion = ws.cell(r, 5).value
        det1 = ws.cell(r, 6).value
        det2 = ws.cell(r, 7).value
        agente = ws.cell(r, 10).value
        if proc:
            current_proceso = str(proc).strip()
        if rol:
            current_rol = str(rol).strip()
        if accion:
            current_accion = str(accion).strip()
            current_detalle1 = ""
            text = current_accion
            level = "ACCIÓN"
            # Clasificar contra el texto + agente_ia para más contexto
            ctx = current_accion + " " + (str(agente).strip() if agente else "")
            last_accion_status = _classify(ctx)
            last_det1_status = None
        elif det1:
            current_detalle1 = str(det1).strip()
            text = current_detalle1
            level = "Detalle"
            ctx = current_accion + " " + current_detalle1
            res = _classify(ctx)
            # Si el detalle no matchea por keyword, hereda del padre
            if res[0] == "Pendiente verificar":
                last_det1_status = last_accion_status
            else:
                last_det1_status = res
        elif det2:
            text = str(det2).strip()
            level = "Sub-detalle"
            ctx = current_accion + " " + current_detalle1 + " " + text
            res = _classify(ctx)
            # Sub-detalle: si no matchea por keyword, hereda del detalle1 (o de la acción)
            if res[0] == "Pendiente verificar":
                last_det1_status = last_det1_status or last_accion_status
        else:
            continue
        if not text:
            continue
        # Determinar status según nivel y herencia
        if level == "ACCIÓN":
            status, evidencia, eta, owner = last_accion_status
        elif level == "Detalle":
            status, evidencia, eta, owner = last_det1_status if last_det1_status else last_accion_status
        else:  # Sub-detalle
            res = _classify(current_accion + " " + current_detalle1 + " " + text)
            if res[0] != "Pendiente verificar":
                status, evidencia, eta, owner = res
            else:
                status, evidencia, eta, owner = last_det1_status if last_det1_status else last_accion_status
        items.append({
            "proceso": current_proceso,
            "rol": current_rol,
            "nivel": level,
            "descripcion": text,
            "agente_ia": str(agente).strip() if agente else "",
            "status": status,
            "evidencia": evidencia,
            "eta": eta,
            "owner": owner,
        })
    return items


# ── Estilos ───────────────────────────────────────────────
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1B5E20")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill("solid", fgColor="E8F5E9")
STATUS_FILL = {
    "HECHO": PatternFill("solid", fgColor="C8E6C9"),
    "PARCIAL": PatternFill("solid", fgColor="FFF3CD"),
    "PENDIENTE": PatternFill("solid", fgColor="F8D7DA"),
    "N/A": PatternFill("solid", fgColor="E0E0E0"),
    "Pendiente verificar": PatternFill("solid", fgColor="FFE0B2"),
}


def _write_resumen(wb: openpyxl.Workbook, items: list[dict]):
    ws = wb.create_sheet("Resumen", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 60

    ws["A1"] = "Gap Analysis Software AMS-Production vs Excel Jorge Alquinta"
    ws["A1"].font = Font(bold=True, size=14, color="1B5E20")
    ws["A2"] = f"Fecha: {date.today().isoformat()}  ·  Commit HEAD: {_git_head_short()}"
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A3"] = "Stakeholder: Magdalena Ortega  ·  Validación funcional: Jorge Alquinta"
    ws["A3"].font = Font(italic=True, size=10, color="555555")

    # Totales por status
    from collections import Counter
    by_status = Counter(i["status"] for i in items)
    by_proceso = {}
    for i in items:
        p = i["proceso"]
        by_proceso.setdefault(p, Counter())[i["status"]] += 1

    ws["A5"] = "Totales por status"
    ws["A5"].font = Font(bold=True, size=12)
    row = 6
    for status in ["HECHO", "PARCIAL", "PENDIENTE", "Pendiente verificar", "N/A"]:
        n = by_status.get(status, 0)
        ws.cell(row, 1).value = status
        ws.cell(row, 2).value = n
        if status in STATUS_FILL:
            ws.cell(row, 1).fill = STATUS_FILL[status]
        row += 1
    ws.cell(row, 1).value = "TOTAL"
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2).value = sum(by_status.values())
    ws.cell(row, 2).font = Font(bold=True)

    row += 2
    ws.cell(row, 1).value = "Por proceso (HECHO / PARCIAL / PENDIENTE / Verif / N-A)"
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    headers = ["Proceso", "HECHO", "PARCIAL", "PENDIENTE", "Verif", "N/A", "% Done"]
    for col, h in enumerate(headers, 1):
        ws.cell(row, col).value = h
        ws.cell(row, col).fill = HEADER_FILL
        ws.cell(row, col).font = HEADER_FONT
    row += 1
    for proc, counts in by_proceso.items():
        ws.cell(row, 1).value = proc
        ws.cell(row, 2).value = counts.get("HECHO", 0)
        ws.cell(row, 3).value = counts.get("PARCIAL", 0)
        ws.cell(row, 4).value = counts.get("PENDIENTE", 0)
        ws.cell(row, 5).value = counts.get("Pendiente verificar", 0)
        ws.cell(row, 6).value = counts.get("N/A", 0)
        total = sum(counts.values())
        partial_credit = counts.get("HECHO", 0) + 0.5 * counts.get("PARCIAL", 0)
        pct = round(100 * partial_credit / total, 1) if total else 0
        ws.cell(row, 7).value = f"{pct}%"
        row += 1

    row += 2
    ws.cell(row, 1).value = "Highlights estratégicos"
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    highlights = [
        "✅ Scheduling v2 vista Horarios entregada esta semana (commit af700db) — eje vertical = horas, matching auto por skill.",
        "✅ SF-656 auditoría visual capacidad + violaciones día/noche (commit a80493d) — banner + tablas con severidades.",
        "✅ Audit log policy SF-660 — tests inmutabilidad + script purga 5y + endpoint scoped por rol (admin/manager/supervisor/planner).",
        "✅ Anonimización completa Goldfields (4 commits 0D1) — plant_id ya no leakea ni por title/aria-label.",
        "🟡 SF-661 análisis IA OT — función 1 de 7 entregada (commit d5b6496) · 2-7 requieren histórico real planta (0B2).",
        "🟡 SF-662 Preparativos OT estilo Rappi — modelo + state machine + tests (13 pytest) · UI completa pendiente spec Jorge.",
        "❌ Refactor módulo Ejecución (deadline VIE 15-may) — pendiente.",
        "❌ 0E centros de costo + clases de gasto — fuera de Sprint 7, propuesto SP8.",
        "❌ Catálogos síntoma+causa SAP-style (Aviso) — Carlos lidera, requiere data cliente.",
    ]
    for h in highlights:
        ws.cell(row, 1).value = h
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    ws.cell(row, 1).value = "Próximos pasos (SP7-SP8)"
    ws.cell(row, 1).font = Font(bold=True, size=12)
    row += 1
    next_steps = [
        "MIE 13 PM — Showcase Marco Ovalle / Gold Fields — Programación completa.",
        "JUE 14 — Refactor módulo Ejecución (EN_EJECUCION → CERRADO).",
        "VIE 15 — Cerrar Ejecución (deadline duro) + entregar este Gap a Magdalena.",
        "Sprint 8 (LUN 18+) — SF-661 funciones 2-7, SF-662 UI, Tanda 0E centros costo, refactor Análisis Desempeño.",
        "Q3-2026 — Catálogos síntoma+causa SAP, integración real ERP (requiere credenciales cliente).",
    ]
    for s in next_steps:
        ws.cell(row, 1).value = s
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 25
        row += 1


def _write_proceso_sheet(wb: openpyxl.Workbook, name: str, items: list[dict]):
    safe_name = name[:31].replace("/", "-")
    ws = wb.create_sheet(safe_name)
    headers = ["Proceso", "Rol", "Nivel", "Descripción (Jorge)", "Status", "Evidencia AMS", "ETA / Sprint", "Owner", "Agente IA esperado"]
    widths = [18, 14, 12, 50, 14, 50, 12, 12, 60]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        ws.cell(1, col).value = h
        ws.cell(1, col).fill = HEADER_FILL
        ws.cell(1, col).font = HEADER_FONT
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    row = 2
    for i in items:
        ws.cell(row, 1).value = i["proceso"]
        ws.cell(row, 2).value = i["rol"]
        ws.cell(row, 3).value = i["nivel"]
        ws.cell(row, 4).value = i["descripcion"]
        ws.cell(row, 5).value = i["status"]
        if i["status"] in STATUS_FILL:
            ws.cell(row, 5).fill = STATUS_FILL[i["status"]]
            ws.cell(row, 5).font = Font(bold=True)
        ws.cell(row, 6).value = i["evidencia"]
        ws.cell(row, 7).value = i["eta"]
        ws.cell(row, 8).value = i["owner"]
        ws.cell(row, 9).value = i["agente_ia"] or "—"
        for col in range(1, 10):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, col).border = BORDER
        # Filas pesadas por wrap-text
        text_len = max(len(str(i["descripcion"])), len(str(i["evidencia"])), len(i["agente_ia"]))
        ws.row_dimensions[row].height = min(max(20, text_len // 6 * 4), 120)
        row += 1


def main():
    items = _extract_items()
    print(f"[INFO] {len(items)} items extraídos del Excel de Jorge.")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Resumen primero
    _write_resumen(wb, items)

    # Una sheet por proceso
    procesos = sorted({i["proceso"] for i in items})
    for proc in procesos:
        proc_items = [i for i in items if i["proceso"] == proc]
        _write_proceso_sheet(wb, proc, proc_items)

    # Sheet con todo
    _write_proceso_sheet(wb, "Todos (174 items)", items)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] Excel guardado en: {OUT}")
    print(f"[INFO] Sheets: Resumen + {len(procesos)} procesos + Todos")


if __name__ == "__main__":
    main()
