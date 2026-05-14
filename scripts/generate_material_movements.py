"""Generate enriched 27_material_movements.xlsx and 37_material_reservations.xlsx.

Creates material movements covering ALL SAP movement types:
  - Entradas (EM): 101, 102, 561
  - Salidas (SM): 201, 261, 601
  - Traslados: 301, 311, 321
  - Devoluciones: 122, 701, 702

Scales to the full 8,379 spare parts inventory from 07_spare_parts_inventory.xlsx.
Also regenerates 37_material_reservations.xlsx with updated material references.
"""

import math
import os
import random
import sys
from datetime import datetime, timedelta

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

random.seed(2026)
AUTHOR = "AMS-Production"
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ---------------------------------------------------------------------------
# Movement type definitions (from Flujos-de-materiales.png)
# ---------------------------------------------------------------------------
MOVEMENT_TYPES = {
    # Entradas de Mercancias (EM)
    101: {
        "desc": "Recepcion de mercancias contra pedido",
        "category": "EM",
        "direction": "IN",
        "weight": 25,  # relative frequency
        "qty_range": (1, 50),
        "needs_order": True,
        "needs_cost_center": False,
    },
    102: {
        "desc": "Anulacion de entrada de mercancias 101",
        "category": "EM",
        "direction": "IN_REVERSAL",
        "weight": 3,
        "qty_range": (1, 20),
        "needs_order": True,
        "needs_cost_center": False,
    },
    561: {
        "desc": "Entrada inicial de stock (carga inicial)",
        "category": "EM",
        "direction": "IN",
        "weight": 2,
        "qty_range": (5, 200),
        "needs_order": False,
        "needs_cost_center": False,
    },
    # Salidas de Mercancias (SM)
    201: {
        "desc": "Salida para centro de coste (consumo interno)",
        "category": "SM",
        "direction": "OUT",
        "weight": 20,
        "qty_range": (1, 30),
        "needs_order": False,
        "needs_cost_center": True,
    },
    261: {
        "desc": "Consumo de material para orden de mantenimiento",
        "category": "SM",
        "direction": "OUT",
        "weight": 30,
        "qty_range": (1, 20),
        "needs_order": True,
        "needs_cost_center": True,
    },
    601: {
        "desc": "Salida de mercancias para entrega de venta",
        "category": "SM",
        "direction": "OUT",
        "weight": 2,
        "qty_range": (1, 10),
        "needs_order": True,
        "needs_cost_center": False,
    },
    # Traslados y Traspasos
    301: {
        "desc": "Traslado entre almacenes de la misma planta",
        "category": "TR",
        "direction": "TRANSFER",
        "weight": 8,
        "qty_range": (1, 50),
        "needs_order": False,
        "needs_cost_center": False,
    },
    311: {
        "desc": "Traspaso inmediato de stock de un almacen a otro",
        "category": "TR",
        "direction": "TRANSFER",
        "weight": 5,
        "qty_range": (1, 30),
        "needs_order": False,
        "needs_cost_center": False,
    },
    321: {
        "desc": "Traspaso de stock control calidad a libre utilizacion",
        "category": "TR",
        "direction": "QC_RELEASE",
        "weight": 4,
        "qty_range": (1, 40),
        "needs_order": False,
        "needs_cost_center": False,
    },
    # Devoluciones y Diferencias
    122: {
        "desc": "Devolucion de mercancias a proveedor",
        "category": "DV",
        "direction": "RETURN",
        "weight": 3,
        "qty_range": (1, 15),
        "needs_order": True,
        "needs_cost_center": False,
    },
    701: {
        "desc": "Entrada por diferencias de inventario fisico",
        "category": "DV",
        "direction": "ADJ_IN",
        "weight": 2,
        "qty_range": (1, 10),
        "needs_order": False,
        "needs_cost_center": False,
    },
    702: {
        "desc": "Salida por diferencias de inventario fisico",
        "category": "DV",
        "direction": "ADJ_OUT",
        "weight": 2,
        "qty_range": (1, 10),
        "needs_order": False,
        "needs_cost_center": False,
    },
}

# Column comments for 27_material_movements.xlsx
COLS_27 = [
    "movement_id",
    "movement_type",
    "movement_type_desc",
    "movement_category",
    "material_code",
    "sap_material_number",
    "description",
    "quantity",
    "unit",
    "order_number",
    "sap_func_loc",
    "sap_func_loc_short",
    "equipment_name",
    "posting_date",
    "cost_center",
    "storage_location_from",
    "storage_location_to",
    "total_value_usd",
    "document_number",
    "reversal_indicator",
]

COMMENTS_27 = {
    "movement_id": (
        "Identificador unico del movimiento de material.\n"
        "Formato: MV-NNNNNNN (secuencial).\n"
        "Generado por AMS para trazabilidad interna."
    ),
    "movement_type": (
        "Tipo de movimiento SAP (BWART).\n"
        "Tabla SAP: MSEG-BWART.\n"
        "Transacciones: MIGO, MB51, MB52.\n"
        "Entradas(EM): 101, 102, 561.\n"
        "Salidas(SM): 201, 261, 601.\n"
        "Traslados(TR): 301, 311, 321.\n"
        "Devoluciones(DV): 122, 701, 702."
    ),
    "movement_type_desc": (
        "Descripcion del tipo de movimiento.\n"
        "Fuente: tabla T156T (textos movimientos).\n"
        "Idioma: espanol."
    ),
    "movement_category": (
        "Categoria del movimiento.\n"
        "EM = Entrada de Mercancias.\n"
        "SM = Salida de Mercancias.\n"
        "TR = Traslado / Traspaso.\n"
        "DV = Devolucion / Diferencia inventario."
    ),
    "material_code": (
        "Codigo interno AMS del material.\n"
        "Formato: S26-MAT-NNNNN.\n"
        "Referencia cruzada con 07_spare_parts_inventory."
    ),
    "sap_material_number": (
        "Numero de material SAP (MATNR).\n"
        "Tabla SAP: MSEG-MATNR.\n"
        "Vinculo directo con maestro de materiales MM60."
    ),
    "description": (
        "Texto breve del material (MAKTX).\n"
        "Fuente: tabla MAKT.\n"
        "Copiado del inventario de repuestos."
    ),
    "quantity": (
        "Cantidad del movimiento.\n"
        "Campo SAP: MSEG-MENGE.\n"
        "Positivo = entrada, puede ser negativo en anulaciones.\n"
        "Unidad definida en campo 'unit'."
    ),
    "unit": (
        "Unidad de medida del movimiento.\n"
        "Campo SAP: MSEG-MEINS.\n"
        "Valores: EA (each), KG, L, M, M2, KIT."
    ),
    "order_number": (
        "Numero de orden SAP asociada.\n"
        "Campo SAP: MSEG-AUFNR.\n"
        "Para 261: orden de mantenimiento (OT).\n"
        "Para 101/102/601: pedido de compras.\n"
        "Para 201: puede ser vacio (consumo directo)."
    ),
    "sap_func_loc": (
        "Ubicacion funcional SAP completa.\n"
        "Campo SAP: MSEG-TPLNR / ILOA-TPLNR.\n"
        "Formato: SN-XXXX-XXXX-XXXX-XXXXXXXXXX.\n"
        "Vincula el consumo al equipo destino."
    ),
    "sap_func_loc_short": (
        "Codigo corto de ubicacion funcional.\n"
        "Ultimos 10 caracteres del sap_func_loc.\n"
        "Referencia rapida al equipo."
    ),
    "equipment_name": (
        "Nombre del equipo destino/origen.\n"
        "Campo SAP: EQKT-EQKTX.\n"
        "Del maestro de equipos (hierarchy)."
    ),
    "posting_date": (
        "Fecha de contabilizacion del movimiento.\n"
        "Campo SAP: MKPF-BUDAT.\n"
        "Formato: YYYY-MM-DD.\n"
        "Rango: 2024-01-01 a 2026-03-31."
    ),
    "cost_center": (
        "Centro de costes receptor.\n"
        "Campo SAP: MSEG-KOSTL.\n"
        "Formato: CC-XXXX (por area de planta).\n"
        "Aplica a movimientos 201 (consumo a CC)\n"
        "y 261 (consumo a orden con CC imputacion)."
    ),
    "storage_location_from": (
        "Almacen origen del movimiento.\n"
        "Campo SAP: MSEG-LGORT.\n"
        "Valores: ALM-CENTRAL, ALM-SEC-01, ALM-RIP-01, ALM-HUM-01.\n"
        "Para entradas (101, 561): vacio (viene de exterior).\n"
        "Para traslados (301, 311): almacen origen."
    ),
    "storage_location_to": (
        "Almacen destino del movimiento.\n"
        "Campo SAP: MSEG-UMLGO.\n"
        "Para entradas: almacen donde se recibe.\n"
        "Para salidas: vacio (sale de planta).\n"
        "Para traslados: almacen destino."
    ),
    "total_value_usd": (
        "Valor total del movimiento en USD.\n"
        "Calculo: quantity x unit_cost_usd.\n"
        "Campo SAP: MSEG-DMBTR.\n"
        "Moneda: USD (MSEG-WAERS)."
    ),
    "document_number": (
        "Numero de documento de material SAP.\n"
        "Campo SAP: MKPF-MBLNR.\n"
        "Formato: 49XXXXXXXX (10 digitos).\n"
        "Transacciones: MB51, MIGO."
    ),
    "reversal_indicator": (
        "Indicador de anulacion.\n"
        "Campo SAP: MKPF-XABLN.\n"
        "Vacio = movimiento normal.\n"
        "Ref doc = referencia al documento anulado.\n"
        "Aplica a tipos 102 (anula 101) y ajustes."
    ),
}

# Column comments for 37_material_reservations.xlsx
COLS_37 = [
    "reservation_id",
    "order_number",
    "operation_number",
    "material_code",
    "sap_material_number",
    "description",
    "quantity_required",
    "quantity_withdrawn",
    "quantity_available",
    "unit",
    "storage_location",
    "availability_status",
    "expected_delivery_date",
    "is_critical",
    "ved_class",
]

COMMENTS_37 = {
    "reservation_id": (
        "Numero de reserva SAP.\n"
        "Formato: RES-YYYY-NNNNN.\n"
        "Campo SAP: RESB-RSNUM.\n"
        "Transaccion: MB21 (crear), MB22 (modificar)."
    ),
    "order_number": (
        "Numero de orden de mantenimiento.\n"
        "Campo SAP: RESB-AUFNR.\n"
        "Vincula la reserva a la OT.\n"
        "Formato: 005XXXXX (8 digitos)."
    ),
    "operation_number": (
        "Numero de operacion dentro de la orden.\n"
        "Campo SAP: RESB-VORNR.\n"
        "Formato: 00XX (multiplos de 10).\n"
        "Indica en que paso se necesita el material."
    ),
    "material_code": (
        "Codigo interno AMS del material.\n"
        "Formato: S26-MAT-NNNNN.\n"
        "Referencia cruzada con 07_spare_parts_inventory."
    ),
    "sap_material_number": (
        "Numero de material SAP (MATNR).\n"
        "Campo SAP: RESB-MATNR.\n"
        "Vinculo directo con maestro de materiales."
    ),
    "description": (
        "Texto breve del material (MAKTX).\n"
        "Fuente: tabla MAKT."
    ),
    "quantity_required": (
        "Cantidad requerida por la operacion.\n"
        "Campo SAP: RESB-BDMNG.\n"
        "Unidad: definida en campo 'unit'."
    ),
    "quantity_withdrawn": (
        "Cantidad ya retirada del almacen.\n"
        "Campo SAP: RESB-ENMNG.\n"
        "Si = quantity_required -> completamente atendida."
    ),
    "quantity_available": (
        "Cantidad disponible en stock actual.\n"
        "Campo SAP: MARD-LABST.\n"
        "Verificada contra quantity_on_hand del inventario."
    ),
    "unit": (
        "Unidad de medida.\n"
        "Campo SAP: RESB-MEINS.\n"
        "Valores: EA, KG, L, M."
    ),
    "storage_location": (
        "Almacen de donde se retira el material.\n"
        "Campo SAP: RESB-LGORT.\n"
        "Valores: ALM-CENTRAL, ALM-SEC-01, ALM-RIP-01, ALM-HUM-01."
    ),
    "availability_status": (
        "Estado de disponibilidad del material.\n"
        "DISPONIBLE: stock >= requerido.\n"
        "PARCIAL: 0 < stock < requerido.\n"
        "NO_DISPONIBLE: stock = 0 o insuficiente.\n"
        "EN_TRANSITO: pedido en camino.\n"
        "Transaccion: CO02 (verificacion disponibilidad)."
    ),
    "expected_delivery_date": (
        "Fecha esperada de entrega si no disponible.\n"
        "Calculada: hoy + lead_time_days del material.\n"
        "Vacio si status = DISPONIBLE.\n"
        "Campo SAP: RESB-BDTER."
    ),
    "is_critical": (
        "Indicador de material critico para la orden.\n"
        "Si = la orden no puede ejecutarse sin este material.\n"
        "No = la orden puede proceder parcialmente.\n"
        "Basado en VED class: VITAL -> siempre critico."
    ),
    "ved_class": (
        "Clasificacion VED del material.\n"
        "Copiado del inventario 07_spare_parts_inventory.\n"
        "VITAL / ESSENTIAL / DESIRABLE.\n"
        "Determina criticidad de la reserva."
    ),
}

WAREHOUSES = ["ALM-CENTRAL", "ALM-SEC-01", "ALM-RIP-01", "ALM-HUM-01"]

# ---------------------------------------------------------------------------
# 1. Load spare parts inventory
# ---------------------------------------------------------------------------
print("Loading spare parts inventory (8,379 materials)...")
wb_inv = openpyxl.load_workbook(
    os.path.join(BASE, "data", "seeds", "07_spare_parts_inventory.xlsx"), read_only=True
)
ws_inv = wb_inv.active
materials = []
for row in ws_inv.iter_rows(min_row=2, values_only=True):
    materials.append({
        "material_code": row[0],
        "sap_material_number": str(row[1]),
        "description": row[2],
        "unit_cost_usd": float(row[13] or 0),
        "unit_of_measure": row[14],
        "warehouse_location": row[16],
        "ved_class": row[5],
        "fsn_class": row[6],
        "abc_class": row[7],
        "quantity_on_hand": int(row[8] or 0),
    })
wb_inv.close()
print(f"  Loaded {len(materials)} materials")

# Index by FSN for weighted sampling
fast = [m for m in materials if m["fsn_class"] == "FAST_MOVING"]
normal = [m for m in materials if m["fsn_class"] == "NORMAL"]
slow = [m for m in materials if m["fsn_class"] == "SLOW_MOVING"]

# ---------------------------------------------------------------------------
# 2. Load equipment hierarchy
# ---------------------------------------------------------------------------
print("Loading equipment hierarchy...")
wb_hier = openpyxl.load_workbook(
    os.path.join(BASE, "data", "seeds", "01_equipment_hierarchy.xlsx"), read_only=True
)
ws_hier = wb_hier.active
equipments = []
for row in ws_hier.iter_rows(min_row=2, values_only=True):
    level = row[4]
    if level and level >= 4 and row[1]:
        fl = str(row[0] or "")
        area = fl.split("-")[1] if len(fl.split("-")) > 1 else "3000"
        equipments.append({
            "fl": fl,
            "fl_short": str(row[1]),
            "name": str(row[3] or ""),
            "area": area,
        })
wb_hier.close()
print(f"  Loaded {len(equipments)} equipment locations")

# ---------------------------------------------------------------------------
# 3. Load work orders for realistic order numbers
# ---------------------------------------------------------------------------
print("Loading work orders...")
wo_path = os.path.join(BASE, "data", "seeds", "06_work_order_history.xlsx")
order_numbers = []
if os.path.exists(wo_path):
    wb_wo = openpyxl.load_workbook(wo_path, read_only=True)
    ws_wo = wb_wo.active
    headers = [c.value for c in next(ws_wo.iter_rows(max_row=1))]
    order_col = None
    for i, h in enumerate(headers):
        if h and "order" in str(h).lower() and "number" in str(h).lower():
            order_col = i
            break
    if order_col is None:
        for i, h in enumerate(headers):
            if h and "order" in str(h).lower():
                order_col = i
                break
    if order_col is not None:
        for row in ws_wo.iter_rows(min_row=2, values_only=True):
            on = row[order_col]
            if on:
                order_numbers.append(str(on))
    wb_wo.close()
if not order_numbers:
    order_numbers = [f"005{random.randint(10000, 99999):05d}" for _ in range(2000)]
print(f"  Loaded {len(order_numbers)} order numbers")

# ---------------------------------------------------------------------------
# 4. Generate material movements
# ---------------------------------------------------------------------------
print("Generating material movements...")

# Target: ~15,000-20,000 movements over 2+ years
# Fast movers: more movements, slow: fewer
DATE_START = datetime(2024, 1, 1)
DATE_END = datetime(2026, 3, 31)
DAYS_SPAN = (DATE_END - DATE_START).days

# Build weighted movement type list
mvt_pool = []
for mvt_type, info in MOVEMENT_TYPES.items():
    mvt_pool.extend([mvt_type] * info["weight"])

movements = []
doc_counter = 4900000000

# For each material, generate movements proportional to FSN
for mat in materials:
    fsn = mat["fsn_class"]
    if fsn == "FAST_MOVING":
        n_movements = random.randint(3, 8)
    elif fsn == "NORMAL":
        n_movements = random.randint(1, 4)
    else:
        n_movements = random.choices([0, 1, 2], weights=[40, 45, 15], k=1)[0]

    if n_movements == 0:
        continue

    for _ in range(n_movements):
        mvt_type = random.choice(mvt_pool)
        info = MOVEMENT_TYPES[mvt_type]

        # Pick equipment
        equip = random.choice(equipments)
        cost_center = f"CC-{equip['area']}"

        # Quantity
        lo, hi = info["qty_range"]
        quantity = random.randint(lo, hi)

        # Value
        total_value = round(quantity * mat["unit_cost_usd"], 2)

        # Date
        posting_date = DATE_START + timedelta(days=random.randint(0, DAYS_SPAN))

        # Order number
        order_num = ""
        if info["needs_order"]:
            order_num = random.choice(order_numbers)

        # Storage locations
        storage_from = ""
        storage_to = ""
        direction = info["direction"]

        if direction == "IN":
            storage_to = mat["warehouse_location"]
        elif direction == "IN_REVERSAL":
            storage_to = mat["warehouse_location"]
            quantity = -quantity
        elif direction == "OUT":
            storage_from = mat["warehouse_location"]
        elif direction == "TRANSFER":
            storage_from = mat["warehouse_location"]
            others = [w for w in WAREHOUSES if w != storage_from]
            storage_to = random.choice(others) if others else WAREHOUSES[0]
        elif direction == "QC_RELEASE":
            storage_from = "QC-INSPECCION"
            storage_to = mat["warehouse_location"]
        elif direction == "RETURN":
            storage_from = mat["warehouse_location"]
            storage_to = "PROVEEDOR"
        elif direction == "ADJ_IN":
            storage_to = mat["warehouse_location"]
        elif direction == "ADJ_OUT":
            storage_from = mat["warehouse_location"]

        # Document number
        doc_counter += 1
        doc_number = str(doc_counter)

        # Reversal indicator
        reversal = ""
        if mvt_type == 102:
            reversal = f"49{random.randint(10000000, 99999999)}"

        movements.append({
            "movement_id": "",  # will fill after sort
            "movement_type": mvt_type,
            "movement_type_desc": info["desc"],
            "movement_category": info["category"],
            "material_code": mat["material_code"],
            "sap_material_number": mat["sap_material_number"],
            "description": mat["description"],
            "quantity": quantity,
            "unit": mat["unit_of_measure"],
            "order_number": order_num,
            "sap_func_loc": equip["fl"],
            "sap_func_loc_short": equip["fl_short"],
            "equipment_name": equip["name"],
            "posting_date": posting_date.strftime("%Y-%m-%d"),
            "cost_center": cost_center if info["needs_cost_center"] else "",
            "storage_location_from": storage_from,
            "storage_location_to": storage_to,
            "total_value_usd": abs(total_value),
            "document_number": doc_number,
            "reversal_indicator": reversal,
        })

# Sort by posting_date
movements.sort(key=lambda m: m["posting_date"])

# Assign sequential IDs
for i, m in enumerate(movements, start=1):
    m["movement_id"] = f"MV-{i:07d}"

print(f"  Generated {len(movements)} movements")

# ---------------------------------------------------------------------------
# 5. Write 27_material_movements.xlsx
# ---------------------------------------------------------------------------
print("Writing 27_material_movements.xlsx...")
wb27 = openpyxl.Workbook()
ws27 = wb27.active
ws27.title = "Material Movements"

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, col_name in enumerate(COLS_27, start=1):
    cell = ws27.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    comment_text = COMMENTS_27.get(col_name, col_name)
    cell.comment = Comment(comment_text, AUTHOR)
    cell.comment.width = 350
    cell.comment.height = 200

for row_idx, mv in enumerate(movements, start=2):
    for col_idx, col_name in enumerate(COLS_27, start=1):
        ws27.cell(row=row_idx, column=col_idx, value=mv[col_name])

# Column widths
for col_idx in range(1, len(COLS_27) + 1):
    col_letter = get_column_letter(col_idx)
    max_len = len(str(COLS_27[col_idx - 1]))
    for row_idx in range(2, min(102, len(movements) + 2)):
        val = ws27.cell(row=row_idx, column=col_idx).value
        if val is not None:
            max_len = max(max_len, min(len(str(val)), 55))
    ws27.column_dimensions[col_letter].width = max_len + 3

ws27.freeze_panes = "A2"
ws27.auto_filter.ref = f"A1:{get_column_letter(len(COLS_27))}{len(movements) + 1}"

output_27 = os.path.join(BASE, "data", "seeds", "27_material_movements.xlsx")
wb27.save(output_27)
print(f"  Wrote {len(movements)} rows to {output_27}")

# ---------------------------------------------------------------------------
# 6. Generate material reservations (37)
# ---------------------------------------------------------------------------
print("\nGenerating material reservations...")

# ~3,000 reservations across active orders
reservations = []
base_date = datetime(2026, 3, 15)

# Sample materials for reservations (biased toward FAST and NORMAL)
res_materials = (
    random.choices(fast, k=min(1200, len(fast)))
    + random.choices(normal, k=min(1200, len(normal)))
    + random.choices(slow, k=min(600, len(slow)))
)
random.shuffle(res_materials)

for i, mat in enumerate(res_materials, start=1):
    order_num = random.choice(order_numbers)
    operation = f"00{random.choice([10, 20, 30, 40, 50, 60]):02d}"

    qty_required = random.randint(1, 20)
    qty_on_hand = mat["quantity_on_hand"]

    # Determine availability
    if qty_on_hand >= qty_required:
        status = "DISPONIBLE"
        qty_available = qty_on_hand
        qty_withdrawn = random.choices(
            [qty_required, random.randint(0, qty_required - 1) if qty_required > 1 else 0],
            weights=[70, 30], k=1
        )[0]
        delivery_date = ""
    elif qty_on_hand > 0:
        status = "PARCIAL"
        qty_available = qty_on_hand
        qty_withdrawn = random.randint(0, qty_on_hand)
        delivery_date = (base_date + timedelta(days=random.randint(5, 60))).strftime("%Y-%m-%d")
    else:
        status = random.choices(
            ["NO_DISPONIBLE", "EN_TRANSITO"], weights=[60, 40], k=1
        )[0]
        qty_available = 0
        qty_withdrawn = 0
        delivery_date = (base_date + timedelta(days=random.randint(7, 90))).strftime("%Y-%m-%d")

    is_critical = "Si" if mat["ved_class"] == "VITAL" else (
        "Si" if mat["ved_class"] == "ESSENTIAL" and random.random() < 0.3 else "No"
    )

    reservations.append({
        "reservation_id": f"RES-2026-{i:05d}",
        "order_number": order_num,
        "operation_number": operation,
        "material_code": mat["material_code"],
        "sap_material_number": mat["sap_material_number"],
        "description": mat["description"],
        "quantity_required": qty_required,
        "quantity_withdrawn": qty_withdrawn,
        "quantity_available": qty_available,
        "unit": mat["unit_of_measure"],
        "storage_location": mat["warehouse_location"],
        "availability_status": status,
        "expected_delivery_date": delivery_date,
        "is_critical": is_critical,
        "ved_class": mat["ved_class"],
    })

print(f"  Generated {len(reservations)} reservations")

# ---------------------------------------------------------------------------
# 7. Write 37_material_reservations.xlsx
# ---------------------------------------------------------------------------
print("Writing 37_material_reservations.xlsx...")
wb37 = openpyxl.Workbook()
ws37 = wb37.active
ws37.title = "Material Reservations"

for col_idx, col_name in enumerate(COLS_37, start=1):
    cell = ws37.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    comment_text = COMMENTS_37.get(col_name, col_name)
    cell.comment = Comment(comment_text, AUTHOR)
    cell.comment.width = 350
    cell.comment.height = 200

for row_idx, res in enumerate(reservations, start=2):
    for col_idx, col_name in enumerate(COLS_37, start=1):
        ws37.cell(row=row_idx, column=col_idx, value=res[col_name])

for col_idx in range(1, len(COLS_37) + 1):
    col_letter = get_column_letter(col_idx)
    max_len = len(str(COLS_37[col_idx - 1]))
    for row_idx in range(2, min(102, len(reservations) + 2)):
        val = ws37.cell(row=row_idx, column=col_idx).value
        if val is not None:
            max_len = max(max_len, min(len(str(val)), 55))
    ws37.column_dimensions[col_letter].width = max_len + 3

ws37.freeze_panes = "A2"
ws37.auto_filter.ref = f"A1:{get_column_letter(len(COLS_37))}{len(reservations) + 1}"

output_37 = os.path.join(BASE, "data", "seeds", "37_material_reservations.xlsx")
wb37.save(output_37)
print(f"  Wrote {len(reservations)} rows to {output_37}")

# ---------------------------------------------------------------------------
# 8. Summary
# ---------------------------------------------------------------------------
print(f"\n{'='*60}")
print("SUMMARY: 27_material_movements.xlsx")
print(f"{'='*60}")
mvt_counts = {}
cat_counts = {}
for m in movements:
    k = f"{m['movement_type']} ({m['movement_type_desc'][:40]})"
    mvt_counts[k] = mvt_counts.get(k, 0) + 1
    cat_counts[m["movement_category"]] = cat_counts.get(m["movement_category"], 0) + 1

print(f"Total movements: {len(movements)}")
print(f"\nBy category:")
for k, v in sorted(cat_counts.items()):
    print(f"  {k}: {v}")
print(f"\nBy type:")
for k, v in sorted(mvt_counts.items()):
    print(f"  {k}: {v}")

values = [m["total_value_usd"] for m in movements]
print(f"\nValue range: ${min(values):,.2f} - ${max(values):,.2f}")
print(f"Total value: ${sum(values):,.2f}")

dates = [m["posting_date"] for m in movements]
print(f"Date range: {min(dates)} to {max(dates)}")
print(f"Unique materials used: {len(set(m['material_code'] for m in movements))}")

print(f"\n{'='*60}")
print("SUMMARY: 37_material_reservations.xlsx")
print(f"{'='*60}")
status_counts = {}
for r in reservations:
    status_counts[r["availability_status"]] = status_counts.get(r["availability_status"], 0) + 1
print(f"Total reservations: {len(reservations)}")
print(f"By status: {dict(sorted(status_counts.items()))}")
crit = sum(1 for r in reservations if r["is_critical"] == "Si")
print(f"Critical: {crit} ({100*crit/len(reservations):.1f}%)")
print(f"Unique materials: {len(set(r['material_code'] for r in reservations))}")
