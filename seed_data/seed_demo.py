"""Demo Seed Script — loads Excel data into the OCP system database.
Hybrid approach: maps SAP codes to system internal codes bidirectionally.
Run inside the backend container or with access to the DB.
"""
import os, sys, json
from datetime import datetime, date

# Add project to path
sys.path.insert(0, "/app")
os.chdir("/app")

# ═══════════════════════════════════════════
# SAP <-> SYSTEM MAPPING CONFIG
# ═══════════════════════════════════════════

SAP_MAPPING = {
    # WO Type: SAP -> Internal
    "wo_type": {
        "PM01": "PM01", "PM02": "PM02", "PM03": "PM03",
        "PM05": "PM05", "PM06": "PM06", "PM07": "PM07",
        "CORRECTIVO": "PM01", "PREVENTIVO": "PM02", "PREDICTIVO": "PM03",
    },
    # Priority: SAP -> Internal
    "priority": {
        "I": "P1", "A": "P2", "M": "P3", "B": "P4",
        "1": "P1", "2": "P2", "3": "P3", "4": "P4",
        "P1": "P1", "P2": "P2", "P3": "P3", "P4": "P4",
    },
    # Status: SAP -> Internal
    "wo_status": {
        "ABIE": "EN_EJECUCION", "CTEC": "CERRADO", "NOTI": "CREADO",
        "LIBE": "PROGRAMADO", "PLAN": "PLANIFICADO", "CANC": "CANCELADO",
        "CREADO": "CREADO", "PLANIFICADO": "PLANIFICADO",
        "PROGRAMADO": "PROGRAMADO", "EN_EJECUCION": "EN_EJECUCION",
        "CERRADO": "CERRADO", "CANCELADO": "CANCELADO",
    },
    # Notification status
    "notif_status": {
        "NOPR": "PENDIENTE", "OSNO": "VALIDATED", "APRO": "APROBADO",
        "RECH": "RECHAZADO", "CANC": "CANCELADO",
    },
    # Specialty: Excel code -> System
    "specialty": {
        "MEC": "MECHANICAL", "ELE": "ELECTRICAL", "INS": "INSTRUMENTATION",
        "LUB": "LUBRICATION", "SOL": "WELDING", "SIN": "INSTRUMENTATION",
        "DCS": "INSTRUMENTATION", "SUP": "GENERAL",
        "MECHANICAL": "MECHANICAL", "ELECTRICAL": "ELECTRICAL",
    },
}

# Save mapping as JSON config
with open("/tmp/sap_mapping.json", "w") as f:
    json.dump(SAP_MAPPING, f, indent=2)
print("Saved sap_mapping.json")


def read_excel(path):
    """Read Excel file and return list of dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for h, v in zip(headers, row):
            if h:
                d[h] = v
        if any(v is not None for v in d.values()):
            rows.append(d)
    wb.close()
    return rows


def safe_str(v, default=""):
    if v is None:
        return default
    return str(v).strip()


def safe_float(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def safe_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v if isinstance(v, datetime) else datetime.combine(v, datetime.min.time())
    try:
        return datetime.fromisoformat(str(v)[:10])
    except:
        return None


# ═══════════════════════════════════════════
# CONNECT TO DB
# ═══════════════════════════════════════════
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

DB_URL = os.getenv("DATABASE_URL", "sqlite:///data/ocp_maintenance.db")
engine = create_engine(DB_URL)
Session = sessionmaker(bind=engine)
db = Session()

print(f"Connected to DB: {DB_URL}")


# ═══════════════════════════════════════════
# 1. CREATE sap_materials TABLE + SEED
# ═══════════════════════════════════════════
print("\n=== 1. SAP Materials ===")

db.execute(text("DROP TABLE IF EXISTS sap_materials"))
db.commit()
db.execute(text("""
    CREATE TABLE sap_materials (
        sap_id VARCHAR(20) PRIMARY KEY,
        description VARCHAR(200),
        category VARCHAR(30),
        unit VARCHAR(10) DEFAULT 'PZ',
        unit_cost_usd FLOAT DEFAULT 0,
        manufacturer VARCHAR(100),
        part_number VARCHAR(50),
        ved_class VARCHAR(20),
        abc_class VARCHAR(5),
        max_stock INTEGER DEFAULT 0,
        min_stock INTEGER DEFAULT 0,
        current_stock INTEGER DEFAULT 0,
        lead_time_days INTEGER DEFAULT 0,
        warehouse VARCHAR(30),
        applicable_equipment TEXT
    )
"""))
db.commit()

rows = read_excel("/app/seed_data/07_spare_parts_inventory.xlsx")
count = 0
for r in rows:
    sap_id = safe_str(r.get("sap_material_number", r.get("material_code", "")))
    if not sap_id:
        continue
    # Check if exists
    existing = db.execute(text("SELECT sap_id FROM sap_materials WHERE sap_id = :sid"), {"sid": sap_id}).fetchone()
    if existing:
        continue
    db.execute(text("""
        INSERT INTO sap_materials (sap_id, description, category, unit, unit_cost_usd, manufacturer,
            part_number, ved_class, abc_class, max_stock, min_stock, current_stock, lead_time_days,
            warehouse, applicable_equipment)
        VALUES (:sid, :desc, :cat, :unit, :cost, :mfr, :pn, :ved, :abc, :max, :min, :cur, :lt, :wh, :equip)
    """), {
        "sid": sap_id,
        "desc": safe_str(r.get("description", "")),
        "cat": safe_str(r.get("material_group", r.get("category", ""))),
        "unit": safe_str(r.get("unit_of_measure", "PZ")),
        "cost": safe_float(r.get("unit_cost_usd", 0)),
        "mfr": safe_str(r.get("manufacturer", "")),
        "pn": safe_str(r.get("part_number", "")),
        "ved": safe_str(r.get("ved_class", "")),
        "abc": safe_str(r.get("abc_class", "")),
        "max": int(safe_float(r.get("max_stock", 0))),
        "min": int(safe_float(r.get("reorder_point", r.get("min_stock", 0)))),
        "cur": int(safe_float(r.get("current_stock", 0))),
        "lt": int(safe_float(r.get("lead_time_days", 0))),
        "wh": safe_str(r.get("warehouse_location", r.get("warehouse", ""))),
        "equip": safe_str(r.get("applicable_equipment_csv", "")),
    })
    count += 1

db.commit()
print(f"  Inserted {count} materials into sap_materials")


# ═══════════════════════════════════════════
# 2. SEED WORKFORCE (200 technicians)
# ═══════════════════════════════════════════
print("\n=== 2. Workforce ===")

rows = read_excel("/app/seed_data/09_workforce.xlsx")
count = 0
for r in rows:
    wid = safe_str(r.get("worker_id", ""))
    if not wid:
        continue
    existing = db.execute(text("SELECT worker_id FROM workforce WHERE worker_id = :wid"), {"wid": wid}).fetchone()
    if existing:
        continue

    specialty_raw = safe_str(r.get("specialty", "MEC"))
    specialty = SAP_MAPPING["specialty"].get(specialty_raw, "MECHANICAL")
    shift_raw = safe_str(r.get("shift_code", "7X7"))

    equip_exp = safe_str(r.get("work_center", ""))
    db.execute(text("""
        INSERT INTO workforce (worker_id, name, specialty, shift, plant_id, available,
            years_experience, safety_training_current, equipment_expertise)
        VALUES (:wid, :name, :spec, :shift, :plant, :avail, :exp, :safety, :equip)
    """), {
        "wid": wid,
        "name": safe_str(r.get("name", r.get("worker_name", "Worker"))),
        "spec": specialty,
        "shift": "MORNING" if "D" in shift_raw.upper() or "7X7" in shift_raw else "NIGHT",
        "plant": safe_str(r.get("plant_id", "OCP-JFC1")),
        "avail": True,
        "exp": int(safe_float(r.get("years_experience", r.get("seniority_years", 5)))),
        "safety": True,
        "equip": equip_exp,
    })
    count += 1

db.commit()
print(f"  Inserted {count} workers into workforce")


# ═══════════════════════════════════════════
# 3. SEED EQUIPMENT HIERARCHY
# ═══════════════════════════════════════════
print("\n=== 3. Equipment Hierarchy ===")

rows = read_excel("/app/seed_data/01_equipment_hierarchy.xlsx")
count = 0
for r in rows:
    equnr = safe_str(r.get("equnr", ""))
    fl = safe_str(r.get("sap_func_loc", ""))
    # Derive tag: use last segment of func loc or build from fl
    tag = "OCP-" + fl.replace("-", "-") if fl else ""
    if not equnr and not tag:
        continue
    if not equnr:
        continue  # Only seed actual equipment, not location nodes
    # Check if exists in hierarchy_nodes
    existing = db.execute(text("SELECT node_id FROM hierarchy_nodes WHERE tag = :tag"), {"tag": tag}).fetchone()
    if existing:
        continue

    from uuid import uuid4
    node_id = str(uuid4())
    tplnr = safe_str(r.get("sap_func_loc", ""))

    db.execute(text("""
        INSERT INTO hierarchy_nodes (node_id, node_type, tag, name, name_fr, code, level, "order",
            plant_id, sap_func_loc, sap_equipment_nr, criticality, status, metadata_json)
        VALUES (:nid, 'EQUIPMENT', :tag, :name, :name, :code, :lvl, :ord, :plant, :tplnr, :equnr, :crit, 'ACTIVE', :meta)
    """), {
        "nid": node_id,
        "tag": tag,
        "name": safe_str(r.get("eqktx", r.get("pltxt", tag))),
        "code": tag,
        "plant": "OCP-JFC1",
        "tplnr": tplnr,
        "equnr": safe_str(r.get("equnr", "")),
        "lvl": int(safe_float(r.get("level", 4))),
        "ord": count + 200,
        "crit": {"1": "AA", "2": "A", "3": "B"}.get(safe_str(r.get("abckz", "2")), "B"),
        "meta": json.dumps({"fl_type": safe_str(r.get("fl_type", "")), "herst": safe_str(r.get("herst", "")), "model": safe_str(r.get("model", "")), "power_kw": safe_str(r.get("power_kw", "")), "install_date": safe_str(r.get("install_date", "")), "planning_group": safe_str(r.get("planning_group", "")), "business_area": safe_str(r.get("business_area", ""))}),
    })
    count += 1

db.commit()
print(f"  Inserted {count} equipment nodes into hierarchy_nodes")


# ═══════════════════════════════════════════
# 4. SEED WORK ORDER HISTORY
# ═══════════════════════════════════════════
print("\n=== 4. Work Order History ===")

rows = read_excel("/app/seed_data/06_work_order_history.xlsx")
count = 0
for r in rows:
    aufnr = safe_str(r.get("aufnr", ""))
    if not aufnr:
        continue
    wo_number = f"OT-{aufnr}"
    existing = db.execute(text("SELECT wo_id FROM managed_work_orders WHERE wo_number = :wn"), {"wn": wo_number}).fetchone()
    if existing:
        continue

    from uuid import uuid4
    wo_id = str(uuid4())
    wo_type = safe_str(r.get("auart", "PM01"))
    priority_raw = safe_str(r.get("priokx", "M"))
    priority = SAP_MAPPING["priority"].get(priority_raw, "P3")
    status_raw = safe_str(r.get("system_status", "CTEC"))
    status = SAP_MAPPING["wo_status"].get(status_raw, "CERRADO")
    work_class = "NO_PROGRAMADO" if priority in ("P1", "P2") else "PROGRAMADO"

    db.execute(text("""
        INSERT INTO managed_work_orders (wo_id, wo_number, plant_id, equipment_id, equipment_tag,
            description, wo_type, priority_code, work_class, status,
            estimated_hours, actual_hours, planned_start, planned_end, actual_start, actual_end,
            labor_cost, material_cost, external_cost, work_center, planning_group,
            is_fast_track, created_at, closed_at, closed_by, completion_pct, budget_approved)
        VALUES (:wid, :wn, :plant, :equip, :equip, :desc, :wtype, :prio, :wclass, :status,
            :est_h, :act_h, :p_start, :p_end, :a_start, :a_end,
            :labor, :mat, :ext, :wc, :pg, :ft, :created, :closed, :closedby, :pct, :ba)
    """), {
        "wid": wo_id,
        "wn": wo_number,
        "plant": "OCP-JFC1",
        "equip": safe_str(r.get("equipment_tag", "")),
        "desc": safe_str(r.get("description", r.get("short_text", ""))),
        "wtype": wo_type,
        "prio": priority,
        "wclass": work_class,
        "status": status,
        "est_h": safe_float(r.get("planned_hours", r.get("duration_planned", 4))),
        "act_h": safe_float(r.get("actual_hours", r.get("duration_actual", 0))),
        "p_start": safe_date(r.get("gstrp", r.get("basic_start", None))),
        "p_end": safe_date(r.get("gltrp", r.get("basic_end", None))),
        "a_start": safe_date(r.get("iedd", r.get("actual_start", None))),
        "a_end": safe_date(r.get("iedt", r.get("actual_end", None))),
        "labor": safe_float(r.get("cost_labour_usd", 0)),
        "mat": safe_float(r.get("cost_materials_usd", 0)),
        "ext": safe_float(r.get("cost_external_usd", 0)),
        "wc": safe_str(r.get("arbpl", r.get("work_center", ""))),
        "pg": safe_str(r.get("planning_group", "")),
        "ft": priority in ("P1", "P2"),
        "created": safe_date(r.get("erdat", r.get("created_date", None))),
        "closed": safe_date(r.get("iedt", None)) if status == "CERRADO" else None,
        "closedby": safe_str(r.get("supervisor_wc", "")),
        "pct": 100.0 if status == "CERRADO" else 0.0,
        "ba": True,
    })
    count += 1

db.commit()
print(f"  Inserted {count} work orders into managed_work_orders")


# ═══════════════════════════════════════════
# 5. SEED NOTIFICATIONS (AVISOS)
# ═══════════════════════════════════════════
print("\n=== 5. Notifications / Avisos ===")

rows = read_excel("/app/seed_data/24_notifications.xlsx")
count = 0
for r in rows:
    qmnum = safe_str(r.get("qmnum", ""))
    if not qmnum:
        continue
    req_id = f"WR-{qmnum}"
    existing = db.execute(text("SELECT request_id FROM work_requests WHERE request_id = :rid"), {"rid": req_id}).fetchone()
    if existing:
        continue

    from uuid import uuid4
    rid = req_id
    priority_raw = safe_str(r.get("priority", "M"))
    priority = SAP_MAPPING["priority"].get(priority_raw, "P3")
    status_raw = safe_str(r.get("user_status", "NOPR"))
    status = SAP_MAPPING["notif_status"].get(status_raw, "PENDIENTE")

    problem_desc = json.dumps({
        "original_text": safe_str(r.get("description", "")),
        "long_text": safe_str(r.get("long_text", "")),
    })
    ai_class = json.dumps({
        "failure_category": safe_str(r.get("damage_code", "")),
        "failure_cause": safe_str(r.get("cause_code", "")),
        "failure_object_part": safe_str(r.get("object_part_code", "")),
        "plant_id": "OCP-JFC1",
        "planning_group": safe_str(r.get("planning_group", "")),
        "work_center": safe_str(r.get("work_center", "")),
        "area": safe_str(r.get("area", "")),
        "notification_catalog": safe_str(r.get("notification_catalog", "")),
    })

    db.execute(text("""
        INSERT INTO work_requests (request_id, equipment_tag, equipment_id, equipment_confidence,
            resolution_method, problem_description,
            priority_code, status, ai_classification, notification_type, created_at,
            reported_by, work_class)
        VALUES (:rid, :equip, :equip, 0.95, 'MANUAL', :prob, :prio, :status, :ai, :ntype, :created, :reporter, :wclass)
    """), {
        "rid": rid,
        "equip": safe_str(r.get("equipment_tag", "")),
        "prob": problem_desc,
        "prio": priority,
        "status": status,
        "ai": ai_class,
        "ntype": safe_str(r.get("qmart", "A1")),
        "created": safe_date(r.get("reported_date", None)),
        "reporter": safe_str(r.get("reported_by", "")),
        "wclass": "NO_PROGRAMADO" if priority in ("P1", "P2") else "PROGRAMADO",
    })
    count += 1

db.commit()
print(f"  Inserted {count} work requests (avisos) into work_requests")


# ═══════════════════════════════════════════
# 6. ADD SHIFT COLUMN IF MISSING
# ═══════════════════════════════════════════
print("\n=== 6. Schema migrations ===")
try:
    db.execute(text("ALTER TABLE managed_work_orders ADD COLUMN shift VARCHAR(10) DEFAULT 'day'"))
    db.commit()
    print("  Added shift column to managed_work_orders")
except:
    print("  shift column already exists")

# Ensure sap_materials table works with the search endpoint
test = db.execute(text("SELECT COUNT(*) FROM sap_materials")).scalar()
print(f"  sap_materials table: {test} rows")

test2 = db.execute(text("SELECT COUNT(*) FROM workforce")).scalar()
print(f"  workforce table: {test2} rows")

test3 = db.execute(text("SELECT COUNT(*) FROM managed_work_orders")).scalar()
print(f"  managed_work_orders table: {test3} rows")

test4 = db.execute(text("SELECT COUNT(*) FROM work_requests")).scalar()
print(f"  work_requests table: {test4} rows")

test5 = db.execute(text("SELECT COUNT(*) FROM hierarchy_nodes WHERE node_type='EQUIPMENT'")).scalar()
print(f"  hierarchy_nodes (EQUIPMENT): {test5} rows")

db.close()
print("\n=== SEED COMPLETE ===")
