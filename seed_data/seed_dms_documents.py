"""Seed dms_documents table from 18_dms_maf_documents.xlsx."""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from api.database.connection import SessionLocal, engine
from api.database.models import Base, DMSDocumentModel

XLSX = os.path.join(os.path.dirname(__file__), "18_dms_maf_documents.xlsx")

Base.metadata.create_all(engine)

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active

headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
print("Columns:", headers)

db = SessionLocal()
existing = {r.document_number for r in db.query(DMSDocumentModel.document_number).all()}

inserted = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    row_dict = dict(zip(headers, row))
    doc_num = str(row_dict.get("document_number") or "").strip()
    if not doc_num or doc_num in existing:
        continue

    doc = DMSDocumentModel(
        document_number=doc_num,
        document_type=str(row_dict.get("document_type") or "").strip(),
        document_desc=str(row_dict.get("document_desc") or "").strip(),
        version=int(row_dict.get("version") or 1),
        sap_func_loc=str(row_dict.get("sap_func_loc") or "").strip() or None,
        sap_func_loc_short=str(row_dict.get("sap_func_loc_short") or "").strip() or None,
        equipment_name=str(row_dict.get("equipment_name") or "").strip() or None,
        eqart=str(row_dict.get("eqart") or "").strip() or None,
        file_path=str(row_dict.get("file_path") or "").strip() or None,
        created_date=str(row_dict.get("created_date") or "").strip() or None,
        created_by=str(row_dict.get("created_by") or "").strip() or None,
        status=str(row_dict.get("status") or "Activo").strip(),
    )
    db.add(doc)
    existing.add(doc_num)
    inserted += 1

db.commit()
db.close()
print(f"Inserted {inserted} DMS documents.")
