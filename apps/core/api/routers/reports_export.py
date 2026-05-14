"""Reports export — XLSX downloads for weekly schedule, closed WOs, KPIs."""

from datetime import datetime, timedelta, date
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from api.database.connection import get_db
from api.database.models import ManagedWorkOrderModel
from api.dependencies.auth import get_current_user

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


router = APIRouter(
    prefix="/reports-export",
    tags=["reports-export"],
    dependencies=[Depends(get_current_user)],
)


_INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell(value):
    """Prevent CSV/XLSX formula injection. Si una cadena empieza por
    =/+/-/@/tab/CR, Excel la interpreta como fórmula (DDE attack vector).
    Anteponemos un apóstrofo para forzar literal text."""
    if isinstance(value, str) and value and value[0] in _INJECTION_PREFIXES:
        return "'" + value
    return value


def _safe_row(values: list) -> list:
    return [_sanitize_cell(v) for v in values]


def _wb_response(wb, filename: str) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _header(ws, columns: list[str]):
    fill = PatternFill("solid", fgColor="1B5E20")
    font = Font(bold=True, color="FFFFFF")
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


@router.get("/weekly-schedule.xlsx")
def weekly_schedule_xlsx(
    plant_id: str | None = None,
    week_start: str | None = None,
    db: Session = Depends(get_db),
):
    """Weekly schedule export: all WOs with planned_start in the given week."""
    if not OPENPYXL_OK:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    if week_start:
        try:
            start = datetime.fromisoformat(week_start).date()
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid week_start")
    else:
        today = date.today()
        start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=7)

    q = db.query(ManagedWorkOrderModel).filter(
        ManagedWorkOrderModel.planned_start >= datetime.combine(start, datetime.min.time()),
        ManagedWorkOrderModel.planned_start < datetime.combine(end, datetime.min.time()),
    )
    if plant_id:
        q = q.filter(ManagedWorkOrderModel.plant_id == plant_id)
    wos = q.order_by(ManagedWorkOrderModel.planned_start.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Week {start.isoformat()}"
    _header(ws, [
        "WO Number", "Type", "Priority", "Equipment", "Description",
        "Planning Group", "Work Center", "Planned Start", "Planned End",
        "Est. Hours", "Shift", "Status", "Techs",
    ])
    for wo in wos:
        techs = ", ".join((w.get("name") or w.get("worker_id") or "") for w in (wo.assigned_workers or []))
        ws.append(_safe_row([
            wo.wo_number, wo.wo_type, wo.priority_code, wo.equipment_tag or "",
            wo.description or "", wo.planning_group or "", wo.work_center or "",
            wo.planned_start.isoformat() if wo.planned_start else "",
            wo.planned_end.isoformat() if wo.planned_end else "",
            float(wo.estimated_hours or 0),
            getattr(wo, "shift", "") or "day",
            wo.status, techs,
        ]))
    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    return _wb_response(wb, f"weekly-schedule-{start.isoformat()}.xlsx")


@router.get("/wos-closed.xlsx")
def wos_closed_xlsx(
    plant_id: str | None = None,
    month: str | None = None,  # YYYY-MM
    db: Session = Depends(get_db),
):
    """Closed WOs for a month. Defaults to current month."""
    if not OPENPYXL_OK:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    today = date.today()
    if month:
        try:
            y, m = [int(x) for x in month.split("-")]
            start = date(y, m, 1)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid month")
    else:
        start = today.replace(day=1)
    end_month = start.replace(day=28) + timedelta(days=4)
    end = end_month.replace(day=1)

    q = db.query(ManagedWorkOrderModel).filter(
        ManagedWorkOrderModel.status == "CERRADO",
        ManagedWorkOrderModel.closed_at >= datetime.combine(start, datetime.min.time()),
        ManagedWorkOrderModel.closed_at < datetime.combine(end, datetime.min.time()),
    )
    if plant_id:
        q = q.filter(ManagedWorkOrderModel.plant_id == plant_id)
    wos = q.order_by(ManagedWorkOrderModel.closed_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Closed {start.strftime('%Y-%m')}"
    _header(ws, [
        "WO Number", "Type", "Priority", "Equipment", "Description",
        "Closed At", "Closed By", "Signed By",
        "Plan HH", "Actual HH", "Variance %",
        "Labor Cost", "Material Cost", "External Cost", "Total Cost",
        "Notes",
    ])
    total_plan = total_actual = total_cost = 0.0
    for wo in wos:
        plan = float(wo.estimated_hours or 0)
        actual = float(wo.actual_hours or 0)
        variance = round(((actual - plan) / plan) * 100, 1) if plan else 0
        cost = float(wo.actual_total_cost or 0) or (float(wo.labor_cost or 0) + float(wo.material_cost or 0) + float(wo.external_cost or 0))
        total_plan += plan
        total_actual += actual
        total_cost += cost
        ws.append(_safe_row([
            wo.wo_number, wo.wo_type, wo.priority_code, wo.equipment_tag or "",
            wo.description or "",
            wo.closed_at.isoformat() if wo.closed_at else "",
            wo.closed_by or "",
            getattr(wo, "closed_by_signature", "") or "",
            plan, actual, variance,
            float(wo.labor_cost or 0), float(wo.material_cost or 0), float(wo.external_cost or 0), cost,
            getattr(wo, "closure_notes", "") or "",
        ]))
    # Totals row
    ws.append([])
    totals_row = ["TOTAL", "", "", "", "", "", "", "", total_plan, total_actual,
                  round(((total_actual - total_plan) / total_plan) * 100, 1) if total_plan else 0,
                  "", "", "", total_cost, ""]
    ws.append(totals_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)
    return _wb_response(wb, f"wos-closed-{start.strftime('%Y-%m')}.xlsx")


@router.get("/kpi-summary.xlsx")
def kpi_summary_xlsx(plant_id: str | None = None, db: Session = Depends(get_db)):
    """Multi-sheet KPI summary: MTBF/MTTR, PM compliance, backlog aging, cost."""
    if not OPENPYXL_OK:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    from api.routers.analytics_dashboards import (
        mtbf_mttr_timeseries, pm_compliance, backlog_aging, cost_per_equipment, summary,
    )
    wb = Workbook()
    # Summary
    ws = wb.active
    ws.title = "Summary"
    s = summary(plant_id=plant_id, db=db)
    _header(ws, ["KPI", "Value"])
    for k, v in s.items():
        ws.append(_safe_row([k, v]))

    # MTBF/MTTR
    ws2 = wb.create_sheet("MTBF-MTTR")
    _header(ws2, ["Month", "Failures", "MTBF (days)", "MTTR (hours)"])
    ts = mtbf_mttr_timeseries(plant_id=plant_id, months=12, db=db)
    for row in ts["series"]:
        ws2.append([row["month_label"], row["failures"], row.get("mtbf_days") or "", row.get("mttr_hours") or ""])

    # PM compliance
    ws3 = wb.create_sheet("PM Compliance")
    _header(ws3, ["Planning Group", "Total", "On-time", "Overdue", "Pending", "Compliance %"])
    pc = pm_compliance(plant_id=plant_id, db=db)
    for g in pc["by_group"]:
        ws3.append([g["group"], g["total"], g["on_time"], g["overdue"], g["pending"], g["compliance_pct"]])

    # Backlog aging
    ws4 = wb.create_sheet("Backlog Aging")
    _header(ws4, ["Age Range", "Count"])
    ba = backlog_aging(plant_id=plant_id, db=db)
    for b in ba["buckets"]:
        ws4.append([b["range"], b["count"]])

    # Cost per equipment
    ws5 = wb.create_sheet("Cost per Equipment")
    _header(ws5, ["Equipment", "WO Count", "Labor", "Material", "External", "Total Cost", "Actual HH"])
    cp = cost_per_equipment(plant_id=plant_id, limit=50, db=db)
    for r in cp["top"]:
        ws5.append([r["equipment_tag"], r["wo_count"], r["labor"], r["material"], r["external"], r["total_cost"], r["actual_hours"]])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    today = date.today().isoformat()
    return _wb_response(wb, f"kpi-summary-{today}.xlsx")


# ── Weekly Digest (PDF-friendly aggregator) ──────────────────────────
# Construye los datos para el reporte semanal de 1 página. El frontend
# lo renderiza como HTML imprimible y el usuario hace "Save as PDF".

@router.get("/weekly-digest")
def weekly_digest(
    plant_id: str | None = None,
    week_start: str | None = None,
    db: Session = Depends(get_db),
):
    """Weekly digest data: counts, KPIs, top equipment, vencidas.
    week_start ISO date (lunes). Default: lunes de la semana actual."""
    today = date.today()
    if week_start:
        try:
            ws_date = datetime.fromisoformat(week_start).date()
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid week_start")
    else:
        ws_date = today - timedelta(days=today.weekday())
    we_date = ws_date + timedelta(days=7)
    ws_dt = datetime.combine(ws_date, datetime.min.time())
    we_dt = datetime.combine(we_date, datetime.min.time())
    now = datetime.now()

    base_q = db.query(ManagedWorkOrderModel)
    if plant_id:
        base_q = base_q.filter(ManagedWorkOrderModel.plant_id == plant_id)

    # Counts dentro de la semana (planned)
    week_q = base_q.filter(
        ManagedWorkOrderModel.planned_start >= ws_dt,
        ManagedWorkOrderModel.planned_start < we_dt,
    )
    week_wos = week_q.all()
    created = sum(1 for w in week_wos if w.created_at and ws_dt <= w.created_at < we_dt)
    scheduled = sum(1 for w in week_wos if w.status in ("PROGRAMADO", "EN_PROGRAMACION"))
    in_exec = sum(1 for w in week_wos if w.status == "EN_EJECUCION")
    closed = sum(1 for w in week_wos if w.status in ("CERRADO", "COMPLETADO", "CLOSED"))
    canceled = sum(1 for w in week_wos if w.status in ("CANCELADO", "CANCELED"))

    # Adherencia / Cumplimiento sobre cerradas en la semana
    closed_in_week = [w for w in week_wos if w.status == "CERRADO" and w.planned_start]
    adherent = compliant = 0
    for w in closed_in_week:
        actual = w.actual_start or w.closed_at
        if actual and w.planned_start:
            delta_h = abs((actual - w.planned_start).total_seconds()) / 3600.0
            if delta_h <= 4.0:
                adherent += 1
            if delta_h <= 24.0 * 7:
                compliant += 1
    total_closed = len(closed_in_week)
    adherence_pct = round(adherent / total_closed * 100, 1) if total_closed else None
    compliance_pct = round(compliant / total_closed * 100, 1) if total_closed else None

    # HH plan vs actual de la semana (todas las cerradas en la ventana)
    hh_plan = sum(float(w.estimated_hours or 0) for w in week_wos)
    hh_actual = sum(float(w.actual_hours or 0) for w in week_wos if w.status in ("CERRADO", "COMPLETADO"))

    # Top 5 equipos por HH consumida en la semana
    by_eq: dict[str, dict] = {}
    for w in week_wos:
        tag = (w.equipment_tag or w.equipment_id or "—")
        if tag not in by_eq:
            by_eq[tag] = {"tag": tag, "wo_count": 0, "hh": 0.0}
        by_eq[tag]["wo_count"] += 1
        by_eq[tag]["hh"] += float(w.actual_hours or w.estimated_hours or 0)
    top_equipment = sorted(by_eq.values(), key=lambda x: -x["hh"])[:5]

    # OTs vencidas (planned_end pasado, status no terminal)
    overdue_q = base_q.filter(
        ManagedWorkOrderModel.planned_end.isnot(None),
        ManagedWorkOrderModel.planned_end < now,
        ~ManagedWorkOrderModel.status.in_(("CERRADO", "COMPLETADO", "CLOSED", "CANCELADO", "CANCELED")),
    )
    overdue_count = overdue_q.count()
    overdue_top = [
        {
            "wo_number": w.wo_number,
            "equipment_tag": w.equipment_tag or "",
            "priority": w.priority_code,
            "planned_end": w.planned_end.isoformat() if w.planned_end else None,
            "days_overdue": (now - w.planned_end).days if w.planned_end else 0,
        }
        for w in overdue_q.order_by(ManagedWorkOrderModel.planned_end.asc()).limit(5).all()
    ]

    # Distribución por prioridad de la semana
    prio_dist: dict[str, int] = {}
    for w in week_wos:
        p = w.priority_code or "—"
        prio_dist[p] = prio_dist.get(p, 0) + 1

    return {
        "plant_id": plant_id or "",
        "week_start": ws_date.isoformat(),
        "week_end": (we_date - timedelta(days=1)).isoformat(),
        "week_iso": f"{ws_date.isocalendar()[0]}-W{ws_date.isocalendar()[1]:02d}",
        "generated_at": now.isoformat(),
        "counts": {
            "created": created,
            "scheduled": scheduled,
            "in_execution": in_exec,
            "closed": closed,
            "canceled": canceled,
            "total_planned": len(week_wos),
        },
        "kpis": {
            "adherence_pct": adherence_pct,
            "compliance_pct": compliance_pct,
            "total_closed": total_closed,
            "hh_plan": round(hh_plan, 1),
            "hh_actual": round(hh_actual, 1),
            "hh_variance_pct": round((hh_actual - hh_plan) / hh_plan * 100, 1) if hh_plan else None,
        },
        "priority_distribution": prio_dist,
        "top_equipment": top_equipment,
        "overdue": {
            "count": overdue_count,
            "top": overdue_top,
        },
    }
