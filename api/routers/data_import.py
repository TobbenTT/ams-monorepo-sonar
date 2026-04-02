"""Data Import router — browse tables, upload Excel/CSV, map columns, execute imports."""

from __future__ import annotations

import csv
import io
import logging
import os
import tempfile
import uuid
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy import text
from sqlalchemy.orm import Session

from api.database.connection import get_db
from api.dependencies.auth import get_current_user

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/data-import",
    tags=["data-import"],
    dependencies=[Depends(get_current_user)],
)

SEED_DATA_DIR = Path(__file__).resolve().parent.parent.parent / "seed_data"
ALLOWED_EXT = {".xlsx", ".csv"}

# ── Helpers ───────────────────────────────────────────────────────

def _ensure_history_table(db: Session):
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS data_import_log (
            import_id TEXT PRIMARY KEY,
            table_name TEXT NOT NULL,
            filename TEXT NOT NULL,
            rows_imported INTEGER DEFAULT 0,
            mode TEXT DEFAULT 'append',
            status TEXT DEFAULT 'pending',
            error_message TEXT,
            created_at TEXT NOT NULL,
            user_id TEXT
        )
    """))
    db.commit()


def _read_xlsx(filepath: str, sheet_name: str | None = None):
    """Read xlsx with openpyxl, return list of dicts per sheet."""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheets = []
    target_sheets = [sheet_name] if sheet_name else wb.sheetnames
    for sn in target_sheets:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheets.append({"name": sn, "columns": [], "row_count": 0, "preview_rows": []})
            continue
        headers = [str(c) if c is not None else "col_%d" % i for i, c in enumerate(rows[0])]
        data_rows = rows[1:]
        preview = []
        for row in data_rows[:5]:
            preview.append({h: (_cell_val(v)) for h, v in zip(headers, row)})
        sheets.append({
            "name": sn,
            "columns": headers,
            "row_count": len(data_rows),
            "preview_rows": preview,
        })
    wb.close()
    return sheets


def _read_csv(filepath: str):
    """Read CSV, return single-sheet structure."""
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = list(reader)
    preview = rows[:5]
    return [{
        "name": "CSV",
        "columns": headers,
        "row_count": len(rows),
        "preview_rows": preview,
    }]


def _cell_val(v):
    """Convert cell value to JSON-safe type."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    return str(v)


# ── Endpoints ─────────────────────────────────────────────────────

@router.get("/tables")
def list_tables(db: Session = Depends(get_db)):
    """List all tables with row counts."""
    tables_raw = db.execute(text(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
    )).fetchall()
    result = []
    for (tname,) in tables_raw:
        count = db.execute(text("SELECT COUNT(*) FROM \"%s\"" % tname)).scalar()
        cols_raw = db.execute(text("PRAGMA table_info(\"%s\")" % tname)).fetchall()
        columns = [{"name": r[1], "type": r[2], "notnull": bool(r[3]), "pk": bool(r[5])} for r in cols_raw]
        result.append({"name": tname, "row_count": count, "columns": columns})
    return result


@router.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload Excel/CSV, return preview with sheets, columns, sample rows."""
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXT:
        raise HTTPException(400, "Unsupported file type: %s. Use .xlsx or .csv" % ext)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext, dir="/tmp")
    content = await file.read()
    tmp.write(content)
    tmp.close()

    try:
        if ext == ".xlsx":
            sheets = _read_xlsx(tmp.name)
        else:
            sheets = _read_csv(tmp.name)
    except Exception as e:
        os.unlink(tmp.name)
        raise HTTPException(400, "Error reading file: %s" % e)

    return {
        "filename": file.filename,
        "temp_path": tmp.name,
        "sheets": sheets,
    }


@router.post("/execute")
async def execute_import(
    file: UploadFile = File(...),
    table_name: str = Form(...),
    mode: str = Form("append"),
    sheet_name: str = Form(None),
    column_mapping: str = Form("{}"),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Execute the import: read file, map columns, insert into target table."""
    import json

    _ensure_history_table(db)

    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXT:
        raise HTTPException(400, "Unsupported: %s" % ext)

    try:
        mapping = json.loads(column_mapping)
    except json.JSONDecodeError:
        raise HTTPException(400, "Invalid column_mapping JSON")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext, dir="/tmp")
    content = await file.read()
    tmp.write(content)
    tmp.close()

    import_id = str(uuid.uuid4())
    now = datetime.utcnow().isoformat()
    user_id = user.get("user_id", "unknown") if isinstance(user, dict) else getattr(user, "user_id", "unknown")

    try:
        if ext == ".xlsx":
            sheets = _read_xlsx(tmp.name, sheet_name)
        else:
            sheets = _read_csv(tmp.name)

        if not sheets:
            raise HTTPException(400, "No data found in file")

        sheet = sheets[0]

        # Re-read all rows (not just preview)
        if ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(tmp.name, read_only=True, data_only=True)
            ws = wb[sheet["name"]]
            all_rows_raw = list(ws.iter_rows(values_only=True))
            wb.close()
            headers = [str(c) if c is not None else "col_%d" % i for i, c in enumerate(all_rows_raw[0])]
            data_rows = all_rows_raw[1:]
        else:
            with open(tmp.name, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                csv_rows = list(reader)
            data_rows = [tuple(row.get(h) for h in headers) for row in csv_rows]

        # Build effective mapping
        if not mapping:
            mapping = {c: c for c in headers}

        db_cols = list(mapping.values())
        src_indices = []
        for src_col in mapping.keys():
            if src_col in headers:
                src_indices.append(headers.index(src_col))
            else:
                raise HTTPException(400, "Source column '%s' not found in file" % src_col)

        # Validate target table
        tbl_check = db.execute(text(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=:t"
        ), {"t": table_name}).fetchone()
        if not tbl_check:
            raise HTTPException(400, "Table '%s' does not exist" % table_name)

        if mode == "replace":
            db.execute(text('DELETE FROM "%s"' % table_name))

        placeholders = ", ".join([":%s" % c for c in db_cols])
        col_names = ", ".join(['"%s"' % c for c in db_cols])
        insert_sql = 'INSERT INTO "%s" (%s) VALUES (%s)' % (table_name, col_names, placeholders)

        inserted = 0
        errors = []
        for i, row in enumerate(data_rows):
            try:
                vals = {}
                for j, db_col in enumerate(db_cols):
                    v = row[src_indices[j]] if src_indices[j] < len(row) else None
                    vals[db_col] = _cell_val(v)
                db.execute(text(insert_sql), vals)
                inserted += 1
            except Exception as e:
                errors.append("Row %d: %s" % (i + 2, e))
                if len(errors) >= 50:
                    break

        db.commit()

        status = "success" if not errors else "partial"
        error_msg = "; ".join(errors[:10]) if errors else None
        db.execute(text("""
            INSERT INTO data_import_log (import_id, table_name, filename, rows_imported, mode, status, error_message, created_at, user_id)
            VALUES (:id, :tbl, :fn, :rows, :mode, :status, :err, :ts, :uid)
        """), {
            "id": import_id, "tbl": table_name, "fn": file.filename,
            "rows": inserted, "mode": mode, "status": status,
            "err": error_msg, "ts": now, "uid": user_id,
        })
        db.commit()

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Import failed")
        try:
            db.execute(text("""
                INSERT INTO data_import_log (import_id, table_name, filename, rows_imported, mode, status, error_message, created_at, user_id)
                VALUES (:id, :tbl, :fn, 0, :mode, 'failed', :err, :ts, :uid)
            """), {
                "id": import_id, "tbl": table_name, "fn": file.filename,
                "mode": mode, "err": str(e), "ts": now, "uid": user_id,
            })
            db.commit()
        except Exception:
            pass
        raise HTTPException(500, "Import failed: %s" % e)
    finally:
        os.unlink(tmp.name)

    return {
        "import_id": import_id,
        "table_name": table_name,
        "rows_imported": inserted,
        "status": status,
        "errors": errors[:10] if errors else [],
    }


@router.get("/history")
def get_data_import_log(db: Session = Depends(get_db)):
    """List import history."""
    _ensure_history_table(db)
    rows = db.execute(text(
        "SELECT import_id, table_name, filename, rows_imported, mode, status, error_message, created_at, user_id "
        "FROM data_import_log ORDER BY created_at DESC LIMIT 50"
    )).fetchall()
    return [
        {
            "import_id": r[0], "table_name": r[1], "filename": r[2],
            "rows_imported": r[3], "mode": r[4], "status": r[5],
            "error_message": r[6], "created_at": r[7], "user_id": r[8],
        }
        for r in rows
    ]


@router.get("/templates")
def list_templates():
    """List available seed_data template files."""
    if not SEED_DATA_DIR.is_dir():
        return []

# -- AI Auto-Configure --

from pydantic import BaseModel
from typing import Any

class AIAnalyzeRequest(BaseModel):
    columns: list[str]
    sample_rows: list[dict[str, Any]]
    tables: list[dict[str, Any]]
    filename: str = ""

# Synonym groups for fuzzy column matching
_SYNONYMS = {
    'worker': {'employee', 'emp', 'staff', 'person', 'technician', 'tecnico'},
    'id': {'number', 'code', 'num', 'key'},
    'name': {'nombre', 'description', 'desc', 'title'},
    'date': {'fecha', 'time', 'timestamp', 'at'},
    'notification': {'aviso', 'request', 'alert', 'notif'},
    'order': {'wo', 'ot', 'workorder'},
    'equipment': {'equipo', 'equnr', 'asset', 'machine', 'tag'},
    'location': {'loc', 'func_loc', 'ubicacion', 'site', 'plant', 'area'},
    'cost': {'costo', 'price', 'amount', 'budget', 'expense'},
    'material': {'part', 'spare', 'component', 'item'},
    'type': {'tipo', 'class', 'category', 'kind'},
    'status': {'estado', 'state', 'condition'},
    'priority': {'prioridad', 'urgency', 'severity'},
    'shift': {'turno', 'pattern'},
    'specialty': {'especialidad', 'skill', 'trade', 'craft'},
    'plan': {'strategy', 'schedule', 'program', 'maintenance'},
    'created': {'creation', 'created_at', 'fecha_creacion'},
    'start': {'begin', 'from', 'inicio'},
    'end': {'finish', 'to', 'fin', 'until'},
    'hours': {'hh', 'duration', 'time', 'horas'},
    'quantity': {'qty', 'cant', 'amount', 'count'},
    'certification': {'certifications', 'cert', 'qualification'},
}

def _word_sim(a: str, b: str) -> float:
    """Score similarity between two column names."""
    import re as _re
    def tok(s):
        return set(_re.split(r'[_\-\s]+', s.lower())) - {''}
    wa, wb = tok(a), tok(b)
    if not wa or not wb:
        return 0
    if a.lower() == b.lower():
        return 1.0
    overlap = wa & wb
    if overlap:
        return len(overlap) / max(len(wa), len(wb))
    for key, syns in _SYNONYMS.items():
        all_w = {key} | syns
        if (wa & all_w) and (wb & all_w):
            return 0.6
    return 0

def _score_table(tinfo, excel_cols, filename):
    """Score how well a DB table matches the Excel file."""
    import re as _re
    tname = tinfo["name"].lower()
    db_cols = []
    for c in (tinfo.get("columns") or []):
        db_cols.append(c.get("name", "") if isinstance(c, dict) else str(c))

    score = 0.0

    # 1. Filename keyword match
    fname = _re.sub(r'^\d+_?', '', filename.lower().replace('.xlsx','').replace('.csv','').replace('-','_'))
    fname_words = set(fname.split('_')) - {'', 'data', 'seed', 'the', 'of', 'and', 'input'}
    tname_words = set(tname.split('_')) - {''}
    name_overlap = fname_words & tname_words
    if name_overlap:
        score += len(name_overlap) * 15

    # 2. Column fuzzy matching
    matched = 0
    total_sim = 0
    for ec in excel_cols:
        best = max((_word_sim(ec, dc) for dc in db_cols), default=0)
        if best >= 0.5:
            matched += 1
        total_sim += best

    if excel_cols:
        score += (matched / len(excel_cols)) * 40
        score += (total_sim / len(excel_cols)) * 20

    # 3. Non-empty table bonus
    if tinfo.get("row_count", 0) > 0:
        score += 5

    return round(score, 1)


@router.post("/ai-analyze")
def ai_analyze(req: AIAnalyzeRequest, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """AI suggests target table, column mapping, and data quality warnings."""
    import anthropic
    import json as _json
    from api.config import settings

    api_key = settings.ANTHROPIC_API_KEY
    if not api_key:
        raise HTTPException(500, "ANTHROPIC_API_KEY not configured")

    # Score and rank all tables
    scored = []
    for t in req.tables:
        s = _score_table(t, req.columns, req.filename)
        col_strs = []
        for c in (t.get("columns") or [])[:20]:
            if isinstance(c, dict):
                col_strs.append(c.get("name", "") + "(" + c.get("type", "") + ")")
            else:
                col_strs.append(str(c))
        scored.append((s, t["name"], ", ".join(col_strs), t.get("row_count", 0)))
    scored.sort(key=lambda x: -x[0])

    # Build table details for prompt (top 10)
    table_details = []
    for s, tname, cols, rcount in scored[:10]:
        table_details.append(f"{tname} (score:{s}, {rcount} rows): {cols}")

    # Get sample from top candidate
    top_sample = ""
    if scored and scored[0][0] > 10:
        try:
            top_name = scored[0][1]
            r = db.execute(text('SELECT * FROM "%s" LIMIT 2' % top_name))
            rows = [dict(row._mapping) for row in r]
            if rows:
                top_sample = "\nSample data in " + top_name + ":\n" + _json.dumps(rows[:2], default=str, ensure_ascii=False)[:500]
        except Exception:
            pass

    sample_json = _json.dumps(req.sample_rows[:3], default=str, ensure_ascii=False)[:1500]
    nl = chr(10)

    prompt = f"""You are a data import assistant for an industrial maintenance system (SAP PM / CMMS).

FILE: {req.filename}
EXCEL COLUMNS: {req.columns}
SAMPLE DATA:
{sample_json}
{top_sample}

TOP MATCHING TABLES (ranked by relevance score):
{nl.join(table_details)}

CRITICAL RULES:
- You MUST pick one of the existing tables above. NEVER set create_table_sql (always null).
- The first table has the highest algorithmic score - strongly prefer it unless data clearly belongs elsewhere.
- column_mapping: map Excel column names to EXISTING column names in the chosen table. Use semantic matching:
  employee_id->worker_id, shift_pattern->shift, certification->certifications, notification_number->wr_code, etc.
- Only include columns that have a reasonable semantic match.
- confidence: 80+ if topic matches and 3+ columns map, 60-79 if partial match, <60 if poor.
- warnings: unmapped columns and type concerns.

Respond ONLY with valid JSON (no markdown):
{{"suggested_table": "name", "confidence": 85, "alternatives": [{{"table": "other", "confidence": 30}}], "column_mapping": {{"excel_col": "db_col"}}, "warnings": ["warn"], "create_table_sql": null}}"""

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        text = response.content[0].text.strip()
        if "```" in text:
            parts = text.split("```")
            text = parts[1] if len(parts) > 1 else parts[0]
            if text.startswith("json"):
                text = text[4:]
        result = _json.loads(text.strip())
        return result
    except _json.JSONDecodeError:
        return {"suggested_table": "", "confidence": 0, "alternatives": [], "column_mapping": {}, "warnings": ["AI response was not valid JSON"], "create_table_sql": None}
    except Exception as e:
        raise HTTPException(500, f"AI analysis failed: {e}")
