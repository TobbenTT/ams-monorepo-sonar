"""Reports export — XLSX downloads for weekly schedule, closed WOs, KPIs."""

from datetime import datetime, timedelta, date
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from api.database.connection import get_db
from api.database.models import ManagedWorkOrderModel
from api.dependencies.auth import get_current_user

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


router = APIRouter(
    prefix="/reports-export",
    tags=["reports-export"],
    dependencies=[Depends(get_current_user)],
)


def _wb_response(wb, filename: str) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _header(ws, columns: list[str]):
    fill = PatternFill("solid", fgColor="1B5E20")
    font = Font(bold=True, color="FFFFFF")
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


@router.get("/weekly-schedule.xlsx")
def weekly_schedule_xlsx(
    plant_id: str | None = None,
    week_start: str | None = None,
    db: Session = Depends(get_db),
):
    """Weekly schedule export: all WOs with planned_start in the given week."""
    if not OPENPYXL_OK:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    if week_start:
        try:
            start = datetime.fromisoformat(week_start).date()
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid week_start")
    else:
        today = date.today()
        start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=7)

    q = db.query(ManagedWorkOrderModel).filter(
        ManagedWorkOrderModel.planned_start >= datetime.combine(start, datetime.min.time()),
        ManagedWorkOrderModel.planned_start < datetime.combine(end, datetime.min.time()),
    )
    if plant_id:
        q = q.filter(ManagedWorkOrderModel.plant_id == plant_id)
    wos = q.order_by(ManagedWorkOrderModel.planned_start.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Week {start.isoformat()}"
    _header(ws, [
        "WO Number", "Type", "Priority", "Equipment", "Description",
        "Planning Group", "Work Center", "Planned Start", "Planned End",
        "Est. Hours", "Shift", "Status", "Techs",
    ])
    for wo in wos:
        techs = ", ".join((w.get("name") or w.get("worker_id") or "") for w in (wo.assigned_workers or []))
        ws.append([
            wo.wo_number, wo.wo_type, wo.priority_code, wo.equipment_tag or "",
            wo.description or "", wo.planning_group or "", wo.work_center or "",
            wo.planned_start.isoformat() if wo.planned_start else "",
            wo.planned_end.isoformat() if wo.planned_end else "",
            float(wo.estimated_hours or 0),
            getattr(wo, "shift", "") or "day",
            wo.status, techs,
        ])
    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    return _wb_response(wb, f"weekly-schedule-{start.isoformat()}.xlsx")


@router.get("/wos-closed.xlsx")
def wos_closed_xlsx(
    plant_id: str | None = None,
    month: str | None = None,  # YYYY-MM
    db: Session = Depends(get_db),
):
    """Closed WOs for a month. Defaults to current month."""
    if not OPENPYXL_OK:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    today = date.today()
    if month:
        try:
            y, m = [int(x) for x in month.split("-")]
            start = date(y, m, 1)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid month")
    else:
        start = today.replace(day=1)
    end_month = start.replace(day=28) + timedelta(days=4)
    end = end_month.replace(day=1)

    q = db.query(ManagedWorkOrderModel).filter(
        ManagedWorkOrderModel.status == "CERRADO",
        ManagedWorkOrderModel.closed_at >= datetime.combine(start, datetime.min.time()),
        ManagedWorkOrderModel.closed_at < datetime.combine(end, datetime.min.time()),
    )
    if plant_id:
        q = q.filter(ManagedWorkOrderModel.plant_id == plant_id)
    wos = q.order_by(ManagedWorkOrderModel.closed_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Closed {start.strftime('%Y-%m')}"
    _header(ws, [
        "WO Number", "Type", "Priority", "Equipment", "Description",
        "Closed At", "Closed By", "Signed By",
        "Plan HH", "Actual HH", "Variance %",
        "Labor Cost", "Material Cost", "External Cost", "Total Cost",
        "Notes",
    ])
    total_plan = total_actual = total_cost = 0.0
    for wo in wos:
        plan = float(wo.estimated_hours or 0)
        actual = float(wo.actual_hours or 0)
        variance = round(((actual - plan) / plan) * 100, 1) if plan else 0
        cost = float(wo.actual_total_cost or 0) or (float(wo.labor_cost or 0) + float(wo.material_cost or 0) + float(wo.external_cost or 0))
        total_plan += plan
        total_actual += actual
        total_cost += cost
        ws.append([
            wo.wo_number, wo.wo_type, wo.priority_code, wo.equipment_tag or "",
            wo.description or "",
            wo.closed_at.isoformat() if wo.closed_at else "",
            wo.closed_by or "",
            getattr(wo, "closed_by_signature", "") or "",
            plan, actual, variance,
            float(wo.labor_cost or 0), float(wo.material_cost or 0), float(wo.external_cost or 0), cost,
            getattr(wo, "closure_notes", "") or "",
        ])
    # Totals row
    ws.append([])
    totals_row = ["TOTAL", "", "", "", "", "", "", "", total_plan, total_actual,
                  round(((total_actual - total_plan) / total_plan) * 100, 1) if total_plan else 0,
                  "", "", "", total_cost, ""]
    ws.append(totals_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)
    return _wb_response(wb, f"wos-closed-{start.strftime('%Y-%m')}.xlsx")


@router.get("/kpi-summary.xlsx")
def kpi_summary_xlsx(plant_id: str | None = None, db: Session = Depends(get_db)):
    """Multi-sheet KPI summary: MTBF/MTTR, PM compliance, backlog aging, cost."""
    if not OPENPYXL_OK:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    from api.routers.analytics_dashboards import (
        mtbf_mttr_timeseries, pm_compliance, backlog_aging, cost_per_equipment, summary,
    )
    wb = Workbook()
    # Summary
    ws = wb.active
    ws.title = "Summary"
    s = summary(plant_id=plant_id, db=db)
    _header(ws, ["KPI", "Value"])
    for k, v in s.items():
        ws.append([k, v])

    # MTBF/MTTR
    ws2 = wb.create_sheet("MTBF-MTTR")
    _header(ws2, ["Month", "Failures", "MTBF (days)", "MTTR (hours)"])
    ts = mtbf_mttr_timeseries(plant_id=plant_id, months=12, db=db)
    for row in ts["series"]:
        ws2.append([row["month_label"], row["failures"], row.get("mtbf_days") or "", row.get("mttr_hours") or ""])

    # PM compliance
    ws3 = wb.create_sheet("PM Compliance")
    _header(ws3, ["Planning Group", "Total", "On-time", "Overdue", "Pending", "Compliance %"])
    pc = pm_compliance(plant_id=plant_id, db=db)
    for g in pc["by_group"]:
        ws3.append([g["group"], g["total"], g["on_time"], g["overdue"], g["pending"], g["compliance_pct"]])

    # Backlog aging
    ws4 = wb.create_sheet("Backlog Aging")
    _header(ws4, ["Age Range", "Count"])
    ba = backlog_aging(plant_id=plant_id, db=db)
    for b in ba["buckets"]:
        ws4.append([b["range"], b["count"]])

    # Cost per equipment
    ws5 = wb.create_sheet("Cost per Equipment")
    _header(ws5, ["Equipment", "WO Count", "Labor", "Material", "External", "Total Cost", "Actual HH"])
    cp = cost_per_equipment(plant_id=plant_id, limit=50, db=db)
    for r in cp["top"]:
        ws5.append([r["equipment_tag"], r["wo_count"], r["labor"], r["material"], r["external"], r["total_cost"], r["actual_hours"]])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    today = date.today().isoformat()
    return _wb_response(wb, f"kpi-summary-{today}.xlsx")
