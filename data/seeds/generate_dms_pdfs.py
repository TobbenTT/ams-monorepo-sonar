"""
Generate pseudo-real DMS PDFs for demo purposes.
One PDF per row in 18_dms_maf_documents.xlsx.
Output: docs/{TYPE}/{number}.pdf  (matching file_path column)
"""

import os, sys, random, re, unicodedata
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

def _safe(text: str) -> str:
    """Transliterate accented chars to ASCII for Helvetica compatibility."""
    return unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")

import openpyxl
from fpdf import FPDF
from fpdf.enums import XPos, YPos

XLSX = os.path.join(os.path.dirname(__file__), "18_dms_maf_documents.xlsx")
OUT_BASE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "docs")

GOLD = (184, 148, 48)
DARK = (30, 30, 30)
GRAY = (100, 100, 100)
LGRAY = (240, 240, 240)
WHITE = (255, 255, 255)

COMPANY = "Goldfields - Salares Norte"
PLANT   = "Planta de Procesamiento - SN"

# ??? helpers ???????????????????????????????????????????????????????????????

def header(pdf: FPDF, doc_num: str, doc_type: str, title: str, version: int):
    # Gold top bar
    pdf.set_fill_color(*GOLD)
    pdf.rect(0, 0, 210, 14, "F")
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(*WHITE)
    pdf.set_xy(8, 3)
    pdf.cell(130, 8, COMPANY, new_x=XPos.RIGHT, new_y=YPos.TOP)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_xy(148, 3)
    pdf.cell(54, 4, f"N° {doc_num}", new_x=XPos.LEFT, new_y=YPos.NEXT, align="R")
    pdf.set_xy(148, 7)
    pdf.cell(54, 4, f"Rev. {version:02d}  |  {doc_type}", new_x=XPos.LEFT, new_y=YPos.NEXT, align="R")

    # Title bar
    pdf.set_fill_color(*DARK)
    pdf.rect(0, 14, 210, 12, "F")
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(*WHITE)
    pdf.set_xy(8, 16)
    short = title[:95] + "..." if len(title) > 95 else title
    pdf.cell(194, 8, short, new_x=XPos.LEFT, new_y=YPos.NEXT)

    pdf.set_text_color(*DARK)
    pdf.set_xy(8, 28)


def section(pdf: FPDF, title: str):
    pdf.ln(3)
    pdf.set_fill_color(*GOLD)
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_text_color(*WHITE)
    pdf.set_x(8)
    pdf.cell(194, 5, f"  {title.upper()}", fill=True,
             new_x=XPos.LEFT, new_y=YPos.NEXT)
    pdf.set_text_color(*DARK)
    pdf.set_font("Helvetica", "", 8)
    pdf.ln(1)


def body(pdf: FPDF, text: str):
    pdf.set_x(8)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*DARK)
    pdf.multi_cell(194, 4.5, text, new_x=XPos.LEFT, new_y=YPos.NEXT)


def row_table(pdf: FPDF, cells: list[tuple], widths: list[float], fill=False):
    pdf.set_x(8)
    if fill:
        pdf.set_fill_color(*LGRAY)
    for i, (txt, w) in enumerate(zip(cells, widths)):
        pdf.set_font("Helvetica", "B" if i == 0 and fill else "", 7.5)
        pdf.cell(w, 5.5, str(txt), border=1, fill=fill,
                 new_x=XPos.RIGHT if i < len(cells)-1 else XPos.LEFT,
                 new_y=YPos.TOP if i < len(cells)-1 else YPos.NEXT)


def footer(pdf: FPDF, doc_num: str, created_by: str, created_date: str):
    pdf.set_y(-14)
    pdf.set_fill_color(*LGRAY)
    pdf.rect(0, pdf.get_y(), 210, 14, "F")
    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(*GRAY)
    pdf.set_x(8)
    pdf.cell(90, 6, f"Elaborado por: {created_by or 'N/A'}   Fecha: {created_date or 'N/A'}", new_x=XPos.RIGHT, new_y=YPos.TOP)
    pdf.cell(60, 6, f"Documento: {doc_num}", new_x=XPos.RIGHT, new_y=YPos.TOP, align="C")
    pdf.cell(44, 6, f"CONFIDENCIAL ? {COMPANY}", new_x=XPos.LEFT, new_y=YPos.NEXT, align="R")
    pdf.set_x(8)
    pdf.cell(194, 5, f"Página {{nb}} de {{nb}}  |  {PLANT}", align="C")


# ??? content generators ????????????????????????????????????????????????????

SAFETY_ITEMS = [
    "Use EPP completo: casco, lentes de seguridad, guantes y calzado de seguridad.",
    "Bloqueo y etiquetado (LOTO) obligatorio antes de iniciar cualquier intervencion.",
    "Verifique ausencia de energia con multimetro calibrado antes de tocar componentes.",
    "Mantenga area de trabajo despejada y senalizada con conos y cinta reflectante.",
    "En caso de derrame de aceite, aplicar kit de derrames y reportar a supervision.",
]

def gen_pro(pdf, title, eq_name, eq_loc, doc_num, created_by, created_date):
    """Procedimiento de trabajo."""
    section(pdf, "1. Alcance y objetivo")
    body(pdf, f"Este procedimiento describe los pasos para ejecutar correctamente la tarea de {title.lower()} "
              f"en el equipo {eq_name or 'indicado'}, ubicado en {eq_loc or 'la planta de procesamiento'}. "
              f"Aplica al personal de mantenimiento calificado con autorización vigente.")

    section(pdf, "2. Equipos y herramientas requeridas")
    tools = ["Llave de impacto 1/2\"", "Torquímetro 0?300 Nm", "Extractor de rodamientos",
             "Calibrador vernier 150 mm", "Multímetro digital Cat III", "Linterna LED ATEX"]
    for t in tools:
        pdf.set_x(12)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(190, 4.5, f"? {t}", new_x=XPos.LEFT, new_y=YPos.NEXT)

    section(pdf, "3. Precauciones de seguridad")
    for s in random.sample(SAFETY_ITEMS, 4):
        pdf.set_x(12)
        pdf.set_font("Helvetica", "", 8)
        pdf.multi_cell(188, 4.5, f"?  {s}", new_x=XPos.LEFT, new_y=YPos.NEXT)

    section(pdf, "4. Pasos del procedimiento")
    steps = [
        "Notificar a supervisión y obtener permiso de trabajo firmado.",
        "Aplicar procedimiento LOTO en el tablero de control correspondiente.",
        "Verificar ausencia de presión/energía con instrumentos calibrados.",
        "Desconectar conexiones mecánicas y eléctricas identificando cada punto.",
        "Retirar componente defectuoso y registrar condición encontrada en formulario de terreno.",
        "Instalar componente nuevo verificando torques y especificaciones del fabricante.",
        "Reconectar en orden inverso al desmontaje. Verificar cada conexión.",
        "Retirar LOTO, energizar equipo y realizar prueba funcional supervisada.",
        "Completar registro de intervención en sistema OT y firma de supervisor.",
    ]
    for i, s in enumerate(steps, 1):
        pdf.set_x(12)
        pdf.set_font("Helvetica", "", 8)
        pdf.multi_cell(188, 4.8, f"{i}.  {s}", new_x=XPos.LEFT, new_y=YPos.NEXT)

    section(pdf, "5. Criterios de aceptación")
    body(pdf, "El equipo debe operar sin ruidos anómalos, vibraciones fuera de rango o fugas. "
              "Temperatura de rodamientos < 70 °C a los 30 min de operación. "
              "Corriente de operación dentro de ±10% del valor nominal en placa.")

    section(pdf, "6. Registros")
    body(pdf, f"? Formulario de terreno (FT-MANT-001)  ? OT N° referenciada  ? Fotografías antes/después  ? Firma supervisor")


def gen_maf(pdf, title, eq_name, eq_loc, doc_num, created_by, created_date):
    """Pauta de mantenimiento."""
    section(pdf, "1. Datos del equipo")
    body(pdf, f"Equipo: {eq_name or 'N/A'}   |   Ubicación: {eq_loc or 'N/A'}   |   Frecuencia: Según PM")

    section(pdf, "2. Tareas de mantenimiento preventivo")
    headers = ("N°", "Descripción de tarea", "Frec.", "Tiempo est.", "Especialidad")
    widths  = [10, 94, 22, 28, 32]
    row_table(pdf, headers, widths, fill=True)
    tasks = [
        ("Inspección visual general ? fugas, fisuras, estado de anclajes", "Mensual", "30 min", "Mecánico"),
        ("Lubricación de rodamientos según ficha técnica del fabricante", "Trimestral", "45 min", "Mecánico"),
        ("Medición de vibraciones y temperatura con analizador portátil", "Mensual", "20 min", "Predictivo"),
        ("Verificación de torques en pernos de fijación y bridas", "Semestral", "1 h", "Mecánico"),
        ("Limpieza de filtros de ventilación y enfriamiento", "Mensual", "20 min", "Operador"),
        ("Calibración de instrumentos de medición y control asociados", "Anual", "2 h", "Instrumentista"),
        ("Inspección de conexiones eléctricas ? apriete y corrosión", "Semestral", "30 min", "Electricista"),
        ("Prueba funcional post-mantenimiento y registro de parámetros", "C/intervención", "30 min", "Mecánico"),
    ]
    for i, (desc, freq, time, spec) in enumerate(tasks):
        row_table(pdf, (i+1, desc, freq, time, spec), widths)

    section(pdf, "3. Materiales y repuestos requeridos")
    mats = ["Grasa de litio EP-2 (según especificación OEM)", "Filtros de aceite hidráulico",
            "Empaquetaduras y sellos de repuesto", "Kit de pernos DIN 933 M16 Gr.8.8"]
    for m in mats:
        pdf.set_x(12)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(190, 4.5, f"? {m}", new_x=XPos.LEFT, new_y=YPos.NEXT)

    section(pdf, "4. Criterios de aceptación y rechazo")
    body(pdf, "Vibración global < 7.1 mm/s RMS (ISO 10816-3 Zona B).  "
              "Temperatura de carcasa < 80 °C.  Nivel de aceite dentro de rango visual.")

    section(pdf, "5. Responsable y firma")
    body(pdf, "Técnico ejecutor: ______________________________   Firma: _________   Fecha: ____________\n"
              "Supervisor revisión: ___________________________   Firma: _________   Fecha: ____________")


def gen_dwg(pdf, title, eq_name, eq_loc, doc_num, created_by, created_date):
    """Plano técnico (representación esquemática)."""
    section(pdf, "1. Información del plano")
    body(pdf, f"Equipo: {eq_name or 'N/A'}   |   Ubicación técnica: {eq_loc or 'N/A'}\n"
              f"Escala: 1:10   |   Formato: A4   |   Sistema de referencia: ISO 128")

    section(pdf, "2. Lista de materiales (BOM)")
    headers = ("Pos.", "Descripción", "Material", "Cant.", "Norma")
    widths  = [14, 90, 40, 18, 32]
    row_table(pdf, headers, widths, fill=True)
    items = [
        ("Cuerpo principal", "Acero ASTM A36", "1", "ASTM A36"),
        ("Brida de conexión DN150", "Acero carbono", "2", "ASME B16.5"),
        ("Perno de anclaje M20", "Acero 8.8", "8", "DIN 933"),
        ("Sello mecánico tipo cartucho", "AISI 316L", "1", "API 682"),
        ("Rodamiento de bolas 6310", "Cromo", "2", "ISO 355"),
        ("Tapa lateral ciega", "Acero A36", "2", "ASTM A36"),
    ]
    for i, (desc, mat, qty, norm) in enumerate(items, 1):
        row_table(pdf, (i, desc, mat, qty, norm), widths)

    section(pdf, "3. Diagrama esquemático")
    # ASCII-style diagram
    pdf.ln(2)
    pdf.set_x(30)
    pdf.set_font("Courier", "", 7)
    pdf.set_fill_color(*LGRAY)
    lines = [
        "  ???????????????????????????????????????????",
        "  ?           VISTA FRONTAL                 ?",
        "  ?   ????????????????????????????????      ?",
        "  ?   ?  ?  RODAMIENTO  ?            ?      ?",
        "  ?   ?     [==EJE PRINCIPAL==]      ?      ?",
        "  ?   ?  ?  RODAMIENTO  ?            ?      ?",
        "  ?   ????????????????????????????????      ?",
        "  ?   ? DN150   ? DN150                     ?",
        "  ?  IN ?       ? OUT                       ?",
        "  ???????????????????????????????????????????",
        "    COTAS EN mm   |   TOLERANCIA: ±0.5 mm   ",
    ]
    for line in lines:
        pdf.set_x(30)
        pdf.cell(150, 5, line, new_x=XPos.LEFT, new_y=YPos.NEXT)

    section(pdf, "4. Notas generales")
    body(pdf, "1. Todas las dimensiones en milímetros salvo indicación contraria.\n"
              "2. Soldaduras según AWS D1.1 ? Electrodo E7018.\n"
              "3. Pintura: fondo epoxi 75?m + acabado poliuretano 50?m, color RAL 5015.\n"
              "4. Verificar interferencias con estructura existente antes de fabricar.")


def gen_man(pdf, title, eq_name, eq_loc, doc_num, created_by, created_date):
    """Manual de operación."""
    section(pdf, "1. Descripción del equipo")
    body(pdf, f"El equipo {eq_name or 'referenciado'} está diseñado para operar en condiciones de proceso continuo "
              f"en {eq_loc or 'la planta de procesamiento'}. Este manual cubre los procedimientos de arranque, "
              f"operación normal, parada programada y parada de emergencia.")

    section(pdf, "2. Especificaciones técnicas de operación")
    headers = ("Parámetro", "Valor nominal", "Rango aceptable", "Unidad")
    widths  = [60, 45, 55, 34]
    row_table(pdf, headers, widths, fill=True)
    params = [
        ("Presión de operación", "6.5", "5.5 ? 7.5", "bar"),
        ("Temperatura proceso", "65", "55 ? 75", "°C"),
        ("Caudal nominal", "120", "100 ? 140", "m³/h"),
        ("Velocidad de rotación", "1480", "1450 ? 1510", "RPM"),
        ("Corriente nominal", "48", "44 ? 52", "A"),
        ("Potencia instalada", "30", "?", "kW"),
        ("Nivel de ruido", "78", "< 85", "dB(A)"),
    ]
    for p, v, r, u in params:
        row_table(pdf, (p, v, r, u), widths)

    section(pdf, "3. Secuencia de arranque")
    steps = [
        "Verificar nivel de aceite lubricante en visor lateral (entre marcas MIN?MAX).",
        "Confirmar que válvulas de succión y descarga estén en posición correcta.",
        "Energizar tablero de control y verificar ausencia de alarmas activas.",
        "Activar sistema de enfriamiento/sello mecánico según corresponda.",
        "Arrancar equipo desde HMI o panel local según modo de operación definido.",
        "Verificar parámetros de operación durante los primeros 5 minutos de marcha.",
        "Registrar valores iniciales en bitácora de operación.",
    ]
    for i, s in enumerate(steps, 1):
        pdf.set_x(12)
        pdf.set_font("Helvetica", "", 8)
        pdf.multi_cell(188, 4.8, f"{i}.  {s}", new_x=XPos.LEFT, new_y=YPos.NEXT)

    section(pdf, "4. Alarmas y acciones")
    headers2 = ("Alarma", "Causa probable", "Acción")
    widths2 = [50, 70, 74]
    row_table(pdf, headers2, widths2, fill=True)
    alarms = [
        ("Alta temperatura rodamiento", "Falta de lubricante / sobrecarga", "Detener ? lubricar ? revisar carga"),
        ("Baja presión succión", "Cavitación / válvula cerrada", "Verificar válvula y nivel de fluido"),
        ("Sobrecorriente motor", "Obstrucción / falla mecánica", "Detener ? inspeccionar ? reportar"),
    ]
    for a, c, ac in alarms:
        row_table(pdf, (a, c, ac), widths2)


def gen_chk(pdf, title, eq_name, eq_loc, doc_num, created_by, created_date):
    """Checklist de inspección."""
    section(pdf, "1. Identificación")
    body(pdf, f"Equipo: {eq_name or 'N/A'}  |  Ubicación: {eq_loc or 'N/A'}  |  Turno: _____  |  Inspector: ____________________")

    section(pdf, "2. Checklist de inspección")
    headers = ("N°", "Ítem a verificar", "OK", "NOK", "N/A", "Observación")
    widths  = [10, 90, 12, 12, 12, 58]
    row_table(pdf, headers, widths, fill=True)
    items_chk = [
        "Estado visual de carcasa y estructura (fisuras, corrosión, deformaciones)",
        "Nivel de lubricante en visor ? dentro de rango MIN/MAX",
        "Fugas de fluido (aceite, agua, proceso) en sellos y juntas",
        "Estado de cables y conexiones eléctricas ? sin daños visibles",
        "Temperatura de carcasa y rodamientos con pirómetro",
        "Vibración audible y/o excesiva durante operación",
        "Estado de guardas de protección mecánica ? instaladas y completas",
        "Señalética de seguridad visible y legible",
        "Instrumentos de control en rango nominal (presión, temperatura, flujo)",
        "Estado de anclaje y pernos de base ? sin aflojamiento",
        "Limpieza general del equipo y área circundante",
        "Funcionamiento de paros de emergencia (prueba mensual)",
    ]
    for i, item in enumerate(items_chk, 1):
        pdf.set_x(8)
        for j, (txt, w) in enumerate(zip([i, item, "", "", "", ""], widths)):
            pdf.set_font("Helvetica", "", 7.5)
            pdf.cell(w, 6, str(txt), border=1,
                     new_x=XPos.RIGHT if j < 5 else XPos.LEFT,
                     new_y=YPos.TOP if j < 5 else YPos.NEXT)

    section(pdf, "3. Observaciones generales")
    for _ in range(3):
        pdf.set_x(8)
        pdf.cell(194, 6, "", border=1, new_x=XPos.LEFT, new_y=YPos.NEXT)

    section(pdf, "4. Firmas")
    body(pdf, "Inspector: ________________________   Firma: _______   Fecha: ____________\n"
              "Supervisor: _______________________   Firma: _______   Fecha: ____________")


def gen_generic(pdf, title, eq_name, eq_loc, doc_num, created_by, created_date):
    """Fallback para tipos desconocidos."""
    section(pdf, "1. Descripción")
    body(pdf, f"Documento técnico: {title}\nEquipo: {eq_name or 'N/A'}\nUbicación: {eq_loc or 'N/A'}")
    section(pdf, "2. Contenido")
    body(pdf, "Consultar con el área de ingeniería de mantenimiento para el contenido detallado de este documento.")


GENERATORS = {
    "PRO": gen_pro,
    "MAF": gen_maf,
    "DWG": gen_dwg,
    "MAN": gen_man,
    "CHK": gen_chk,
}


def generate_pdf(row: dict) -> str:
    file_path = row["file_path"] or ""
    if not file_path:
        return None

    # e.g. /docs/DWG/000001.pdf  ? docs/DWG/000001.pdf
    rel = file_path.lstrip("/")
    out_path = os.path.join(OUT_BASE, rel.replace("/", os.sep))
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    doc_num    = _safe(row["document_number"] or "")
    doc_type   = _safe(row["document_type"] or "").upper()
    title      = _safe(row["document_desc"] or "Documento tecnico")
    version    = int(row["version"] or 1)
    eq_name    = _safe(row["equipment_name"] or "")
    eq_loc     = _safe(row["sap_func_loc"] or "")
    created_by = _safe(row["created_by"] or "")
    created_date = _safe(row["created_date"] or "")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()
    pdf.set_margins(8, 8, 8)

    header(pdf, doc_num, doc_type, title, version)

    gen_fn = GENERATORS.get(doc_type, gen_generic)
    gen_fn(pdf, title, eq_name, eq_loc, doc_num, created_by, created_date)

    footer(pdf, doc_num, created_by, created_date)

    pdf.output(out_path)
    return out_path


# ??? main ??????????????????????????????????????????????????????????????????
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active
headers_row = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

ok, skip = 0, 0
for row_vals in ws.iter_rows(min_row=2, values_only=True):
    row = dict(zip(headers_row, row_vals))
    if not row.get("file_path"):
        skip += 1
        continue
    try:
        generate_pdf(row)
        ok += 1
        if ok % 100 == 0:
            print(f"  {ok} PDFs generados...")
    except Exception as e:
        print(f"  ERROR {row.get('document_number')}: {e}")
        skip += 1

print(f"\nListo: {ok} PDFs generados, {skip} omitidos.")
print(f"Directorio: {OUT_BASE}")
