"""
rebuild_failure_modes_v2.py
===========================
RCM Functional-Analysis-driven failure mode generator (v2).

Methodology:
    Equipment → Catalog Parts (MIs) → classify_mi() → applicable FM codes
    → FM-MASTER-REFERENCE-ES for Spanish mechanism/cause/evidence/detection
    → RPN calculation → output 30-column 03_failure_modes.xlsx

Usage:
    python scripts/rebuild_failure_modes_v2.py          # full run
    python scripts/rebuild_failure_modes_v2.py --pilot   # only 3110PU0061 & 2110CV0001
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import argparse
import os
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Resolve paths relative to project root
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

sys.path.insert(0, str(SCRIPT_DIR))

from mi_class_rules import classify_mi, get_failure_modes_for_mi, MI_CLASS_RULES   # noqa: E402
from rcm_rules import (                                    # noqa: E402
    GFSN_CONTEXT,
    classify_subsystem,
    apply_altitude_factor,
)
from equipment_bom_templates import complete_equipment_mis  # noqa: E402

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
parser = argparse.ArgumentParser(description="RCM failure-mode generator v2")
parser.add_argument("--pilot", action="store_true",
                    help="Generate only for tags 3110PU0061 and 2110CV0001")
args = parser.parse_args()

PILOT_TAGS = {"3110PU0061", "2110CV0001"}
PILOT_AREAS = {"2110", "3110"}  # Include all equipment in these areas

# ===========================================================================
# EQART → CATALOG PROFILE MAPPING  (from v1 – kept identical)
# ===========================================================================
EQART_TO_PROFILE = {
    'BLAX': 'VENT_AX', 'BLSC': 'COMP_TOR', 'BULL': 'BULL_BUL',
    'CCTV': 'CÁM_CCTV', 'CELO': 'CELOSÌA', 'ELHE': 'CAL_ELÉC',
    'EMBT': 'MOT_ELÉC', 'HBFF': 'SOP_CENT', 'HECQ': 'INT_ENFR',
    'HLCN': 'PUEN_GRÚ', 'LDLU': 'SIST_LUB', 'MSST': 'AGIT_MEC',
    'MXAG': 'AGIT_MEC', 'PCDT': 'DET_META', 'POOL': 'PIS_SUM',
    'PUCE': 'BOMB_CEN', 'PURE': 'BOMB_REC', 'PURO': 'BOMB_ROT',
    'PUTU': 'BOMB_TUR', 'PUVA': 'BOMB_VAC', 'SERV': 'SERVIDOR',
    'SFPR': 'FIL_PREN', 'SHEW': 'ESPESADO', 'SRCV': 'CORR_TRA',
    'SRFE': 'ALIM_BAN', 'VEBN': 'TOLVA', 'VEBX': 'CAJÒN',
    'VECH': 'CHUTE', 'VEDR': 'SECA_AIR', 'VELN': 'LIN_PROC',
    'VETA': 'EST_AGIT', 'VIBR': 'VIBRADOR',
    'ACOU': 'UNID_AC', 'AEID': 'INCH_DRI', 'AESS': 'SIST_LAN',
    'AETS': 'SOPORTES', 'AEWM': 'MON_LAVA', 'BCOM': 'TAB_COMU',
    'BLEH': 'CAMP_EXT', 'BLFC': 'VENT_CEN', 'BSCI': 'DET_FH',
    'BUVA': 'LIN_PROC', 'CEDE': 'BOMB_CEN', 'CFOF': 'DIAL_ACE',
    'CMCS': 'CORE_SWI', 'COHI': 'COMP_HID', 'CPLC': 'GAB_PLC',
    'CRIO': 'GAB_RIO', 'CSSP': 'PANE_SCA', 'CTDP': 'TAB_CEL',
    'DECH': 'GAB_MANG', 'DECO': 'CONS_ESC', 'DEFI': 'EST_MANG',
    'DEHY': 'HIDRANTE', 'EABA': 'BANC_BAT', 'EABC': 'CARG_BAT',
    'EAUP': 'UPS', 'ELCB': 'CEL_IPOT', 'ELDI': 'CEL_IPOT',
    'ELIN': 'CEL_PMOT', 'ELPS': 'CEL_PMOT', 'EMRE': 'CARG_FRO',
    'EYWA': 'DUCH_EME', 'FIPS': 'TAB_SCI', 'FIWA': 'FIREWALL',
    'GAVA': 'LIN_PROC', 'HBEH': 'HORN_ELÉ', 'HEPL': 'INT_PLAC',
    'HIWA': 'HIDR_AGU', 'HMIB': 'GAB_PLC', 'HTEL': 'TECL_ELE',
    'HTHA': 'TECL_ELE', 'HYPU': 'UNID_HID', 'INDP': 'TAB_DIST',
    'INPR': 'IND_PROC', 'KNVA': 'VALVULAS', 'KVMS': 'KVMSWITC',
    'LEBT': 'CAM_PLUM', 'LEFL': 'MONT_CAR', 'LIDP': 'TAB_DIST',
    'LIIN': 'TECL_ELE', 'LOFR': 'CARG_FRO', 'LUME': 'PAQ_QUIM',
    'MBAY': 'MUES_TRA', 'MBSC': 'BASCULA', 'MCSC': 'HARN_EST',
    'MFSU': 'SUPR_MF', 'MIRI': 'IND_PROC', 'MOPL': 'PLAT_MÓV',
    'MTTR': 'CARR_MNT', 'OCRD': 'PUER_MOT', 'PCCC': 'CICL_POT',
    'PCCO': 'COL_POLV', 'PCPT': 'TRAN_POT', 'PREU': 'UNI_PRES',
    'PRLI': 'DET_FH', 'PTWW': 'PTA_AR', 'PUSC': 'BOMB_HAR',
    'PWDP': 'TAB_DIST', 'ROVA': 'VALVULAS', 'RPHA': 'REM_PERN',
    'SAPR': 'MUES_ROT', 'SASE': 'SEN_PROC', 'SAWS': 'MUES_ROT',
    'SBDP': 'TAB_DIST', 'SBLR': 'ARRA_LÍQ', 'SECT': 'CÁM_CCTV',
    'SEFL': 'SEN_PROC', 'SFCF': 'FIL_COAL', 'SFCL': 'FIL_HOJA',
    'SMST': 'CHIMENEA', 'SRRT': 'MESA_ROD', 'SSEP': 'TAB_DIST',
    'STIT': 'CAM_RIEG', 'SUPP': 'SOPORTES', 'TAPR': 'EST_ALMA',
    'TOLH': 'MANI_COR', 'TRDE': 'LIN_PROC', 'TRFL': 'LIN_PROC',
    'TRLE': 'TOLVA', 'TRPR': 'TRAN_SEÑ', 'TRTE': 'TRAN_SEÑ',
    'UPDP': 'TAB_DIST', 'USAP': 'TAB_CEL', 'USCT': 'TAB_CEL',
    'USPS': 'TAB_CEL', 'VABU': 'VALVULAS', 'VACO': 'VALVULAS',
    'VAKN': 'VALVULAS', 'VAPI': 'VALVULAS', 'VEAL': 'EST_ALMA',
    'VEAR': 'ACUM_AIR', 'VEMO': 'CASC_MOL', 'VEPC': 'COL_PROC',
    'VETR': 'RACK_BAN', 'VEWR': 'PLAN_POT', 'VFBT': 'CELD_VDF',
    'VVCP': 'SUPR_POL', 'DUCT': 'DUCTOS', 'HETU': 'BOMB_TUR',
    'ROBL': 'MESA_ROD', 'SRPL': 'PLAT_MÓV', 'MCTR': 'TRN_PROC',
    'LUBR': 'SIST_LUB',
    # --- Unmapped EQART (52 codes, 1,141 equipment) ---
    # Válvulas (284 equipos)
    'BAVA': 'VALVULAS', 'COVA': 'VALVULAS', 'DAVA': 'VALVULAS',
    'DIVA': 'VALVULAS', 'DRVA': 'VALVULAS', 'GLVA': 'VALVULAS',
    'KGVA': 'VALVULAS', 'PIVA': 'VALVULAS', 'PSVA': 'VALVULAS',
    # Switches (411 equipos)
    'COSE': 'SEN_PROC', 'SSHA': 'SEN_PROC',
    'SWFL': 'SEN_PROC', 'SWLE': 'SEN_PROC', 'SWMI': 'SEN_PROC',
    'SWPD': 'SEN_PROC', 'SWPO': 'SEN_PROC', 'SWPR': 'SEN_PROC',
    'SWSB': 'SEN_PROC', 'SWSD': 'SEN_PROC', 'SWSP': 'SEN_PROC',
    'SWTE': 'SEN_PROC', 'SWTO': 'SEN_PROC',
    # Transmisores (85 equipos)
    'TRAN': 'TRAN_SEÑ', 'TRC2': 'TRAN_SEÑ', 'TRCN': 'TRAN_SEÑ',
    'TRGA': 'TRAN_SEÑ', 'TRPD': 'TRAN_SEÑ', 'TRPH': 'TRAN_SEÑ',
    'TRPO': 'TRAN_SEÑ', 'TRSP': 'TRAN_SEÑ', 'TRTO': 'TRAN_SEÑ',
    'TRVI': 'TRAN_SEÑ', 'TRWE': 'TRAN_SEÑ', 'TRWQ': 'TRAN_SEÑ',
    # Sensores (51 equipos)
    'LESE': 'SEN_PROC', 'SETE': 'SEN_PROC', 'SESP': 'SEN_PROC',
    'SEME': 'DET_META',
    # Indicadores (17 equipos)
    'INNI': 'IND_PROC', 'INTE': 'IND_PROC', 'VILE': 'IND_PROC',
    # Señalización (96 equipos)
    'SIBE': 'SEÑAL', 'SIRE': 'SEÑAL',
    # Filtros (10 equipos)
    'CFAI': 'FIL_AIRE', 'CFRE': 'FIL_REFR', 'CFWA': 'FIL_AGUA',
    # Contraincendios (40 equipos)
    'FIMO': 'DET_FH', 'SCIG': 'DET_FH',
    # Otros (24 equipos)
    'BURN': 'HORN_ELÉ', 'SOFU': 'PLAN_POT', 'LECT': 'PUEN_GRÚ',
    'WEME': 'BASCULA',
}

# ===========================================================================
# FUNCTION TEMPLATES (equipment-level RCM function description)
# ===========================================================================
FUNCTION_TEMPLATES = {
    'BOMB_CEN': "Bombear fluido de proceso a caudal de diseño en {sys}, {area}",
    'BOMB_REC': "Dosificar fluido a presión en {sys}, {area}",
    'BOMB_ROT': "Bombear fluido manteniendo presión constante en {sys}, {area}",
    'BOMB_TUR': "Bombear fluido a caudal de diseño en {sys}, {area}",
    'BOMB_VAC': "Generar vacío para proceso en {sys}, {area}",
    'BOMB_HAR': "Bombear pulpa del harnero en {sys}, {area}",
    'VALVULAS': "Controlar flujo de proceso manteniendo regulación en {sys}, {area}",
    'MOT_ELÉC': "Accionar equipo proporcionando potencia mecánica de diseño en {sys}, {area}",
    'CELD_VDF': "Regular velocidad del equipo accionado mediante variación de frecuencia en {sys}, {area}",
    'INCH_DRI': "Proporcionar giro lento para mantenimiento en {sys}, {area}",
    'CICL_POT': "Convertir potencia eléctrica para accionamiento del equipo en {sys}, {area}",
    'ARRA_LÍQ': "Arrancar equipo limitando corriente mediante resistencia líquida en {sys}, {area}",
    'COMP_TOR': "Comprimir aire a presión de diseño para sistema neumático en {sys}, {area}",
    'ACUM_AIR': "Almacenar aire comprimido a presión de diseño en {sys}, {area}",
    'SECA_AIR': "Secar aire comprimido removiendo humedad en {sys}, {area}",
    'SOP_CENT': "Suministrar aire de proceso a caudal de diseño en {sys}, {area}",
    'CORR_TRA': "Transportar material a capacidad de diseño en {sys}, {area}",
    'ALIM_BAN': "Alimentar material de forma controlada en {sys}, {area}",
    'HARN_EST': "Clasificar material por tamaño según apertura de malla en {sys}, {area}",
    'EST_AGIT': "Contener y agitar fluido manteniendo homogeneidad en {sys}, {area}",
    'EST_ALMA': "Almacenar fluido a capacidad nominal en {sys}, {area}",
    'TOLVA': "Almacenar material proveyendo descarga controlada en {sys}, {area}",
    'CAJÒN': "Contener y distribuir fluido hacia equipos aguas abajo en {sys}, {area}",
    'CHUTE': "Transferir material por gravedad entre equipos en {sys}, {area}",
    'PIS_SUM': "Contener fluido para recirculación o tratamiento en {sys}, {area}",
    'ESPESADO': "Espesar pulpa separando sólidos de líquido en {sys}, {area}",
    'FIL_PREN': "Filtrar fluido separando sólidos hasta humedad de diseño en {sys}, {area}",
    'FIL_HOJA': "Filtrar fluido removiendo sólidos suspendidos en {sys}, {area}",
    'FIL_COAL': "Filtrar fluido removiendo partículas finas en {sys}, {area}",
    'INT_ENFR': "Intercambiar calor enfriando fluido en {sys}, {area}",
    'INT_PLAC': "Intercambiar calor entre fluidos mediante placas en {sys}, {area}",
    'PUEN_GRÚ': "Izar y posicionar cargas para mantenimiento en {sys}, {area}",
    'TECL_ELE': "Izar cargas para mantenimiento en {sys}, {area}",
    'MANI_COR': "Manipular revestimientos durante mantenimiento de molino en {sys}, {area}",
    'AGIT_MEC': "Agitar fluido manteniendo sólidos en suspensión en {sys}, {area}",
    'TAB_CEL': "Distribuir y proteger suministro eléctrico en {sys}, {area}",
    'TAB_DIST': "Distribuir energía eléctrica a cargas del {sys}, {area}",
    'TAB_COMU': "Proveer comunicación de datos entre equipos del {sys}, {area}",
    'CEL_IPOT': "Proteger y seccionar circuitos de potencia en {sys}, {area}",
    'CEL_PMOT': "Arrancar y proteger motores eléctricos del {sys}, {area}",
    'TRAN_POT': "Transformar voltaje para distribución en {sys}, {area}",
    'UPS': "Proveer energía ininterrumpida a cargas críticas del {sys}, {area}",
    'BANC_BAT': "Almacenar energía eléctrica para respaldo en {sys}, {area}",
    'CARG_BAT': "Cargar banco de baterías del {sys}, {area}",
    'GAB_PLC': "Controlar proceso mediante lógica programable en {sys}, {area}",
    'GAB_RIO': "Adquirir señales de campo para PLC en {sys}, {area}",
    'PANE_SCA': "Visualizar y operar proceso desde sala de control en {sys}, {area}",
    'CONS_ESC': "Operar y monitorear proceso desde estación de trabajo en {sys}, {area}",
    'IND_PROC': "Indicar variable de proceso localmente en {sys}, {area}",
    'SEN_PROC': "Medir variable de proceso y transmitir señal en {sys}, {area}",
    'TRN_PROC': "Transmitir señal de medición al sistema de control en {sys}, {area}",
    'TRAN_SEÑ': "Transmitir señal de medición al sistema de control en {sys}, {area}",
    'SERVIDOR': "Procesar y almacenar datos del sistema de control en {sys}, {area}",
    'CORE_SWI': "Proveer conectividad de red entre equipos en {sys}, {area}",
    'FIREWALL': "Proteger red de control contra acceso no autorizado en {sys}, {area}",
    'KVMSWITC': "Compartir teclado/video/mouse entre servidores en {sys}, {area}",
    'DET_FH': "Detectar fuego y humo activando alarma en {sys}, {area}",
    'TAB_SCI': "Controlar sistema contraincendios en {sys}, {area}",
    'SUPR_MF': "Suprimir fuego mediante descarga de agente extintor en {sys}, {area}",
    'HIDRANTE': "Proveer agua contra incendios a presión en {sys}, {area}",
    'EST_MANG': "Proveer manguera contra incendios en {sys}, {area}",
    'GAB_MANG': "Almacenar y proteger mangueras contra incendios en {sys}, {area}",
    'DUCH_EME': "Proveer ducha y lavaojos de emergencia en {sys}, {area}",
    'UNID_AC': "Climatizar espacio manteniendo temperatura de diseño en {sys}, {area}",
    'UNI_PRES': "Presurizar sala evitando ingreso de polvo en {sys}, {area}",
    'CAL_ELÉC': "Calentar espacio manteniendo temperatura de confort en {sys}, {area}",
    'VENT_AX': "Ventilar espacio extrayendo aire viciado en {sys}, {area}",
    'VENT_CEN': "Ventilar espacio mediante extracción centralizada en {sys}, {area}",
    'CAMP_EXT': "Capturar y extraer gases/polvo del área de trabajo en {sys}, {area}",
    'DUCTOS': "Conducir aire/gases entre equipos de ventilación en {sys}, {area}",
    'CHIMENEA': "Descargar gases tratados a la atmósfera en {sys}, {area}",
    'LIN_PROC': "Transportar fluido manteniendo integridad y estanqueidad en {sys}, {area}",
    'SIST_LUB': "Lubricar componentes mecánicos manteniendo película de aceite en {sys}, {area}",
    'DIAL_ACE': "Filtrar y purificar aceite de lubricación en {sys}, {area}",
    'UNID_HID': "Proveer potencia hidráulica a presión de diseño en {sys}, {area}",
    'COMP_HID': "Controlar flujo mediante apertura/cierre hidráulico en {sys}, {area}",
    'CÁM_CCTV': "Monitorear visualmente área para seguridad en {sys}, {area}",
    'SOPORTES': "Soportar equipos y tuberías manteniendo alineación en {sys}, {area}",
    'DET_META': "Detectar metales protegiendo equipos aguas abajo en {sys}, {area}",
    'SUPR_POL': "Suprimir polvo fugitivo mediante aspersión en {sys}, {area}",
    'VIBRADOR': "Fluidizar material evitando apelmazamiento en {sys}, {area}",
    'COL_POLV': "Capturar y remover polvo del aire en {sys}, {area}",
    'COL_PROC': "Realizar proceso de contacto contracorriente en {sys}, {area}",
    'BULL_BUL': "Empujar y apilar material en zona de acopio en {sys}, {area}",
    'CAM_PLUM': "Izar y posicionar cargas mediante pluma hidráulica en {sys}, {area}",
    'CAM_RIEG': "Regar caminos suprimiendo polvo fugitivo en {sys}, {area}",
    'CARG_FRO': "Cargar y transportar material en {sys}, {area}",
    'HORN_ELÉ': "Calentar material a temperatura de proceso en {sys}, {area}",
    'MUES_TRA': "Obtener muestra representativa para análisis en {sys}, {area}",
    'MUES_ROT': "Obtener muestra representativa para análisis en {sys}, {area}",
    'BASCULA': "Pesar material con precisión de diseño en {sys}, {area}",
    'CELOSÌA': "Proveer soporte estructural y contención en {sys}, {area}",
    'PUER_MOT': "Permitir acceso vehicular mediante apertura/cierre en {sys}, {area}",
    'PLAT_MÓV': "Proveer plataforma de trabajo móvil en {sys}, {area}",
    'CARR_MNT': "Transportar herramientas para mantenimiento en {sys}, {area}",
    'REM_PERN': "Remover pernos durante mantenimiento de molino en {sys}, {area}",
    'CASC_MOL': "Conducir metal fundido desde horno hasta moldeo en {sys}, {area}",
    'RACK_BAN': "Soportar y organizar bandejas de proceso en {sys}, {area}",
    'MON_LAVA': "Lavar equipos mediante chorro de agua a presión en {sys}, {area}",
    'SIST_LAN': "Inyectar agua/aire a presión para limpieza en {sys}, {area}",
    'PAQ_QUIM': "Dosificar químicos de proceso a tasa controlada en {sys}, {area}",
    'PLAN_POT': "Potabilizar agua cumpliendo normas sanitarias en {sys}, {area}",
    'PTA_AR': "Tratar agua residual cumpliendo normas de descarga en {sys}, {area}",
    'HIDR_AGU': "Presurizar y distribuir agua potable en {sys}, {area}",
    'MONT_CAR': "Elevar y transportar cargas en {sys}, {area}",
    'MESA_ROD': "Transportar material sobre rodillos en {sys}, {area}",
    'ALIM_PLA': "Alimentar material mediante placas a caudal controlado en {sys}, {area}",
}

# ===========================================================================
# LOAD DATA
# ===========================================================================
print("=" * 60, flush=True)
print("RCM FAILURE-MODE GENERATOR v2", flush=True)
print("=" * 60, flush=True)

# --- 1. Equipment Hierarchy ---
print("Loading 01_equipment_hierarchy.xlsx ...", flush=True)
wb_hier = openpyxl.load_workbook(
    str(PROJECT_ROOT / "seed_data" / "01_equipment_hierarchy.xlsx"), data_only=True
)
ws_hier = wb_hier["Equipment Hierarchy"]
h_headers = [ws_hier.cell(1, c).value for c in range(1, ws_hier.max_column + 1)]
h_idx = {h: i + 1 for i, h in enumerate(h_headers)}

# Area / system name lookup
area_names = {}
system_names = {}
for row in range(2, ws_hier.max_row + 1):
    level = ws_hier.cell(row, h_idx["level"]).value
    fl = ws_hier.cell(row, h_idx["sap_func_loc"]).value
    pltxt = ws_hier.cell(row, h_idx["pltxt"]).value
    if level == 1 and fl:
        area_names[fl] = pltxt
    elif level == 3 and fl:
        system_names[fl] = pltxt

# L4 equipment
equipment_l4 = []
for row in range(2, ws_hier.max_row + 1):
    row_level = ws_hier.cell(row, h_idx["level"]).value
    if row_level not in (4, 5):
        continue
    equnr = ws_hier.cell(row, h_idx["equnr"]).value
    if not equnr:
        continue
    fl = ws_hier.cell(row, h_idx["sap_func_loc"]).value or ""
    parts = fl.split("-")
    area_fl = "-".join(parts[:2]) if len(parts) >= 2 else ""
    sys_fl = "-".join(parts[:4]) if len(parts) >= 4 else ""
    sap_func_loc_short = (
        ws_hier.cell(row, h_idx["sap_func_loc_short"]).value or fl.split("-")[-1]
    )
    equipment_l4.append(
        {
            "equnr": equnr,
            "sap_func_loc": fl,
            "sap_func_loc_short": sap_func_loc_short,
            "pltxt": ws_hier.cell(row, h_idx["pltxt"]).value,
            "eqart": ws_hier.cell(row, h_idx["eqart"]).value,
            "eqart_desc": ws_hier.cell(row, h_idx["eqart_desc"]).value,
            "abckz": ws_hier.cell(row, h_idx["abckz"]).value,
            "area_name": area_names.get(area_fl, parts[1] if len(parts) > 1 else ""),
            "system_name": system_names.get(sys_fl, ""),
        }
    )

print(f"  L4 equipment loaded: {len(equipment_l4)}", flush=True)

# --- 2. Catalog profiles (partes / causas / sintomas by profile code) ---
print("Loading 15_catalog_profiles.xlsx ...", flush=True)
wb_cat = openpyxl.load_workbook(
    str(PROJECT_ROOT / "seed_data" / "15_catalog_profiles.xlsx"), data_only=True
)
ws_cat = wb_cat[wb_cat.sheetnames[1]]  # "Catálogos" sheet

catalog_data = defaultdict(lambda: {"causas": [], "partes": [], "sintomas": []})
for r in range(2, ws_cat.max_row + 1):
    pc = ws_cat.cell(r, 1).value       # Cód. PC
    tipo = ws_cat.cell(r, 3).value      # Tipo catálogo
    valor = ws_cat.cell(r, 5).value     # Síntomas/ causas/ partes
    if not pc or not valor:
        continue
    tipo_str = str(tipo).strip()
    if "Causas" in tipo_str:
        catalog_data[pc]["causas"].append(str(valor))
    elif "Partes" in tipo_str:
        catalog_data[pc]["partes"].append(str(valor))
    elif "ntomas" in tipo_str:          # Síntomas (handles encoding)
        catalog_data[pc]["sintomas"].append(str(valor))

print(f"  Catalog profiles loaded: {len(catalog_data)}", flush=True)

# --- 3. FM-MASTER-REFERENCE-ES (Spanish) ---
print("Loading FM-MASTER-REFERENCE-ES.xlsx ...", flush=True)
wb_fm = openpyxl.load_workbook(
    str(
        PROJECT_ROOT
        / "skills"
        / "00-knowledge-base"
        / "data-models"
        / "failure-modes"
        / "FM-MASTER-REFERENCE-ES.xlsx"
    ),
    data_only=True,
)
ws_fm = wb_fm[wb_fm.sheetnames[0]]  # "72 Modos de Falla"
fm_headers = [ws_fm.cell(1, c).value for c in range(1, ws_fm.max_column + 1)]
fm_idx = {h: i + 1 for i, h in enumerate(fm_headers)}

fm_data = {}
for r in range(2, ws_fm.max_row + 1):
    fm_num = ws_fm.cell(r, fm_idx["FM#"]).value
    if not fm_num:
        break
    fm_data[fm_num] = {
        "mechanism_es": ws_fm.cell(r, fm_idx["Mecanismo"]).value or "",
        "cause_es": ws_fm.cell(r, fm_idx["Causa"]).value or "",
        "pattern": ws_fm.cell(r, fm_idx["Patron"]).value or "B",
        "detection": ws_fm.cell(r, fm_idx["Tecnica CBM Principal"]).value or "",
        "evidence": ws_fm.cell(r, fm_idx["Principales Condiciones P"]).value or "",
    }

print(f"  FM-MASTER entries loaded: {len(fm_data)}", flush=True)

# --- 4. Criticality ---
print("Loading 02_criticality_assessment.xlsx ...", flush=True)
wb_crit = openpyxl.load_workbook(
    str(PROJECT_ROOT / "seed_data" / "02_criticality_assessment.xlsx"), data_only=True
)
ws_crit = wb_crit[wb_crit.sheetnames[0]]
c_headers = [ws_crit.cell(1, c).value for c in range(1, ws_crit.max_column + 1)]
c_idx = {h: i + 1 for i, h in enumerate(c_headers)}

# Find downtime column
downtime_col = None
for h in c_headers:
    if h and "downtime" in str(h).lower():
        downtime_col = c_idx[h]
        break

criticality_by_fl = {}
for r in range(2, ws_crit.max_row + 1):
    fl = ws_crit.cell(r, c_idx.get("sap_func_loc", 1)).value
    if fl:
        crit = ws_crit.cell(
            r, c_idx.get("criticality_level", c_idx.get("abckz", 1))
        ).value
        dt = ws_crit.cell(r, downtime_col).value if downtime_col else None
        criticality_by_fl[fl] = {"criticality": crit, "downtime": dt or 12}

print(f"  Criticality entries: {len(criticality_by_fl)}", flush=True)

# ===========================================================================
# RPN HELPER MAPS
# ===========================================================================
SEVERITY_MAP = {
    1: 8, 2: 5, 3: 3,
    1.0: 8, 2.0: 5, 3.0: 3,
    "1": 8, "2": 5, "3": 3,
    "A": 8, "B": 5, "C": 3,
}

OCCURRENCE_BY_PATTERN = {"A": 3, "B": 5, "C": 6, "D": 4, "E": 7, "F": 8}

DETECTION_SCORE_MAP = {
    "vibracion": 3, "vibration": 3,
    "aceite": 4, "oil": 4, "lubricante": 4,
    "termogra": 4, "thermo": 4, "temperatura": 4, "insul": 4,
    "resistencia": 4, "megger": 4,
    "visual": 6, "inspect": 6, "evaluacion": 6,
    "mpi": 5, "ut": 5, "ultrasonido": 5, "ndt": 5, "espesor": 5,
    "calibracion": 5, "calibration": 5,
    "diferencial": 5, "presion": 5,
    "corriente": 4, "current": 4,
    "torque": 5,
    "dureza": 5,
    "monitor": 3,
    "online": 3,
}


def rpn_detection_score(detection_text: str) -> int:
    """Map detection technique text to RPN detection score (lower = easier to detect)."""
    dt = (detection_text or "").lower()
    for keyword, score in DETECTION_SCORE_MAP.items():
        if keyword in dt:
            return score
    return 7  # default — hard to detect


def get_consequence(abckz) -> str:
    """Map criticality to failure consequence."""
    if abckz in (1, 1.0, "1", "A"):
        return "OPERACIONAL"
    elif abckz in (3, 3.0, "3", "C"):
        return "NO OPERACIONAL"
    return "OPERACIONAL"


# ===========================================================================
# SYMPTOM MATCHING
# ===========================================================================
SYMPTOM_KEYWORDS = {
    "afloj": ["floj", "vibra", "ruido"],
    "corros": ["corrosi", "óxido", "oxidaci", "material"],
    "desgast": ["desgast", "abras", "erosi", "material"],
    "fisura": ["grieta", "fisura", "fractur", "rotura"],
    "fuga": ["fuga", "goteo", "derrame", "pérdida"],
    "vibra": ["vibra", "ruido", "oscila"],
    "sobrecal": ["temperat", "calent", "calor", "humo"],
    "fatiga": ["grieta", "fractur", "rotura"],
    "contamin": ["contamin", "suciedad", "partícul"],
    "erosi": ["erosi", "desgast", "material"],
    "bloqu": ["bloqu", "atasc", "obstrucc"],
    "cortocircuit": ["eléctric", "cortocircuit", "quemad"],
    "degradac": ["degradac", "deterioro", "envejecim"],
}


def find_best_symptom(cause_text: str, symptoms_list: list[str]) -> str:
    """Pick the most relevant symptom for a given cause."""
    if not symptoms_list:
        return "Condición anormal detectada"
    c_lower = (cause_text or "").lower()
    for prefix, kws in SYMPTOM_KEYWORDS.items():
        if prefix in c_lower:
            for sym in symptoms_list:
                s_lower = sym.lower()
                for kw in kws:
                    if kw in s_lower:
                        return sym
    return symptoms_list[0]


# ===========================================================================
# GENERATE FAILURE MODES
# ===========================================================================
print("\n" + "=" * 60, flush=True)
print("GENERATING FAILURE MODES (RCM Functional Analysis)", flush=True)
print("=" * 60, flush=True)

HEADERS = [
    "equipment_tag",
    "sap_func_loc_short",
    "equipment_name",
    "equnr",
    "sap_func_loc",
    "area",
    "equipment_function_description",
    "equipment_functional_failure",
    "function_type",
    "failure_type",
    "subunit",
    "maintainable_item",
    "maintainable_item_function_description",
    "maintainable_item_functional_failure",
    "partes_falla",
    "sintomas_falla",
    "causas_falla",
    "fm_what",
    "fm_mechanism",
    "fm_cause",
    "fm_number",
    "failure_pattern",
    "failure_consequence",
    "evidence",
    "downtime_hours",
    "detection_method",
    "rpn_severity",
    "rpn_occurrence",
    "rpn_detection",
    "rpn_total",
]

wb_out = Workbook(write_only=True)
ws_out = wb_out.create_sheet("failure_modes")
ws_out.append(HEADERS)

total_rows = 0
eq_processed = 0
eq_skipped_no_profile = 0

# Statistics collectors
mech_dist = defaultdict(int)
pattern_dist = defaultdict(int)
equip_with_fms = set()
bom_autocomplete_total = 0
bom_autocomplete_equip = 0

for eq in equipment_l4:
    tag = eq["sap_func_loc_short"] or ""

    # Pilot filter
    if args.pilot and tag[:4] not in PILOT_AREAS:
        continue

    eqart = eq["eqart"]
    profile_code = EQART_TO_PROFILE.get(eqart)

    # Get catalog data for this equipment type
    if profile_code and profile_code in catalog_data:
        cat = catalog_data[profile_code]
    elif eqart and eqart in catalog_data:
        cat = catalog_data[eqart]
        profile_code = eqart
    else:
        eq_skipped_no_profile += 1
        cat = {
            "partes": ["Componente general"],
            "causas": ["Desgaste", "Corrosión", "Fatiga"],
            "sintomas": ["Condición anormal detectada"],
        }

    partes = cat["partes"] or []
    # For simple equipment (instruments, switches, valves, balizas, sirenas)
    # with no catalog parts, use the equipment name itself as the MI
    if not partes or partes == ["Componente general"]:
        equip_name = eq["pltxt"] or eq["eqktx"] or eqart or "Componente general"
        partes = [equip_name]

    # --- BOM auto-complete: add missing MIs from equipment_bom_templates ---
    if profile_code:
        partes, added_mis = complete_equipment_mis(
            tag, profile_code, partes, equipment_l4
        )
        if added_mis:
            bom_autocomplete_total += len(added_mis)
            bom_autocomplete_equip += 1

    sintomas = cat["sintomas"] or ["Condición anormal detectada"]

    # Equipment function (RCM Moubray level)
    area_name = eq["area_name"] or "planta"
    sys_name = eq["system_name"] or eq["pltxt"] or "sistema"
    template = FUNCTION_TEMPLATES.get(
        EQART_TO_PROFILE.get(eqart, ""),
        f"Cumplir función operativa de {eq['eqart_desc'] or eqart} en {{sys}}, {{area}}",
    )
    func_desc = template.replace("{sys}", sys_name).replace("{area}", f"área {area_name}")
    func_failure = f"Incapaz de {func_desc[0].lower()}{func_desc[1:]}"

    # Criticality
    crit_info = criticality_by_fl.get(
        eq["sap_func_loc"], {"criticality": eq["abckz"] or "2", "downtime": 12}
    )
    abckz = crit_info["criticality"] or eq["abckz"] or "2"
    downtime = crit_info.get("downtime", 12)
    severity = SEVERITY_MAP.get(abckz, 5)
    consequence = get_consequence(abckz)

    # ---------------------------------------------------------------
    # For each MI (parte), classify and generate FMs
    # ---------------------------------------------------------------
    for mi_name in partes:
        # Use RCM functional analysis: get_failure_modes_for_mi returns
        # the full chain: function → functional failure → failure modes
        fm_list = get_failure_modes_for_mi(mi_name)
        subunit = classify_subsystem(mi_name)

        for fm in fm_list:
            fm_code = fm.get("fm_ref", "")
            mechanism_es = fm.get("mechanism_es", "")
            cause_es = fm.get("cause_es", "")
            pattern = fm.get("pattern", "B")
            detection = fm.get("detection", "")
            mi_function_es = fm.get("function_es", "")
            mi_failure_es = fm.get("failure_es", "")

            # Look up evidence from FM-MASTER
            fm_master = fm_data.get(fm_code, {})
            evidence = fm_master.get("evidence", fm.get("why", ""))
            if evidence and len(str(evidence)) > 200:
                evidence = str(evidence)[:197] + "..."

            # Functional failure from RCM analysis
            mi_functional_failure = mi_failure_es if mi_failure_es else f"No logra {mi_function_es}"

            # Match a symptom from catalog
            best_symptom = find_best_symptom(cause_es, sintomas)

            # Match a causa from catalog
            best_causa = cause_es
            cause_lower = cause_es.lower() if cause_es else ""
            for cat_causa in cat.get("causas", []):
                if cat_causa.lower() in cause_lower or cause_lower in cat_causa.lower():
                    best_causa = cat_causa
                    break

            failure_type = "TOTAL" if pattern in ("E", "F") else "PARCIAL"

            # RPN
            occurrence = OCCURRENCE_BY_PATTERN.get(pattern, 5)
            det_score = rpn_detection_score(detection)

            # GFSN altitude adjustment for non-instrumented detection
            if det_score >= 6:
                det_score = min(det_score + 1, 10)

            rpn_total = severity * occurrence * det_score

            ws_out.append(
                [
                    tag,                        # equipment_tag
                    tag,                        # sap_func_loc_short
                    eq["pltxt"],                # equipment_name
                    eq["equnr"],                # equnr
                    eq["sap_func_loc"],         # sap_func_loc
                    area_name,                  # area
                    func_desc,                  # equipment_function_description
                    func_failure,               # equipment_functional_failure
                    "PRIMARY",                  # function_type
                    failure_type,               # failure_type
                    subunit,                    # subunit
                    mi_name,                    # maintainable_item
                    mi_function_es,             # maintainable_item_function_description
                    mi_functional_failure,      # maintainable_item_functional_failure
                    mi_name,                    # partes_falla
                    best_symptom,               # sintomas_falla
                    best_causa,                 # causas_falla
                    mi_name,                    # fm_what
                    mechanism_es,               # fm_mechanism
                    cause_es,                   # fm_cause
                    fm_code,                    # fm_number
                    pattern,                    # failure_pattern
                    consequence,                # failure_consequence
                    evidence,                   # evidence
                    downtime or 12,             # downtime_hours
                    detection,                  # detection_method
                    severity,                   # rpn_severity
                    occurrence,                 # rpn_occurrence
                    det_score,                  # rpn_detection
                    rpn_total,                  # rpn_total
                ]
            )
            total_rows += 1
            mech_dist[mechanism_es] += 1
            pattern_dist[pattern] += 1

    equip_with_fms.add(tag)
    eq_processed += 1
    if eq_processed % 200 == 0:
        print(
            f"  Processed {eq_processed} equipment... ({total_rows} FMs)",
            flush=True,
        )

# ===========================================================================
# SAVE
# ===========================================================================
if args.pilot:
    out_path = str(PROJECT_ROOT / "seed_data" / "03_failure_modes_v2.xlsx")
else:
    out_path = str(PROJECT_ROOT / "seed_data" / "03_failure_modes.xlsx")

print(f"\nSaving to {out_path} ...", flush=True)
wb_out.save(out_path)
print("SAVED!", flush=True)

# ===========================================================================
# SUMMARY
# ===========================================================================
print("\n" + "=" * 60, flush=True)
print("SUMMARY", flush=True)
print("=" * 60, flush=True)
print(f"  Total equipment processed : {eq_processed}", flush=True)
print(f"  Equipment without catalog : {eq_skipped_no_profile}", flush=True)
print(f"  Equipment with FMs        : {len(equip_with_fms)}", flush=True)
print(f"  Total failure modes       : {total_rows}", flush=True)
print(f"  BOM auto-completed MIs   : {bom_autocomplete_total} (across {bom_autocomplete_equip} equipment)", flush=True)
print(
    f"  Coverage                  : {len(equip_with_fms)}/{eq_processed}"
    f" ({100*len(equip_with_fms)/max(eq_processed,1):.1f}%)",
    flush=True,
)

print(f"\n  Distribution by mechanism ({len(mech_dist)} unique):", flush=True)
for mech, count in sorted(mech_dist.items(), key=lambda x: -x[1])[:15]:
    print(f"    {mech:50s} {count:6d}", flush=True)
if len(mech_dist) > 15:
    print(f"    ... and {len(mech_dist) - 15} more", flush=True)

print(f"\n  Distribution by pattern:", flush=True)
for pat in sorted(pattern_dist.keys()):
    print(f"    Pattern {pat}: {pattern_dist[pat]:6d}", flush=True)

print(f"\n  GFSN context applied:", flush=True)
print(f"    Altitude factor        : {GFSN_CONTEXT['D1_ALTITUDE_INTERVAL_FACTOR']}", flush=True)
print(f"    UV degradation factor  : {GFSN_CONTEXT['D7_UV_DEGRADATION_FACTOR']}", flush=True)
print(f"    Motor derating severity: {GFSN_CONTEXT['D6_MOTOR_DERATING_SEVERITY']}", flush=True)

print("\nDone.", flush=True)
