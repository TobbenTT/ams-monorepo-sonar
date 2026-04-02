"""Generate enriched 07_spare_parts_inventory.xlsx from MM60 feedback data.

Uses the 17-column structure from client/seed_data/07_spare_parts_inventory-init.xlsx
and enriches ALL ZREP records from seed_data/feedback/MM60 06 01 16.xlsx.

Every header cell gets an openpyxl Comment explaining the field.
"""

import math
import os
import random
import sys

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
random.seed(2026)
AUTHOR = "AMS-Production"

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MM60_PATH = os.path.join(BASE, "seed_data", "feedback", "MM60 06 01 16.xlsx")
HIER_PATH = os.path.join(BASE, "seed_data", "01_equipment_hierarchy.xlsx")
OUTPUT_PATH = os.path.join(BASE, "seed_data", "07_spare_parts_inventory.xlsx")

# ---------------------------------------------------------------------------
# Column definitions and comments (data dictionary)
# ---------------------------------------------------------------------------
COLUMNS = [
    "material_code",
    "sap_material_number",
    "description",
    "manufacturer",
    "part_number",
    "ved_class",
    "fsn_class",
    "abc_class",
    "quantity_on_hand",
    "min_stock",
    "max_stock",
    "reorder_point",
    "lead_time_days",
    "unit_cost_usd",
    "unit_of_measure",
    "applicable_equipment_csv",
    "warehouse_location",
]

COMMENTS = {
    "material_code": (
        "Codigo interno de material Salares Norte.\n"
        "Formato: S26-MAT-NNNNN (secuencial, 5 digitos).\n"
        "No es un campo SAP — identificador local de seed data.\n"
        "Prefijo S26 = Salares Norte site code."
    ),
    "sap_material_number": (
        "Numero de material SAP (MATNR).\n"
        "Fuente: MM60 Master Data, columna 'Material'.\n"
        "Tabla SAP: MARA. Transacciones: MM03, MM60.\n"
        "Formato: numerico, hasta 18 caracteres."
    ),
    "description": (
        "Texto breve del material (MAKTX).\n"
        "Fuente: MM60 'Texto breve de material'.\n"
        "Extraido: primer segmento antes del punto y coma.\n"
        "Tabla SAP: MAKT. Idioma: espanol (ES).\n"
        "Formato SAP original: DESC;FABRICANTE;NUM_PARTE."
    ),
    "manufacturer": (
        "Nombre del fabricante / marca.\n"
        "Parseado del 2do segmento del texto breve SAP.\n"
        "Formato original: DESC;MANUFACTURER;PART_NUMBER.\n"
        "Campo SAP relacionado: MARA-MFRNR.\n"
        "Si no parseable: 'GENERICO'.\n"
        "Ejemplos: SKF, 3M, LETOURNEAU, FLENDER, P&H, ABB."
    ),
    "part_number": (
        "Numero de parte del fabricante / catalogo.\n"
        "Parseado del 3er segmento del texto breve SAP.\n"
        "Usado para referencia cruzada con catalogos de proveedor.\n"
        "Si no parseable: se usa la especificacion tecnica o 'N/A'.\n"
        "Ejemplos: 3309AC3, 62062ZC3, 22211EKC3."
    ),
    "ved_class": (
        "Clasificacion VED (Vital / Essential / Desirable).\n"
        "VITAL: falla detiene produccion, sin sustituto.\n"
        "ESSENTIAL: existe workaround, impacto moderado.\n"
        "DESIRABLE: sin impacto en produccion.\n"
        "Asignado segun costo unitario y grupo de material.\n"
        "Ref: SparePartsEngine.classify_ved()"
    ),
    "fsn_class": (
        "Clasificacion FSN por frecuencia de consumo.\n"
        "FAST_MOVING: >12 movimientos/ano.\n"
        "NORMAL: 4-12 movimientos/ano.\n"
        "SLOW_MOVING: <4 movimientos/ano.\n"
        "Asignado segun grupo de articulos SAP.\n"
        "Ref: SparePartsEngine.classify_fsn()"
    ),
    "abc_class": (
        "Clasificacion ABC por valor (Pareto).\n"
        "A: top items que suman 80% del costo total.\n"
        "B: siguiente 15% del costo acumulado.\n"
        "C: ultimo 5% del costo acumulado.\n"
        "Recalculado desde unit_cost_usd (MM60 tenia todos 'B').\n"
        "Ref: SparePartsEngine.classify_abc()"
    ),
    "quantity_on_hand": (
        "Stock actual no restringido.\n"
        "Generado sinteticamente en rango [0, max_stock].\n"
        "Campo SAP: MARD-LABST.\n"
        "Transacciones: MMBE, MB52.\n"
        "Unidad: misma que unit_of_measure."
    ),
    "min_stock": (
        "Stock de seguridad / stock minimo.\n"
        "Formula: Z x sigma x raiz(lead_time), Z=1.645 (95% SL).\n"
        "Campo SAP: MARC-MINBE.\n"
        "Calculado via SparePartsEngine.calculate_stock_levels().\n"
        "Unidad: misma que unit_of_measure."
    ),
    "max_stock": (
        "Nivel maximo de stock.\n"
        "Formula: reorder_point + EOQ.\n"
        "Campo SAP: MARC-MABST.\n"
        "Calculado via SparePartsEngine.calculate_stock_levels().\n"
        "Unidad: misma que unit_of_measure."
    ),
    "reorder_point": (
        "Punto de reorden que dispara aprovisionamiento.\n"
        "Formula: consumo_diario x lead_time + safety_stock.\n"
        "Campo SAP: MARC-MINBE.\n"
        "Transaccion: MD04 (lista de necesidades).\n"
        "Calculado via SparePartsEngine.calculate_stock_levels()."
    ),
    "lead_time_days": (
        "Tiempo de entrega planificado en dias calendario.\n"
        "Asignado segun ABC class (patron existente catalogo).\n"
        "Campo SAP: MARC-PLIFZ (Info Record) / EINA-APLFZ.\n"
        "A: 45-180d (avg 79), B: 14-90d (avg 35), C: 7-45d (avg 21)."
    ),
    "unit_cost_usd": (
        "Costo unitario en USD.\n"
        "Fuente: MM60 columna 'Precio' cuando disponible.\n"
        "Campo SAP: MBEW-STPRS o MBEW-VERPR.\n"
        "Moneda: USD (confirmada desde columna 'Moneda').\n"
        "Transaccion: MR21 (cambio de precio).\n"
        "Sin precio MM60 -> estimacion por grupo articulos\n"
        "(distribucion log-normal calibrada con datos reales)."
    ),
    "unit_of_measure": (
        "Unidad de medida base.\n"
        "Fuente: MM60 'Unidad medida base'.\n"
        "Mapeado: UN -> EA (each), CA -> EA.\n"
        "Campo SAP: MARA-MEINS.\n"
        "Valores validos: EA, KG, L, M, M2, SET, KIT."
    ),
    "applicable_equipment_csv": (
        "Codigos de ubicacion funcional de equipos aplicables.\n"
        "Separados por coma (sufijo _csv = valores multiples).\n"
        "Asignados por afinidad: descripcion del repuesto\n"
        "se cruza con tipo de equipo (eqart) del hierarchy.\n"
        "En SAP: vinculado via BOM (CS01) o enlace equipo-material.\n"
        "Formato: codigo corto de sap_func_loc_short."
    ),
    "warehouse_location": (
        "Codigo de ubicacion de almacen.\n"
        "Campo SAP: MARC-LGORT (almacen por defecto).\n"
        "Transaccion: MMBE.\n"
        "ALM-CENTRAL (principal), ALM-SEC-01 (secundario),\n"
        "ALM-RIP-01 (repuestos criticos), ALM-HUM-01 (humedad ctrl).\n"
        "Asignado segun VED class y tipo de material."
    ),
}

# ---------------------------------------------------------------------------
# Price estimation calibrated from real MM60 data (log-normal per group)
# ---------------------------------------------------------------------------
# Real price stats from MM60: grupo -> (n, min, max, avg, median)
PRICE_STATS = {
    "2010": (5, 85, 459, 188, 147),
    "2020": (59, 2, 10751, 432, 83),
    "2021": (153, 2, 17945, 1266, 123),
    "2030": (23, 10, 3607, 396, 142),
    "2040": (32, 5, 15096, 2771, 273),
    "2041": (109, 2, 19861, 1791, 293),
    "2042": (165, 2, 20413, 1610, 490),
    "2043": (46, 18, 5057, 1023, 537),
    "2050": (19, 17, 16363, 1034, 58),
    "2051": (500, 1, 16913, 808, 196),
}
DEFAULT_PRICE_STAT = (0, 10, 5000, 600, 150)


def estimate_price(grupo):
    """Log-normal price estimation calibrated to real MM60 distribution."""
    g = str(grupo or "")[:4]
    _, pmin, pmax, pavg, pmed = PRICE_STATS.get(g, DEFAULT_PRICE_STAT)
    # Use log-normal with median as mu, spread from real range
    mu = math.log(max(pmed, 1))
    sigma = min(1.5, math.log(max(pmax, 2) / max(pmed, 1)) / 2.5)
    price = random.lognormvariate(mu, sigma)
    price = max(pmin, min(pmax * 1.1, price))
    return round(price, 2)


# ---------------------------------------------------------------------------
# Keyword -> equipment type affinity mapping
# ---------------------------------------------------------------------------
# Maps keywords in spare part descriptions to likely equipment types (eqart)
KEYWORD_EQART_MAP = {
    # Bearings / Rodamientos
    "RODAMIENTO": ["BUVA", "BAVA", "TRPR", "EMBT", "TRFL"],
    "BEARING": ["BUVA", "BAVA", "TRPR", "EMBT", "TRFL"],
    # Motors / Electrical
    "MOTOR": ["EMBT", "ELHE", "PUCE"],
    "CONTACTOR": ["SWPO", "SWFL", "SWLE", "PUCE"],
    "FUSIBLE": ["SWPO", "SWFL", "PUCE"],
    "FUSE": ["SWPO", "SWFL", "PUCE"],
    "INTERRUPTOR": ["SWPO", "SWFL", "SWLE"],
    "SWITCH": ["SWPO", "SWFL", "SWLE"],
    "CIRCUIT": ["SWPO", "SWFL", "PUCE"],
    "RELAY": ["SWPO", "SWFL", "PUCE"],
    "RELE": ["SWPO", "SWFL", "PUCE"],
    "CABLE": ["PUCE", "INPR", "ELHE"],
    "TRANSFORMER": ["PUCE", "ELHE"],
    "TRANSFORMADOR": ["PUCE", "ELHE"],
    "RECTIFIER": ["PUCE", "ELHE"],
    # Pumps
    "BOMBA": ["BUVA", "BAVA", "KNVA"],
    "PUMP": ["BUVA", "BAVA", "KNVA"],
    "IMPELLER": ["BUVA", "BAVA"],
    "IMPULSOR": ["BUVA", "BAVA"],
    # Valves
    "VALVULA": ["KNVA", "GAVA", "BUVA"],
    "VALVE": ["KNVA", "GAVA", "BUVA"],
    # Filters
    "FILTRO": ["BUVA", "TRPR", "TRLE", "INPR"],
    "FILTER": ["BUVA", "TRPR", "TRLE", "INPR"],
    # Conveyors / Transport
    "CORREA": ["TRLE", "TRTE", "TRFL", "TRPR"],
    "BELT": ["TRLE", "TRTE", "TRFL", "TRPR"],
    "POLEA": ["TRLE", "TRTE", "TRFL"],
    "PULLEY": ["TRLE", "TRTE", "TRFL"],
    "RODILLO": ["TRLE", "TRTE", "TRFL"],
    "ROLLER": ["TRLE", "TRTE", "TRFL"],
    # Hydraulic
    "HIDRAULIC": ["BUVA", "BAVA", "KNVA", "GAVA"],
    "HYDRAULIC": ["BUVA", "BAVA", "KNVA", "GAVA"],
    "CILINDRO": ["BUVA", "BAVA", "KNVA"],
    "CYLINDER": ["BUVA", "BAVA", "KNVA"],
    "MANGUERA": ["BUVA", "BAVA", "KNVA", "GAVA"],
    "HOSE": ["BUVA", "BAVA", "KNVA", "GAVA"],
    # Sensors / Instruments
    "SENSOR": ["INPR", "CELO", "SIBE"],
    "TRANSMISOR": ["INPR", "CELO"],
    "TRANSMITTER": ["INPR", "CELO"],
    "MODULO": ["INPR", "PUCE", "SWPO"],
    "MODULE": ["INPR", "PUCE", "SWPO"],
    "MEDIDOR": ["INPR", "CELO"],
    # Seals / Gaskets
    "SELLO": ["BUVA", "BAVA", "KNVA", "GAVA"],
    "SEAL": ["BUVA", "BAVA", "KNVA", "GAVA"],
    "JUNTA": ["BUVA", "BAVA", "KNVA"],
    "GASKET": ["BUVA", "BAVA", "KNVA"],
    "ORING": ["BUVA", "BAVA", "KNVA"],
    # Crusher / Screening
    "CHANCADOR": ["CRIO", "SIBE"],
    "CRUSHER": ["CRIO", "SIBE"],
    "HARNERO": ["SIBE", "CRIO"],
    "SCREEN": ["SIBE", "CRIO"],
    "REVESTIMIENTO": ["CRIO", "SIBE"],
    "LINER": ["CRIO", "SIBE"],
    # General mechanical
    "ACOPLAMIENTO": ["TRPR", "EMBT", "BUVA"],
    "COUPLING": ["TRPR", "EMBT", "BUVA"],
    "CADENA": ["TRLE", "TRTE"],
    "CHAIN": ["TRLE", "TRTE"],
    "PERNO": ["BUVA", "BAVA", "TRPR", "TRLE", "CRIO"],
    "BOLT": ["BUVA", "BAVA", "TRPR", "TRLE", "CRIO"],
    "TUERCA": ["BUVA", "BAVA", "TRPR", "CRIO"],
    "NUT": ["BUVA", "BAVA", "TRPR", "CRIO"],
    "PLACA": ["CRIO", "SIBE", "TRLE"],
    "PLATE": ["CRIO", "SIBE", "TRLE"],
    # Vehicles / Heavy equipment
    "NEUMATICO": ["VEAL", "VECH"],
    "TIRE": ["VEAL", "VECH"],
    "FRENO": ["VEAL", "VECH", "TRLE"],
    "BRAKE": ["VEAL", "VECH", "TRLE"],
}

# Fallback eqarts for items with no keyword match
FALLBACK_EQARTS = [
    "BUVA", "INPR", "PUCE", "TRPR", "BAVA", "TRLE", "TRTE",
    "SWPO", "GAVA", "EMBT", "KNVA", "TRFL", "CELO", "SWFL",
]


def find_affinity_eqarts(description):
    """Find equipment types with affinity to a spare part description."""
    desc_upper = description.upper()
    matched = set()
    for keyword, eqarts in KEYWORD_EQART_MAP.items():
        if keyword in desc_upper:
            matched.update(eqarts)
    return list(matched) if matched else None


# ---------------------------------------------------------------------------
# Warehouse assignment by VED + material type
# ---------------------------------------------------------------------------
WAREHOUSES = ["ALM-CENTRAL", "ALM-SEC-01", "ALM-RIP-01", "ALM-HUM-01"]

def assign_warehouse(ved, description):
    """Assign warehouse following existing catalog patterns."""
    desc_upper = description.upper()
    # Critical items -> ALM-RIP-01 (repuestos criticos)
    if ved == "VITAL":
        return random.choices(
            WAREHOUSES, weights=[25, 15, 50, 10], k=1
        )[0]
    # Humidity-sensitive items
    if any(kw in desc_upper for kw in ("SENSOR", "TRANSMISOR", "MODULO", "CIRCUIT", "PCB", "TARJETA")):
        return random.choices(
            WAREHOUSES, weights=[20, 15, 15, 50], k=1
        )[0]
    # Essential -> mix
    if ved == "ESSENTIAL":
        return random.choices(
            WAREHOUSES, weights=[35, 30, 25, 10], k=1
        )[0]
    # Desirable -> mostly central
    return random.choices(
        WAREHOUSES, weights=[50, 25, 15, 10], k=1
    )[0]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
UOM_MAP = {
    "UN": "EA", "KG": "KG", "L": "L", "M": "M",
    "M2": "M2", "KIT": "KIT", "SET": "SET", "CA": "EA",
}


def calculate_stock_levels(daily_consumption, lead_time_days):
    """Same logic as SparePartsEngine.calculate_stock_levels."""
    if daily_consumption <= 0:
        return {"safety_stock": 0, "reorder_point": 0, "min_stock": 0, "max_stock": 1}
    z = 1.645
    sigma = daily_consumption * 0.3
    safety_stock = max(1, round(z * sigma * math.sqrt(lead_time_days)))
    reorder_point = round(daily_consumption * lead_time_days + safety_stock)
    eoq = max(1, round(math.sqrt(2 * daily_consumption * 365 * 10 / 1)))
    max_stock = reorder_point + eoq
    return {
        "safety_stock": safety_stock,
        "reorder_point": reorder_point,
        "min_stock": safety_stock,
        "max_stock": max_stock,
    }


# ---------------------------------------------------------------------------
# 1. Load equipment hierarchy
# ---------------------------------------------------------------------------
print("Loading equipment hierarchy...")
wb_hier = openpyxl.load_workbook(HIER_PATH, read_only=True)
ws_hier = wb_hier.active

# Build lookup: eqart -> list of func_loc_short
equip_by_eqart = {}
all_func_locs = []
for row in ws_hier.iter_rows(min_row=2, values_only=True):
    level = row[4]
    fl_short = row[1]
    eqart = str(row[8] or "")
    if level is not None and level >= 4 and fl_short:
        fl = str(fl_short)
        all_func_locs.append(fl)
        if eqart:
            equip_by_eqart.setdefault(eqart, []).append(fl)
wb_hier.close()
print(f"  Equipment locations: {len(all_func_locs)} (types: {len(equip_by_eqart)})")

# ---------------------------------------------------------------------------
# 2. Load ALL ZREP from MM60 Master Data
# ---------------------------------------------------------------------------
print("Loading MM60 Master Data (all ZREP)...")
wb_mm = openpyxl.load_workbook(MM60_PATH, read_only=True)
ws_mm = wb_mm["Master data"]

candidates = []
seen_materials = set()

for row in ws_mm.iter_rows(min_row=2, values_only=True):
    material = str(row[0] or "")
    desc_raw = str(row[3] or "").strip()
    mat_type = row[5]
    grupo = str(row[6] or "")
    uom = str(row[7] or "UN").strip()
    precio = row[13]

    if mat_type != "ZREP":
        continue
    if not desc_raw or not material:
        continue
    if material in seen_materials:
        continue
    seen_materials.add(material)

    # Parse description based on semicolons
    semi_count = desc_raw.count(";")
    if semi_count >= 2:
        parts = desc_raw.split(";")
        description = parts[0].strip()
        manufacturer = parts[1].strip()
        part_number = ";".join(parts[2:]).strip()
    elif semi_count == 1:
        parts = desc_raw.split(";")
        description = parts[0].strip()
        spec = parts[1].strip()
        # Try to detect if spec is a manufacturer or a dimension
        if any(c.isdigit() for c in spec) and "X" in spec.upper():
            manufacturer = "GENERICO"
            part_number = spec
        else:
            manufacturer = spec if len(spec) <= 20 else "GENERICO"
            part_number = spec if manufacturer == "GENERICO" else "N/A"
    else:
        description = desc_raw
        manufacturer = "GENERICO"
        part_number = "N/A"

    if not description:
        continue

    has_price = precio is not None and isinstance(precio, (int, float)) and precio > 0
    unit_cost = round(float(precio), 2) if has_price else estimate_price(grupo)

    uom_mapped = UOM_MAP.get(uom.upper(), "EA")

    candidates.append({
        "material": material,
        "description": description,
        "manufacturer": manufacturer,
        "part_number": part_number,
        "unit_cost_usd": unit_cost,
        "unit_of_measure": uom_mapped,
        "grupo": grupo,
        "price_source": "MM60" if has_price else "synthetic",
    })

wb_mm.close()
n_real_price = sum(1 for c in candidates if c["price_source"] == "MM60")
print(f"  Total ZREP loaded: {len(candidates)} ({n_real_price} with MM60 price, {len(candidates)-n_real_price} synthetic)")

# ---------------------------------------------------------------------------
# 3. Assign VED (weighted by cost tier, matching existing catalog)
# ---------------------------------------------------------------------------
def assign_ved(cost, grupo):
    g = str(grupo)[:4]
    # High-value items more likely vital
    if cost > 5000:
        return random.choices(["VITAL", "ESSENTIAL", "DESIRABLE"], weights=[60, 30, 10], k=1)[0]
    if cost > 1000:
        return random.choices(["VITAL", "ESSENTIAL", "DESIRABLE"], weights=[25, 50, 25], k=1)[0]
    if cost > 100:
        return random.choices(["VITAL", "ESSENTIAL", "DESIRABLE"], weights=[10, 35, 55], k=1)[0]
    return random.choices(["VITAL", "ESSENTIAL", "DESIRABLE"], weights=[5, 25, 70], k=1)[0]


# ---------------------------------------------------------------------------
# 4. Assign FSN (weighted by material group)
# ---------------------------------------------------------------------------
def assign_fsn(grupo):
    g = str(grupo)[:4]
    if g in ("2050", "2030", "2031"):  # consumables, filters
        return random.choices(["FAST_MOVING", "NORMAL", "SLOW_MOVING"], weights=[40, 45, 15], k=1)[0]
    if g in ("2021", "2020"):  # electrical
        return random.choices(["FAST_MOVING", "NORMAL", "SLOW_MOVING"], weights=[15, 50, 35], k=1)[0]
    if g in ("2041", "2042", "2043"):  # mechanical / heavy
        return random.choices(["FAST_MOVING", "NORMAL", "SLOW_MOVING"], weights=[10, 40, 50], k=1)[0]
    if g in ("2040",):  # heavy mechanical
        return random.choices(["FAST_MOVING", "NORMAL", "SLOW_MOVING"], weights=[5, 35, 60], k=1)[0]
    if g in ("2051", "2010"):  # instruments
        return random.choices(["FAST_MOVING", "NORMAL", "SLOW_MOVING"], weights=[12, 48, 40], k=1)[0]
    return random.choices(["FAST_MOVING", "NORMAL", "SLOW_MOVING"], weights=[10, 45, 45], k=1)[0]


# ---------------------------------------------------------------------------
# 5. Build all rows
# ---------------------------------------------------------------------------
print("Generating enriched data for all records...")
rows = []
for i, item in enumerate(candidates, start=1):
    cost = item["unit_cost_usd"]
    ved = assign_ved(cost, item["grupo"])
    fsn = assign_fsn(item["grupo"])

    # Lead time by ABC-like tiers (assign preliminary based on cost,
    # will get final ABC from Pareto later)
    # Following existing catalog: A(high cost) -> long LT, C(low) -> short LT
    if cost > 3000:
        lead_time = random.choice([45, 60, 75, 90, 120, 150, 180])
    elif cost > 500:
        lead_time = random.choice([14, 21, 30, 45, 60, 90])
    elif cost > 50:
        lead_time = random.choice([7, 14, 21, 30, 45])
    else:
        lead_time = random.choice([7, 14, 21, 30])

    # Daily consumption based on FSN
    if fsn == "FAST_MOVING":
        daily = random.uniform(0.3, 2.5)
    elif fsn == "NORMAL":
        daily = random.uniform(0.03, 0.4)
    else:
        daily = random.uniform(0.003, 0.04)

    levels = calculate_stock_levels(daily, lead_time)
    qty = random.randint(0, max(1, levels["max_stock"]))

    # Equipment assignment: use keyword affinity
    affinity_eqarts = find_affinity_eqarts(item["description"])
    n_equip = random.randint(1, min(5, 3 if fsn == "SLOW_MOVING" else 5))

    assigned_locs = []
    if affinity_eqarts:
        # Pick from affinity eqarts
        for eqart in random.sample(affinity_eqarts, min(len(affinity_eqarts), n_equip)):
            locs = equip_by_eqart.get(eqart, [])
            if locs:
                assigned_locs.append(random.choice(locs))
    # Fill remaining with general pool if needed
    while len(assigned_locs) < n_equip and all_func_locs:
        loc = random.choice(all_func_locs)
        if loc not in assigned_locs:
            assigned_locs.append(loc)
    equip_csv = ", ".join(assigned_locs) if assigned_locs else random.choice(all_func_locs)

    warehouse = assign_warehouse(ved, item["description"])

    rows.append({
        "material_code": f"S26-MAT-{i:05d}",
        "sap_material_number": item["material"],
        "description": item["description"],
        "manufacturer": item["manufacturer"],
        "part_number": item["part_number"],
        "ved_class": ved,
        "fsn_class": fsn,
        "abc_class": "",  # placeholder — Pareto below
        "quantity_on_hand": qty,
        "min_stock": levels["min_stock"],
        "max_stock": levels["max_stock"],
        "reorder_point": levels["reorder_point"],
        "lead_time_days": lead_time,
        "unit_cost_usd": cost,
        "unit_of_measure": item["unit_of_measure"],
        "applicable_equipment_csv": equip_csv,
        "warehouse_location": warehouse,
    })

print(f"  Generated {len(rows)} rows")

# ---------------------------------------------------------------------------
# 6. ABC Pareto reclassification
# ---------------------------------------------------------------------------
print("Applying ABC Pareto classification...")
rows.sort(key=lambda r: r["unit_cost_usd"], reverse=True)
total_cost = sum(r["unit_cost_usd"] for r in rows)
cumulative = 0.0
for r in rows:
    cumulative += r["unit_cost_usd"]
    pct = (cumulative / total_cost) * 100 if total_cost > 0 else 100
    if pct <= 80:
        r["abc_class"] = "A"
    elif pct <= 95:
        r["abc_class"] = "B"
    else:
        r["abc_class"] = "C"

# Adjust lead_time to match ABC pattern from existing catalog
# A: 45-180 (avg 79), B: 14-90 (avg 35), C: 7-45 (avg 21)
for r in rows:
    abc = r["abc_class"]
    if abc == "A" and r["lead_time_days"] < 45:
        r["lead_time_days"] = random.choice([45, 60, 75, 90, 120])
    elif abc == "C" and r["lead_time_days"] > 45:
        r["lead_time_days"] = random.choice([7, 14, 21, 30, 45])

# Re-sort by material_code
rows.sort(key=lambda r: r["material_code"])

# ---------------------------------------------------------------------------
# 7. Write Excel
# ---------------------------------------------------------------------------
print("Writing Excel file...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Spare Parts Inventory"

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, col_name in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    comment_text = COMMENTS.get(col_name, col_name)
    cell.comment = Comment(comment_text, AUTHOR)
    cell.comment.width = 350
    cell.comment.height = 200

for row_idx, row_data in enumerate(rows, start=2):
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])

# Column widths (sample first 100 rows)
for col_idx in range(1, len(COLUMNS) + 1):
    col_letter = get_column_letter(col_idx)
    max_len = len(str(COLUMNS[col_idx - 1]))
    for row_idx in range(2, min(102, len(rows) + 2)):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is not None:
            max_len = max(max_len, min(len(str(val)), 55))
    ws.column_dimensions[col_letter].width = max_len + 3

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

wb.save(OUTPUT_PATH)
print(f"\nDone! Wrote {len(rows)} rows to {OUTPUT_PATH}")

# ---------------------------------------------------------------------------
# 8. Summary
# ---------------------------------------------------------------------------
abc_counts = {}
ved_counts = {}
fsn_counts = {}
wh_counts = {}
for r in rows:
    abc_counts[r["abc_class"]] = abc_counts.get(r["abc_class"], 0) + 1
    ved_counts[r["ved_class"]] = ved_counts.get(r["ved_class"], 0) + 1
    fsn_counts[r["fsn_class"]] = fsn_counts.get(r["fsn_class"], 0) + 1
    wh_counts[r["warehouse_location"]] = wh_counts.get(r["warehouse_location"], 0) + 1

prices = [r["unit_cost_usd"] for r in rows]
mm60_prices = [r["unit_cost_usd"] for r in rows if any(
    c["material"] == r["sap_material_number"] and c["price_source"] == "MM60" for c in candidates
)]

print(f"\n{'='*60}")
print(f"SUMMARY: {len(rows)} spare parts")
print(f"{'='*60}")
print(f"ABC: {dict(sorted(abc_counts.items()))}")
print(f"VED: {dict(sorted(ved_counts.items()))}")
print(f"FSN: {dict(sorted(fsn_counts.items()))}")
print(f"Warehouses: {dict(sorted(wh_counts.items()))}")
print(f"Price range: ${min(prices):,.2f} - ${max(prices):,.2f}")
print(f"Price mean: ${sum(prices)/len(prices):,.2f}")
print(f"Unique manufacturers: {len(set(r['manufacturer'] for r in rows))}")
print(f"Unique SAP materials: {len(set(r['sap_material_number'] for r in rows))}")

# Lead time by ABC
for abc_val in ["A", "B", "C"]:
    lts = [r["lead_time_days"] for r in rows if r["abc_class"] == abc_val]
    if lts:
        print(f"Lead time {abc_val}: min={min(lts)}, max={max(lts)}, avg={sum(lts)/len(lts):.0f}")
