"""
Rebuild 03_failure_modes.xlsx with 28-column RCM/Moubray format.
Uses write_only mode for performance with large datasets.
Strategy: For each equipment, iterate ALL parts, assign 1-3 most relevant causes per part.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl import Workbook
from collections import defaultdict
import random
random.seed(42)

print("=" * 60, flush=True)
print("LOADING SOURCE DATA", flush=True)
print("=" * 60, flush=True)

# ============================================================
# EQART → CATALOG PROFILE MAPPING
# ============================================================
EQART_TO_PROFILE = {
    'BLAX': 'VENT_AX', 'BLSC': 'COMP_TOR', 'BULL': 'BULL_BUL',
    'CCTV': 'CÁM_CCTV', 'CELO': 'CELOSÌA', 'ELHE': 'CAL_ELÉC',
    'EMBT': 'MOT_ELÉC', 'HBFF': 'SOP_CENT', 'HECQ': 'INT_ENFR',
    'HLCN': 'PUEN_GRÚ', 'LDLU': 'SIST_LUB', 'MSST': 'AGIT_MEC',
    'MXAG': 'AGIT_MEC', 'PCDT': 'DET_META', 'POOL': 'PIS_SUM',
    'PUCE': 'BOMB_CEN', 'PURE': 'BOMB_REC', 'PURO': 'BOMB_ROT',
    'PUTU': 'BOMB_TUR', 'PUVA': 'BOMB_VAC', 'SERV': 'SERVIDOR',
    'SFPR': 'FIL_PREN', 'SHEW': 'ESPESADO', 'SRCV': 'CORR_TRA',
    'SRFE': 'ALIM_BAN', 'VEBN': 'TOLVA', 'VEBX': 'CAJÒN',
    'VECH': 'CHUTE', 'VEDR': 'SECA_AIR', 'VELN': 'LIN_PROC',
    'VETA': 'EST_AGIT', 'VIBR': 'VIBRADOR',
    'ACOU': 'UNID_AC', 'AEID': 'INCH_DRI', 'AESS': 'SIST_LAN',
    'AETS': 'SOPORTES', 'AEWM': 'MON_LAVA', 'BCOM': 'TAB_COMU',
    'BLEH': 'CAMP_EXT', 'BLFC': 'VENT_CEN', 'BSCI': 'DET_FH',
    'BUVA': 'LIN_PROC', 'CEDE': 'BOMB_CEN', 'CFOF': 'DIAL_ACE',
    'CMCS': 'CORE_SWI', 'COHI': 'COMP_HID', 'CPLC': 'GAB_PLC',
    'CRIO': 'GAB_RIO', 'CSSP': 'PANE_SCA', 'CTDP': 'TAB_CEL',
    'DECH': 'GAB_MANG', 'DECO': 'CONS_ESC', 'DEFI': 'EST_MANG',
    'DEHY': 'HIDRANTE', 'EABA': 'BANC_BAT', 'EABC': 'CARG_BAT',
    'EAUP': 'UPS', 'ELCB': 'CEL_IPOT', 'ELDI': 'CEL_IPOT',
    'ELIN': 'CEL_PMOT', 'ELPS': 'CEL_PMOT', 'EMRE': 'CARG_FRO',
    'EYWA': 'DUCH_EME', 'FIPS': 'TAB_SCI', 'FIWA': 'FIREWALL',
    'GAVA': 'LIN_PROC', 'HBEH': 'HORN_ELÉ', 'HEPL': 'INT_PLAC',
    'HIWA': 'HIDR_AGU', 'HMIB': 'GAB_PLC', 'HTEL': 'TECL_ELE',
    'HTHA': 'TECL_ELE', 'HYPU': 'UNID_HID', 'INDP': 'TAB_DIST',
    'INPR': 'IND_PROC', 'KNVA': 'VALVULAS', 'KVMS': 'KVMSWITC',
    'LEBT': 'CAM_PLUM', 'LEFL': 'MONT_CAR', 'LIDP': 'TAB_DIST',
    'LIIN': 'TECL_ELE', 'LOFR': 'CARG_FRO', 'LUME': 'PAQ_QUIM',
    'MBAY': 'MUES_TRA', 'MBSC': 'BASCULA', 'MCSC': 'HARN_EST',
    'MFSU': 'SUPR_MF', 'MIRI': 'IND_PROC', 'MOPL': 'PLAT_MÓV',
    'MTTR': 'CARR_MNT', 'OCRD': 'PUER_MOT', 'PCCC': 'CICL_POT',
    'PCCO': 'COL_POLV', 'PCPT': 'TRAN_POT', 'PREU': 'UNI_PRES',
    'PRLI': 'DET_FH', 'PTWW': 'PTA_AR', 'PUSC': 'BOMB_HAR',
    'PWDP': 'TAB_DIST', 'ROVA': 'VALVULAS', 'RPHA': 'REM_PERN',
    'SAPR': 'MUES_ROT', 'SASE': 'SEN_PROC', 'SAWS': 'MUES_ROT',
    'SBDP': 'TAB_DIST', 'SBLR': 'ARRA_LÍQ', 'SECT': 'CÁM_CCTV',
    'SEFL': 'SEN_PROC', 'SFCF': 'FIL_COAL', 'SFCL': 'FIL_HOJA',
    'SMST': 'CHIMENEA', 'SRRT': 'MESA_ROD', 'SSEP': 'TAB_DIST',
    'STIT': 'CAM_RIEG', 'SUPP': 'SOPORTES', 'TAPR': 'EST_ALMA',
    'TOLH': 'MANI_COR', 'TRDE': 'LIN_PROC', 'TRFL': 'LIN_PROC',
    'TRLE': 'TOLVA', 'TRPR': 'TRAN_SEÑ', 'TRTE': 'TRAN_SEÑ',
    'UPDP': 'TAB_DIST', 'USAP': 'TAB_CEL', 'USCT': 'TAB_CEL',
    'USPS': 'TAB_CEL', 'VABU': 'VALVULAS', 'VACO': 'VALVULAS',
    'VAKN': 'VALVULAS', 'VAPI': 'VALVULAS', 'VEAL': 'EST_ALMA',
    'VEAR': 'ACUM_AIR', 'VEMO': 'CASC_MOL', 'VEPC': 'COL_PROC',
    'VETR': 'RACK_BAN', 'VEWR': 'PLAN_POT', 'VFBT': 'CELD_VDF',
    'VVCP': 'SUPR_POL', 'DUCT': 'DUCTOS', 'HETU': 'BOMB_TUR',
    'ROBL': 'MESA_ROD', 'SRPL': 'PLAT_MÓV', 'MCTR': 'TRN_PROC',
    'LUBR': 'SIST_LUB',
}

# ============================================================
# CAUSE_MAP: Spanish catalog cause → FM-MASTER
# ============================================================
CAUSE_MAP = {
    'Aflojamiento': ('Looses Preload', 'Vibration', 'FM-47', 'B'),
    'Aflojamiento de pernos': ('Looses Preload', 'Vibration', 'FM-47', 'B'),
    'Atascamiento': ('Immobilised (binds/jams)', 'Contamination', 'FM-43', 'C'),
    'Bloqueo': ('Blocks', 'Contamination', 'FM-02', 'C'),
    'Cavitación': ('Wears', 'Entrained air', 'FM-65', 'E'),
    'Contaminación': ('Blocks', 'Contamination', 'FM-02', 'C'),
    'Corrosión': ('Corrodes', 'Corrosive environment', 'FM-10', 'B'),
    'Corrosión interna': ('Corrodes', 'Chemical attack', 'FM-09', 'B'),
    'Corrosión externa': ('Corrodes', 'Exposure to atmosphere', 'FM-13', 'B'),
    'Cortocircuito': ('Short-Circuits', 'Breakdown in insulation', 'FM-58', 'B'),
    'Daño mecánico': ('Breaks/Fracture/Separates', 'Mechanical overload', 'FM-06', 'E'),
    'Degradación': ('Degrades', 'Age', 'FM-25', 'B'),
    'Degradación de aislamiento': ('Degrades', 'Age', 'FM-25', 'B'),
    'Degradación material': ('Degrades', 'Chemical attack', 'FM-26', 'B'),
    'Desalineamiento': ('Wears', 'Relative movement between contacting surfaces', 'FM-72', 'B'),
    'Desbalance': ('Wears', 'Relative movement between contacting surfaces', 'FM-72', 'B'),
    'Desgaste': ('Wears', 'Relative movement between contacting surfaces', 'FM-72', 'B'),
    'Desgaste abrasivo': ('Wears', 'Impact/shock loading', 'FM-67', 'B'),
    'Desgaste adhesivo': ('Wears', 'Metal to metal contact', 'FM-71', 'B'),
    'Desgaste erosivo': ('Wears', 'Excessive fluid velocity', 'FM-66', 'B'),
    'Desgaste por fatiga': ('Breaks/Fracture/Separates', 'Cyclic loading (thermal/mechanical)', 'FM-05', 'B'),
    'Distorsión': ('Distorts', 'Mechanical overload', 'FM-34', 'E'),
    'Erosión': ('Wears', 'Excessive fluid velocity', 'FM-66', 'B'),
    'Envejecimiento': ('Degrades', 'Age', 'FM-25', 'B'),
    'Falla eléctrica': ('Arcs', 'Breakdown in insulation', 'FM-01', 'B'),
    'Fatiga': ('Breaks/Fracture/Separates', 'Cyclic loading (thermal/mechanical)', 'FM-05', 'B'),
    'Fisuramiento': ('Cracks', 'Cyclic loading (thermal/mechanical)', 'FM-20', 'B'),
    'Fisuramiento / rotura': ('Cracks', 'Cyclic loading (thermal/mechanical)', 'FM-20', 'B'),
    'Fractura': ('Breaks/Fracture/Separates', 'Cyclic loading (thermal/mechanical)', 'FM-05', 'B'),
    'Fuga': ('Severs (cut, tear, hole)', 'Mechanical overload', 'FM-57', 'E'),
    'Fuga externa': ('Severs (cut, tear, hole)', 'Mechanical overload', 'FM-57', 'E'),
    'Fuga interna': ('Severs (cut, tear, hole)', 'Mechanical overload', 'FM-57', 'E'),
    'Gripado': ('Immobilised (binds/jams)', 'Lack of lubrication', 'FM-44', 'C'),
    'Impacto': ('Breaks/Fracture/Separates', 'Mechanical overload', 'FM-06', 'E'),
    'Obstrucción': ('Blocks', 'Contamination', 'FM-02', 'C'),
    'Oxidación': ('Corrodes', 'Exposure to atmosphere', 'FM-13', 'B'),
    'Pérdida de lubricación': ('Overheats/Melts', 'Lack of lubrication', 'FM-51', 'B'),
    'Rotura': ('Breaks/Fracture/Separates', 'Mechanical overload', 'FM-06', 'E'),
    'Sobrecalentamiento': ('Overheats/Melts', 'Mechanical overload', 'FM-52', 'E'),
    'Sobrecarga': ('Thermally Overloads (burns, overheats, melts)', 'Overcurrent', 'FM-61', 'E'),
    'Sobrecarga eléctrica': ('Thermally Overloads (burns, overheats, melts)', 'Overcurrent', 'FM-61', 'E'),
    'Sobrecarga mecánica': ('Overheats/Melts', 'Mechanical overload', 'FM-52', 'E'),
    'Taponamiento': ('Blocks', 'Contamination', 'FM-02', 'C'),
    'Vibración': ('Looses Preload', 'Vibration', 'FM-47', 'B'),
    'Vibración excesiva': ('Looses Preload', 'Vibration', 'FM-47', 'B'),
    'Desconexión': ('Open-Circuit', 'Electrical overload', 'FM-48', 'E'),
    'Error de calibración': ('Drifts', 'Use', 'FM-41', 'B'),
    'Deriva': ('Drifts', 'Use', 'FM-41', 'B'),
    'Deformación': ('Distorts', 'Mechanical overload', 'FM-34', 'E'),
    'Abrasión': ('Severs (cut, tear, hole)', 'Abrasion', 'FM-55', 'B'),
    'Pandeo': ('Distorts', 'Off-center loading', 'FM-35', 'C'),
    'Descarga eléctrica': ('Arcs', 'Breakdown in insulation', 'FM-01', 'B'),
    'Pérdida de presión': ('Severs (cut, tear, hole)', 'Mechanical overload', 'FM-57', 'E'),
    'Ruido excesivo': ('Wears', 'Relative movement between contacting surfaces', 'FM-72', 'B'),
    'Recalentamiento': ('Overheats/Melts', 'Mechanical overload', 'FM-52', 'E'),
    'Falla de sellado': ('Severs (cut, tear, hole)', 'Mechanical overload', 'FM-57', 'E'),
    'Oxidación de contactos': ('Corrodes', 'Exposure to atmosphere', 'FM-13', 'B'),
    'Humedad': ('Degrades', 'Contamination', 'FM-28', 'C'),
    'Error de señal': ('Drifts', 'Use', 'FM-41', 'B'),
    'Contaminación de aceite': ('Wears', 'Lubricant contamination (particles)', 'FM-69', 'C'),
    'Pérdida de material': ('Wears', 'Excessive fluid velocity', 'FM-66', 'B'),
    'Agrietamiento': ('Cracks', 'Cyclic loading (thermal/mechanical)', 'FM-20', 'B'),
    'Desgaste de revestimiento': ('Wears', 'Impact/shock loading', 'FM-67', 'B'),
}
DEFAULT_FM = ('Degrades', 'Age', 'FM-25', 'B')

# ============================================================
# PART-TYPE CATEGORIES for cause relevance filtering
# ============================================================
PART_CATEGORIES = {
    'mechanical': ['eje', 'rodamiento', 'engranaje', 'acoplamiento', 'polea', 'piñón',
                    'corona', 'cadena', 'correa', 'revestimiento', 'impulsor', 'rotor',
                    'estator', 'álabe', 'aspa', 'paleta', 'tornillo', 'husillo', 'biela',
                    'pistón', 'cilindro', 'vástago', 'buje', 'casquillo', 'camisa',
                    'cojinete', 'balero', 'perno', 'tuerca', 'arandela', 'chaveta',
                    'cuña', 'pasador', 'resorte', 'muelle', 'placa', 'bastidor',
                    'estructura', 'soporte', 'base', 'chasis', 'mandíbula', 'cono',
                    'manto', 'faldón', 'liner', 'martillo', 'bola', 'barra', 'yugo',
                    'brazo', 'eslabón', 'rueda', 'tambor', 'freno', 'embrague', 'disco'],
    'sealing': ['sello', 'empaquetadura', 'retén', 'o-ring', 'junta', 'diafragma',
                'membrana', 'prensa estopa', 'anillo', 'asiento', 'prensaestopa'],
    'piping': ['tubería', 'línea', 'cañería', 'manguera', 'conector', 'niple',
               'fitting', 'brida', 'acople', 'codo', 'tee', 'reducción', 'unión',
               'abrazadera', 'vejiga', 'fuelle'],
    'electrical': ['motor', 'bobina', 'devanado', 'cable', 'conductor', 'bornera',
                   'contactor', 'relé', 'fusible', 'interruptor', 'disyuntor',
                   'transformador', 'capacitor', 'resistencia', 'tiristor', 'diodo',
                   'rectificador', 'variador', 'arrancador', 'escobilla', 'carbón',
                   'anillo rozante', 'terminal', 'conmutador', 'batería', 'celda'],
    'instrumentation': ['sensor', 'transmisor', 'indicador', 'medidor', 'detector',
                        'analizador', 'switch', 'presostato', 'termostato', 'flotador',
                        'placa orificio', 'rotámetro', 'caudalímetro', 'termocupla',
                        'termopar', 'encoder', 'PLC', 'tarjeta', 'módulo', 'pantalla',
                        'display', 'HMI'],
    'fluid_handling': ['válvula', 'bomba', 'filtro', 'tamiz', 'malla', 'cartucho',
                       'elemento filtrante', 'canasta', 'difusor', 'boquilla',
                       'inyector', 'tobera', 'venturi', 'orificio', 'regulador'],
    'structural': ['estructura', 'viga', 'columna', 'plataforma', 'escalera',
                   'barandilla', 'pasamanos', 'techo', 'pared', 'piso', 'cubierta',
                   'tolva', 'chute', 'ducto', 'canaleta', 'bandeja', 'rack', 'gabinete',
                   'carcasa', 'cuerpo', 'tanque', 'recipiente', 'contenedor'],
    'lubrication': ['aceite', 'lubricante', 'grasa', 'lubricación', 'engrase',
                    'filtro aceite', 'enfriador aceite', 'bomba lubricación',
                    'sistema lubricación', 'reservorio'],
}

# Causes that apply to each part category
CAUSE_RELEVANCE = {
    'mechanical': ['Desgaste', 'Fatiga', 'Vibración', 'Aflojamiento', 'Rotura',
                   'Fisuramiento', 'Sobrecarga mecánica', 'Desalineamiento', 'Corrosión',
                   'Abrasión', 'Erosión', 'Impacto', 'Deformación'],
    'sealing': ['Desgaste', 'Fuga', 'Envejecimiento', 'Degradación', 'Corrosión',
                'Falla de sellado', 'Sobrecalentamiento'],
    'piping': ['Corrosión', 'Erosión', 'Fuga', 'Vibración', 'Fatiga', 'Obstrucción',
               'Aflojamiento', 'Desgaste'],
    'electrical': ['Cortocircuito', 'Sobrecarga eléctrica', 'Degradación de aislamiento',
                   'Sobrecalentamiento', 'Falla eléctrica', 'Corrosión', 'Aflojamiento',
                   'Vibración', 'Envejecimiento'],
    'instrumentation': ['Error de calibración', 'Deriva', 'Degradación', 'Vibración',
                        'Corrosión', 'Contaminación', 'Humedad', 'Error de señal',
                        'Cortocircuito'],
    'fluid_handling': ['Desgaste', 'Corrosión', 'Obstrucción', 'Fuga', 'Cavitación',
                       'Erosión', 'Contaminación', 'Taponamiento', 'Bloqueo'],
    'structural': ['Corrosión', 'Fatiga', 'Desgaste', 'Fisuramiento', 'Deformación',
                   'Aflojamiento', 'Vibración', 'Impacto'],
    'lubrication': ['Contaminación de aceite', 'Degradación', 'Fuga', 'Obstrucción',
                    'Sobrecalentamiento', 'Pérdida de lubricación'],
}

def classify_part(part_name):
    """Classify a part into a category."""
    p = part_name.lower()
    for cat, keywords in PART_CATEGORIES.items():
        for kw in keywords:
            if kw in p:
                return cat
    return 'mechanical'  # default

def get_relevant_causes(part_name, all_causes):
    """Get causes relevant to this part type."""
    cat = classify_part(part_name)
    relevant = CAUSE_RELEVANCE.get(cat, CAUSE_RELEVANCE['mechanical'])
    # Filter catalog causes to only relevant ones
    result = []
    for causa in all_causes:
        causa_lower = causa.lower()
        for rel in relevant:
            if rel.lower() in causa_lower or causa_lower in rel.lower():
                result.append(causa)
                break
    # Always include at least 2 causes
    if len(result) < 2:
        result = all_causes[:3]
    return result

# ============================================================
# SYMPTOM MATCHING
# ============================================================
SYMPTOM_KEYWORDS = {
    'afloj': ['floj', 'vibra', 'ruido'],
    'corros': ['corrosi', 'óxido', 'oxidaci', 'material'],
    'desgast': ['desgast', 'abras', 'erosi', 'material'],
    'fisura': ['grieta', 'fisura', 'fractur', 'rotura'],
    'fuga': ['fuga', 'goteo', 'derrame', 'pérdida'],
    'vibra': ['vibra', 'ruido', 'oscila'],
    'sobrecal': ['temperat', 'calent', 'calor', 'humo'],
    'fatiga': ['grieta', 'fractur', 'rotura'],
    'contamin': ['contamin', 'suciedad', 'partícul'],
    'erosi': ['erosi', 'desgast', 'material'],
    'bloqu': ['bloqu', 'atasc', 'obstrucc'],
    'cortocircuit': ['eléctric', 'cortocircuit', 'quemad'],
    'degradac': ['degradac', 'deterioro', 'envejecim'],
    'rotura': ['rotura', 'fractur', 'separac'],
    'obstrucc': ['obstrucc', 'bloqu', 'tapona'],
    'sobrecarg': ['sobrecarg', 'eléctric', 'temperat'],
}

def find_best_symptom(cause_text, symptoms_list):
    if not symptoms_list:
        return "Condición anormal detectada"
    c_lower = (cause_text or '').lower()
    for prefix, kws in SYMPTOM_KEYWORDS.items():
        if prefix in c_lower:
            for sym in symptoms_list:
                s_lower = sym.lower()
                for kw in kws:
                    if kw in s_lower:
                        return sym
    return symptoms_list[0]

# ============================================================
# FUNCTION TEMPLATES
# ============================================================
FUNCTION_TEMPLATES = {
    'BOMB_CEN': "Bombear fluido de proceso a caudal de diseño en {sys}, {area}",
    'BOMB_REC': "Dosificar fluido a presión en {sys}, {area}",
    'BOMB_ROT': "Bombear fluido manteniendo presión constante en {sys}, {area}",
    'BOMB_TUR': "Bombear fluido a caudal de diseño en {sys}, {area}",
    'BOMB_VAC': "Generar vacío para proceso en {sys}, {area}",
    'BOMB_HAR': "Bombear pulpa del harnero en {sys}, {area}",
    'VALVULAS': "Controlar flujo de proceso manteniendo regulación en {sys}, {area}",
    'MOT_ELÉC': "Accionar equipo proporcionando potencia mecánica de diseño en {sys}, {area}",
    'CELD_VDF': "Regular velocidad del equipo accionado mediante variación de frecuencia en {sys}, {area}",
    'INCH_DRI': "Proporcionar giro lento para mantenimiento en {sys}, {area}",
    'CICL_POT': "Convertir potencia eléctrica para accionamiento del equipo en {sys}, {area}",
    'ARRA_LÍQ': "Arrancar equipo limitando corriente mediante resistencia líquida en {sys}, {area}",
    'COMP_TOR': "Comprimir aire a presión de diseño para sistema neumático en {sys}, {area}",
    'ACUM_AIR': "Almacenar aire comprimido a presión de diseño en {sys}, {area}",
    'SECA_AIR': "Secar aire comprimido removiendo humedad en {sys}, {area}",
    'SOP_CENT': "Suministrar aire de proceso a caudal de diseño en {sys}, {area}",
    'CHAN_MAN': "Reducir tamaño de mineral al rango de diseño en {sys}, {area}",
    'CORR_TRA': "Transportar material a capacidad de diseño en {sys}, {area}",
    'ALIM_BAN': "Alimentar material de forma controlada en {sys}, {area}",
    'ALIM_PLA': "Alimentar material mediante placas a caudal controlado en {sys}, {area}",
    'HARN_EST': "Clasificar material por tamaño según apertura de malla en {sys}, {area}",
    'HIDROCIC': "Clasificar pulpa por tamaño mediante fuerza centrífuga en {sys}, {area}",
    'EST_AGIT': "Contener y agitar fluido manteniendo homogeneidad en {sys}, {area}",
    'EST_ALMA': "Almacenar fluido a capacidad nominal en {sys}, {area}",
    'TOLVA': "Almacenar material proveyendo descarga controlada en {sys}, {area}",
    'CAJÒN': "Contener y distribuir fluido hacia equipos aguas abajo en {sys}, {area}",
    'CHUTE': "Transferir material por gravedad entre equipos en {sys}, {area}",
    'PIS_SUM': "Contener fluido para recirculación o tratamiento en {sys}, {area}",
    'ESPESADO': "Espesar pulpa separando sólidos de líquido en {sys}, {area}",
    'FIL_PREN': "Filtrar fluido separando sólidos hasta humedad de diseño en {sys}, {area}",
    'FIL_HOJA': "Filtrar fluido removiendo sólidos suspendidos en {sys}, {area}",
    'FIL_COAL': "Filtrar fluido removiendo partículas finas en {sys}, {area}",
    'INT_ENFR': "Intercambiar calor enfriando fluido en {sys}, {area}",
    'INT_PLAC': "Intercambiar calor entre fluidos mediante placas en {sys}, {area}",
    'PUEN_GRÚ': "Izar y posicionar cargas para mantenimiento en {sys}, {area}",
    'TECL_ELE': "Izar cargas para mantenimiento en {sys}, {area}",
    'MANI_COR': "Manipular revestimientos durante mantenimiento de molino en {sys}, {area}",
    'AGIT_MEC': "Agitar fluido manteniendo sólidos en suspensión en {sys}, {area}",
    'TAB_CEL': "Distribuir y proteger suministro eléctrico en {sys}, {area}",
    'TAB_DIST': "Distribuir energía eléctrica a cargas del {sys}, {area}",
    'TAB_COMU': "Proveer comunicación de datos entre equipos del {sys}, {area}",
    'CEL_IPOT': "Proteger y seccionar circuitos de potencia en {sys}, {area}",
    'CEL_PMOT': "Arrancar y proteger motores eléctricos del {sys}, {area}",
    'TRAN_POT': "Transformar voltaje para distribución en {sys}, {area}",
    'UPS': "Proveer energía ininterrumpida a cargas críticas del {sys}, {area}",
    'BANC_BAT': "Almacenar energía eléctrica para respaldo en {sys}, {area}",
    'CARG_BAT': "Cargar banco de baterías del {sys}, {area}",
    'GAB_PLC': "Controlar proceso mediante lógica programable en {sys}, {area}",
    'GAB_RIO': "Adquirir señales de campo para PLC en {sys}, {area}",
    'PANE_SCA': "Visualizar y operar proceso desde sala de control en {sys}, {area}",
    'CONS_ESC': "Operar y monitorear proceso desde estación de trabajo en {sys}, {area}",
    'IND_PROC': "Indicar variable de proceso localmente en {sys}, {area}",
    'SEN_PROC': "Medir variable de proceso y transmitir señal en {sys}, {area}",
    'TRN_PROC': "Transmitir señal de medición al sistema de control en {sys}, {area}",
    'TRAN_SEÑ': "Transmitir señal de medición al sistema de control en {sys}, {area}",
    'SERVIDOR': "Procesar y almacenar datos del sistema de control en {sys}, {area}",
    'CORE_SWI': "Proveer conectividad de red entre equipos en {sys}, {area}",
    'FIREWALL': "Proteger red de control contra acceso no autorizado en {sys}, {area}",
    'KVMSWITC': "Compartir teclado/video/mouse entre servidores en {sys}, {area}",
    'DET_FH': "Detectar fuego y humo activando alarma en {sys}, {area}",
    'TAB_SCI': "Controlar sistema contraincendios en {sys}, {area}",
    'SUPR_MF': "Suprimir fuego mediante descarga de agente extintor en {sys}, {area}",
    'HIDRANTE': "Proveer agua contra incendios a presión en {sys}, {area}",
    'EST_MANG': "Proveer manguera contra incendios en {sys}, {area}",
    'GAB_MANG': "Almacenar y proteger mangueras contra incendios en {sys}, {area}",
    'DUCH_EME': "Proveer ducha y lavaojos de emergencia en {sys}, {area}",
    'UNID_AC': "Climatizar espacio manteniendo temperatura de diseño en {sys}, {area}",
    'UNI_PRES': "Presurizar sala evitando ingreso de polvo en {sys}, {area}",
    'CAL_ELÉC': "Calentar espacio manteniendo temperatura de confort en {sys}, {area}",
    'VENT_AX': "Ventilar espacio extrayendo aire viciado en {sys}, {area}",
    'VENT_CEN': "Ventilar espacio mediante extracción centralizada en {sys}, {area}",
    'CAMP_EXT': "Capturar y extraer gases/polvo del área de trabajo en {sys}, {area}",
    'DUCTOS': "Conducir aire/gases entre equipos de ventilación en {sys}, {area}",
    'CHIMENEA': "Descargar gases tratados a la atmósfera en {sys}, {area}",
    'LIN_PROC': "Transportar fluido manteniendo integridad y estanqueidad en {sys}, {area}",
    'SIST_LUB': "Lubricar componentes mecánicos manteniendo película de aceite en {sys}, {area}",
    'DIAL_ACE': "Filtrar y purificar aceite de lubricación en {sys}, {area}",
    'UNID_HID': "Proveer potencia hidráulica a presión de diseño en {sys}, {area}",
    'COMP_HID': "Controlar flujo mediante apertura/cierre hidráulico en {sys}, {area}",
    'BULL_BUL': "Empujar y apilar material en zona de acopio en {sys}, {area}",
    'CAM_PLUM': "Izar y posicionar cargas mediante pluma hidráulica en {sys}, {area}",
    'CAM_RIEG': "Regar caminos suprimiendo polvo fugitivo en {sys}, {area}",
    'CARG_FRO': "Cargar y transportar material en {sys}, {area}",
    'HORN_ELÉ': "Calentar material a temperatura de proceso en {sys}, {area}",
    'MUES_TRA': "Obtener muestra representativa para análisis en {sys}, {area}",
    'MUES_ROT': "Obtener muestra representativa para análisis en {sys}, {area}",
    'BASCULA': "Pesar material con precisión de diseño en {sys}, {area}",
    'CELOSÌA': "Proveer soporte estructural y contención en {sys}, {area}",
    'SOPORTES': "Soportar equipos y tuberías manteniendo alineación en {sys}, {area}",
    'PUER_MOT': "Permitir acceso vehicular mediante apertura/cierre en {sys}, {area}",
    'CÁM_CCTV': "Monitorear visualmente área para seguridad en {sys}, {area}",
    'PLAT_MÓV': "Proveer plataforma de trabajo móvil en {sys}, {area}",
    'CARR_MNT': "Transportar herramientas para mantenimiento en {sys}, {area}",
    'REM_PERN': "Remover pernos durante mantenimiento de molino en {sys}, {area}",
    'CASC_MOL': "Conducir metal fundido desde horno hasta moldeo en {sys}, {area}",
    'RACK_BAN': "Soportar y organizar bandejas de proceso en {sys}, {area}",
    'COL_POLV': "Capturar y remover polvo del aire en {sys}, {area}",
    'DET_META': "Detectar metales protegiendo equipos aguas abajo en {sys}, {area}",
    'SUPR_POL': "Suprimir polvo fugitivo mediante aspersión en {sys}, {area}",
    'MON_LAVA': "Lavar equipos mediante chorro de agua a presión en {sys}, {area}",
    'SIST_LAN': "Inyectar agua/aire a presión para limpieza en {sys}, {area}",
    'PAQ_QUIM': "Dosificar químicos de proceso a tasa controlada en {sys}, {area}",
    'PLAN_POT': "Potabilizar agua cumpliendo normas sanitarias en {sys}, {area}",
    'PTA_AR': "Tratar agua residual cumpliendo normas de descarga en {sys}, {area}",
    'HIDR_AGU': "Presurizar y distribuir agua potable en {sys}, {area}",
    'COL_PROC': "Realizar proceso de contacto contracorriente en {sys}, {area}",
    'VIBRADOR': "Fluidizar material evitando apelmazamiento en {sys}, {area}",
    'MONT_CAR': "Elevar y transportar cargas en {sys}, {area}",
    'HARN_MÓV': "Clasificar material por tamaño en operación móvil en {sys}, {area}",
    'MOL_SAG': "Moler mineral semi-autógenamente hasta P80 de diseño en {sys}, {area}",
    'MOL_BOLA': "Moler mineral mediante bolas hasta P80 de diseño en {sys}, {area}",
    'TRAN_MED': "Transformar señales de medición para protección en {sys}, {area}",
    'HORN_COM': "Calentar material mediante combustión controlada en {sys}, {area}",
    'ANAL_PAR': "Analizar distribución de tamaño de partículas en {sys}, {area}",
    'INT_TUBO': "Intercambiar calor entre fluidos mediante tubos en {sys}, {area}",
    'GRÚA_TRA': "Izar y transportar cargas en {sys}, {area}",
    'MESA_ROD': "Transportar material sobre rodillos en {sys}, {area}",
    'PESOMETR': "Medir flujo másico en correa transportadora en {sys}, {area}",
    'SILO': "Almacenar material a granel proveyendo descarga controlada en {sys}, {area}",
    'EST_MEZC': "Mezclar fluido manteniendo homogeneidad en {sys}, {area}",
    'TAB_HID': "Controlar actuadores hidráulicos en {sys}, {area}",
    'TAB_SCI': "Controlar sistema contraincendios en {sys}, {area}",
    'MUES_ALA': "Obtener muestra de alambre para análisis en {sys}, {area}",
    'ROMP_ROC': "Fragmentar roca sobredimensionada en parrilla en {sys}, {area}",
    'TORN_ALI': "Dosificar material mediante tornillo helicoidal en {sys}, {area}",
}

def get_mi_function(mi_name):
    """Generate RCM function description for a maintainable item."""
    mi_lower = (mi_name or '').lower()
    templates = {
        'rodamiento': "Soportar carga radial/axial permitiendo rotación libre",
        'sello': "Contener fluido evitando fuga al exterior",
        'impulsor': "Transferir energía al fluido generando presión",
        'motor': "Proporcionar torque y velocidad de diseño",
        'eje': "Transmitir torque al componente rotativo",
        'engranaje': "Transmitir potencia reduciendo velocidad",
        'acoplamiento': "Transmitir torque compensando desalineamiento",
        'correa': "Transmitir potencia entre poleas",
        'polea': "Guiar y tensar correa de transmisión",
        'revestimiento': "Proteger superficie contra desgaste",
        'válvula': "Controlar flujo manteniendo regulación",
        'filtro': "Retener partículas manteniendo flujo limpio",
        'tubería': "Conducir fluido manteniendo estanqueidad",
        'perno': "Mantener unión mecánica con torque especificado",
        'junta': "Sellar unión evitando fuga",
        'manguera': "Conducir fluido manteniendo flexibilidad",
        'cable': "Conducir energía eléctrica sin sobrecalentamiento",
        'contactor': "Energizar/desenergizar circuito de potencia",
        'relé': "Proteger circuito ante condiciones anormales",
        'sensor': "Medir variable con precisión de diseño",
        'batería': "Almacenar energía para respaldo",
        'membrana': "Separar medios manteniendo integridad",
        'cilindro': "Proveer fuerza lineal mediante presión",
        'bobina': "Generar campo electromagnético de diseño",
        'resistencia': "Disipar energía de forma controlada",
        'transformador': "Transformar voltaje con mínimas pérdidas",
        'interruptor': "Seccionar circuito para aislación segura",
        'fusible': "Interrumpir circuito ante sobrecorriente",
        'bomba': "Bombear fluido a caudal de diseño",
        'ventilador': "Mover aire/gas a caudal de diseño",
        'grasa': "Lubricar superficies reduciendo fricción",
        'aceite': "Lubricar y refrigerar componentes mecánicos",
        'piñón': "Transmitir potencia mediante engrane",
        'corona': "Transmitir potencia al componente rotativo",
        'camisa': "Proteger superficie interna del cilindro",
        'pistón': "Convertir presión en fuerza lineal",
        'difusor': "Convertir energía cinética en presión",
        'voluta': "Conducir fluido desde impulsor a descarga",
    }
    for key, func in templates.items():
        if key in mi_lower:
            return func
    return f"Cumplir función operativa según diseño"

# ============================================================
# LOAD ALL DATA
# ============================================================

# --- Hierarchy ---
print("Loading hierarchy...", flush=True)
wb_hier = openpyxl.load_workbook('seed_data/01_equipment_hierarchy.xlsx', data_only=True)
ws_hier = wb_hier['Equipment Hierarchy']
h_headers = [ws_hier.cell(1, c).value for c in range(1, ws_hier.max_column + 1)]
h_idx = {h: i + 1 for i, h in enumerate(h_headers)}

# Area/system name lookup
area_names = {}
system_names = {}
for row in range(2, ws_hier.max_row + 1):
    level = ws_hier.cell(row, h_idx['level']).value
    fl = ws_hier.cell(row, h_idx['sap_func_loc']).value
    pltxt = ws_hier.cell(row, h_idx['pltxt']).value
    if level == 1 and fl:
        area_names[fl] = pltxt
    elif level == 3 and fl:
        system_names[fl] = pltxt

# L4 equipment
equipment_l4 = []
for row in range(2, ws_hier.max_row + 1):
    if ws_hier.cell(row, h_idx['level']).value != 4:
        continue
    fl = ws_hier.cell(row, h_idx['sap_func_loc']).value or ''
    parts = fl.split('-')
    area_fl = '-'.join(parts[:2]) if len(parts) >= 2 else ''
    sys_fl = '-'.join(parts[:4]) if len(parts) >= 4 else ''

    sap_func_loc_short = ws_hier.cell(row, h_idx['sap_func_loc_short']).value or fl.split('-')[-1]
    equipment_l4.append({
        'equnr': ws_hier.cell(row, h_idx['equnr']).value,
        'sap_func_loc': fl,
        'sap_func_loc_short': sap_func_loc_short,
        'pltxt': ws_hier.cell(row, h_idx['pltxt']).value,
        'eqart': ws_hier.cell(row, h_idx['eqart']).value,
        'eqart_desc': ws_hier.cell(row, h_idx['eqart_desc']).value,
        'abckz': ws_hier.cell(row, h_idx['abckz']).value,
        'area_name': area_names.get(area_fl, parts[1] if len(parts) > 1 else ''),
        'system_name': system_names.get(sys_fl, ''),
    })

print(f"  L4 equipment: {len(equipment_l4)}", flush=True)

# --- BOM children ---
ws_bom = wb_hier['Equipment BOM']
bom_headers = [ws_bom.cell(1, c).value for c in range(1, ws_bom.max_column + 1)]
b_idx = {h: i + 1 for i, h in enumerate(bom_headers)}
children_by_parent = defaultdict(list)
for row in range(2, ws_bom.max_row + 1):
    parent = ws_bom.cell(row, b_idx['parent_equnr']).value
    comp = ws_bom.cell(row, b_idx['component_name']).value
    if parent and comp:
        children_by_parent[parent].append(comp)

# --- Catalog ---
print("Loading catalog...", flush=True)
wb_cat = openpyxl.load_workbook('seed_data/15_catalog_profiles.xlsx', data_only=True)
ws_cat = wb_cat['Catálogos']
catalog_data = defaultdict(lambda: {'causas': [], 'partes': [], 'sintomas': []})
for r in range(2, ws_cat.max_row + 1):
    pc = ws_cat.cell(r, 1).value
    tipo = ws_cat.cell(r, 3).value
    valor = ws_cat.cell(r, 5).value
    if not pc or not valor:
        continue
    if tipo == 'Causas de falla':
        catalog_data[pc]['causas'].append(valor)
    elif tipo == 'Partes de falla':
        catalog_data[pc]['partes'].append(valor)
    elif tipo == 'Síntomas de falla':
        catalog_data[pc]['sintomas'].append(valor)

print(f"  Catalog profiles: {len(catalog_data)}", flush=True)

# --- FM-MASTER ---
print("Loading FM-MASTER...", flush=True)
wb_fm = openpyxl.load_workbook('skills/00-knowledge-base/data-models/failure-modes/FM-MASTER-REFERENCE.xlsx', data_only=True)
ws_fm = wb_fm['72 Failure Modes']
fm_headers = [ws_fm.cell(1, c).value for c in range(1, ws_fm.max_column + 1)]
fm_idx = {h: i + 1 for i, h in enumerate(fm_headers)}
fm_data = {}
for r in range(2, ws_fm.max_row + 1):
    fm_num = ws_fm.cell(r, fm_idx['FM#']).value
    if not fm_num:
        break
    fm_data[fm_num] = {
        'mechanism': ws_fm.cell(r, fm_idx['Mechanism']).value,
        'cause': ws_fm.cell(r, fm_idx['Cause']).value,
        'pattern': ws_fm.cell(r, fm_idx['Pattern']).value,
        'detection': ws_fm.cell(r, fm_idx['Primary CBM Technique']).value,
        'evidence': ws_fm.cell(r, fm_idx['Top P-Conditions']).value,
    }

# --- Criticality ---
print("Loading criticality...", flush=True)
wb_crit = openpyxl.load_workbook('seed_data/02_criticality_assessment.xlsx', data_only=True)
ws_crit = wb_crit[wb_crit.sheetnames[0]]
c_headers = [ws_crit.cell(1, c).value for c in range(1, ws_crit.max_column + 1)]
c_idx = {h: i + 1 for i, h in enumerate(c_headers)}
downtime_col = None
for h in c_headers:
    if h and 'downtime' in str(h).lower():
        downtime_col = c_idx[h]
        break

criticality_by_fl = {}
for r in range(2, ws_crit.max_row + 1):
    fl = ws_crit.cell(r, c_idx.get('sap_func_loc', 1)).value
    if fl:
        crit = ws_crit.cell(r, c_idx.get('criticality_level', c_idx.get('abckz', 1))).value
        dt = ws_crit.cell(r, downtime_col).value if downtime_col else None
        criticality_by_fl[fl] = {'criticality': crit, 'downtime': dt or random.choice([4, 8, 12, 24, 48])}

print(f"  Criticality entries: {len(criticality_by_fl)}", flush=True)

# ============================================================
# GENERATE FAILURE MODES
# ============================================================
print("\n" + "=" * 60, flush=True)
print("GENERATING FAILURE MODES", flush=True)
print("=" * 60, flush=True)

HEADERS = [
    'equipment_tag', 'equnr', 'sap_func_loc', 'area',
    'equipment_function_description', 'equipment_functional_failure',
    'function_type', 'failure_type', 'subunit', 'maintainable_item',
    'maintainable_item_function_description', 'maintainable_item_functional_failure',
    'partes_falla', 'sintomas_falla', 'causas_falla',
    'fm_what', 'fm_mechanism', 'fm_cause', 'fm_number',
    'failure_pattern', 'failure_consequence', 'evidence',
    'downtime_hours', 'detection_method',
    'rpn_severity', 'rpn_occurrence', 'rpn_detection', 'rpn_total'
]

# Use write_only for performance
wb_out = Workbook(write_only=True)
ws_out = wb_out.create_sheet('failure_modes')
ws_out.append(HEADERS)

total_rows = 0
eq_count = 0
no_profile = 0

for eq in equipment_l4:
    eqart = eq['eqart']
    profile_code = EQART_TO_PROFILE.get(eqart)

    # Get catalog data
    if profile_code and profile_code in catalog_data:
        cat = catalog_data[profile_code]
    elif eqart in catalog_data:
        cat = catalog_data[eqart]
        profile_code = eqart
    else:
        no_profile += 1
        bom = children_by_parent.get(eq['equnr'], ['Componente general'])
        cat = {
            'partes': bom if bom else ['Componente general'],
            'causas': ['Desgaste', 'Corrosión', 'Fatiga', 'Vibración', 'Sobrecalentamiento'],
            'sintomas': ['Condición anormal detectada']
        }

    partes = cat['partes'] or ['Componente general']
    causas = cat['causas'] or ['Desgaste', 'Corrosión', 'Fatiga']
    sintomas = cat['sintomas'] or ['Condición anormal detectada']

    # Equipment function (RCM Moubray)
    area_name = eq['area_name'] or 'planta'
    sys_name = eq['system_name'] or eq['pltxt'] or 'sistema'
    template = FUNCTION_TEMPLATES.get(
        EQART_TO_PROFILE.get(eqart, ''),
        f"Cumplir función operativa de {eq['eqart_desc'] or eqart} en {{sys}}, {{area}}"
    )
    func_desc = template.replace('{sys}', sys_name).replace('{area}', f"área {area_name}")
    func_failure = f"Incapaz de {func_desc[0].lower()}{func_desc[1:]}"

    # Criticality
    crit = criticality_by_fl.get(eq['sap_func_loc'], {'criticality': eq['abckz'] or '2', 'downtime': 12})
    abckz = crit['criticality'] or eq['abckz'] or '2'
    downtime = crit.get('downtime', 12)

    # BOM children for subunit
    bom_children = children_by_parent.get(eq['equnr'], [])

    # RPN severity
    sev_map = {1: 8, 2: 5, 3: 3, '1': 8, '2': 5, '3': 3, 'A': 8, 'B': 5, 'C': 3}
    severity = sev_map.get(abckz, 5)

    # Consequence
    if abckz in (1, '1', 'A'):
        consequence = 'OPERATIONAL'
    elif abckz in (3, '3', 'C'):
        consequence = 'NON-OPERATIONAL'
    else:
        consequence = 'OPERATIONAL'

    # For each PART, assign relevant causes (filtered by part type)
    for parte in partes:
        relevant_causes = get_relevant_causes(parte, causas)

        for causa in relevant_causes:
            # Map to FM-MASTER
            fm_match = CAUSE_MAP.get(causa)
            if not fm_match:
                for key in CAUSE_MAP:
                    if key.lower() in causa.lower() or causa.lower() in key.lower():
                        fm_match = CAUSE_MAP[key]
                        break
            if not fm_match:
                fm_match = DEFAULT_FM

            fm_mechanism, fm_cause, fm_number, pattern = fm_match
            fm_info = fm_data.get(fm_number, {})

            # Best symptom
            best_symptom = find_best_symptom(causa, sintomas)

            # Subunit
            subunit = ''
            for child in bom_children:
                if any(w in child.lower() for w in parte.lower().split()[:2] if len(w) > 3):
                    subunit = child
                    break
            if not subunit:
                subunit = bom_children[0] if bom_children else parte

            # MI function
            mi_func = get_mi_function(parte)
            mi_failure = f"No logra {mi_func[0].lower()}{mi_func[1:]}"

            # RPN
            occ_map = {'A': 3, 'B': 5, 'C': 6, 'D': 4, 'E': 7, 'F': 8}
            occurrence = occ_map.get(pattern, 5)

            det_lower = ((fm_info.get('detection') or '') + '').lower()
            if 'vibra' in det_lower or 'online' in det_lower or 'monitor' in det_lower:
                detection = 3
            elif 'oil' in det_lower or 'thermo' in det_lower or 'insul' in det_lower:
                detection = 4
            elif 'visual' in det_lower or 'inspect' in det_lower:
                detection = 6
            elif 'mpi' in det_lower or 'ndt' in det_lower:
                detection = 5
            else:
                detection = 7

            evidence = fm_info.get('evidence', 'Condición anormal observable')
            if evidence and len(str(evidence)) > 150:
                evidence = str(evidence)[:147] + '...'

            failure_type = 'TOTAL' if pattern in ('E', 'F') else 'PARTIAL'

            ws_out.append([
                eq['sap_func_loc_short'],  # equipment_tag
                eq['equnr'],           # equnr
                eq['sap_func_loc'],    # sap_func_loc
                area_name,             # area
                func_desc,             # equipment_function_description
                func_failure,          # equipment_functional_failure
                'PRIMARY',             # function_type
                failure_type,          # failure_type
                subunit,               # subunit
                parte,                 # maintainable_item
                mi_func,               # maintainable_item_function_description
                mi_failure,            # maintainable_item_functional_failure
                parte,                 # partes_falla
                best_symptom,          # sintomas_falla
                causa,                 # causas_falla
                parte,                 # fm_what
                fm_mechanism,          # fm_mechanism
                fm_cause,              # fm_cause
                fm_number,             # fm_number
                pattern,               # failure_pattern
                consequence,           # failure_consequence
                evidence,              # evidence
                downtime or 12,        # downtime_hours
                fm_info.get('detection', 'Inspección visual'),  # detection_method
                severity,              # rpn_severity
                occurrence,            # rpn_occurrence
                detection,             # rpn_detection
                severity * occurrence * detection,  # rpn_total
            ])
            total_rows += 1

    eq_count += 1
    if eq_count % 200 == 0:
        print(f"  Processed {eq_count}/{len(equipment_l4)} equipment... ({total_rows} FMs)", flush=True)

print(f"\n  Total equipment: {eq_count}", flush=True)
print(f"  Without catalog: {no_profile}", flush=True)
print(f"  Total failure modes: {total_rows}", flush=True)

# Save
print("\nSaving 03_failure_modes.xlsx...", flush=True)
wb_out.save('seed_data/03_failure_modes.xlsx')
print("SAVED!", flush=True)

# Verification
unique_tags = set()
for eq in equipment_l4:
    unique_tags.add(eq['equnr'])
print(f"\nVERIFICATION:", flush=True)
print(f"  Columns: {len(HEADERS)}", flush=True)
print(f"  Data rows: {total_rows}", flush=True)
print(f"  Equipment coverage: {eq_count}/{len(equipment_l4)} ({eq_count/len(equipment_l4)*100:.1f}%)", flush=True)
print("DONE!", flush=True)
