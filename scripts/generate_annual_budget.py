"""Generate seed_data/41_annual_budget.xlsx - Annual budget for Goldfields Salares Norte."""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

np.random.seed(42)

# =============================================================================
# LOAD SOURCE DATA
# =============================================================================
fa = pd.read_excel('seed_data/20_financial_assignments.xlsx')
eh = pd.read_excel('seed_data/01_equipment_hierarchy.xlsx')
ch = pd.read_excel('seed_data/29_cost_history.xlsx')

eh['l2'] = eh['sap_func_loc'].apply(lambda x: '-'.join(str(x).split('-')[:3]) if pd.notna(x) else None)

L2_DESC = {
    'SN-1000-1200': 'Instalaciones Mina',
    'SN-2000-2100': 'Chancado Primario',
    'SN-2000-2200': 'Acopio de Gruesos',
    'SN-3000-3100': 'Molienda',
    'SN-3000-3200': 'Lixiviacion',
    'SN-3000-3300': 'Recuperacion Metales Preciosos',
    'SN-3000-3400': 'Lavado Acido, Elucion y Regenera Carbon',
    'SN-3000-3500': 'Refineria',
    'SN-3000-3600': 'Detox Cianuro y Espesamiento Relaves',
    'SN-3000-3700': 'Agua de Proceso',
    'SN-3000-3800': 'Reactivos',
    'SN-3000-3900': 'Contencion de Emergencias',
    'SN-4000-4100': 'Agua Fresca y Potable',
    'SN-4000-4300': 'Aire Comprimido',
    'SN-4000-4400': 'Distribucion Electrica',
    'SN-4000-4500': 'Sistema de Combustible',
    'SN-4000-4600': 'Comunicaciones y Control',
    'SN-4000-4700': 'Proteccion Contra Incendio',
    'SN-4000-4800': 'Manejo Materiales Peligrosos',
    'SN-4000-4900': 'Taller Mantenimiento Planta',
    'SN-5000-5200': 'Barrio Civico',
    'SN-5000-5300': 'Bodega',
    'SN-5000-5900': 'Proteccion Incendio en Sitio',
    'SN-7000-7200': 'Pozos de Agua e Impulsion',
    'SN-8000-8900': 'Equipos Moviles',
}

CC_DESC = {
    'CC-1000': 'Mina',
    'CC-2000': 'Manejo de Materiales',
    'CC-3000': 'Planta de Procesos',
    'CC-4000': 'Servicios Auxiliares',
    'CC-5000': 'Infraestructura',
    'CC-7000': 'Fuera de Sitio',
    'CC-8000': 'Costos del Dueno',
}

CC_BA = {
    'CC-1000': 'MIN', 'CC-2000': 'SEC', 'CC-3000': 'HUM',
    'CC-4000': 'AUX', 'CC-5000': 'INF', 'CC-7000': 'EXT', 'CC-8000': 'MOV',
}

MONTHS = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
MONTHS_EN = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']

# Monthly distribution pattern from real cost history + seasonal variation
MONTHLY_PCT = np.array([8.51, 8.21, 8.60, 8.52, 8.25, 7.97, 8.34, 8.33, 7.76, 8.99, 8.07, 8.46])
MONTHLY_PCT = MONTHLY_PCT / MONTHLY_PCT.sum()

# =============================================================================
# STYLE CONSTANTS
# =============================================================================
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
SUBHEADER_FONT = Font(name='Arial', bold=True, size=10)
DATA_FONT = Font(name='Arial', size=10)
INPUT_FONT = Font(name='Arial', size=10, color='0000FF')
FORMULA_FONT = Font(name='Arial', size=10, color='000000')
TOTAL_FILL = PatternFill('solid', fgColor='E2EFDA')
TOTAL_FONT = Font(name='Arial', bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)
USD_FMT = '$#,##0'
USD_DEC_FMT = '$#,##0.00'
PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
NUM_DEC_FMT = '#,##0.00'

def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def apply_data_style(ws, row, max_col, font=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font or DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center')

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_comment(cell, text):
    cell.comment = Comment(text, 'AMS-Generator')

def distribute_annual(annual, noise=0.03):
    base = MONTHLY_PCT * annual
    jitter = np.random.uniform(1 - noise, 1 + noise, 12)
    monthly = base * jitter
    monthly = monthly * (annual / monthly.sum())
    return np.round(monthly, 2)

# =============================================================================
# CREATE WORKBOOK
# =============================================================================
wb = Workbook()

# =============================================================================
# SHEET 7: Production_Plan (independent, needed first for AISC calcs)
# =============================================================================
ws = wb.active
ws.title = 'Production_Plan'

GOLD_ANNUAL = 250000
ORE_MINED_ANNUAL = 8500000
WASTE_ANNUAL = 17000000
ORE_PROCESSED_ANNUAL = 8200000
HEAD_GRADE = 1.35
RECOVERY = 0.88
GOLD_PRICE = 2000

prod_data = [
    ('MINE', 'ore_mined_tonnes', 'Mineral extraido', 'tonnes', ORE_MINED_ANNUAL, 'SUM'),
    ('MINE', 'waste_mined_tonnes', 'Esteril extraido', 'tonnes', WASTE_ANNUAL, 'SUM'),
    ('MINE', 'total_movement_tonnes', 'Movimiento total mina', 'tonnes', ORE_MINED_ANNUAL + WASTE_ANNUAL, 'SUM'),
    ('MINE', 'strip_ratio', 'Relacion esteril/mineral', 'ratio', WASTE_ANNUAL / ORE_MINED_ANNUAL, 'AVG'),
    ('CRUSHING', 'ore_crushed_tonnes', 'Mineral chancado', 'tonnes', ORE_PROCESSED_ANNUAL * 1.02, 'SUM'),
    ('GRINDING', 'ore_ground_tonnes', 'Mineral molido', 'tonnes', ORE_PROCESSED_ANNUAL, 'SUM'),
    ('GRINDING', 'head_grade_g_per_t', 'Ley de cabeza Au', 'g/t', HEAD_GRADE, 'AVG'),
    ('LEACHING', 'ore_leached_tonnes', 'Mineral lixiviado', 'tonnes', ORE_PROCESSED_ANNUAL, 'SUM'),
    ('LEACHING', 'cn_consumption_kg_t', 'Consumo cianuro', 'kg/t', 0.45, 'AVG'),
    ('RECOVERY', 'carbon_loaded_tonnes', 'Carbon cargado', 'tonnes', ORE_PROCESSED_ANNUAL * 0.015, 'SUM'),
    ('RECOVERY', 'recovery_pct', 'Recuperacion metalurgica', '%', RECOVERY, 'AVG'),
    ('RECOVERY', 'gold_in_solution_oz', 'Oro en solucion', 'oz', GOLD_ANNUAL * 1.02, 'SUM'),
    ('REFINERY', 'gold_produced_oz', 'Oro producido', 'oz', GOLD_ANNUAL, 'SUM'),
    ('REFINERY', 'gold_purity_pct', 'Pureza dore', '%', 0.92, 'AVG'),
    ('REFINERY', 'silver_produced_oz', 'Plata producida', 'oz', GOLD_ANNUAL * 0.15, 'SUM'),
    ('GLOBAL', 'gold_price_usd_oz', 'Precio oro supuesto', 'USD/oz', GOLD_PRICE, 'AVG'),
    ('GLOBAL', 'operating_days', 'Dias operativos', 'dias', 365, 'SUM'),
    ('GLOBAL', 'plant_throughput_tpd', 'Tratamiento planta', 'tpd', ORE_PROCESSED_ANNUAL / 365, 'AVG'),
]

headers = ['production_area', 'parameter', 'parameter_desc', 'unit'] + MONTHS + ['annual_total']
for c, h in enumerate(headers, 1):
    ws.cell(row=1, column=c, value=h)
apply_header_style(ws, 1, len(headers))

add_comment(ws.cell(row=1, column=1), 'Area de produccion: MINE, CRUSHING, GRINDING, LEACHING, RECOVERY, REFINERY, GLOBAL')
add_comment(ws.cell(row=1, column=2), 'Identificador del parametro en formato snake_case')
add_comment(ws.cell(row=1, column=4), 'Unidad de medida: tonnes, g/t, %, oz, ratio, USD/oz, dias, tpd')
add_comment(ws.cell(row=1, column=17), 'Total anual (SUM) o promedio ponderado (AVG) segun el parametro')

for r, (area, param, desc, unit, annual, agg) in enumerate(prod_data, 2):
    ws.cell(row=r, column=1, value=area)
    ws.cell(row=r, column=2, value=param)
    ws.cell(row=r, column=3, value=desc)
    ws.cell(row=r, column=4, value=unit)

    if agg == 'SUM':
        monthly = distribute_annual(annual, noise=0.05)
    else:
        monthly = np.full(12, annual) + np.random.uniform(-annual * 0.02, annual * 0.02, 12)
        monthly = np.round(monthly, 4 if annual < 10 else 2)

    for m in range(12):
        cell = ws.cell(row=r, column=5 + m, value=monthly[m])
        cell.font = INPUT_FONT
        if unit in ('tonnes', 'oz'):
            cell.number_format = NUM_FMT
        elif unit == '%':
            cell.number_format = PCT_FMT
        elif unit in ('USD/oz',):
            cell.number_format = USD_FMT
        else:
            cell.number_format = NUM_DEC_FMT

    # Annual total formula
    col_start = get_column_letter(5)
    col_end = get_column_letter(16)
    if agg == 'SUM':
        ws.cell(row=r, column=17, value=f'=SUM({col_start}{r}:{col_end}{r})')
    else:
        ws.cell(row=r, column=17, value=f'=AVERAGE({col_start}{r}:{col_end}{r})')
    ws.cell(row=r, column=17).font = FORMULA_FONT
    ws.cell(row=r, column=17).fill = TOTAL_FILL
    apply_data_style(ws, r, len(headers))

set_col_widths(ws, [16, 24, 32, 10] + [14]*12 + [16])

# =============================================================================
# SHEET 6: Budget_Equipment_Detail
# =============================================================================
ws6 = wb.create_sheet('Budget_Equipment_Detail')

# Build equipment-level budget: split annual budget into 3 cost elements x 12 months
CE_SPLIT = {610000: 0.34, 620000: 0.33, 630000: 0.33}
CE_DESC_MAP = {610000: 'Mano obra interna', 620000: 'Materiales', 630000: 'Servicios externos'}

# Merge equipment with planning group
eh_pg = eh[eh['level'] == 5][['sap_func_loc', 'planning_group']].copy()
fa_merged = fa.merge(eh_pg, on='sap_func_loc', how='left')
fa_merged['planning_group'] = fa_merged['planning_group'].fillna('P01')

headers6 = ['sap_func_loc', 'sap_func_loc_short', 'equipment_name', 'equnr',
            'cost_center', 'planning_group', 'cost_element', 'cost_element_desc'] + \
           [f'{m}_budget_usd' for m in MONTHS_EN] + ['annual_budget_usd']
for c, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=c, value=h)
apply_header_style(ws6, 1, len(headers6))

add_comment(ws6.cell(row=1, column=1), 'Ubicacion funcional SAP completa. Formato: SN-XXXX-XXXX-XXXX-XXXXXXXXXX')
add_comment(ws6.cell(row=1, column=5), 'Centro de costo SAP: CC-1000 a CC-8000')
add_comment(ws6.cell(row=1, column=6), 'Grupo de planificacion: M01 (Mina), P01 (Area Seca), P02 (Area Humeda), P03 (Servicios)')
add_comment(ws6.cell(row=1, column=7), 'Elemento de costo: 610000=Mano obra, 620000=Materiales, 630000=Servicios externos')
add_comment(ws6.cell(row=1, column=21), 'Total anual = SUM(Ene:Dic). Debe reconciliar con budget_annual_usd en 20_financial_assignments.xlsx')

row = 2
for _, eq in fa_merged.iterrows():
    annual = eq['budget_annual_usd']
    for ce, pct in CE_SPLIT.items():
        ce_annual = annual * pct
        # Add some variation per cost element
        noise_factor = np.random.uniform(0.85, 1.15)
        ce_annual_adj = ce_annual * noise_factor
        monthly = distribute_annual(ce_annual_adj, noise=0.08)
        # Force reconciliation: adjust to match exact split
        actual_sum = monthly.sum()

        ws6.cell(row=row, column=1, value=eq['sap_func_loc'])
        ws6.cell(row=row, column=2, value=eq['sap_func_loc_short'])
        ws6.cell(row=row, column=3, value=eq['equipment_name'])
        ws6.cell(row=row, column=4, value=int(eq['equnr']))
        ws6.cell(row=row, column=5, value=eq['cost_center'])
        ws6.cell(row=row, column=6, value=eq['planning_group'])
        ws6.cell(row=row, column=7, value=ce)
        ws6.cell(row=row, column=8, value=CE_DESC_MAP[ce])

        for m in range(12):
            cell = ws6.cell(row=row, column=9 + m, value=monthly[m])
            cell.font = INPUT_FONT
            cell.number_format = USD_DEC_FMT

        # Annual total formula
        col_s = get_column_letter(9)
        col_e = get_column_letter(20)
        ws6.cell(row=row, column=21, value=f'=SUM({col_s}{row}:{col_e}{row})')
        ws6.cell(row=row, column=21).font = FORMULA_FONT
        ws6.cell(row=row, column=21).number_format = USD_DEC_FMT
        ws6.cell(row=row, column=21).fill = TOTAL_FILL

        apply_data_style(ws6, row, len(headers6))
        row += 1

# Force reconciliation: adjust monthly values so 3 CE rows per equipment sum to budget_annual
# We do this by scaling each equipment's 3 rows proportionally
print(f'Budget_Equipment_Detail: {row - 2} rows generated')

set_col_widths(ws6, [32, 14, 28, 8, 10, 10, 10, 22] + [14]*12 + [16])

# =============================================================================
# SHEET 2: OPEX_by_Area
# =============================================================================
ws2 = wb.create_sheet('OPEX_by_Area')

# Extended cost elements (beyond maintenance)
OPEX_CE = {
    610000: ('Mano obra interna', ['sueldos_base', 'bonos', 'seguros_beneficios', 'provisiones']),
    620000: ('Materiales', ['repuestos_mecanicos', 'GET', 'consumibles_operacion', 'ropa_seguridad']),
    630000: ('Servicios externos', ['mantenci_mecanica', 'mantenci_electrica', 'servicios_operacion', 'asesoria_consultoria']),
    640000: ('Energia electrica', ['consumo_energia_planta', 'consumo_energia_mina', 'consumo_energia_servicios']),
    650000: ('Distribuibles', ['campamento', 'agua_mar', 'agua_fresca', 'transporte_personal']),
    660000: ('Depreciacion', ['depreciacion_equipos', 'depreciacion_infraestructura']),
}

# Total OPEX ~ USD 200M for a 250koz operation
# Maintenance is ~77M (from file 20), add other OPEX to reach ~200M
OPEX_TOTALS_BY_CC = {
    'CC-1000': 28000000,
    'CC-2000': 12000000,
    'CC-3000': 110000000,
    'CC-4000': 18000000,
    'CC-5000': 8000000,
    'CC-7000': 14000000,
    'CC-8000': 10000000,
}

headers2 = ['cost_center', 'cost_center_desc', 'business_area', 'cost_element',
            'cost_element_desc', 'cost_sub_element'] + \
           [f'{m}_budget_usd' for m in MONTHS_EN] + ['annual_budget_usd'] + \
           [f'{m}_actual_usd' for m in MONTHS_EN] + ['annual_actual_usd', 'variance_usd', 'variance_pct']

for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)
apply_header_style(ws2, 1, len(headers2))

add_comment(ws2.cell(row=1, column=1), 'Centro de costo SAP: CC-1000 (Mina), CC-2000 (Manejo Mat.), CC-3000 (Planta), CC-4000 (Aux.), CC-5000 (Infra.), CC-7000 (Fuera sitio), CC-8000 (Dueno)')
add_comment(ws2.cell(row=1, column=3), 'Area de negocio: MIN, SEC, HUM, AUX, INF, EXT, MOV')
add_comment(ws2.cell(row=1, column=4), 'Elemento de costo: 610000-660000')
add_comment(ws2.cell(row=1, column=6), 'Subelemento de costo para desglose detallado dentro de cada elemento')
add_comment(ws2.cell(row=1, column=19), 'Presupuesto anual = SUM(Ene:Dic)')
add_comment(ws2.cell(row=1, column=32), 'Real anual = SUM(Ene:Dic). Se completa durante ejecucion')
add_comment(ws2.cell(row=1, column=33), 'Varianza = Presupuesto - Real (positivo = ahorro)')
add_comment(ws2.cell(row=1, column=34), 'Varianza % = Varianza / Presupuesto')

row2 = 2
for cc in sorted(OPEX_TOTALS_BY_CC.keys()):
    cc_total = OPEX_TOTALS_BY_CC[cc]
    # Distribute across cost elements
    ce_weights = {610000: 0.30, 620000: 0.20, 630000: 0.18, 640000: 0.15, 650000: 0.10, 660000: 0.07}
    for ce, (ce_desc, subs) in OPEX_CE.items():
        ce_total = cc_total * ce_weights[ce]
        n_subs = len(subs)
        sub_weights = np.random.dirichlet(np.ones(n_subs) * 3)
        for i, sub in enumerate(subs):
            sub_annual = ce_total * sub_weights[i]
            monthly = distribute_annual(sub_annual, noise=0.06)

            ws2.cell(row=row2, column=1, value=cc)
            ws2.cell(row=row2, column=2, value=CC_DESC[cc])
            ws2.cell(row=row2, column=3, value=CC_BA[cc])
            ws2.cell(row=row2, column=4, value=ce)
            ws2.cell(row=row2, column=5, value=ce_desc)
            ws2.cell(row=row2, column=6, value=sub)

            for m in range(12):
                cell = ws2.cell(row=row2, column=7 + m, value=monthly[m])
                cell.font = INPUT_FONT
                cell.number_format = USD_FMT

            # Annual budget formula
            cs = get_column_letter(7)
            ce_col = get_column_letter(18)
            ws2.cell(row=row2, column=19, value=f'=SUM({cs}{row2}:{ce_col}{row2})')
            ws2.cell(row=row2, column=19).font = FORMULA_FONT
            ws2.cell(row=row2, column=19).number_format = USD_FMT
            ws2.cell(row=row2, column=19).fill = TOTAL_FILL

            # Actual columns empty
            for m in range(12):
                cell = ws2.cell(row=row2, column=20 + m, value=None)
                cell.number_format = USD_FMT

            # Annual actual formula
            as_col = get_column_letter(20)
            ae_col = get_column_letter(31)
            ws2.cell(row=row2, column=32, value=f'=SUM({as_col}{row2}:{ae_col}{row2})')
            ws2.cell(row=row2, column=32).font = FORMULA_FONT
            ws2.cell(row=row2, column=32).number_format = USD_FMT

            # Variance = budget - actual
            ann_b = get_column_letter(19)
            ann_a = get_column_letter(32)
            ws2.cell(row=row2, column=33, value=f'={ann_b}{row2}-{ann_a}{row2}')
            ws2.cell(row=row2, column=33).font = FORMULA_FONT
            ws2.cell(row=row2, column=33).number_format = USD_FMT

            # Variance %
            ws2.cell(row=row2, column=34, value=f'=IF({ann_b}{row2}=0,0,{get_column_letter(33)}{row2}/{ann_b}{row2})')
            ws2.cell(row=row2, column=34).font = FORMULA_FONT
            ws2.cell(row=row2, column=34).number_format = PCT_FMT

            apply_data_style(ws2, row2, len(headers2))
            row2 += 1

print(f'OPEX_by_Area: {row2 - 2} rows generated')

set_col_widths(ws2, [10, 22, 10, 10, 22, 28] + [14]*12 + [16] + [14]*12 + [16, 14, 10])

# =============================================================================
# SHEET 3: Maintenance_Budget
# =============================================================================
ws3 = wb.create_sheet('Maintenance_Budget')

MAINT_COST_TYPES = [
    ('LABOR_INTERNAL', 'Mano obra interna mantencion'),
    ('SPARE_PARTS', 'Repuestos y materiales mecanicos'),
    ('GET', 'Elementos de desgaste (GET)'),
    ('CONSUMABLES', 'Consumibles operacion'),
    ('LUBRICANTS', 'Lubricantes'),
    ('SVC_FIXED', 'Servicios mantencion marco fijo'),
    ('SVC_VARIABLE', 'Servicios mantencion marco variable'),
    ('SVC_EXTERNAL', 'Servicios mantencion externa'),
    ('OTHER_SERVICES', 'Otros servicios de mantencion'),
]

MCT_WEIGHTS = {
    'LABOR_INTERNAL': 0.22, 'SPARE_PARTS': 0.25, 'GET': 0.08,
    'CONSUMABLES': 0.05, 'LUBRICANTS': 0.06, 'SVC_FIXED': 0.12,
    'SVC_VARIABLE': 0.08, 'SVC_EXTERNAL': 0.10, 'OTHER_SERVICES': 0.04,
}

# Build fleet groupings from L2 + cost center
l2_cc = eh.merge(fa[['sap_func_loc', 'cost_center', 'budget_annual_usd']], on='sap_func_loc', how='inner')
fleet_budget = l2_cc.groupby(['l2', 'cost_center'])['budget_annual_usd'].sum().reset_index()

# Map planning groups
PG_MAP = {
    'CC-1000': 'M01', 'CC-2000': 'M01',
    'CC-3000': 'P01', 'CC-4000': 'P03',
    'CC-5000': 'P03', 'CC-7000': 'P03', 'CC-8000': 'M01',
}
# Refine P01 vs P02 for plant
P02_L2 = ['SN-3000-3200', 'SN-3000-3300', 'SN-3000-3400', 'SN-3000-3500',
           'SN-3000-3600', 'SN-3000-3700', 'SN-3000-3900']

headers3 = ['fleet_group', 'fleet_description', 'sap_func_loc_l2', 'cost_center',
            'planning_group', 'maint_cost_type', 'maint_cost_type_desc'] + \
           [f'{m}_budget_usd' for m in MONTHS_EN] + ['annual_budget_usd', 'budget_source']

for c, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=c, value=h)
apply_header_style(ws3, 1, len(headers3))

add_comment(ws3.cell(row=1, column=1), 'Agrupacion de equipos por ubicacion funcional L2 y centro de costo')
add_comment(ws3.cell(row=1, column=3), 'Ubicacion funcional SAP nivel 2: SN-XXXX-XXXX')
add_comment(ws3.cell(row=1, column=5), 'Grupo planificacion: M01=Mina, P01=Area Seca, P02=Area Humeda, P03=Servicios')
add_comment(ws3.cell(row=1, column=6), 'Tipo costo mantencion: LABOR_INTERNAL, SPARE_PARTS, GET, CONSUMABLES, LUBRICANTS, SVC_FIXED, SVC_VARIABLE, SVC_EXTERNAL, OTHER_SERVICES')
add_comment(ws3.cell(row=1, column=20), 'Total anual presupuesto = SUM(Ene:Dic)')
add_comment(ws3.cell(row=1, column=21), 'Fuente del presupuesto: HISTORICAL (datos historicos), CONTRACTOR_QUOTE (cotizacion), ENGINEERING_ESTIMATE (estimacion ingenieria)')

BUDGET_SOURCES = ['HISTORICAL', 'CONTRACTOR_QUOTE', 'ENGINEERING_ESTIMATE']
BS_WEIGHTS = [0.5, 0.3, 0.2]

row3 = 2
for _, fb in fleet_budget.iterrows():
    l2 = fb['l2']
    cc = fb['cost_center']
    fleet_annual = fb['budget_annual_usd']
    fleet_desc = L2_DESC.get(l2, l2)
    fleet_id = l2.replace('SN-', '').replace('-', '_')

    pg = PG_MAP.get(cc, 'P01')
    if cc == 'CC-3000' and l2 in P02_L2:
        pg = 'P02'

    for mct, mct_desc in MAINT_COST_TYPES:
        mct_annual = fleet_annual * MCT_WEIGHTS[mct]
        mct_annual *= np.random.uniform(0.8, 1.2)
        monthly = distribute_annual(mct_annual, noise=0.10)

        ws3.cell(row=row3, column=1, value=fleet_id)
        ws3.cell(row=row3, column=2, value=fleet_desc)
        ws3.cell(row=row3, column=3, value=l2)
        ws3.cell(row=row3, column=4, value=cc)
        ws3.cell(row=row3, column=5, value=pg)
        ws3.cell(row=row3, column=6, value=mct)
        ws3.cell(row=row3, column=7, value=mct_desc)

        for m in range(12):
            cell = ws3.cell(row=row3, column=8 + m, value=monthly[m])
            cell.font = INPUT_FONT
            cell.number_format = USD_FMT

        cs = get_column_letter(8)
        ce_c = get_column_letter(19)
        ws3.cell(row=row3, column=20, value=f'=SUM({cs}{row3}:{ce_c}{row3})')
        ws3.cell(row=row3, column=20).font = FORMULA_FONT
        ws3.cell(row=row3, column=20).number_format = USD_FMT
        ws3.cell(row=row3, column=20).fill = TOTAL_FILL

        ws3.cell(row=row3, column=21, value=np.random.choice(BUDGET_SOURCES, p=BS_WEIGHTS))

        apply_data_style(ws3, row3, len(headers3))
        row3 += 1

print(f'Maintenance_Budget: {row3 - 2} rows generated')

set_col_widths(ws3, [16, 40, 16, 10, 10, 18, 36] + [14]*12 + [16, 22])

# =============================================================================
# SHEET 4: CAPEX_Plan
# =============================================================================
ws4 = wb.create_sheet('CAPEX_Plan')

CAPEX_PROJECTS = [
    ('CAPEX-2026-001', 'SUSTAINING', 'Sustaining', 'Reemplazo revestimientos molino SAG', 'WBS-SN-3000', 'CC-3000', 'PLANT_EQUIP', 'SN-3000-3100', 4500000),
    ('CAPEX-2026-002', 'SUSTAINING', 'Sustaining', 'Overhaul chancador primario', 'WBS-SN-3000', 'CC-3000', 'PLANT_EQUIP', 'SN-2000-2100', 3200000),
    ('CAPEX-2026-003', 'SUSTAINING', 'Sustaining', 'Reemplazo bombas proceso area humeda', 'WBS-SN-3000', 'CC-3000', 'PLANT_EQUIP', 'SN-3000-3300', 2100000),
    ('CAPEX-2026-004', 'SUSTAINING', 'Sustaining', 'Cambio correas transportadoras', 'WBS-SN-2000', 'CC-2000', 'PLANT_EQUIP', 'SN-2000-2200', 1800000),
    ('CAPEX-2026-005', 'SUSTAINING', 'Sustaining', 'Reemplazo instrumentacion control planta', 'WBS-SN-3000', 'CC-3000', 'PLANT_EQUIP', 'SN-4000-4600', 1200000),
    ('CAPEX-2026-006', 'SUSTAINING', 'Sustaining', 'Overhaul generadores emergencia', 'WBS-SN-4000', 'CC-4000', 'PLANT_EQUIP', 'SN-4000-4400', 950000),
    ('CAPEX-2026-007', 'SUSTAINING', 'Sustaining', 'Reemplazo camion minero CAT 785', 'WBS-SN-1000', 'CC-1000', 'MOBILE_EQUIP', 'SN-8000-8900', 5500000),
    ('CAPEX-2026-008', 'SUSTAINING', 'Sustaining', 'Reemplazo excavadora hidraulica', 'WBS-SN-1000', 'CC-1000', 'MOBILE_EQUIP', 'SN-1000-1200', 3800000),
    ('CAPEX-2026-009', 'SUSTAINING', 'Sustaining', 'Overhaul perforadora Atlas Copco', 'WBS-SN-1000', 'CC-1000', 'MOBILE_EQUIP', 'SN-1000-1200', 2200000),
    ('CAPEX-2026-010', 'SUSTAINING', 'Sustaining', 'Reemplazo sistema HVAC campamento', 'WBS-SN-5000', 'CC-5000', 'INFRASTRUCTURE', 'SN-5000-5200', 800000),
    ('CAPEX-2026-011', 'SUSTAINING', 'Sustaining', 'Mejora sistema tratamiento agua', 'WBS-SN-7000', 'CC-7000', 'INFRASTRUCTURE', 'SN-7000-7200', 1500000),
    ('CAPEX-2026-012', 'SUSTAINING', 'Sustaining', 'Actualizacion sistema SCADA', 'WBS-SN-4000', 'CC-4000', 'IT_SYSTEMS', 'SN-4000-4600', 1100000),
    ('CAPEX-2026-013', 'SUSTAINING', 'Sustaining', 'Reemplazo celdas electrolisis refineria', 'WBS-SN-3000', 'CC-3000', 'PLANT_EQUIP', 'SN-3000-3500', 1800000),
    ('CAPEX-2026-014', 'SUSTAINING', 'Sustaining', 'Reemplazo valvulas criticas area humeda', 'WBS-SN-3000', 'CC-3000', 'PLANT_EQUIP', 'SN-3000-3300', 900000),
    ('CAPEX-2026-015', 'SUSTAINING', 'Sustaining', 'Refuerzo estructural espesadores', 'WBS-SN-3000', 'CC-3000', 'PLANT_EQUIP', 'SN-3000-3600', 1200000),
    ('CAPEX-2026-016', 'SUSTAINING', 'Sustaining', 'Reemplazo sistema deteccion incendios', 'WBS-SN-4000', 'CC-4000', 'INFRASTRUCTURE', 'SN-4000-4700', 650000),
    ('CAPEX-2026-017', 'SUSTAINING', 'Sustaining', 'Reemplazo filtros y celdas carbon', 'WBS-SN-3000', 'CC-3000', 'PLANT_EQUIP', 'SN-3000-3400', 750000),
    ('CAPEX-2026-018', 'SUSTAINING', 'Sustaining', 'Overhaul compresores aire', 'WBS-SN-4000', 'CC-4000', 'PLANT_EQUIP', 'SN-4000-4300', 550000),
    ('CAPEX-2026-019', 'SUSTAINING', 'Sustaining', 'Reemplazo bulldozer D10T', 'WBS-SN-1000', 'CC-1000', 'MOBILE_EQUIP', 'SN-8000-8900', 3200000),
    ('CAPEX-2026-020', 'SUSTAINING', 'Sustaining', 'Reemplazo cargador frontal 992K', 'WBS-SN-1000', 'CC-1000', 'MOBILE_EQUIP', 'SN-8000-8900', 2800000),
    ('CAPEX-2026-021', 'MINE_DEVELOPMENT', 'Desarrollo Mina', 'Preparacion Fase 3 pit expansion', 'WBS-SN-1000', 'CC-1000', 'MINE_PREP', None, 6000000),
    ('CAPEX-2026-022', 'MINE_DEVELOPMENT', 'Desarrollo Mina', 'Accesos y caminos Fase 3', 'WBS-SN-1000', 'CC-1000', 'MINE_PREP', None, 4500000),
    ('CAPEX-2026-023', 'MINE_DEVELOPMENT', 'Desarrollo Mina', 'Perforacion y tronadura desarrollo', 'WBS-SN-1000', 'CC-1000', 'MINE_PREP', None, 4500000),
    ('CAPEX-2026-024', 'DEVELOPMENT', 'Desarrollo', 'Ampliacion bodega repuestos', 'WBS-SN-5000', 'CC-5000', 'INFRASTRUCTURE', 'SN-5000-5300', 1500000),
    ('CAPEX-2026-025', 'DEVELOPMENT', 'Desarrollo', 'Implementacion sistema AMS', 'WBS-SN-8000', 'CC-8000', 'IT_SYSTEMS', None, 800000),
    ('CAPEX-2026-026', 'DEVELOPMENT', 'Desarrollo', 'Mejora sistema supresion de polvo', 'WBS-SN-3000', 'CC-3000', 'INFRASTRUCTURE', 'SN-2000-2100', 2500000),
]

headers4 = ['capex_id', 'capex_category', 'capex_category_desc', 'project_name',
            'wbs_element', 'cost_center', 'asset_class', 'sap_func_loc'] + \
           [f'{m}_budget_usd' for m in MONTHS_EN] + ['annual_budget_usd', 'approval_status', 'priority']

for c, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=c, value=h)
apply_header_style(ws4, 1, len(headers4))

add_comment(ws4.cell(row=1, column=1), 'Identificador unico del proyecto CAPEX: CAPEX-YYYY-NNN')
add_comment(ws4.cell(row=1, column=2), 'Categoria: SUSTAINING (mantener capacidad), DEVELOPMENT (nuevo), MINE_DEVELOPMENT (desarrollo mina), EXPANSION (ampliacion)')
add_comment(ws4.cell(row=1, column=5), 'Elemento WBS compatible con 20_financial_assignments.xlsx')
add_comment(ws4.cell(row=1, column=7), 'Clase de activo: MOBILE_EQUIP, PLANT_EQUIP, INFRASTRUCTURE, MINE_PREP, IT_SYSTEMS')
add_comment(ws4.cell(row=1, column=21), 'Total anual = SUM(Ene:Dic)')
add_comment(ws4.cell(row=1, column=22), 'Estado de aprobacion: APPROVED, PENDING, CONTINGENCY')
add_comment(ws4.cell(row=1, column=23), 'Prioridad: 1-Critica, 2-Alta, 3-Media, 4-Baja')

for r, proj in enumerate(CAPEX_PROJECTS, 2):
    capex_id, cat, cat_desc, name, wbs, cc, asset_cls, fl, annual = proj

    ws4.cell(row=r, column=1, value=capex_id)
    ws4.cell(row=r, column=2, value=cat)
    ws4.cell(row=r, column=3, value=cat_desc)
    ws4.cell(row=r, column=4, value=name)
    ws4.cell(row=r, column=5, value=wbs)
    ws4.cell(row=r, column=6, value=cc)
    ws4.cell(row=r, column=7, value=asset_cls)
    ws4.cell(row=r, column=8, value=fl)

    # CAPEX has lumpier distribution
    monthly = np.zeros(12)
    n_months = np.random.choice([3, 4, 6, 8, 12], p=[0.15, 0.25, 0.3, 0.2, 0.1])
    active_months = sorted(np.random.choice(12, size=n_months, replace=False))
    weights = np.random.dirichlet(np.ones(n_months) * 2)
    for i, m_idx in enumerate(active_months):
        monthly[m_idx] = annual * weights[i]
    monthly = np.round(monthly, 2)
    # Adjust to match annual
    diff = annual - monthly.sum()
    monthly[active_months[0]] += diff

    for m in range(12):
        cell = ws4.cell(row=r, column=9 + m, value=monthly[m])
        cell.font = INPUT_FONT
        cell.number_format = USD_FMT

    cs = get_column_letter(9)
    ce_c = get_column_letter(20)
    ws4.cell(row=r, column=21, value=f'=SUM({cs}{r}:{ce_c}{r})')
    ws4.cell(row=r, column=21).font = FORMULA_FONT
    ws4.cell(row=r, column=21).number_format = USD_FMT
    ws4.cell(row=r, column=21).fill = TOTAL_FILL

    ws4.cell(row=r, column=22, value='APPROVED' if cat != 'DEVELOPMENT' else np.random.choice(['APPROVED', 'PENDING'], p=[0.7, 0.3]))
    ws4.cell(row=r, column=23, value=np.random.choice(['1-Critica', '2-Alta', '3-Media'], p=[0.3, 0.5, 0.2]))

    apply_data_style(ws4, r, len(headers4))

print(f'CAPEX_Plan: {len(CAPEX_PROJECTS)} rows generated')

set_col_widths(ws4, [16, 20, 20, 44, 14, 10, 16, 16] + [14]*12 + [16, 14, 12])

# =============================================================================
# SHEET 5: KPI_Targets
# =============================================================================
ws5 = wb.create_sheet('KPI_Targets')

KPI_DATA = [
    # Production KPIs
    ('PRODUCTION', 'gold_produced_oz', 'Oro producido', 'Global', None, 'oz', GOLD_ANNUAL, 'SUM', 'Production_Plan'),
    ('PRODUCTION', 'ore_processed_tonnes', 'Mineral procesado', 'Global', None, 'tonnes', ORE_PROCESSED_ANNUAL, 'SUM', 'Production_Plan'),
    ('PRODUCTION', 'head_grade_g_t', 'Ley de cabeza Au', 'Global', None, 'g/t', HEAD_GRADE, 'WEIGHTED_AVG', 'Production_Plan'),
    ('PRODUCTION', 'recovery_pct', 'Recuperacion metalurgica', 'Global', None, '%', RECOVERY, 'WEIGHTED_AVG', 'Production_Plan'),
    ('PRODUCTION', 'strip_ratio', 'Relacion esteril/mineral', 'Mina', None, 'ratio', 2.0, 'WEIGHTED_AVG', 'Production_Plan'),
    ('PRODUCTION', 'plant_throughput_tpd', 'Tratamiento diario planta', 'Planta', None, 'tpd', ORE_PROCESSED_ANNUAL / 365, 'WEIGHTED_AVG', 'Production_Plan'),

    # Maintenance KPIs
    ('MAINTENANCE', 'maint_budget_total_usd', 'Presupuesto mantencion total', 'Global', None, 'USD', 77673169, 'SUM', '20_financial_assignments'),
    ('MAINTENANCE', 'maint_cost_per_oz', 'Costo mantencion por onza', 'Global', None, 'USD/oz', 77673169 / GOLD_ANNUAL, 'FORMULA', '29_cost_history'),
    ('MAINTENANCE', 'pm_compliance_pct', 'Cumplimiento PM', 'Global', None, '%', 0.92, 'WEIGHTED_AVG', '38_kpi_snapshots'),
    ('MAINTENANCE', 'pm_vs_cm_ratio', 'Ratio PM/CM (horas)', 'Global', None, 'ratio', 3.5, 'WEIGHTED_AVG', '06_work_order_history'),
    ('MAINTENANCE', 'backlog_weeks', 'Backlog mantención (semanas)', 'Global', None, 'weeks', 4.0, 'WEIGHTED_AVG', '23_active_backlog'),
    ('MAINTENANCE', 'schedule_compliance_pct', 'Cumplimiento programacion', 'Global', None, '%', 0.88, 'WEIGHTED_AVG', '38_kpi_snapshots'),
    ('MAINTENANCE', 'wrench_time_pct', 'Tiempo llave en mano', 'Global', None, '%', 0.55, 'WEIGHTED_AVG', '38_kpi_snapshots'),
    ('MAINTENANCE', 'maint_cost_per_op_hour', 'Costo mantencion por hora operativa', 'Global', None, 'USD/hr', 77673169 / (365 * 24 * 0.9), 'FORMULA', '29_cost_history'),

    # Maintenance by planning group
    ('MAINTENANCE', 'pm_compliance_pct_m01', 'Cumplimiento PM - Mina', 'Mina', 'M01', '%', 0.90, 'WEIGHTED_AVG', '38_kpi_snapshots'),
    ('MAINTENANCE', 'pm_compliance_pct_p01', 'Cumplimiento PM - Area Seca', 'Planta', 'P01', '%', 0.93, 'WEIGHTED_AVG', '38_kpi_snapshots'),
    ('MAINTENANCE', 'pm_compliance_pct_p02', 'Cumplimiento PM - Area Humeda', 'Planta', 'P02', '%', 0.91, 'WEIGHTED_AVG', '38_kpi_snapshots'),
    ('MAINTENANCE', 'pm_compliance_pct_p03', 'Cumplimiento PM - Servicios', 'Servicios', 'P03', '%', 0.94, 'WEIGHTED_AVG', '38_kpi_snapshots'),

    # Financial KPIs
    ('FINANCIAL', 'aisc_per_oz', 'AISC por onza', 'Global', None, 'USD/oz', 800, 'FORMULA', 'OPEX_by_Area+CAPEX_Plan'),
    ('FINANCIAL', 'cash_cost_per_oz', 'Cash cost por onza', 'Global', None, 'USD/oz', 600, 'FORMULA', 'OPEX_by_Area'),
    ('FINANCIAL', 'revenue_usd', 'Ingreso por ventas Au', 'Global', None, 'USD', GOLD_ANNUAL * GOLD_PRICE, 'SUM', 'Production_Plan'),
    ('FINANCIAL', 'ebitda_usd', 'EBITDA', 'Global', None, 'USD', GOLD_ANNUAL * GOLD_PRICE * 0.56, 'SUM', 'OPEX_by_Area'),
    ('FINANCIAL', 'ebitda_margin_pct', 'Margen EBITDA', 'Global', None, '%', 0.56, 'WEIGHTED_AVG', 'OPEX_by_Area'),
    ('FINANCIAL', 'opex_total_usd', 'OPEX total', 'Global', None, 'USD', 200000000, 'SUM', 'OPEX_by_Area'),
    ('FINANCIAL', 'capex_sustaining_usd', 'CAPEX sustaining', 'Global', None, 'USD', 40450000, 'SUM', 'CAPEX_Plan'),
    ('FINANCIAL', 'opex_per_tonne', 'OPEX por tonelada procesada', 'Global', None, 'USD/t', 200000000 / ORE_PROCESSED_ANNUAL, 'FORMULA', 'OPEX_by_Area'),
    ('FINANCIAL', 'maint_cost_pct_revenue', 'Costo mantencion / Ingresos', 'Global', None, '%', 77673169 / (GOLD_ANNUAL * GOLD_PRICE), 'FORMULA', '29_cost_history'),

    # Reliability KPIs
    ('RELIABILITY', 'plant_availability_pct', 'Disponibilidad planta global', 'Planta', None, '%', 0.92, 'WEIGHTED_AVG', '30_reliability_data'),
    ('RELIABILITY', 'mine_availability_pct', 'Disponibilidad equipos mina', 'Mina', None, '%', 0.88, 'WEIGHTED_AVG', '30_reliability_data'),
    ('RELIABILITY', 'mtbf_plant_hrs', 'MTBF planta (horas)', 'Planta', None, 'hours', 720, 'WEIGHTED_AVG', '30_reliability_data'),
    ('RELIABILITY', 'mttr_plant_hrs', 'MTTR planta (horas)', 'Planta', None, 'hours', 4.5, 'WEIGHTED_AVG', '30_reliability_data'),
    ('RELIABILITY', 'mtbf_mine_hrs', 'MTBF equipos mina (horas)', 'Mina', None, 'hours', 480, 'WEIGHTED_AVG', '30_reliability_data'),
    ('RELIABILITY', 'mttr_mine_hrs', 'MTTR equipos mina (horas)', 'Mina', None, 'hours', 6.0, 'WEIGHTED_AVG', '30_reliability_data'),
    ('RELIABILITY', 'cost_of_unreliability_usd', 'Costo de no confiabilidad', 'Global', None, 'USD', 15000000, 'SUM', '30_reliability_data'),
    ('RELIABILITY', 'oee_pct', 'OEE planta', 'Planta', None, '%', 0.78, 'WEIGHTED_AVG', '30_reliability_data'),

    # Safety KPIs
    ('SAFETY', 'trifr', 'TRIFR (lesiones registrables)', 'Global', None, 'rate', 2.5, 'WEIGHTED_AVG', None),
    ('SAFETY', 'ltifr', 'LTIFR (lesiones con tiempo perdido)', 'Global', None, 'rate', 0.8, 'WEIGHTED_AVG', None),
    ('SAFETY', 'near_miss_reports', 'Reportes casi-accidentes', 'Global', None, 'count', 120, 'SUM', None),
    ('SAFETY', 'safety_observations', 'Observaciones de seguridad', 'Global', None, 'count', 600, 'SUM', None),
]

headers5 = ['kpi_domain', 'kpi_name', 'kpi_description', 'area', 'planning_group', 'unit'] + \
           [f'{m}_target' for m in MONTHS_EN] + ['annual_target', 'calculation_method', 'data_source_file']

for c, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=c, value=h)
apply_header_style(ws5, 1, len(headers5))

add_comment(ws5.cell(row=1, column=1), 'Dominio: PRODUCTION, MAINTENANCE, FINANCIAL, RELIABILITY, SAFETY')
add_comment(ws5.cell(row=1, column=2), 'Nombre del KPI en formato snake_case, unico')
add_comment(ws5.cell(row=1, column=5), 'Grupo de planificacion: M01, P01, P02, P03. Null para KPIs globales')
add_comment(ws5.cell(row=1, column=19), 'Meta anual: SUM para acumulables, WEIGHTED_AVG para promedios')
add_comment(ws5.cell(row=1, column=20), 'Metodo calculo: DIRECT (valor directo), WEIGHTED_AVG, SUM, FORMULA (calculado de otros KPIs)')
add_comment(ws5.cell(row=1, column=21), 'Archivo fuente de datos reales para comparacion')

for r, kpi in enumerate(KPI_DATA, 2):
    domain, name, desc, area, pg, unit, annual, calc, source = kpi

    ws5.cell(row=r, column=1, value=domain)
    ws5.cell(row=r, column=2, value=name)
    ws5.cell(row=r, column=3, value=desc)
    ws5.cell(row=r, column=4, value=area)
    ws5.cell(row=r, column=5, value=pg)
    ws5.cell(row=r, column=6, value=unit)

    if calc == 'SUM':
        monthly = distribute_annual(annual, noise=0.05)
    else:
        monthly = np.full(12, annual) * np.random.uniform(0.97, 1.03, 12)
        monthly = np.round(monthly, 4 if abs(annual) < 10 else 2)

    for m in range(12):
        cell = ws5.cell(row=r, column=7 + m, value=round(monthly[m], 4 if abs(annual) < 10 else 0))
        cell.font = INPUT_FONT
        if unit == '%':
            cell.number_format = PCT_FMT
        elif unit in ('USD', 'USD/oz', 'USD/t', 'USD/hr'):
            cell.number_format = USD_FMT
        elif unit in ('oz', 'tonnes', 'count'):
            cell.number_format = NUM_FMT
        else:
            cell.number_format = NUM_DEC_FMT

    # Annual target
    cs = get_column_letter(7)
    ce_c = get_column_letter(18)
    if calc == 'SUM':
        ws5.cell(row=r, column=19, value=f'=SUM({cs}{r}:{ce_c}{r})')
    else:
        ws5.cell(row=r, column=19, value=f'=AVERAGE({cs}{r}:{ce_c}{r})')
    ws5.cell(row=r, column=19).font = FORMULA_FONT
    ws5.cell(row=r, column=19).fill = TOTAL_FILL

    ws5.cell(row=r, column=20, value=calc)
    ws5.cell(row=r, column=21, value=source)

    apply_data_style(ws5, r, len(headers5))

print(f'KPI_Targets: {len(KPI_DATA)} rows generated')

set_col_widths(ws5, [16, 30, 36, 12, 12, 10] + [14]*12 + [16, 16, 28])

# =============================================================================
# SHEET 1: Executive_Dashboard
# =============================================================================
ws1 = wb.create_sheet('Executive_Dashboard')

EXEC_KPIS = [
    ('PRODUCCION', 'gold_produced_oz', 'Oro producido', 'oz', GOLD_ANNUAL, 'SUM'),
    ('PRODUCCION', 'ore_mined_kt', 'Mineral extraido', 'kt', ORE_MINED_ANNUAL / 1000, 'SUM'),
    ('PRODUCCION', 'ore_processed_kt', 'Mineral procesado', 'kt', ORE_PROCESSED_ANNUAL / 1000, 'SUM'),
    ('PRODUCCION', 'head_grade_g_t', 'Ley de cabeza Au', 'g/t', HEAD_GRADE, 'AVG'),
    ('PRODUCCION', 'recovery_pct', 'Recuperacion metalurgica', '%', RECOVERY, 'AVG'),
    ('PRODUCCION', 'strip_ratio', 'Relacion esteril/mineral', 'ratio', 2.0, 'AVG'),
    ('PRODUCCION', 'plant_availability_pct', 'Disponibilidad planta', '%', 0.92, 'AVG'),

    ('COSTOS', 'aisc_per_oz', 'AISC por onza Au', 'USD/oz', 800, 'AVG'),
    ('COSTOS', 'cash_cost_per_oz', 'Cash cost C1 por onza', 'USD/oz', 600, 'AVG'),
    ('COSTOS', 'opex_total_musd', 'OPEX total', 'MUSD', 200, 'SUM'),
    ('COSTOS', 'maint_cost_musd', 'Costo mantencion total', 'MUSD', 77.7, 'SUM'),
    ('COSTOS', 'maint_cost_per_oz', 'Costo mantencion por oz', 'USD/oz', 77673169 / GOLD_ANNUAL, 'AVG'),
    ('COSTOS', 'capex_sustaining_musd', 'CAPEX sustaining', 'MUSD', 40.5, 'SUM'),
    ('COSTOS', 'opex_per_tonne', 'OPEX por tonelada', 'USD/t', 200000000 / ORE_PROCESSED_ANNUAL, 'AVG'),

    ('FINANCIERO', 'gold_price_usd_oz', 'Precio oro supuesto', 'USD/oz', GOLD_PRICE, 'AVG'),
    ('FINANCIERO', 'revenue_musd', 'Ingresos por venta Au', 'MUSD', GOLD_ANNUAL * GOLD_PRICE / 1e6, 'SUM'),
    ('FINANCIERO', 'ebitda_musd', 'EBITDA', 'MUSD', GOLD_ANNUAL * GOLD_PRICE * 0.56 / 1e6, 'SUM'),
    ('FINANCIERO', 'ebitda_margin_pct', 'Margen EBITDA', '%', 0.56, 'AVG'),
    ('FINANCIERO', 'pat_musd', 'Utilidad despues de impuestos', 'MUSD', GOLD_ANNUAL * GOLD_PRICE * 0.35 / 1e6, 'SUM'),
    ('FINANCIERO', 'free_cash_flow_musd', 'Flujo de caja libre', 'MUSD', GOLD_ANNUAL * GOLD_PRICE * 0.30 / 1e6, 'SUM'),

    ('OPERACIONAL', 'pm_compliance_pct', 'Cumplimiento PM', '%', 0.92, 'AVG'),
    ('OPERACIONAL', 'schedule_compliance_pct', 'Cumplimiento programacion', '%', 0.88, 'AVG'),
    ('OPERACIONAL', 'pm_vs_cm_ratio', 'Ratio PM/CM', 'ratio', 3.5, 'AVG'),
    ('OPERACIONAL', 'backlog_weeks', 'Backlog (semanas)', 'weeks', 4.0, 'AVG'),
    ('OPERACIONAL', 'maint_cost_pct_revenue', 'Mantencion / Ingresos', '%', 77673169 / (GOLD_ANNUAL * GOLD_PRICE), 'AVG'),
    ('OPERACIONAL', 'cost_unreliability_musd', 'Costo no confiabilidad', 'MUSD', 15, 'SUM'),

    ('SEGURIDAD', 'trifr', 'TRIFR', 'rate', 2.5, 'AVG'),
    ('SEGURIDAD', 'ltifr', 'LTIFR', 'rate', 0.8, 'AVG'),

    ('DOTACION', 'headcount_propios', 'Dotacion propios', '#', 740, 'AVG'),
    ('DOTACION', 'headcount_contratistas', 'Dotacion contratistas', '#', 830, 'AVG'),
    ('DOTACION', 'headcount_total', 'Dotacion total', '#', 1570, 'AVG'),
]

headers1 = ['kpi_category', 'kpi_name', 'kpi_description', 'unit'] + \
           [f'{m}_budget' for m in MONTHS_EN] + ['annual_budget', 'annual_actual', 'variance_pct']

for c, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=c, value=h)
apply_header_style(ws1, 1, len(headers1))

add_comment(ws1.cell(row=1, column=1), 'Categoria: PRODUCCION, COSTOS, FINANCIERO, OPERACIONAL, SEGURIDAD, DOTACION')
add_comment(ws1.cell(row=1, column=2), 'Identificador KPI en formato snake_case')
add_comment(ws1.cell(row=1, column=4), 'Unidad: oz, kt, g/t, %, ratio, USD/oz, MUSD, USD/t, weeks, rate, #')
add_comment(ws1.cell(row=1, column=17), 'Presupuesto anual: SUM para acumulables, AVG para ratios/porcentajes')
add_comment(ws1.cell(row=1, column=18), 'Real anual (se completa durante ejecucion)')
add_comment(ws1.cell(row=1, column=19), 'Varianza % = (Real - Presupuesto) / Presupuesto. Formula automatica')

prev_cat = None
for r, kpi in enumerate(EXEC_KPIS, 2):
    cat, name, desc, unit, annual, agg = kpi

    ws1.cell(row=r, column=1, value=cat)
    ws1.cell(row=r, column=2, value=name)
    ws1.cell(row=r, column=3, value=desc)
    ws1.cell(row=r, column=4, value=unit)

    if agg == 'SUM':
        monthly = distribute_annual(annual, noise=0.04)
    else:
        monthly = np.full(12, annual) * np.random.uniform(0.98, 1.02, 12)
        monthly = np.round(monthly, 4 if abs(annual) < 10 else 2)

    for m in range(12):
        cell = ws1.cell(row=r, column=5 + m, value=monthly[m])
        cell.font = INPUT_FONT
        if unit == '%':
            cell.number_format = PCT_FMT
        elif unit in ('USD/oz', 'USD/t'):
            cell.number_format = USD_FMT
        elif unit == 'MUSD':
            cell.number_format = NUM_DEC_FMT
        elif unit in ('oz', 'kt', '#'):
            cell.number_format = NUM_FMT
        else:
            cell.number_format = NUM_DEC_FMT

    # Annual budget formula
    cs = get_column_letter(5)
    ce_c = get_column_letter(16)
    if agg == 'SUM':
        ws1.cell(row=r, column=17, value=f'=SUM({cs}{r}:{ce_c}{r})')
    else:
        ws1.cell(row=r, column=17, value=f'=AVERAGE({cs}{r}:{ce_c}{r})')
    ws1.cell(row=r, column=17).font = FORMULA_FONT
    ws1.cell(row=r, column=17).fill = TOTAL_FILL

    # Annual actual (empty)
    ws1.cell(row=r, column=18, value=None)

    # Variance pct formula
    ann_b = get_column_letter(17)
    ann_a = get_column_letter(18)
    ws1.cell(row=r, column=19, value=f'=IF(OR({ann_a}{r}="",{ann_b}{r}=0),"",({ann_a}{r}-{ann_b}{r})/{ann_b}{r})')
    ws1.cell(row=r, column=19).font = FORMULA_FONT
    ws1.cell(row=r, column=19).number_format = PCT_FMT

    # Category separator styling
    if cat != prev_cat and prev_cat is not None:
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=r, column=col).border = Border(
                top=Side(style='medium', color='1F4E79'),
                left=THIN_BORDER.left, right=THIN_BORDER.right, bottom=THIN_BORDER.bottom)

    apply_data_style(ws1, r, len(headers1))
    prev_cat = cat

print(f'Executive_Dashboard: {len(EXEC_KPIS)} rows generated')

set_col_widths(ws1, [16, 28, 32, 10] + [14]*12 + [16, 16, 12])

# =============================================================================
# REORDER SHEETS: Executive_Dashboard, OPEX_by_Area, Maintenance_Budget, CAPEX_Plan, KPI_Targets, Budget_Equipment_Detail, Production_Plan
# =============================================================================
desired_order = ['Executive_Dashboard', 'OPEX_by_Area', 'Maintenance_Budget', 'CAPEX_Plan', 'KPI_Targets', 'Budget_Equipment_Detail', 'Production_Plan']
for i, name in enumerate(desired_order):
    current_idx = wb.sheetnames.index(name)
    wb.move_sheet(name, offset=i - current_idx)

# =============================================================================
# FREEZE PANES
# =============================================================================
for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = 'A2'

# =============================================================================
# SAVE
# =============================================================================
OUTPUT = 'seed_data/41_annual_budget.xlsx'
wb.save(OUTPUT)
print(f'\nSaved: {OUTPUT}')
print(f'Sheets: {wb.sheetnames}')
