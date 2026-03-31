"""Seed remaining data: failure_modes, work_centers, active_backlog."""
import os, sys, json
from datetime import datetime, date

sys.path.insert(0, "/app")
os.chdir("/app")

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from uuid import uuid4

DB_URL = os.getenv("DATABASE_URL", "sqlite:///data/ocp_maintenance.db")
engine = create_engine(DB_URL)
Session = sessionmaker(bind=engine)
db = Session()

def read_excel(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {h: v for h, v in zip(headers, row) if h}
        if any(v is not None for v in d.values()):
            rows.append(d)
    wb.close()
    return rows

def s(v, d=""): return str(v).strip() if v is not None else d
def f(v, d=0.0):
    try: return float(v)
    except: return d
def dt(v):
    if v is None: return None
    if isinstance(v, (datetime, date)):
        return v if isinstance(v, datetime) else datetime.combine(v, datetime.min.time())
    try: return datetime.fromisoformat(str(v)[:10])
    except: return None


# ═══════════════════════════════════════════
# 1. FAILURE MODES (03_failure_modes.xlsx)
# ═══════════════════════════════════════════
print("=== 1. Failure Modes ===")

# Check if failure_modes table exists or use the ORM table
tables = [r[0] for r in db.execute(text("SELECT name FROM sqlite_master WHERE type='table'")).fetchall()]
print(f"  Tables available: {len(tables)}")

target = None  # Always use failure_catalog (failure_modes has different schema)

rows = read_excel("/app/seed_data/03_failure_modes.xlsx")
print(f"  Excel rows: {len(rows)}")

if not target:
    # Create a dedicated failure_catalog table for SAP PM failure codes
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS failure_catalog (
            catalog_id VARCHAR(50) PRIMARY KEY,
            equipment_tag VARCHAR(100),
            equipment_name VARCHAR(200),
            sap_func_loc VARCHAR(50),
            area VARCHAR(50),
            mechanism VARCHAR(50),
            cause VARCHAR(50),
            failure_pattern VARCHAR(30),
            failure_consequence VARCHAR(200),
            evidence VARCHAR(200),
            detection_method VARCHAR(50),
            downtime_hours FLOAT DEFAULT 0,
            rpn_severity INTEGER DEFAULT 0,
            rpn_occurrence INTEGER DEFAULT 0,
            rpn_detection INTEGER DEFAULT 0,
            rpn_total INTEGER DEFAULT 0,
            subunit VARCHAR(100),
            maintainable_item VARCHAR(100)
        )
    """))
    db.commit()
    target = "failure_catalog"
    print("  Created failure_catalog table")

count = 0
for r in rows:
    cid = str(uuid4())[:8]
    equip = s(r.get("equipment_tag", ""))
    mechanism = s(r.get("mechanism", ""))
    cause = s(r.get("cause", ""))
    if not mechanism and not cause:
        continue

    existing = db.execute(text(f"SELECT catalog_id FROM {target} WHERE equipment_tag = :eq AND mechanism = :m AND cause = :c"),
        {"eq": equip, "m": mechanism, "c": cause}).fetchone()
    if existing:
        continue

    try:
        db.execute(text(f"""
            INSERT INTO {target} (catalog_id, equipment_tag, equipment_name, sap_func_loc, area,
                mechanism, cause, failure_pattern, failure_consequence, evidence, detection_method,
                downtime_hours, rpn_severity, rpn_occurrence, rpn_detection, rpn_total,
                subunit, maintainable_item)
            VALUES (:cid, :equip, :ename, :fl, :area, :mech, :cause, :pat, :cons, :evid, :det,
                :dt, :rpns, :rpno, :rpnd, :rpnt, :sub, :mi)
        """), {
            "cid": cid, "equip": equip,
            "ename": s(r.get("equipment_function_description", r.get("eqktx", ""))),
            "fl": s(r.get("sap_func_loc", "")),
            "area": s(r.get("area", "")),
            "mech": mechanism,
            "cause": cause,
            "pat": s(r.get("failure_pattern", "")),
            "cons": s(r.get("failure_consequence", "")),
            "evid": s(r.get("evidence", "")),
            "det": s(r.get("detection_method", "")),
            "dt": f(r.get("downtime_hours", 0)),
            "rpns": int(f(r.get("rpn_severity", 0))),
            "rpno": int(f(r.get("rpn_occurrence", 0))),
            "rpnd": int(f(r.get("rpn_detection", 0))),
            "rpnt": int(f(r.get("rpn_total", 0))),
            "sub": s(r.get("subunit", "")),
            "mi": s(r.get("maintainable_item", "")),
        })
        count += 1
    except Exception as e:
        print(f"  Error: {e}")
        db.rollback()
        break

db.commit()
print(f"  Inserted {count} failure modes")


# ═══════════════════════════════════════════
# 2. WORK CENTERS (11_work_centers.xlsx)
# ═══════════════════════════════════════════
print("\n=== 2. Work Centers ===")

if "work_centers" not in tables:
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS work_centers (
            wc_code VARCHAR(20) PRIMARY KEY,
            wc_name VARCHAR(100),
            wc_type VARCHAR(20),
            area_code VARCHAR(10),
            area_name VARCHAR(50),
            planning_group VARCHAR(10),
            cost_center VARCHAR(20),
            capacity_hours_day FLOAT DEFAULT 8,
            shift_code VARCHAR(10),
            shift_start VARCHAR(10),
            shift_end VARCHAR(10)
        )
    """))
    db.commit()
    print("  Created work_centers table")

rows = read_excel("/app/seed_data/11_work_centers.xlsx")
count = 0
for r in rows:
    code = s(r.get("wc_code", r.get("work_center", "")))
    if not code:
        continue
    existing = db.execute(text("SELECT wc_code FROM work_centers WHERE wc_code = :c"), {"c": code}).fetchone()
    if existing:
        continue
    try:
        db.execute(text("""
            INSERT INTO work_centers (wc_code, wc_name, wc_type, area_code, area_name,
                planning_group, cost_center, capacity_hours_day, shift_code, shift_start, shift_end)
            VALUES (:code, :name, :type, :ac, :an, :pg, :cc, :cap, :sc, :ss, :se)
        """), {
            "code": code,
            "name": s(r.get("wc_name", r.get("description", ""))),
            "type": s(r.get("wc_type", "")),
            "ac": s(r.get("area_code", "")),
            "an": s(r.get("area_name", "")),
            "pg": s(r.get("planning_group", "")),
            "cc": s(r.get("cost_center", "")),
            "cap": f(r.get("capacity_hours_day", 8)),
            "sc": s(r.get("shift_code", "")),
            "ss": s(r.get("shift_start", "")),
            "se": s(r.get("shift_end", "")),
        })
        count += 1
    except Exception as e:
        print(f"  Error: {e}")
        db.rollback()
        break

db.commit()
print(f"  Inserted {count} work centers")


# ═══════════════════════════════════════════
# 3. ACTIVE BACKLOG (23_active_backlog.xlsx)
# ═══════════════════════════════════════════
print("\n=== 3. Active Backlog ===")

rows = read_excel("/app/seed_data/23_active_backlog.xlsx")
print(f"  Excel rows: {len(rows)}")

# Backlog items are WOs that are open — insert as managed_work_orders with CREADO/PLANIFICADO status
count = 0
for r in rows:
    aufnr = s(r.get("aufnr", r.get("wo_number", "")))
    if not aufnr:
        continue
    wo_number = f"BL-{aufnr}"
    existing = db.execute(text("SELECT wo_id FROM managed_work_orders WHERE wo_number = :wn"), {"wn": wo_number}).fetchone()
    if existing:
        continue

    PRIO_MAP = {"I": "P1", "A": "P2", "M": "P3", "B": "P4", "P1": "P1", "P2": "P2", "P3": "P3", "P4": "P4"}
    STATUS_MAP = {"ABIE": "EN_EJECUCION", "CTEC": "CERRADO", "NOTI": "CREADO", "LIBE": "PROGRAMADO", "PLAN": "PLANIFICADO"}

    priority = PRIO_MAP.get(s(r.get("priokx", r.get("priority", "M"))), "P3")
    status_raw = s(r.get("system_status", r.get("status", "NOTI")))
    status = STATUS_MAP.get(status_raw, "CREADO")
    wo_type = s(r.get("auart", r.get("wo_type", "PM01")))

    try:
        db.execute(text("""
            INSERT INTO managed_work_orders (wo_id, wo_number, plant_id, equipment_id, equipment_tag,
                description, wo_type, priority_code, work_class, status,
                estimated_hours, actual_hours, planned_start, planned_end,
                work_center, planning_group, is_fast_track, created_at,
                completion_pct, budget_approved)
            VALUES (:wid, :wn, 'OCP-JFC1', :equip, :equip, :desc, :wtype, :prio, :wclass, :status,
                :est_h, 0, :p_start, :p_end, :wc, :pg, :ft, :created, 0, 1)
        """), {
            "wid": str(uuid4()),
            "wn": wo_number,
            "equip": s(r.get("equipment_tag", "")),
            "desc": s(r.get("description", r.get("short_text", ""))),
            "wtype": wo_type,
            "prio": priority,
            "wclass": "NO_PROGRAMADO" if priority in ("P1", "P2") else "PROGRAMADO",
            "status": status,
            "est_h": f(r.get("planned_hours", r.get("duration_planned", 4))),
            "p_start": dt(r.get("gstrp", r.get("basic_start", None))),
            "p_end": dt(r.get("gltrp", r.get("basic_end", None))),
            "wc": s(r.get("arbpl", r.get("work_center", ""))),
            "pg": s(r.get("planning_group", "")),
            "ft": priority in ("P1", "P2"),
            "created": dt(r.get("erdat", r.get("created_date", None))) or datetime.now(),
        })
        count += 1
    except Exception as e:
        print(f"  Error on {wo_number}: {e}")
        db.rollback()
        continue

db.commit()
print(f"  Inserted {count} backlog items as WOs")


# ═══════════════════════════════════════════
# SUMMARY
# ═══════════════════════════════════════════
print("\n=== FINAL COUNTS ===")
for tbl in ["sap_materials", "workforce", "hierarchy_nodes", "managed_work_orders", "work_requests", "failure_catalog", "work_centers"]:
    try:
        cnt = db.execute(text(f"SELECT COUNT(*) FROM {tbl}")).scalar()
        print(f"  {tbl}: {cnt}")
    except:
        print(f"  {tbl}: does not exist")

db.close()
print("\n=== SEED REMAINING COMPLETE ===")
