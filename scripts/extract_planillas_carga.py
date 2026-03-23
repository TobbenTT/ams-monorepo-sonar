"""
Extract data from PLANILLAS DE CARGA XLSX files into JSON for the AMS.
Source: Gold Fields Salares Norte (MGSN) SAP PM data loading templates.

Generates:
  api/data/failure_catalog_sap.json       — Full catalog by equipment profile (167 profiles)
  api/data/failure_catalog_by_discipline.json — Grouped by discipline for mobile UI
  api/data/equipment_types_sap.json       — 207 SAP equipment type codes
  api/data/taxonomy_sap.json              — 4,880 hierarchy nodes
  api/data/equipment_records_sap.json     — 5,073 equipment records
  api/data/criticality_sap.json           — 1,243 criticality assessments
"""

import sys
import os
import glob
import json

sys.stdout.reconfigure(encoding="utf-8")

import openpyxl

BASE = "C:/Users/Tobbe/Downloads/Practica/SecondBrain/Proyectos Por Subir/Memoria/Toda la informacion/AA. PLANILLAS DE CARGA/"
OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "api", "data")
os.makedirs(OUT_DIR, exist_ok=True)


def find_file(pattern):
    matches = glob.glob(BASE + pattern)
    if not matches:
        print(f"WARNING: No file matching {pattern}")
        return None
    return matches[0]


def save_json(filename, data):
    path = os.path.join(OUT_DIR, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  Saved {filename}")
    return path


# ======================================================================
# 1. FR-00002: FAILURE CATALOG (Perfiles Catalogo)
# ======================================================================
def extract_failure_catalog():
    print("\n=== FR-00002: Failure Catalog ===")
    f = find_file("*FR-00002*")
    if not f:
        return
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

    # Profiles
    profiles = {}
    ws1 = wb["Perfil_Catálogo"]
    for row in ws1.iter_rows(min_row=2, values_only=True):
        cells = list(row) if row else []
        if len(cells) >= 4 and cells[0]:
            code = str(cells[0]).strip()
            profiles[code] = str(cells[1] or "").strip()

    # Catalog entries
    ws2 = wb["Catálogos"]
    catalog = {}
    all_causes = set()
    all_symptoms = set()

    for row in ws2.iter_rows(min_row=2, values_only=True):
        cells = list(row) if row else []
        if len(cells) < 5 or not cells[0]:
            continue
        pc = str(cells[0]).strip()
        tipo = str(cells[2] or "").strip()
        value = str(cells[4] or "").strip()
        if not tipo or not value:
            continue

        if pc not in catalog:
            catalog[pc] = {"causes": [], "symptoms": [], "parts": []}

        if "Causa" in tipo:
            if value not in catalog[pc]["causes"]:
                catalog[pc]["causes"].append(value)
            all_causes.add(value)
        elif "nto" in tipo:  # Síntomas
            if value not in catalog[pc]["symptoms"]:
                catalog[pc]["symptoms"].append(value)
            all_symptoms.add(value)
        elif "Parte" in tipo:
            if value not in catalog[pc]["parts"]:
                catalog[pc]["parts"].append(value)

    wb.close()

    # Full catalog by profile
    full_catalog = {
        "profiles": {},
        "global_causes": sorted(all_causes),
        "global_symptoms": sorted(all_symptoms),
    }
    for pc, data in catalog.items():
        full_catalog["profiles"][pc] = {
            "code": pc,
            "name": profiles.get(pc, pc),
            "causes": sorted(data["causes"]),
            "symptoms": sorted(data["symptoms"]),
            "parts": sorted(data["parts"]),
        }

    save_json("failure_catalog_sap.json", full_catalog)
    print(f"  {len(full_catalog['profiles'])} profiles, {len(all_causes)} causes, {len(all_symptoms)} symptoms")

    # Discipline-grouped version for mobile UI
    ELECTRICO_KEYWORDS = [
        "motor el", "tablero", "transform", "variador", "interruptor",
        "cable", "celda", "arrancador", "ups", "bater", "cargador bat",
        "contactor", "rele protec", "desconectador", "potencia",
    ]
    INSTRUM_KEYWORDS = [
        "sensor", "transmisor", "analizador", "indicador", "switch",
        "plc", "dcs", "scada", "instrument", "detector", "medidor",
        "cctv", "server", "firewall",
    ]

    discipline_catalog = {
        "MECANICO": {"label": "Mecánico", "color": "#6366F1", "symptoms": set(), "parts": set(), "causes": set()},
        "ELECTRICO": {"label": "Eléctrico", "color": "#F59E0B", "symptoms": set(), "parts": set(), "causes": set()},
        "INSTRUMENTACION": {"label": "Instrumentación", "color": "#06B6D4", "symptoms": set(), "parts": set(), "causes": set()},
    }

    for pc, data in catalog.items():
        profile_name = profiles.get(pc, "").lower()

        # Classify profile into discipline
        disc = "MECANICO"
        if any(w in profile_name for w in ELECTRICO_KEYWORDS):
            disc = "ELECTRICO"
        elif any(w in profile_name for w in INSTRUM_KEYWORDS):
            disc = "INSTRUMENTACION"

        for c in data["causes"]:
            discipline_catalog[disc]["causes"].add(c)
        for s in data["symptoms"]:
            discipline_catalog[disc]["symptoms"].add(s)
        for p in data["parts"]:
            discipline_catalog[disc]["parts"].add(p)

    # All causes and symptoms are shared across disciplines (they're generic)
    for d in discipline_catalog:
        discipline_catalog[d]["causes"] = sorted(discipline_catalog[d]["causes"])
        discipline_catalog[d]["symptoms"] = sorted(discipline_catalog[d]["symptoms"])
        discipline_catalog[d]["parts"] = sorted(discipline_catalog[d]["parts"])

    save_json("failure_catalog_by_discipline.json", discipline_catalog)
    for d, v in discipline_catalog.items():
        print(f"  {d}: {len(v['symptoms'])} symptoms, {len(v['parts'])} parts, {len(v['causes'])} causes")


# ======================================================================
# 2. FR-00007: EQUIPMENT TYPES
# ======================================================================
def extract_equipment_types():
    print("\n=== FR-00007: Equipment Types ===")
    f = find_file("*FR-00007*")
    if not f:
        return
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb["Tipo_equipos"]
    types = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row) if row else []
        if len(cells) >= 2 and cells[0]:
            types[str(cells[0]).strip()] = str(cells[1] or "").strip()
    wb.close()
    save_json("equipment_types_sap.json", types)
    print(f"  {len(types)} equipment types")


# ======================================================================
# 3. FR-00001: TAXONOMY (Hierarchy)
# ======================================================================
def extract_taxonomy():
    print("\n=== FR-00001: Taxonomy ===")
    f = find_file("*FR-00001*")
    if not f:
        return
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

    # Use "Taxonomía" sheet (4880 rows)
    ws = wb["Taxonomía"]
    nodes = []
    seen = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row) if row else []
        if len(cells) < 15:
            continue

        # Extract hierarchy levels
        sap_level1 = str(cells[0] or "").strip()  # Plant code
        plant_code = str(cells[1] or "").strip()
        plant_name = str(cells[2] or "").strip()
        sap_level2 = str(cells[3] or "").strip()  # Area
        area_code = str(cells[4] or "").strip()
        area_name = str(cells[5] or "").strip()
        sap_level3 = str(cells[6] or "").strip()  # Sub-Area
        subarea_code = str(cells[7] or "").strip()
        subarea_name = str(cells[8] or "").strip()
        sap_level4 = str(cells[9] or "").strip()  # System
        system_code = str(cells[10] or "").strip()
        system_name = str(cells[11] or "").strip()
        sap_level5 = str(cells[12] or "").strip()  # Equipment UT
        tag_ut = str(cells[13] or "").strip()

        # Build hierarchy entries (deduplicate)
        if plant_code and plant_code not in seen:
            seen.add(plant_code)
            nodes.append({
                "sap_func_loc": sap_level1,
                "code": plant_code,
                "name": plant_name,
                "node_type": "PLANT",
                "level": 1,
                "parent_code": None,
            })

        if area_code and sap_level2 not in seen:
            seen.add(sap_level2)
            nodes.append({
                "sap_func_loc": sap_level2,
                "code": area_code,
                "name": area_name,
                "node_type": "AREA",
                "level": 2,
                "parent_code": plant_code,
            })

        if subarea_code and sap_level3 not in seen:
            seen.add(sap_level3)
            nodes.append({
                "sap_func_loc": sap_level3,
                "code": subarea_code,
                "name": subarea_name,
                "node_type": "AREA",
                "level": 3,
                "parent_code": area_code,
            })

        if system_code and sap_level4 not in seen:
            seen.add(sap_level4)
            nodes.append({
                "sap_func_loc": sap_level4,
                "code": system_code,
                "name": system_name,
                "node_type": "SYSTEM",
                "level": 4,
                "parent_code": subarea_code,
            })

        if tag_ut and sap_level5 not in seen:
            seen.add(sap_level5)
            nodes.append({
                "sap_func_loc": sap_level5,
                "code": tag_ut,
                "name": tag_ut,
                "tag": tag_ut,
                "node_type": "EQUIPMENT",
                "level": 5,
                "parent_code": system_code,
            })

    wb.close()
    save_json("taxonomy_sap.json", nodes)
    print(f"  {len(nodes)} hierarchy nodes")


# ======================================================================
# 4. FR-00003: EQUIPMENT RECORDS
# ======================================================================
def extract_equipment_records():
    print("\n=== FR-00003: Equipment Records ===")
    f = find_file("*FR-00003*")
    if not f:
        return
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

    # Use Planilla_equipos sheet
    ws = wb["Planilla_equipos"]
    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row) if row else []
        if len(cells) < 14:
            continue
        obj_class = str(cells[4] or "").strip()
        denomination = str(cells[6] or "").strip()
        classification_tag = str(cells[10] or "").strip()
        tech_location = str(cells[12] or "").strip()
        superior_equip = str(cells[13] or "").strip()

        if not classification_tag:
            continue

        records.append({
            "object_class": obj_class,
            "denomination": denomination,
            "tag": classification_tag,
            "sap_func_loc": tech_location,
            "parent_tag": superior_equip,
        })

    wb.close()
    save_json("equipment_records_sap.json", records)
    print(f"  {len(records)} equipment records")


# ======================================================================
# 5. IP-00001: CRITICALITY ASSESSMENTS
# ======================================================================
def extract_criticality():
    print("\n=== IP-00001: Criticality ===")
    f = find_file("*IP-00001*")
    if not f:
        return
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

    ws = wb["Plantilla criticidad de activos"]
    assessments = []
    header_found = False

    for row in ws.iter_rows(min_row=1, values_only=True):
        cells = list(row) if row else []
        if len(cells) < 15:
            continue

        # Find the data header row
        if not header_found:
            if cells[0] and "Sub" in str(cells[0]):
                header_found = True
            continue

        sub_area = str(cells[0] or "").strip()
        system = str(cells[1] or "").strip()
        desc_ut = str(cells[2] or "").strip()
        tag_ut = str(cells[3] or "").strip()
        desc_ut2 = str(cells[4] or "").strip()
        main_function = str(cells[5] or "").strip()
        failure_event = str(cells[6] or "").strip()
        failure_consequence = str(cells[7] or "").strip()

        if not tag_ut:
            continue

        # Extract numeric fields safely
        def safe_num(val):
            try:
                return float(val) if val else None
            except (ValueError, TypeError):
                return None

        installed = safe_num(cells[8])
        required = safe_num(cells[9])
        redundancy = str(cells[10] or "").strip()
        redundancy_type = str(cells[11] or "").strip()
        downtime_days = safe_num(cells[12])

        assessments.append({
            "sub_area": sub_area,
            "system": system,
            "tag": tag_ut,
            "description": desc_ut2 or desc_ut,
            "main_function": main_function,
            "failure_event": failure_event,
            "failure_consequence": failure_consequence,
            "installed_count": installed,
            "required_count": required,
            "redundancy": redundancy,
            "redundancy_type": redundancy_type,
            "downtime_days": downtime_days,
        })

    wb.close()
    save_json("criticality_sap.json", assessments)
    print(f"  {len(assessments)} criticality assessments")


# ======================================================================
# MAIN
# ======================================================================
if __name__ == "__main__":
    print("Extracting PLANILLAS DE CARGA data...")
    print(f"Source: {BASE}")
    print(f"Output: {OUT_DIR}")

    extract_failure_catalog()
    extract_equipment_types()
    extract_taxonomy()
    extract_equipment_records()
    extract_criticality()

    print("\nDone! JSON files saved to api/data/")
